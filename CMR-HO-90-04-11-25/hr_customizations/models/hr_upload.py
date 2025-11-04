import io
import logging

import xlrd
from dateutil.relativedelta import relativedelta

from odoo import models,fields,api,_
import base64
from io import BytesIO
import openpyxl
import pytz
from datetime import datetime, timedelta, time, date
from odoo.exceptions import UserError
from odoo.tools import DEFAULT_SERVER_DATE_FORMAT, DEFAULT_SERVER_DATETIME_FORMAT
import calendar
import traceback
_logger = logging.getLogger(__name__)


class Hrupload(models.Model):
    _name = "hr.upload"


    date = fields.Date('Date')
    employee_code = fields.Char('Employee Code')
    # your_datetime = fields.Datetime(string='Check in', compute='_compute_your_datetime',store=True)
    # your_checkout_datetime = fields.Datetime(string='Check-Out', compute='_compute_your_checkout_datetime',
    #                                          store=True)
    your_datetime = fields.Datetime(string='Check in', compute='_compute_your_datetime')
    your_checkout_datetime = fields.Datetime(string='Check-Out', compute='_compute_your_checkout_datetime',
                                             store=True)
    employee_name = fields.Many2one('hr.employee', string="Employee Name", required=True)
    check_in_attendance = fields.Char(string="Check In")
    check_out_attendance = fields.Char(string="Check Out")
    difference_check_in = fields.Float(string="IN Difference (mins)", compute='_compute_attendance', store=True)
    difference_check_out = fields.Float(string="OUT Difference (mins)", compute='_compute_attendance', store=True)
    total_working_hours = fields.Char(string="Total Working Hours", compute='_compute_attendance', store=True)
    attendance_status = fields.Selection([
        ('grace', 'Grace'),
        ('early', 'Early'),
        ('late', 'Late'),
        ('others', 'Others')
    ], string="Attendance Status", compute='_compute_attendance', store=True)
    late_deduction = fields.Float(string="Late Deduction(Rupees)", compute="_compute_late_deduction", store=True)
    # morning_session = fields.Selection([
    #     ('Present', 'Present'),
    #     ('Absent', 'Absent')
    # ], string="Morning Session", compute='_compute_sessions', store=True)
    # afternoon_session = fields.Selection([
    #     ('Present', 'Present'),
    #     ('Absent', 'Absent')
    # ], string="Afternoon Session", compute='_compute_sessions', store=True)
    overtime_amount = fields.Float(string="Overtime Amount (₹)", compute="_compute_overtime_amount", store=True)
    morning_session = fields.Selection([
        ('Present', 'Present'),
        ('Absent', 'Absent')
    ], string="Morning Session", compute='_compute_full_day_status', store=True)
    afternoon_session = fields.Selection([
        ('Present', 'Present'),
        ('Absent', 'Absent')
    ], string="Afternoon Session", compute='_compute_full_day_status', store=True)

    full_day_status = fields.Char(string="Full Day Status", compute='_compute_full_day_status', store=True)

    ctc_type = fields.Selection([
        ('with_bonus', 'With Bonus'),
        ('without_bonus', 'Without Bonus'),
        ('non_ctc', 'Non Ctc'),
    ], string='CTC Type', related='employee_name.ctc_type', store=True, readonly=True)
    designation_id = fields.Many2one('hr.job', string="Designation", related='employee_name.job_id', store=True,
                                     readonly=True)
    department_id = fields.Many2one('hr.department', string="Department", related='employee_name.department_id', store=True,readonly=True)

    company_id = fields.Many2one('res.company', string="Company", related='employee_name.company_id', store=True,
                                 readonly=True)
    division_id = fields.Many2one(
        'product.category',
        string='Division',
        domain=[('parent_id', '=', False)],
        related='employee_name.division_id', store=True, readonly=True
    )

    @api.depends('check_out_attendance')
    def _compute_overtime_amount(self):
        for rec in self:
            rec.overtime_amount = 0.0
            if not rec.check_out_attendance:
                continue

            try:
                # Convert check_out_attendance (HH:MM) to datetime.time
                check_out_time = datetime.strptime(rec.check_out_attendance, "%H:%M").time()
            except Exception:
                rec.overtime_amount = 0.0
                continue

            # Fetch all overtime rules
            overtime_rules = self.env['hr.overtime.master'].search([])

            for rule in overtime_rules:
                try:
                    start_time = datetime.strptime(rule.start_time, "%H:%M").time()
                    end_time = datetime.strptime(rule.end_time, "%H:%M").time()
                except Exception:
                    continue

                # ✅ If check-out is between start and end → assign overtime
                if start_time <= check_out_time <= end_time:
                    rec.overtime_amount = rule.overtime_amount
                    break


    @api.depends(
        'check_in_attendance',
        'check_out_attendance',
        'date',
        'employee_name',
        'employee_name.resource_calendar_id',
        'employee_name.resource_calendar_id.attendance_ids',
        'employee_name.resource_calendar_id.attendance_ids.hour_from',
        'employee_name.resource_calendar_id.attendance_ids.hour_to',
        'employee_name.resource_calendar_id.attendance_ids.day_period',
    )
    def _compute_full_day_status(self):
        for rec in self:
            rec.full_day_status = 'Absent (Full Day)'
            rec.morning_session = 'Absent'
            rec.afternoon_session = 'Absent'

            if not rec.employee_name or not rec.date:
                _logger.debug("hr.upload[%s]: missing employee or date", rec.id)
                continue

            calendar = rec.employee_name.resource_calendar_id
            if not calendar:
                _logger.debug("hr.upload[%s]: employee %s has no calendar", rec.id, rec.employee_name.name)
                continue

            # parse check in/out times robustly (allow 'HH:MM' or 'HH:MM:SS')
            def parse_time_str(tstr):
                if not tstr:
                    return None
                for fmt in ("%H:%M:%S", "%H:%M"):
                    try:
                        return datetime.strptime(tstr, fmt).time()
                    except Exception:
                        continue
                # if already a time/datetime object, handle gracefully
                if isinstance(tstr, datetime):
                    return tstr.time()
                if isinstance(tstr, time):
                    return tstr
                return None

            tin = parse_time_str(rec.check_in_attendance)
            tout = parse_time_str(rec.check_out_attendance)
            if not tin or not tout:
                _logger.debug("hr.upload[%s]: could not parse times: in=%s out=%s", rec.id, rec.check_in_attendance,
                              rec.check_out_attendance)
                continue

            # combine with rec.date so all datetimes share same date
            check_in_dt = datetime.combine(rec.date, tin)
            check_out_dt = datetime.combine(rec.date, tout)

            weekday = rec.date.weekday()
            sessions = calendar.attendance_ids.filtered(lambda a: int(a.dayofweek) == weekday)
            if not sessions:
                _logger.debug("hr.upload[%s]: no sessions for weekday %s", rec.id, weekday)
                continue

            # thresholds
            MORNING_MIN_HOURS = 1.5
            AFTERNOON_MIN_HOURS = 4.5
            EPS = 1e-6  # float tolerance

            for session in sessions:
                # compute session start/end on the same rec.date
                start_time = time(int(session.hour_from), int((session.hour_from % 1) * 60))
                end_time = time(int(session.hour_to), int((session.hour_to % 1) * 60))
                session_start = datetime.combine(rec.date, start_time)
                session_end = datetime.combine(rec.date, end_time)

                # If session_end <= session_start skip invalid slot
                if session_end <= session_start:
                    _logger.warning("hr.upload[%s]: invalid session times %s-%s", rec.id, session.hour_from,
                                    session.hour_to)
                    continue

                # compute overlap (inclusive at boundaries)
                overlap_start = max(check_in_dt, session_start)
                overlap_end = min(check_out_dt, session_end)
                worked_seconds = (overlap_end - overlap_start).total_seconds() if overlap_end > overlap_start else 0.0
                worked_hours = worked_seconds / 3600.0

                _logger.debug(
                    "hr.upload[%s] emp=%s period=%s session=%s-%s check=%s-%s worked=%.3f",
                    rec.id, rec.employee_name.name, session.day_period,
                    session_start.time(), session_end.time(),
                    check_in_dt.time(), check_out_dt.time(), worked_hours
                )

                # apply threshold per session
                if session.day_period == 'morning':
                    if worked_hours + EPS >= MORNING_MIN_HOURS:
                        rec.morning_session = 'Present'
                elif session.day_period == 'afternoon':
                    if worked_hours + EPS >= AFTERNOON_MIN_HOURS:
                        rec.afternoon_session = 'Present'

            # decide final full_day_status
            if rec.morning_session == 'Present' and rec.afternoon_session == 'Present':
                rec.full_day_status = 'Present (Full Day)'
            elif rec.morning_session == 'Present' and rec.afternoon_session == 'Absent':
                rec.full_day_status = 'First Session Present and Second Session Absent'
            elif rec.morning_session == 'Absent' and rec.afternoon_session == 'Present':
                rec.full_day_status = 'First Session Absent and Second Session Present'
            else:
                rec.full_day_status = 'Absent (Full Day)'

    # @api.depends(
    #     'check_in_attendance',
    #     'check_out_attendance',
    #     'date',
    #     'employee_name',
    #     'employee_name.resource_calendar_id',
    #     'employee_name.resource_calendar_id.attendance_ids',
    #     'employee_name.resource_calendar_id.attendance_ids.hour_from',
    #     'employee_name.resource_calendar_id.attendance_ids.hour_to',
    #     'employee_name.resource_calendar_id.attendance_ids.day_period',
    # )
    # def _compute_full_day_status(self):
    #     for rec in self:
    #         rec.full_day_status = 'Absent (Full Day)'
    #         rec.morning_session = 'Absent'
    #         rec.afternoon_session = 'Absent'
    #
    #         if not rec.employee_name or not rec.date:
    #             _logger.debug("hr.upload[%s]: missing employee or date", rec.id)
    #             continue
    #
    #         calendar = rec.employee_name.resource_calendar_id
    #         if not calendar:
    #             _logger.debug("hr.upload[%s]: employee %s has no calendar", rec.id, rec.employee_name.name)
    #             continue
    #
    #         # parse check in/out times robustly (allow 'HH:MM' or 'HH:MM:SS')
    #         def parse_time_str(tstr):
    #             if not tstr:
    #                 return None
    #             for fmt in ("%H:%M:%S", "%H:%M"):
    #                 try:
    #                     return datetime.strptime(tstr, fmt).time()
    #                 except Exception:
    #                     continue
    #             # if already a time/datetime object, handle gracefully
    #             if isinstance(tstr, datetime):
    #                 return tstr.time()
    #             if isinstance(tstr, time):
    #                 return tstr
    #             return None
    #
    #         tin = parse_time_str(rec.check_in_attendance)
    #         tout = parse_time_str(rec.check_out_attendance)
    #         if not tin or not tout:
    #             _logger.debug("hr.upload[%s]: could not parse times: in=%s out=%s", rec.id, rec.check_in_attendance,
    #                           rec.check_out_attendance)
    #             continue
    #
    #         # combine with rec.date so all datetimes share same date
    #         check_in_dt = datetime.combine(rec.date, tin)
    #         check_out_dt = datetime.combine(rec.date, tout)
    #
    #         weekday = rec.date.weekday()
    #         sessions = calendar.attendance_ids.filtered(lambda a: int(a.dayofweek) == weekday)
    #         if not sessions:
    #             _logger.debug("hr.upload[%s]: no sessions for weekday %s", rec.id, weekday)
    #             continue
    #
    #         # thresholds
    #         MORNING_MIN_HOURS = 1.5
    #         AFTERNOON_MIN_HOURS = 4.5
    #         EPS = 1e-6  # float tolerance
    #
    #         for session in sessions:
    #             # compute session start/end on the same rec.date
    #             start_time = time(int(session.hour_from), int((session.hour_from % 1) * 60))
    #             end_time = time(int(session.hour_to), int((session.hour_to % 1) * 60))
    #             session_start = datetime.combine(rec.date, start_time)
    #             session_end = datetime.combine(rec.date, end_time)
    #
    #             # If session_end <= session_start skip invalid slot
    #             if session_end <= session_start:
    #                 _logger.warning("hr.upload[%s]: invalid session times %s-%s", rec.id, session.hour_from,
    #                                 session.hour_to)
    #                 continue
    #
    #             # compute overlap (inclusive at boundaries)
    #             overlap_start = max(check_in_dt, session_start)
    #             overlap_end = min(check_out_dt, session_end)
    #             worked_seconds = (overlap_end - overlap_start).total_seconds() if overlap_end > overlap_start else 0.0
    #             worked_hours = worked_seconds / 3600.0
    #
    #             _logger.debug(
    #                 "hr.upload[%s] emp=%s period=%s session=%s-%s check=%s-%s worked=%.3f",
    #                 rec.id, rec.employee_name.name, session.day_period,
    #                 session_start.time(), session_end.time(),
    #                 check_in_dt.time(), check_out_dt.time(), worked_hours
    #             )
    #
    #             # apply threshold per session
    #             if session.day_period == 'morning':
    #                 if worked_hours + EPS >= MORNING_MIN_HOURS:
    #                     rec.morning_session = 'Present'
    #             elif session.day_period == 'afternoon':
    #                 if worked_hours + EPS >= AFTERNOON_MIN_HOURS:
    #                     rec.afternoon_session = 'Present'
    #
    #         # decide final full_day_status
    #         if rec.morning_session == 'Present' and rec.afternoon_session == 'Present':
    #             rec.full_day_status = 'Present (Full Day)'
    #         elif rec.morning_session == 'Present' and rec.afternoon_session == 'Absent':
    #             rec.full_day_status = 'First Session Present and Second Session Absent'
    #         elif rec.morning_session == 'Absent' and rec.afternoon_session == 'Present':
    #             rec.full_day_status = 'First Session Absent and Second Session Present'
    #         else:
    #             rec.full_day_status = 'Absent (Full Day)'

    # @api.depends('total_working_hours', 'morning_session', 'afternoon_session', 'check_out_attendance')
    # def _compute_full_day_status(self):
    #     for rec in self:
    #         # Convert working hours safely to float
    #         try:
    #             total_hours = float(rec.total_working_hours or 0.0)
    #         except (ValueError, TypeError):
    #             total_hours = 0.0
    #
    #         # If no checkout → mark afternoon as absent
    #         if not rec.check_out_attendance:
    #             rec.afternoon_session = 'Absent'
    #
    #         # Determine full day status
    #         if total_hours >= 9.0:
    #             rec.full_day_status = 'Present (Full Day)'
    #         else:
    #             morning = rec.morning_session or 'Absent'
    #             afternoon = rec.afternoon_session or 'Absent'
    #
    #             if morning == 'Present' and afternoon == 'Present':
    #                 rec.full_day_status = 'FIRST SESSION PRESENT AND SECOND SESSION PRESENT'
    #             elif morning == 'Present' and afternoon == 'Absent':
    #                 rec.full_day_status = 'FIRST SESSION PRESENT AND SECOND SESSION ABSENT'
    #             elif morning == 'Absent' and afternoon == 'Present':
    #                 rec.full_day_status = 'FIRST SESSION ABSENT AND SECOND SESSION PRESENT'
    #             else:
    #                 rec.full_day_status = 'Absent (Full Day)'

    @api.depends('check_in_attendance', 'check_out_attendance')
    def _compute_attendance(self):
        shift_start = datetime.strptime('10:00', '%H:%M')
        grace_end = datetime.strptime('10:15', '%H:%M')
        shift_end = datetime.strptime('19:00', '%H:%M')

        for rec in self:
            rec.difference_check_in = 0.0
            rec.difference_check_out = 0.0
            rec.total_working_hours = "0:00"
            rec.attendance_status = 'others'  # Default only in compute

            # --- Check-in minutes & status ---
            if rec.check_in_attendance:
                try:
                    check_in_time = datetime.strptime(rec.check_in_attendance, '%H:%M')
                    if check_in_time <= shift_start:
                        rec.difference_check_in = (shift_start - check_in_time).total_seconds() / 60
                        rec.attendance_status = 'early'
                    elif shift_start < check_in_time <= grace_end:
                        rec.difference_check_in = 0.0
                        rec.attendance_status = 'grace'
                    else:
                        rec.difference_check_in = -((check_in_time - grace_end).total_seconds() / 60)
                        rec.attendance_status = 'late'
                except:
                    rec.attendance_status = 'others'
            else:
                # No check-in → mark as others
                rec.attendance_status = 'others'

            # --- Check-out minutes ---
            if rec.check_out_attendance:
                try:
                    check_out_time = datetime.strptime(rec.check_out_attendance, '%H:%M')
                    if check_out_time >= shift_end:
                        rec.difference_check_out = (check_out_time - shift_end).total_seconds() / 60
                    else:
                        rec.difference_check_out = -((shift_end - check_out_time).total_seconds() / 60)
                except:
                    rec.difference_check_out = 0.0

            # --- Total working hours ---
            if rec.check_in_attendance and rec.check_out_attendance:
                try:
                    check_in_time = datetime.strptime(rec.check_in_attendance, '%H:%M')
                    check_out_time = datetime.strptime(rec.check_out_attendance, '%H:%M')
                    total_seconds = (check_out_time - check_in_time).total_seconds()
                    if total_seconds < 0:
                        total_seconds += 24 * 3600
                    hours = int(total_seconds // 3600)
                    minutes = int((total_seconds % 3600) // 60)
                    rec.total_working_hours = f"{hours}:{minutes:02d}"
                except:
                    rec.total_working_hours = "0:00"

    @api.depends('check_in_attendance')
    def _compute_late_deduction(self):
        for rec in self:
            rec.late_deduction = 0.0
            if rec.check_in_attendance:
                try:
                    # Convert check_in_attendance (HH:MM) to datetime.time
                    check_in_time = datetime.strptime(rec.check_in_attendance, "%H:%M").time()

                    masters = self.env['hr.late.deduction.master'].search([])
                    for master in masters:
                        start_time = datetime.strptime(master.start_time, "%H:%M").time()
                        end_time = datetime.strptime(master.end_time, "%H:%M").time()

                        if start_time <= check_in_time <= end_time:
                            rec.late_deduction = master.deduction_amount
                            break
                except Exception:
                    rec.late_deduction = 0.0
            else:
                rec.late_deduction = 0.0

    # @api.depends('check_in_attendance')
    # def _compute_sessions(self):
    #     for rec in self:
    #         if rec.check_in_attendance:
    #             try:
    #                 check_in_time = datetime.strptime(rec.check_in_attendance, "%H:%M")
    #                 shift_start = datetime.strptime("10:00", "%H:%M")
    #                 grace_end = datetime.strptime("10:15", "%H:%M")
    #                 first_session_absent_time = grace_end + timedelta(minutes=75)  # 11:30 AM
    #                 second_session_absent_time = datetime.strptime("14:15", "%H:%M")  # 2:15 PM
    #
    #                 # Determine sessions
    #                 if check_in_time > second_session_absent_time:
    #                     rec.morning_session = 'Absent'
    #                     rec.afternoon_session = 'Absent'
    #                 elif check_in_time > first_session_absent_time:
    #                     rec.morning_session = 'Absent'
    #                     rec.afternoon_session = 'Present'
    #                 else:
    #                     rec.morning_session = 'Present'
    #                     rec.afternoon_session = 'Present'
    #
    #                 if not rec.check_out_attendance:
    #                     rec.afternoon_session = 'Absent'
    #             except ValueError:
    #                 rec.morning_session = 'Absent'
    #                 rec.afternoon_session = 'Absent'
    #         else:
    #             rec.morning_session = 'Absent'
    #             rec.afternoon_session = 'Absent'


    # @api.depends('date', 'check_in_attendance')
    # def _compute_your_datetime(self):
    #     for record in self:
    #         if record.date:
    #             try:
    #                 # Use check_in time if provided, otherwise default to "00:00"
    #                 check_in_time = record.check_in_attendance if record.check_in_attendance else "00:00"
    #                 # Combine date and time into a string
    #                 datetime_str = f"{record.date} {check_in_time}:00"
    #                 # Parse the combined string into a naive datetime object
    #                 naive_datetime = datetime.strptime(datetime_str, DEFAULT_SERVER_DATETIME_FORMAT)
    #                 # Convert the naive datetime to the user's timezone
    #                 user_tz = self.env.user.tz or 'UTC'
    #                 local_tz = pytz.timezone(user_tz)
    #                 local_dt = local_tz.localize(naive_datetime, is_dst=None)
    #                 # Convert the localized datetime to UTC
    #                 utc_dt = local_dt.astimezone(pytz.UTC)
    #                 # Assign the UTC datetime to the field
    #                 record.your_datetime = fields.Datetime.to_string(utc_dt)
    #             except ValueError:
    #                 record.your_datetime = False
    #         else:
    #             # If date is not provided, set your_datetime to False
    #             record.your_datetime = False
    #
    # @api.depends('date', 'check_out_attendance')
    # def _compute_your_checkout_datetime(self):
    #     for record in self:
    #         if record.date:
    #             try:
    #                 # Use check_out_attendance if available, otherwise default to "00:00"
    #                 checkout_time = record.check_out_attendance or "00:00"
    #
    #                 # Combine date and time into a string
    #                 datetime_str = f"{record.date} {checkout_time}:00"
    #
    #                 # Parse the combined string into a naive datetime object
    #                 naive_datetime = datetime.strptime(datetime_str, DEFAULT_SERVER_DATETIME_FORMAT)
    #
    #                 # Convert to user's timezone and then to UTC
    #                 user_tz = self.env.user.tz or 'UTC'
    #                 local_tz = pytz.timezone(user_tz)
    #                 local_dt = local_tz.localize(naive_datetime, is_dst=None)
    #                 utc_dt = local_dt.astimezone(pytz.UTC)
    #
    #                 # Assign the UTC datetime
    #                 record.your_checkout_datetime = fields.Datetime.to_string(utc_dt)
    #             except ValueError:
    #                 record.your_checkout_datetime = False
    #         else:
    #             record.your_checkout_datetime = False

    DEFAULT_SERVER_DATETIME_FORMAT = "%Y-%m-%d %H:%M:%S "

    @api.depends('date', 'check_in_attendance')
    def _compute_your_datetime(self):
        for record in self:
            if record.date:
                try:
                    check_in_time = record.check_in_attendance or "00:00"

                    # Normalize both formats HH:MM and HH:MM:SS
                    parts = check_in_time.split(':')
                    if len(parts) == 2:  # Only hours and minutes
                        check_in_time = f"{check_in_time}:00"
                    elif len(parts) == 1:  # Just hour (edge case)
                        check_in_time = f"{check_in_time}:00:00"

                    datetime_str = f"{record.date} {check_in_time}"
                    naive_datetime = datetime.strptime(datetime_str, DEFAULT_SERVER_DATETIME_FORMAT)

                    # Convert to UTC based on user's timezone
                    user_tz = self.env.user.tz or 'UTC'
                    local_tz = pytz.timezone(user_tz)
                    local_dt = local_tz.localize(naive_datetime, is_dst=None)
                    utc_dt = local_dt.astimezone(pytz.UTC)

                    record.your_datetime = fields.Datetime.to_string(utc_dt)
                except Exception:
                    record.your_datetime = False
            else:
                record.your_datetime = False

    @api.depends('date', 'check_out_attendance')
    def _compute_your_checkout_datetime(self):
        for record in self:
            if record.date:
                try:
                    checkout_time = record.check_out_attendance or "00:00"

                    # Normalize time to HH:MM:SS
                    parts = checkout_time.split(':')
                    if len(parts) == 2:
                        checkout_time = f"{checkout_time}:00"
                    elif len(parts) == 1:
                        checkout_time = f"{checkout_time}:00:00"

                    datetime_str = f"{record.date} {checkout_time}"
                    naive_datetime = datetime.strptime(datetime_str, DEFAULT_SERVER_DATETIME_FORMAT)

                    # Convert to UTC based on user's timezone
                    user_tz = self.env.user.tz or 'UTC'
                    local_tz = pytz.timezone(user_tz)
                    local_dt = local_tz.localize(naive_datetime, is_dst=None)
                    utc_dt = local_dt.astimezone(pytz.UTC)

                    record.your_checkout_datetime = fields.Datetime.to_string(utc_dt)
                except Exception:
                    record.your_checkout_datetime = False
            else:
                record.your_checkout_datetime = False


    def open_upload_wizard(self):
        return {
            'name': 'Upload HR Data',
            'type': 'ir.actions.act_window',
            'res_model': 'hr.upload.wizard',
            'view_mode': 'form',
            'target': 'new',
        }



class HrUploadWizard(models.TransientModel):
    _name = "hr.upload.wizard"
    _description = "HR Upload Wizard"

    file = fields.Binary(string="File", required=True)
    filename = fields.Char(string="File Name")

    def action_upload(self):
        """Uploads attendance Excel file and creates hr.upload, hr.attendance, and hr.leave records
        with EL/LOP rules and hybrid working schedule–based attendance logic."""
        if not self.file:
            raise UserError(_("Please upload a file."))

        try:
            import base64
            import openpyxl
            from io import BytesIO
            from datetime import datetime, date, time, timedelta
            import pytz

            wb = openpyxl.load_workbook(
                filename=BytesIO(base64.b64decode(self.file)),
                read_only=True
            )
            ws = wb.active

            HrUpload = self.env['hr.upload']
            Employee = self.env['hr.employee']
            Attendance = self.env['hr.attendance']
            Leave = self.env['hr.leave']
            Allocation = self.env['hr.leave.allocation']
            LeaveType = self.env['hr.leave.type']

            # === Leave Type Lookup by Work Entry Code ===
            el_type = LeaveType.search([('work_entry_type_id.code', '=', 'LEAVE120')], limit=1)
            lop_type = LeaveType.search([('work_entry_type_id.code', '=', 'LEAVE90')], limit=1)
            if not el_type or not lop_type:
                raise UserError(
                    _("Please configure Time Off types with work entry codes LEAVE120 (EL) and LEAVE90 (LOP)."))

            # === Set timezone ===
            user_tz = self.env.user.tz or 'Asia/Kolkata'
            tz = pytz.timezone(user_tz)

            def normalize_time(time_value):
                """Convert Excel time or float value to HH:MM format, ignoring seconds."""
                if not time_value:
                    return None
                try:
                    if isinstance(time_value, (datetime, time)):
                        return time_value.strftime("%H:%M")
                    if isinstance(time_value, (int, float)):
                        from openpyxl.utils.datetime import from_excel
                        dt = from_excel(time_value)
                        return dt.strftime("%H:%M")
                    time_str = str(time_value).strip().replace(';', ':').replace('.', ':')
                    parts = time_str.split(':')
                    if len(parts) >= 2:
                        return f"{parts[0].zfill(2)}:{parts[1].zfill(2)}"
                    return None
                except Exception:
                    return None

            # === Process each row ===
            row_number = 1
            for row in ws.iter_rows(min_row=2, values_only=True):
                row_number += 1

                employee_code = row[0]
                employee_name = row[1]
                date_value = row[2]
                check_in = row[3]
                check_out = row[4]

                if not employee_code:
                    raise UserError(_("Row %d: Employee code is missing.") % row_number)
                if not employee_name:
                    raise UserError(_("Row %d: Employee name is missing.") % row_number)

                employee_name = str(employee_name).strip()
                try:
                    employee_code_str = str(int(employee_code))
                except Exception:
                    employee_code_str = str(employee_code).strip()

                employee = Employee.search([
                    ('cmr_code', '=', employee_code_str),
                    ('name', '=', employee_name)
                ], limit=1)
                if not employee:
                    raise UserError(_("Row %d: No employee found with code '%s' and name '%s'.") %
                                    (row_number, employee_code_str, employee_name))

                # === Date Parsing ===
                from openpyxl.utils.datetime import from_excel

                formatted_date = False
                if isinstance(date_value, datetime):
                    formatted_date = date_value.date()
                elif isinstance(date_value, (int, float)):
                    try:
                        formatted_date = from_excel(date_value).date()
                    except Exception:
                        pass
                elif isinstance(date_value, str):
                    date_value = date_value.strip()
                    for fmt in [
                        "%d-%m-%Y", "%d/%m/%Y", "%Y-%m-%d", "%d-%b-%Y",
                        "%d/%b/%Y", "%b-%d-%Y", "%d.%m.%Y", "%m/%d/%Y", "%d/%m/%y"
                    ]:
                        try:
                            formatted_date = datetime.strptime(date_value, fmt).date()
                            break
                        except Exception:
                            continue

                if not formatted_date:
                    raise UserError(_("Row %d: Invalid date format '%s'.") % (row_number, date_value))

                check_in = normalize_time(check_in)
                check_out = normalize_time(check_out)

                # === Create hr.upload Record ===
                upload_rec = HrUpload.create({
                    'employee_name': employee.id,
                    'ctc_type': employee.ctc_type,
                    'employee_code': employee_code_str,
                    'date': formatted_date,
                    'check_in_attendance': check_in,
                    'check_out_attendance': check_out,
                })

                if hasattr(upload_rec, '_compute_sessions'):
                    upload_rec._compute_sessions()

                morning = getattr(upload_rec, 'morning_session', '') or ''
                afternoon = getattr(upload_rec, 'afternoon_session', '') or ''
                leave_duration = 0.0
                half_day = False

                if morning == 'Absent' and afternoon == 'Absent':
                    leave_duration = 1.0
                elif morning == 'Absent' or afternoon == 'Absent':
                    leave_duration = 0.5
                    half_day = True

                # === Leave Logic ===
                if leave_duration > 0:
                    existing_leave = Leave.search([
                        ('employee_id', '=', employee.id),
                        ('request_date_from', '=', formatted_date),
                        ('request_date_to', '=', formatted_date),
                        ('state', '!=', 'refuse')
                    ], limit=1)
                    if existing_leave:
                        continue

                    if employee.ctc_type != 'non_ctc':
                        leave_type_to_use = lop_type
                        remaining_el = 0.0
                    else:
                        el_allocations = Allocation.search([
                            ('employee_id', '=', employee.id),
                            ('holiday_status_id', '=', el_type.id),
                            ('state', '=', 'validate')
                        ], limit=1)

                        remaining_el = 0.0
                        if el_allocations:
                            alloc = el_allocations[0]
                            start_date = alloc.date_from
                            accrual_per_month = 0.0

                            if alloc.accrual_plan_id:
                                accrual_level = self.env['hr.leave.accrual.level'].search([
                                    ('accrual_plan_id', '=', alloc.accrual_plan_id.id)
                                ], limit=1)
                                if accrual_level:
                                    accrual_per_month = accrual_level.added_value

                            if not accrual_per_month:
                                raise UserError(_(
                                    "Row %d: No accrual value defined in Accrual Plan for %s."
                                ) % (row_number, employee.name))

                            if start_date and formatted_date >= start_date:
                                months_passed = (formatted_date.year - start_date.year) * 12 + (
                                        formatted_date.month - start_date.month) + 1
                                total_allocated = months_passed * accrual_per_month
                            else:
                                total_allocated = 0.0

                            el_taken = sum(Leave.search([
                                ('employee_id', '=', employee.id),
                                ('holiday_status_id', '=', el_type.id),
                                ('state', 'in', ['validate', 'confirm']),
                                ('request_date_from', '<=', formatted_date)
                            ]).mapped('number_of_days'))

                            remaining_el = max(total_allocated - el_taken, 0.0)
                        else:
                            raise UserError(_(
                                "Row %d: No validated EL allocation found for %s."
                            ) % (row_number, employee.name))

                    def create_and_validate_leave(leave_type, days, half_day=False, period=False):
                        vals = {
                            'name': f"Auto Leave {formatted_date}",
                            'employee_id': employee.id,
                            'holiday_status_id': leave_type.id,
                            'request_date_from': formatted_date,
                            'request_date_to': formatted_date,
                            'number_of_days': days,
                        }
                        if half_day:
                            vals.update({
                                'request_unit_half': True,
                                'request_date_from_period': period,
                            })
                        leave_rec = Leave.create(vals)
                        if leave_rec.state == 'draft':
                            try:
                                leave_rec.action_confirm()
                            except Exception:
                                pass
                        if leave_rec.state in ['confirm', 'validate1', 'validate']:
                            try:
                                leave_rec.action_validate()
                            except Exception:
                                pass

                    if employee.ctc_type == 'non_ctc':
                        if remaining_el >= leave_duration:
                            create_and_validate_leave(
                                el_type, leave_duration, half_day,
                                'am' if half_day and morning == 'Absent' else 'pm' if half_day else False
                            )
                        elif remaining_el > 0:
                            create_and_validate_leave(el_type, remaining_el)
                            create_and_validate_leave(lop_type, leave_duration - remaining_el)
                        else:
                            create_and_validate_leave(
                                lop_type, leave_duration, half_day,
                                'am' if half_day and morning == 'Absent' else 'pm' if half_day else False
                            )
                    else:
                        create_and_validate_leave(
                            lop_type, leave_duration, half_day,
                            'am' if half_day and morning == 'Absent' else 'pm' if half_day else False
                        )

                # === Attendance creation (Hybrid logic) ===
                calendar = employee.resource_calendar_id
                if not calendar:
                    continue

                day_of_week = formatted_date.strftime("%A")
                working_hours = calendar.attendance_ids.filtered(
                    lambda a: a.display_name and day_of_week.lower() in a.display_name.lower()
                )
                if not working_hours:
                    continue

                morning_sched = working_hours.filtered(lambda a: 'morning' in a.display_name.lower())
                afternoon_sched = working_hours.filtered(lambda a: 'afternoon' in a.display_name.lower())

                def to_utc_datetime(date_part, hour_float):
                    hours = int(hour_float)
                    minutes = int(round((hour_float - hours) * 60))
                    local_dt = datetime.combine(date_part, time(hours, minutes))
                    local_dt = tz.localize(local_dt)
                    return local_dt.astimezone(pytz.UTC)

                def to_datetime_local(date_part, time_str):
                    try:
                        t = datetime.strptime(time_str, "%H:%M").time()
                        local_dt = datetime.combine(date_part, t)
                        local_dt = tz.localize(local_dt)
                        return local_dt.astimezone(pytz.UTC)
                    except Exception:
                        return None

                check_in_dt = check_out_dt = None

                # Morning present: actual in, schedule out
                if morning != 'Absent' and morning_sched:
                    sched_out = to_utc_datetime(formatted_date, max(morning_sched.mapped('hour_to')))
                    actual_in = to_datetime_local(formatted_date, check_in) if check_in else None
                    check_in_dt = actual_in or to_utc_datetime(formatted_date, min(morning_sched.mapped('hour_from')))
                    check_out_dt = sched_out

                # Afternoon present: schedule in, actual out
                if afternoon != 'Absent' and afternoon_sched:
                    sched_in = to_utc_datetime(formatted_date, min(afternoon_sched.mapped('hour_from')))
                    actual_out = to_datetime_local(formatted_date, check_out) if check_out else None

                    if check_in_dt and check_out_dt:
                        # merge morning+afternoon
                        check_out_dt = actual_out or to_utc_datetime(formatted_date,
                                                                     max(afternoon_sched.mapped('hour_to')))
                    else:
                        check_in_dt = sched_in
                        check_out_dt = actual_out or to_utc_datetime(formatted_date,
                                                                     max(afternoon_sched.mapped('hour_to')))

                if check_in_dt and check_out_dt and check_out_dt > check_in_dt:
                    open_attendance = Attendance.search([
                        ('employee_id', '=', employee.id),
                        ('check_out', '=', False)
                    ], limit=1)

                    if open_attendance:
                        open_attendance.write({'check_out': fields.Datetime.to_string(check_out_dt)})
                    else:
                        Attendance.create({
                            'employee_id': employee.id,
                            'date': formatted_date,
                            'check_in': fields.Datetime.to_string(check_in_dt),
                            'check_out': fields.Datetime.to_string(check_out_dt),
                        })

        except Exception as e:
            raise UserError(_('Upload failed: %s') % e)

    #nov 3rd
    # def action_upload(self):
    #     """Uploads attendance Excel file and creates hr.upload, hr.attendance, and hr.leave records
    #     with EL/LOP rules based on dynamic accrual setup and employee type."""
    #     if not self.file:
    #         raise UserError(_("Please upload a file."))
    #
    #     try:
    #         import base64
    #         import openpyxl
    #         from io import BytesIO
    #         from datetime import datetime, date, time, timedelta
    #         import pytz
    #
    #         wb = openpyxl.load_workbook(
    #             filename=BytesIO(base64.b64decode(self.file)),
    #             read_only=True
    #         )
    #         ws = wb.active
    #
    #         HrUpload = self.env['hr.upload']
    #         Employee = self.env['hr.employee']
    #         Attendance = self.env['hr.attendance']
    #         Leave = self.env['hr.leave']
    #         Allocation = self.env['hr.leave.allocation']
    #         LeaveType = self.env['hr.leave.type']
    #
    #         # === Leave Type Lookup by Work Entry Code ===
    #         el_type = LeaveType.search([('work_entry_type_id.code', '=', 'LEAVE120')], limit=1)
    #         lop_type = LeaveType.search([('work_entry_type_id.code', '=', 'LEAVE90')], limit=1)
    #         if not el_type or not lop_type:
    #             raise UserError(
    #                 _("Please configure Time Off types with work entry codes LEAVE120 (EL) and LEAVE90 (LOP)."))
    #
    #         # === Set timezone ===
    #         user_tz = self.env.user.tz or 'Asia/Kolkata'
    #         tz = pytz.timezone(user_tz)
    #
    #         def normalize_time(time_value):
    #             """Convert Excel time or float value to HH:MM format, ignoring seconds."""
    #             if not time_value:
    #                 return None
    #             try:
    #                 if isinstance(time_value, (datetime, time)):
    #                     return time_value.strftime("%H:%M")
    #                 if isinstance(time_value, (int, float)):
    #                     from openpyxl.utils.datetime import from_excel
    #                     dt = from_excel(time_value)
    #                     return dt.strftime("%H:%M")
    #                 time_str = str(time_value).strip().replace(';', ':').replace('.', ':')
    #                 parts = time_str.split(':')
    #                 if len(parts) >= 2:
    #                     return f"{parts[0].zfill(2)}:{parts[1].zfill(2)}"
    #                 return None
    #             except Exception:
    #                 return None
    #
    #         # === Process each row in Excel ===
    #         row_number = 1
    #         for row in ws.iter_rows(min_row=2, values_only=True):
    #             row_number += 1
    #
    #             employee_code = row[0]
    #             employee_name = row[1]
    #             date_value = row[2]
    #             check_in = row[3]
    #             check_out = row[4]
    #
    #             if not employee_code:
    #                 raise UserError(_("Row %d: Employee code is missing.") % row_number)
    #             if not employee_name:
    #                 raise UserError(_("Row %d: Employee name is missing.") % row_number)
    #
    #             employee_name = str(employee_name).strip()
    #             try:
    #                 employee_code_str = str(int(employee_code))
    #             except Exception:
    #                 employee_code_str = str(employee_code).strip()
    #
    #             employee = Employee.search([
    #                 ('cmr_code', '=', employee_code_str),
    #                 ('name', '=', employee_name)
    #             ], limit=1)
    #             if not employee:
    #                 raise UserError(_("Row %d: No employee found with code '%s' and name '%s'.") %
    #                                 (row_number, employee_code_str, employee_name))
    #
    #             # === Convert Date (supports all formats + Excel serials) ===
    #             from openpyxl.utils.datetime import from_excel
    #
    #             formatted_date = False
    #             if isinstance(date_value, datetime):
    #                 formatted_date = date_value.date()
    #             elif isinstance(date_value, (int, float)):
    #                 try:
    #                     formatted_date = from_excel(date_value).date()
    #                 except Exception:
    #                     pass
    #             elif isinstance(date_value, str):
    #                 date_value = date_value.strip()
    #                 possible_formats = [
    #                     "%d-%m-%Y", "%d/%m/%Y", "%Y-%m-%d", "%d-%b-%Y",
    #                     "%d/%b/%Y", "%b-%d-%Y", "%d.%m.%Y", "%m/%d/%Y", "%d/%m/%y"
    #                 ]
    #                 for fmt in possible_formats:
    #                     try:
    #                         formatted_date = datetime.strptime(date_value, fmt).date()
    #                         break
    #                     except Exception:
    #                         continue
    #
    #             if not formatted_date:
    #                 raise UserError(_("Row %d: Invalid date format '%s'.") % (row_number, date_value))
    #
    #             check_in = normalize_time(check_in)
    #             check_out = normalize_time(check_out)
    #
    #             # === Create hr.upload Record ===
    #             upload_rec = HrUpload.create({
    #                 'employee_name': employee.id,
    #                 'ctc_type': employee.ctc_type,
    #                 'employee_code': employee_code_str,
    #                 'date': formatted_date,
    #                 'check_in_attendance': check_in,
    #                 'check_out_attendance': check_out,
    #             })
    #
    #             if hasattr(upload_rec, '_compute_sessions'):
    #                 upload_rec._compute_sessions()
    #
    #             morning = getattr(upload_rec, 'morning_session', '') or ''
    #             afternoon = getattr(upload_rec, 'afternoon_session', '') or ''
    #             leave_duration = 0.0
    #             half_day = False
    #
    #             if morning == 'Absent' and afternoon == 'Absent':
    #                 leave_duration = 1.0
    #             elif morning == 'Absent' or afternoon == 'Absent':
    #                 leave_duration = 0.5
    #                 half_day = True
    #
    #             # === Leave Logic ===
    #             if leave_duration > 0:
    #                 existing_leave = Leave.search([
    #                     ('employee_id', '=', employee.id),
    #                     ('request_date_from', '=', formatted_date),
    #                     ('request_date_to', '=', formatted_date),
    #                     ('state', '!=', 'refuse')
    #                 ], limit=1)
    #                 if existing_leave:
    #                     continue
    #
    #                 # === For CTC employees, assign LOP directly ===
    #                 if employee.ctc_type != 'non_ctc':
    #                     leave_type_to_use = lop_type
    #                     remaining_el = 0.0
    #                 else:
    #                     # === Dynamic Monthly Accrual from Accrual Plan ===
    #                     el_allocations = Allocation.search([
    #                         ('employee_id', '=', employee.id),
    #                         ('holiday_status_id', '=', el_type.id),
    #                         ('state', '=', 'validate')
    #                     ], limit=1)
    #
    #                     remaining_el = 0.0
    #                     if el_allocations:
    #                         alloc = el_allocations[0]
    #                         start_date = alloc.date_from
    #                         accrual_per_month = 0.0
    #
    #                         if alloc.accrual_plan_id:
    #                             accrual_level = self.env['hr.leave.accrual.level'].search([
    #                                 ('accrual_plan_id', '=', alloc.accrual_plan_id.id)
    #                             ], limit=1)
    #                             if accrual_level:
    #                                 accrual_per_month = accrual_level.added_value
    #
    #                         if not accrual_per_month:
    #                             raise UserError(_(
    #                                 "Row %d: No accrual value defined in Accrual Plan for %s."
    #                             ) % (row_number, employee.name))
    #
    #                         if start_date and formatted_date >= start_date:
    #                             months_passed = (formatted_date.year - start_date.year) * 12 + (
    #                                     formatted_date.month - start_date.month) + 1
    #                             total_allocated = months_passed * accrual_per_month
    #                         else:
    #                             total_allocated = 0.0
    #
    #                         el_taken = sum(Leave.search([
    #                             ('employee_id', '=', employee.id),
    #                             ('holiday_status_id', '=', el_type.id),
    #                             ('state', 'in', ['validate', 'confirm']),
    #                             ('request_date_from', '<=', formatted_date)
    #                         ]).mapped('number_of_days'))
    #
    #                         remaining_el = max(total_allocated - el_taken, 0.0)
    #                     else:
    #                         raise UserError(_(
    #                             "Row %d: No validated EL allocation found for %s."
    #                         ) % (row_number, employee.name))
    #
    #                 def create_and_validate_leave(leave_type, days, half_day=False, period=False):
    #                     vals = {
    #                         'name': f"Auto Leave {formatted_date}",
    #                         'employee_id': employee.id,
    #                         'holiday_status_id': leave_type.id,
    #                         'request_date_from': formatted_date,
    #                         'request_date_to': formatted_date,
    #                         'number_of_days': days,
    #                     }
    #                     if half_day:
    #                         vals.update({
    #                             'request_unit_half': True,
    #                             'request_date_from_period': period,
    #                         })
    #                     leave_rec = Leave.create(vals)
    #                     if leave_rec.state == 'draft':
    #                         try:
    #                             leave_rec.action_confirm()
    #                         except Exception:
    #                             pass
    #                     if leave_rec.state in ['confirm', 'validate1', 'validate']:
    #                         try:
    #                             leave_rec.action_validate()
    #                         except Exception:
    #                             pass
    #
    #                 # === Apply ELs and LOPs ===
    #                 if employee.ctc_type == 'non_ctc':
    #                     if remaining_el >= leave_duration:
    #                         create_and_validate_leave(
    #                             el_type, leave_duration, half_day,
    #                             'am' if half_day and morning == 'Absent' else 'pm' if half_day else False
    #                         )
    #                     elif remaining_el > 0:
    #                         create_and_validate_leave(el_type, remaining_el)
    #                         create_and_validate_leave(lop_type, leave_duration - remaining_el)
    #                     else:
    #                         create_and_validate_leave(
    #                             lop_type, leave_duration, half_day,
    #                             'am' if half_day and morning == 'Absent' else 'pm' if half_day else False
    #                         )
    #                 else:
    #                     create_and_validate_leave(
    #                         lop_type, leave_duration, half_day,
    #                         'am' if half_day and morning == 'Absent' else 'pm' if half_day else False
    #                     )
    #
    #             # === Attendance creation ===
    #             if check_in and check_out:
    #                 def to_datetime_local(date_part, time_value):
    #                     if isinstance(time_value, str):
    #                         try:
    #                             t = datetime.strptime(time_value, "%H:%M").time()
    #                             local_dt = datetime.combine(date_part, t)
    #                         except Exception:
    #                             return False
    #                     else:
    #                         return False
    #                     local_dt = tz.localize(local_dt)
    #                     return local_dt.astimezone(pytz.UTC)
    #
    #                 check_in_dt = to_datetime_local(formatted_date, check_in)
    #                 check_out_dt = to_datetime_local(formatted_date, check_out)
    #
    #                 if not (check_in_dt and check_out_dt):
    #                     continue
    #
    #                 open_attendance = Attendance.search([
    #                     ('employee_id', '=', employee.id),
    #                     ('check_out', '=', False)
    #                 ], limit=1)
    #
    #                 if open_attendance:
    #                     open_attendance.write({'check_out': fields.Datetime.to_string(check_out_dt)})
    #                 else:
    #                     Attendance.create({
    #                         'employee_id': employee.id,
    #                         'date': formatted_date,
    #                         'check_in': fields.Datetime.to_string(check_in_dt),
    #                         'check_out': fields.Datetime.to_string(check_out_dt),
    #                     })
    #
    #     except Exception as e:
    #         raise UserError(_('Upload failed: %s') % e)

    # #only one date format allowing
    # def action_upload(self):
    #     """Uploads attendance Excel file and creates hr.upload, hr.attendance, and hr.leave records
    #     with EL/LOP rules based on dynamic accrual setup and employee type."""
    #     if not self.file:
    #         raise UserError(_("Please upload a file."))
    #
    #     try:
    #         import base64
    #         import openpyxl
    #         from io import BytesIO
    #         from datetime import datetime, date, time, timedelta
    #         import pytz
    #
    #         wb = openpyxl.load_workbook(
    #             filename=BytesIO(base64.b64decode(self.file)),
    #             read_only=True
    #         )
    #         ws = wb.active
    #
    #         HrUpload = self.env['hr.upload']
    #         Employee = self.env['hr.employee']
    #         Attendance = self.env['hr.attendance']
    #         Leave = self.env['hr.leave']
    #         Allocation = self.env['hr.leave.allocation']
    #         LeaveType = self.env['hr.leave.type']
    #
    #         # === Leave Type Lookup by Work Entry Code ===
    #         el_type = LeaveType.search([('work_entry_type_id.code', '=', 'LEAVE120')], limit=1)
    #         lop_type = LeaveType.search([('work_entry_type_id.code', '=', 'LEAVE90')], limit=1)
    #         if not el_type or not lop_type:
    #             raise UserError(
    #                 _("Please configure Time Off types with work entry codes LEAVE120 (EL) and LEAVE90 (LOP)."))
    #
    #         # === Set timezone ===
    #         user_tz = self.env.user.tz or 'Asia/Kolkata'
    #         tz = pytz.timezone(user_tz)
    #
    #         def normalize_time(time_value):
    #             """Convert Excel time or float value to HH:MM format."""
    #             if not time_value:
    #                 return None
    #             try:
    #                 if isinstance(time_value, (datetime, time)):
    #                     return time_value.strftime("%H:%M")
    #                 if isinstance(time_value, (int, float)):
    #                     from openpyxl.utils.datetime import from_excel
    #                     dt = from_excel(time_value)
    #                     return dt.strftime("%H:%M")
    #                 time_str = str(time_value).strip().replace(';', ':').replace('.', ':')
    #                 parts = time_str.split(':')
    #                 if len(parts) >= 2:
    #                     return f"{parts[0].zfill(2)}:{parts[1].zfill(2)}"
    #                 return time_str
    #             except Exception:
    #                 return None
    #
    #         # === Process each row in Excel ===
    #         row_number = 1
    #         for row in ws.iter_rows(min_row=2, values_only=True):
    #             row_number += 1
    #
    #             employee_code = row[0]
    #             employee_name = row[1]
    #             date_value = row[2]
    #             check_in = row[3]
    #             check_out = row[4]
    #
    #             if not employee_code:
    #                 raise UserError(_("Row %d: Employee code is missing.") % row_number)
    #             if not employee_name:
    #                 raise UserError(_("Row %d: Employee name is missing.") % row_number)
    #
    #             employee_name = str(employee_name).strip()
    #             try:
    #                 employee_code_str = str(int(employee_code))
    #             except Exception:
    #                 employee_code_str = str(employee_code).strip()
    #
    #             employee = Employee.search([
    #                 ('cmr_code', '=', employee_code_str),
    #                 ('name', '=', employee_name)
    #             ], limit=1)
    #             if not employee:
    #                 raise UserError(_("Row %d: No employee found with code '%s' and name '%s'.") %
    #                                 (row_number, employee_code_str, employee_name))
    #
    #             # === Convert Date ===
    #             formatted_date = False
    #             if isinstance(date_value, datetime):
    #                 formatted_date = date_value.date()
    #             elif isinstance(date_value, str):
    #                 for fmt in ("%d-%b-%Y", "%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d/%m/%y"):
    #                     try:
    #                         formatted_date = datetime.strptime(date_value, fmt).date()
    #                         break
    #                     except Exception:
    #                         continue
    #             if not formatted_date:
    #                 raise UserError(_("Row %d: Invalid date format '%s'.") % (row_number, date_value))
    #
    #             check_in = normalize_time(check_in)
    #             check_out = normalize_time(check_out)
    #
    #             # === Create hr.upload Record ===
    #             upload_rec = HrUpload.create({
    #                 'employee_name': employee.id,
    #                 'ctc_type': employee.ctc_type,
    #                 'employee_code': employee_code_str,
    #                 'date': formatted_date,
    #                 'check_in_attendance': check_in,
    #                 'check_out_attendance': check_out,
    #             })
    #
    #             if hasattr(upload_rec, '_compute_sessions'):
    #                 upload_rec._compute_sessions()
    #
    #             morning = getattr(upload_rec, 'morning_session', '') or ''
    #             afternoon = getattr(upload_rec, 'afternoon_session', '') or ''
    #             leave_duration = 0.0
    #             half_day = False
    #
    #             if morning == 'Absent' and afternoon == 'Absent':
    #                 leave_duration = 1.0
    #             elif morning == 'Absent' or afternoon == 'Absent':
    #                 leave_duration = 0.5
    #                 half_day = True
    #
    #             # === Leave Logic ===
    #             if leave_duration > 0:
    #                 existing_leave = Leave.search([
    #                     ('employee_id', '=', employee.id),
    #                     ('request_date_from', '=', formatted_date),
    #                     ('request_date_to', '=', formatted_date),
    #                     ('state', '!=', 'refuse')
    #                 ], limit=1)
    #                 if existing_leave:
    #                     continue
    #
    #                 # === For CTC employees, assign LOP directly ===
    #                 if employee.ctc_type != 'non_ctc':
    #                     leave_type_to_use = lop_type
    #                     remaining_el = 0.0
    #                 else:
    #                     # === Dynamic Monthly Accrual from Accrual Plan ===
    #                     el_allocations = Allocation.search([
    #                         ('employee_id', '=', employee.id),
    #                         ('holiday_status_id', '=', el_type.id),
    #                         ('state', '=', 'validate')
    #                     ], limit=1)
    #
    #                     remaining_el = 0.0
    #                     if el_allocations:
    #                         alloc = el_allocations[0]
    #                         start_date = alloc.date_from
    #                         accrual_per_month = 0.0
    #
    #                         # 🔹 FIXED FIELD NAME: use 'accrual_plan_id' instead of 'plan_id'
    #                         if alloc.accrual_plan_id:
    #                             accrual_level = self.env['hr.leave.accrual.level'].search([
    #                                 ('accrual_plan_id', '=', alloc.accrual_plan_id.id)
    #                             ], limit=1)
    #                             if accrual_level:
    #                                 accrual_per_month = accrual_level.added_value
    #
    #                         if not accrual_per_month:
    #                             raise UserError(_(
    #                                 "Row %d: No accrual value defined in Accrual Plan for %s."
    #                             ) % (row_number, employee.name))
    #
    #                         if start_date and formatted_date >= start_date:
    #                             months_passed = (formatted_date.year - start_date.year) * 12 + (
    #                                     formatted_date.month - start_date.month) + 1
    #                             total_allocated = months_passed * accrual_per_month
    #                         else:
    #                             total_allocated = 0.0
    #
    #                         el_taken = sum(Leave.search([
    #                             ('employee_id', '=', employee.id),
    #                             ('holiday_status_id', '=', el_type.id),
    #                             ('state', 'in', ['validate', 'confirm']),
    #                             ('request_date_from', '<=', formatted_date)
    #                         ]).mapped('number_of_days'))
    #
    #                         remaining_el = max(total_allocated - el_taken, 0.0)
    #                     else:
    #                         raise UserError(_(
    #                             "Row %d: No validated EL allocation found for %s."
    #                         ) % (row_number, employee.name))
    #
    #                 def create_and_validate_leave(leave_type, days, half_day=False, period=False):
    #                     vals = {
    #                         'name': f"Auto Leave {formatted_date}",
    #                         'employee_id': employee.id,
    #                         'holiday_status_id': leave_type.id,
    #                         'request_date_from': formatted_date,
    #                         'request_date_to': formatted_date,
    #                         'number_of_days': days,
    #                     }
    #                     if half_day:
    #                         vals.update({
    #                             'request_unit_half': True,
    #                             'request_date_from_period': period,
    #                         })
    #                     leave_rec = Leave.create(vals)
    #                     if leave_rec.state == 'draft':
    #                         try:
    #                             leave_rec.action_confirm()
    #                         except Exception:
    #                             pass
    #                     if leave_rec.state in ['confirm', 'validate1', 'validate']:
    #                         try:
    #                             leave_rec.action_validate()
    #                         except Exception:
    #                             pass
    #
    #                 # === Apply ELs and LOPs ===
    #                 if employee.ctc_type == 'non_ctc':
    #                     if remaining_el >= leave_duration:
    #                         create_and_validate_leave(
    #                             el_type, leave_duration, half_day,
    #                             'am' if half_day and morning == 'Absent' else 'pm' if half_day else False
    #                         )
    #                     elif remaining_el > 0:
    #                         create_and_validate_leave(el_type, remaining_el)
    #                         create_and_validate_leave(lop_type, leave_duration - remaining_el)
    #                     else:
    #                         create_and_validate_leave(
    #                             lop_type, leave_duration, half_day,
    #                             'am' if half_day and morning == 'Absent' else 'pm' if half_day else False
    #                         )
    #                 else:
    #                     create_and_validate_leave(
    #                         lop_type, leave_duration, half_day,
    #                         'am' if half_day and morning == 'Absent' else 'pm' if half_day else False
    #                     )
    #
    #             # === Attendance creation ===
    #             if check_in and check_out:
    #                 def to_datetime_local(date_part, time_value):
    #                     if isinstance(time_value, str):
    #                         try:
    #                             t = datetime.strptime(time_value, "%H:%M").time()
    #                             local_dt = datetime.combine(date_part, t)
    #                         except Exception:
    #                             return False
    #                     else:
    #                         return False
    #                     local_dt = tz.localize(local_dt)
    #                     return local_dt.astimezone(pytz.UTC)
    #
    #                 check_in_dt = to_datetime_local(formatted_date, check_in)
    #                 check_out_dt = to_datetime_local(formatted_date, check_out)
    #
    #                 if not (check_in_dt and check_out_dt):
    #                     continue
    #
    #                 open_attendance = Attendance.search([
    #                     ('employee_id', '=', employee.id),
    #                     ('check_out', '=', False)
    #                 ], limit=1)
    #
    #                 if open_attendance:
    #                     open_attendance.write({'check_out': fields.Datetime.to_string(check_out_dt)})
    #                 else:
    #                     Attendance.create({
    #                         'employee_id': employee.id,
    #                         'date': formatted_date,
    #                         'check_in': fields.Datetime.to_string(check_in_dt),
    #                         'check_out': fields.Datetime.to_string(check_out_dt),
    #                     })
    #
    #     except Exception as e:
    #         raise UserError(_('Upload failed: %s') % e)

    # perfect code1
    # def action_upload(self):
    #     """Uploads attendance Excel file and creates hr.upload, hr.attendance, and hr.leave records with EL/LOP rules."""
    #     if not self.file:
    #         raise UserError(_("Please upload a file."))
    #
    #     try:
    #         import base64
    #         import openpyxl
    #         from io import BytesIO
    #         from datetime import datetime, date, time, timedelta
    #         import pytz
    #
    #         wb = openpyxl.load_workbook(
    #             filename=BytesIO(base64.b64decode(self.file)),
    #             read_only=True
    #         )
    #         ws = wb.active
    #
    #         hr_upload_obj = self.env['hr.upload']
    #         employee_obj = self.env['hr.employee']
    #         attendance_obj = self.env['hr.attendance']
    #         leave_obj = self.env['hr.leave']
    #         allocation_obj = self.env['hr.leave.allocation']
    #         leave_type_obj = self.env['hr.leave.type']
    #
    #         # === Get Leave Types by Work Entry Code ===
    #         el_type = leave_type_obj.search([('work_entry_type_id.code', '=', 'LEAVE120')], limit=1)
    #         lop_type = leave_type_obj.search([('work_entry_type_id.code', '=', 'LEAVE90')], limit=1)
    #         if not el_type or not lop_type:
    #             raise UserError(
    #                 _("Please configure leave types with work entry codes 'LEAVE120' (EL) and 'LEAVE90' (LOP)."))
    #
    #         user_tz = self.env.user.tz or 'Asia/Kolkata'
    #         tz = pytz.timezone(user_tz)
    #
    #         def normalize_time(time_value):
    #             """Converts Excel time to HH:MM."""
    #             if not time_value:
    #                 return None
    #             try:
    #                 if isinstance(time_value, (datetime, time)):
    #                     return time_value.strftime("%H:%M")
    #                 if isinstance(time_value, (int, float)):
    #                     from openpyxl.utils.datetime import from_excel
    #                     dt = from_excel(time_value)
    #                     return dt.strftime("%H:%M")
    #                 time_str = str(time_value).strip().replace(';', ':').replace('.', ':')
    #                 parts = time_str.split(':')
    #                 if len(parts) >= 2:
    #                     return f"{parts[0].zfill(2)}:{parts[1].zfill(2)}"
    #                 return time_str
    #             except Exception:
    #                 return None
    #
    #         row_number = 1
    #         for row in ws.iter_rows(min_row=2, values_only=True):
    #             row_number += 1
    #
    #             employee_code = row[0]
    #             employee_name = row[1]
    #             date_value = row[2]
    #             check_in = row[3]
    #             check_out = row[4]
    #
    #             if not employee_code or not employee_name:
    #                 continue
    #
    #             employee_name = str(employee_name).strip()
    #             try:
    #                 employee_code_str = str(int(employee_code))
    #             except Exception:
    #                 employee_code_str = str(employee_code).strip()
    #
    #             employee = employee_obj.search([
    #                 ('cmr_code', '=', employee_code_str),
    #                 ('name', '=', employee_name)
    #             ], limit=1)
    #             if not employee:
    #                 raise UserError(_(
    #                     "Row %d: No employee found with code '%s' and name '%s'."
    #                 ) % (row_number, employee_code_str, employee_name))
    #
    #             # ===== Date Conversion =====
    #             formatted_date = False
    #             if isinstance(date_value, datetime):
    #                 formatted_date = date_value.date()
    #             elif isinstance(date_value, str):
    #                 for fmt in ("%d-%b-%Y", "%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d/%m/%y"):
    #                     try:
    #                         formatted_date = datetime.strptime(date_value, fmt).date()
    #                         break
    #                     except Exception:
    #                         continue
    #             if not formatted_date:
    #                 raise UserError(_("Row %d: Invalid date format '%s'.") % (row_number, date_value))
    #
    #             check_in = normalize_time(check_in)
    #             check_out = normalize_time(check_out)
    #
    #             # ===== Create hr.upload Record =====
    #             upload_rec = hr_upload_obj.create({
    #                 'employee_name': employee.id,
    #                 'ctc_type': employee.ctc_type,
    #                 'employee_code': employee_code_str,
    #                 'date': formatted_date,
    #                 'check_in_attendance': check_in,
    #                 'check_out_attendance': check_out,
    #             })
    #
    #             if hasattr(upload_rec, '_compute_sessions'):
    #                 upload_rec._compute_sessions()
    #
    #             morning = getattr(upload_rec, 'morning_session', '') or ''
    #             afternoon = getattr(upload_rec, 'afternoon_session', '') or ''
    #             leave_duration = 0.0
    #             half_day = False
    #
    #             if morning == 'Absent' and afternoon == 'Absent':
    #                 leave_duration = 1.0
    #             elif morning == 'Absent' or afternoon == 'Absent':
    #                 leave_duration = 0.5
    #                 half_day = True
    #
    #             # ===== Leave Creation Logic =====
    #             if leave_duration > 0:
    #                 existing_leave = leave_obj.search([
    #                     ('employee_id', '=', employee.id),
    #                     ('request_date_from', '=', formatted_date),
    #                     ('request_date_to', '=', formatted_date),
    #                     ('state', '!=', 'refuse')
    #                 ], limit=1)
    #                 if existing_leave:
    #                     continue
    #
    #                 # --- Calculate EL balance based on attendance month ---
    #                 el_allocations = allocation_obj.search([
    #                     ('employee_id', '=', employee.id),
    #                     ('holiday_status_id', '=', el_type.id),
    #                     ('state', '=', 'validate')
    #                 ], limit=1)
    #
    #                 remaining_el = 0.0
    #                 if el_allocations:
    #                     alloc = el_allocations[0]
    #                     start_date = alloc.date_from
    #                     if start_date and formatted_date >= start_date:
    #                         months_passed = (formatted_date.year - start_date.year) * 12 + (
    #                                 formatted_date.month - start_date.month
    #                         ) + 1
    #                         total_allocated = months_passed * 3
    #                     else:
    #                         total_allocated = 0.0
    #
    #                     el_taken = sum(leave_obj.search([
    #                         ('employee_id', '=', employee.id),
    #                         ('holiday_status_id', '=', el_type.id),
    #                         ('state', 'in', ['validate', 'confirm']),
    #                         ('request_date_from', '<=', formatted_date)
    #                     ]).mapped('number_of_days'))
    #
    #                     remaining_el = max(total_allocated - el_taken, 0.0)
    #
    #                 # === CTC Logic ===
    #                 # Non-CTC → use EL & LOP logic
    #                 # CTC → all LOP
    #                 if employee.ctc_type != 'non_ctc':
    #                     remaining_el = 0.0
    #
    #                 def create_and_validate_leave(leave_type, days, half_day=False, period=False):
    #                     vals = {
    #                         'name': f"Auto Leave {formatted_date}",
    #                         'employee_id': employee.id,
    #                         'holiday_status_id': leave_type.id,
    #                         'request_date_from': formatted_date,
    #                         'request_date_to': formatted_date,
    #                         'number_of_days': days,
    #                     }
    #                     if half_day:
    #                         vals.update({
    #                             'request_unit_half': True,
    #                             'request_date_from_period': period,
    #                         })
    #                     leave_rec = leave_obj.create(vals)
    #                     if leave_rec.state == 'draft':
    #                         try:
    #                             leave_rec.action_confirm()
    #                         except Exception:
    #                             pass
    #                     if leave_rec.state in ['confirm', 'validate1', 'validate']:
    #                         try:
    #                             leave_rec.action_validate()
    #                         except Exception:
    #                             pass
    #
    #                 # --- Apply EL or LOP based on balance ---
    #                 if remaining_el >= leave_duration:
    #                     create_and_validate_leave(
    #                         el_type, leave_duration, half_day,
    #                         'am' if half_day and morning == 'Absent' else 'pm' if half_day else False
    #                     )
    #                 elif remaining_el > 0:
    #                     create_and_validate_leave(el_type, remaining_el)
    #                     create_and_validate_leave(lop_type, leave_duration - remaining_el)
    #                 else:
    #                     create_and_validate_leave(
    #                         lop_type, leave_duration, half_day,
    #                         'am' if half_day and morning == 'Absent' else 'pm' if half_day else False
    #                     )
    #
    #             # ===== Create Attendance =====
    #             if check_in and check_out:
    #                 def to_datetime_local(date_part, time_value):
    #                     if isinstance(time_value, str):
    #                         try:
    #                             t = datetime.strptime(time_value, "%H:%M").time()
    #                             local_dt = datetime.combine(date_part, t)
    #                         except Exception:
    #                             return False
    #                     else:
    #                         return False
    #                     local_dt = tz.localize(local_dt)
    #                     return local_dt.astimezone(pytz.UTC)
    #
    #                 check_in_dt = to_datetime_local(formatted_date, check_in)
    #                 check_out_dt = to_datetime_local(formatted_date, check_out)
    #                 if not (check_in_dt and check_out_dt):
    #                     continue
    #
    #                 open_attendance = attendance_obj.search([
    #                     ('employee_id', '=', employee.id),
    #                     ('check_out', '=', False)
    #                 ], limit=1)
    #                 if open_attendance:
    #                     open_attendance.write({'check_out': fields.Datetime.to_string(check_out_dt)})
    #                 else:
    #                     attendance_obj.create({
    #                         'employee_id': employee.id,
    #                         'date': formatted_date,
    #                         'check_in': fields.Datetime.to_string(check_in_dt),
    #                         'check_out': fields.Datetime.to_string(check_out_dt),
    #                     })
    #
    #         # --- No popup, just finish ---
    #         return True
    #
    #     except Exception as e:
    #         raise UserError(_('Upload failed: %s') % e)

    # perfect code
    # def action_upload(self):
    #     """Uploads attendance Excel file and creates hr.upload, hr.attendance, and hr.leave records with EL/LOP rules."""
    #     if not self.file:
    #         raise UserError(_("Please upload a file."))
    #
    #     try:
    #         import base64
    #         import openpyxl
    #         from io import BytesIO
    #         from datetime import datetime, date, time, timedelta
    #         import pytz
    #
    #         wb = openpyxl.load_workbook(
    #             filename=BytesIO(base64.b64decode(self.file)),
    #             read_only=True
    #         )
    #         ws = wb.active
    #
    #         hr_upload_obj = self.env['hr.upload']
    #         employee_obj = self.env['hr.employee']
    #         attendance_obj = self.env['hr.attendance']
    #         leave_obj = self.env['hr.leave']
    #         allocation_obj = self.env['hr.leave.allocation']
    #
    #         # === Verify Leave Types Exist ===
    #         el_type = self.env['hr.leave.type'].search([('name', '=', 'EL')], limit=1)
    #         lop_type = self.env['hr.leave.type'].search([('name', '=', 'LOP')], limit=1)
    #         if not el_type or not lop_type:
    #             raise UserError(_("Please configure both 'EL' and 'LOP' Time Off Types."))
    #
    #         user_tz = self.env.user.tz or 'Asia/Kolkata'
    #         tz = pytz.timezone(user_tz)
    #
    #         def normalize_time(time_value):
    #             """Removes seconds if present, handles Excel float/time objects."""
    #             if not time_value:
    #                 return None
    #             try:
    #                 if isinstance(time_value, (datetime, time)):
    #                     return time_value.strftime("%H:%M")
    #                 if isinstance(time_value, (int, float)):
    #                     from openpyxl.utils.datetime import from_excel
    #                     dt = from_excel(time_value)
    #                     return dt.strftime("%H:%M")
    #                 time_str = str(time_value).strip().replace(';', ':').replace('.', ':')
    #                 parts = time_str.split(':')
    #                 if len(parts) >= 2:
    #                     return f"{parts[0].zfill(2)}:{parts[1].zfill(2)}"
    #                 return time_str
    #             except Exception:
    #                 return None
    #
    #         row_number = 1
    #         for row in ws.iter_rows(min_row=2, values_only=True):
    #             row_number += 1
    #
    #             employee_code = row[0]
    #             employee_name = row[1]
    #             date_value = row[2]
    #             check_in = row[3]
    #             check_out = row[4]
    #
    #             if not employee_code:
    #                 raise UserError(_("Row %d: Employee code is missing.") % row_number)
    #             if not employee_name:
    #                 raise UserError(_("Row %d: Employee name is missing.") % row_number)
    #
    #             employee_name = str(employee_name).strip()
    #             try:
    #                 employee_code_str = str(int(employee_code))
    #             except Exception:
    #                 employee_code_str = str(employee_code).strip()
    #
    #             employee = employee_obj.search([
    #                 ('cmr_code', '=', employee_code_str),
    #                 ('name', '=', employee_name)
    #             ], limit=1)
    #             if not employee:
    #                 raise UserError(_(
    #                     "Row %d: No employee found with code '%s' and name '%s'."
    #                 ) % (row_number, employee_code_str, employee_name))
    #
    #             # ===== Date Conversion =====
    #             formatted_date = False
    #             if isinstance(date_value, datetime):
    #                 formatted_date = date_value.date()
    #             elif isinstance(date_value, str):
    #                 for fmt in ("%d-%b-%Y", "%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d/%m/%y"):
    #                     try:
    #                         formatted_date = datetime.strptime(date_value, fmt).date()
    #                         break
    #                     except Exception:
    #                         continue
    #             if not formatted_date:
    #                 raise UserError(_("Row %d: Invalid date format '%s'.") % (row_number, date_value))
    #
    #             check_in = normalize_time(check_in)
    #             check_out = normalize_time(check_out)
    #
    #             # ===== Create hr.upload Record =====
    #             upload_rec = hr_upload_obj.create({
    #                 'employee_name': employee.id,
    #                 'ctc_type': employee.ctc_type,
    #                 'employee_code': employee_code_str,
    #                 'date': formatted_date,
    #                 'check_in_attendance': check_in,
    #                 'check_out_attendance': check_out,
    #             })
    #
    #             if hasattr(upload_rec, '_compute_sessions'):
    #                 upload_rec._compute_sessions()
    #
    #             morning = getattr(upload_rec, 'morning_session', '') or ''
    #             afternoon = getattr(upload_rec, 'afternoon_session', '') or ''
    #             leave_duration = 0.0
    #             half_day = False
    #
    #             if morning == 'Absent' and afternoon == 'Absent':
    #                 leave_duration = 1.0
    #             elif morning == 'Absent' or afternoon == 'Absent':
    #                 leave_duration = 0.5
    #                 half_day = True
    #
    #             # ===== Leave Creation Logic =====
    #             if leave_duration > 0:
    #                 existing_leave = leave_obj.search([
    #                     ('employee_id', '=', employee.id),
    #                     ('request_date_from', '=', formatted_date),
    #                     ('request_date_to', '=', formatted_date),
    #                     ('state', '!=', 'refuse')
    #                 ], limit=1)
    #                 if existing_leave:
    #                     continue
    #
    #                 # --- Calculate EL balance based on attendance month ---
    #                 el_allocations = allocation_obj.search([
    #                     ('employee_id', '=', employee.id),
    #                     ('holiday_status_id', '=', el_type.id),
    #                     ('state', '=', 'validate')
    #                 ], limit=1)
    #
    #                 remaining_el = 0.0
    #                 if el_allocations:
    #                     alloc = el_allocations[0]
    #                     start_date = alloc.date_from
    #                     if start_date and formatted_date >= start_date:
    #                         # Calculate how many months passed between start and attendance date
    #                         months_passed = (formatted_date.year - start_date.year) * 12 + (
    #                                     formatted_date.month - start_date.month) + 1
    #                         # Each month accrues 3 ELs
    #                         total_allocated = months_passed * 3
    #                     else:
    #                         total_allocated = 0.0
    #
    #                     # Calculate ELs already taken till that date
    #                     el_taken = sum(leave_obj.search([
    #                         ('employee_id', '=', employee.id),
    #                         ('holiday_status_id', '=', el_type.id),
    #                         ('state', 'in', ['validate', 'confirm']),
    #                         ('request_date_from', '<=', formatted_date)
    #                     ]).mapped('number_of_days'))
    #
    #                     remaining_el = max(total_allocated - el_taken, 0.0)
    #
    #                 def create_and_validate_leave(leave_type, days, half_day=False, period=False):
    #                     vals = {
    #                         'name': f"Auto Leave {formatted_date}",
    #                         'employee_id': employee.id,
    #                         'holiday_status_id': leave_type.id,
    #                         'request_date_from': formatted_date,
    #                         'request_date_to': formatted_date,
    #                         'number_of_days': days,
    #                     }
    #                     if half_day:
    #                         vals.update({
    #                             'request_unit_half': True,
    #                             'request_date_from_period': period,
    #                         })
    #                     leave_rec = leave_obj.create(vals)
    #
    #                     # Safe state transitions
    #                     if leave_rec.state == 'draft':
    #                         try:
    #                             leave_rec.action_confirm()
    #                         except Exception:
    #                             pass
    #                     if leave_rec.state in ['confirm', 'validate1', 'validate']:
    #                         try:
    #                             leave_rec.action_validate()
    #                         except Exception:
    #                             pass
    #
    #                 # --- Apply EL or LOP based on month-based balance ---
    #                 if remaining_el >= leave_duration:
    #                     create_and_validate_leave(
    #                         el_type, leave_duration, half_day,
    #                         'am' if half_day and morning == 'Absent' else 'pm' if half_day else False
    #                     )
    #                 elif remaining_el > 0:
    #                     create_and_validate_leave(el_type, remaining_el)
    #                     create_and_validate_leave(lop_type, leave_duration - remaining_el)
    #                 else:
    #                     create_and_validate_leave(
    #                         lop_type, leave_duration, half_day,
    #                         'am' if half_day and morning == 'Absent' else 'pm' if half_day else False
    #                     )
    #
    #             # ===== Create Attendance =====
    #             if check_in and check_out:
    #                 def to_datetime_local(date_part, time_value):
    #                     if isinstance(time_value, str):
    #                         try:
    #                             t = datetime.strptime(time_value, "%H:%M").time()
    #                             local_dt = datetime.combine(date_part, t)
    #                         except Exception:
    #                             return False
    #                     else:
    #                         return False
    #                     local_dt = tz.localize(local_dt)
    #                     return local_dt.astimezone(pytz.UTC)
    #
    #                 check_in_dt = to_datetime_local(formatted_date, check_in)
    #                 check_out_dt = to_datetime_local(formatted_date, check_out)
    #
    #                 if not (check_in_dt and check_out_dt):
    #                     continue
    #
    #                 open_attendance = attendance_obj.search([
    #                     ('employee_id', '=', employee.id),
    #                     ('check_out', '=', False)
    #                 ], limit=1)
    #
    #                 if open_attendance:
    #                     open_attendance.write({'check_out': fields.Datetime.to_string(check_out_dt)})
    #                 else:
    #                     attendance_obj.create({
    #                         'employee_id': employee.id,
    #                         'date': formatted_date,
    #                         'check_in': fields.Datetime.to_string(check_in_dt),
    #                         'check_out': fields.Datetime.to_string(check_out_dt),
    #                     })
    #
    #         # --- Optional success popup ---
    #         return {
    #             'type': 'ir.actions.client',
    #             'tag': 'display_notification',
    #             'params': {
    #                 'title': _("Success"),
    #                 'message': _("Attendance and leave records uploaded successfully."),
    #                 'type': 'success',
    #                 'sticky': False,
    #             },
    #         }
    #
    #     except Exception as e:
    #         raise UserError(_('Upload failed: %s') % e)

    # Now
    # def action_upload(self):
    #     """Uploads attendance Excel file and creates hr.upload, hr.attendance, and hr.leave records with EL/LOP rules."""
    #     if not self.file:
    #         raise UserError(_("Please upload a file."))
    #
    #     try:
    #         import base64
    #         import openpyxl
    #         from io import BytesIO
    #         from datetime import datetime, date, time, timedelta
    #         from dateutil.relativedelta import relativedelta
    #         import pytz
    #
    #         wb = openpyxl.load_workbook(
    #             filename=BytesIO(base64.b64decode(self.file)),
    #             read_only=True
    #         )
    #         ws = wb.active
    #
    #         hr_upload_obj = self.env['hr.upload']
    #         employee_obj = self.env['hr.employee']
    #         attendance_obj = self.env['hr.attendance']
    #         leave_obj = self.env['hr.leave']
    #         allocation_obj = self.env['hr.leave.allocation']
    #
    #         # === Verify Leave Types Exist ===
    #         el_type = self.env['hr.leave.type'].search([('name', '=', 'EL')], limit=1)
    #         lop_type = self.env['hr.leave.type'].search([('name', '=', 'LOP')], limit=1)
    #         if not el_type or not lop_type:
    #             raise UserError(_("Please configure both 'EL' and 'LOP' Time Off Types."))
    #
    #         user_tz = self.env.user.tz or 'Asia/Kolkata'
    #         tz = pytz.timezone(user_tz)
    #
    #         def normalize_time(time_value):
    #             """Removes seconds if present, handles Excel float/time objects."""
    #             if not time_value:
    #                 return None
    #             try:
    #                 if isinstance(time_value, (datetime, time)):
    #                     return time_value.strftime("%H:%M")
    #                 if isinstance(time_value, (int, float)):
    #                     from openpyxl.utils.datetime import from_excel
    #                     dt = from_excel(time_value)
    #                     return dt.strftime("%H:%M")
    #                 time_str = str(time_value).strip().replace(';', ':').replace('.', ':')
    #                 parts = time_str.split(':')
    #                 if len(parts) >= 2:
    #                     return f"{parts[0].zfill(2)}:{parts[1].zfill(2)}"
    #                 return time_str
    #             except Exception:
    #                 return None
    #
    #         row_number = 1
    #         for row in ws.iter_rows(min_row=2, values_only=True):
    #             row_number += 1
    #
    #             employee_code = row[0]
    #             employee_name = row[1]
    #             date_value = row[2]
    #             check_in = row[3]
    #             check_out = row[4]
    #
    #             if not employee_code:
    #                 raise UserError(_("Row %d: Employee code is missing.") % row_number)
    #             if not employee_name:
    #                 raise UserError(_("Row %d: Employee name is missing.") % row_number)
    #
    #             employee_name = str(employee_name).strip()
    #             try:
    #                 employee_code_str = str(int(employee_code))
    #             except Exception:
    #                 employee_code_str = str(employee_code).strip()
    #
    #             employee = employee_obj.search([
    #                 ('cmr_code', '=', employee_code_str),
    #                 ('name', '=', employee_name)
    #             ], limit=1)
    #             if not employee:
    #                 raise UserError(_(
    #                     "Row %d: No employee found with code '%s' and name '%s'."
    #                 ) % (row_number, employee_code_str, employee_name))
    #
    #             # ===== Date Conversion =====
    #             formatted_date = False
    #             if isinstance(date_value, datetime):
    #                 formatted_date = date_value.date()
    #             elif isinstance(date_value, str):
    #                 for fmt in ("%d-%b-%Y", "%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d/%m/%y"):
    #                     try:
    #                         formatted_date = datetime.strptime(date_value, fmt).date()
    #                         break
    #                     except Exception:
    #                         continue
    #             if not formatted_date:
    #                 raise UserError(_("Row %d: Invalid date format '%s'.") % (row_number, date_value))
    #
    #             check_in = normalize_time(check_in)
    #             check_out = normalize_time(check_out)
    #
    #             # ===== Create hr.upload Record =====
    #             upload_rec = hr_upload_obj.create({
    #                 'employee_name': employee.id,
    #                 'ctc_type': employee.ctc_type,
    #                 'employee_code': employee_code_str,
    #                 'date': formatted_date,
    #                 'check_in_attendance': check_in,
    #                 'check_out_attendance': check_out,
    #             })
    #
    #             if hasattr(upload_rec, '_compute_sessions'):
    #                 upload_rec._compute_sessions()
    #
    #             morning = getattr(upload_rec, 'morning_session', '') or ''
    #             afternoon = getattr(upload_rec, 'afternoon_session', '') or ''
    #             leave_duration = 0.0
    #             half_day = False
    #
    #             if morning == 'Absent' and afternoon == 'Absent':
    #                 leave_duration = 1.0
    #             elif morning == 'Absent' or afternoon == 'Absent':
    #                 leave_duration = 0.5
    #                 half_day = True
    #
    #             # ===== Leave Creation Logic =====
    #             if leave_duration > 0:
    #                 existing_leave = leave_obj.search([
    #                     ('employee_id', '=', employee.id),
    #                     ('request_date_from', '=', formatted_date),
    #                     ('request_date_to', '=', formatted_date),
    #                     ('state', '!=', 'refuse')
    #                 ], limit=1)
    #                 if existing_leave:
    #                     continue
    #
    #                 # --- Fetch available EL balance dynamically ---
    #                 el_allocations = allocation_obj.search([
    #                     ('employee_id', '=', employee.id),
    #                     ('holiday_status_id', '=', el_type.id),
    #                     ('state', '=', 'validate')
    #                 ])
    #                 total_allocated = sum(a.number_of_days_display for a in el_allocations)
    #                 el_taken = sum(leave_obj.search([
    #                     ('employee_id', '=', employee.id),
    #                     ('holiday_status_id', '=', el_type.id),
    #                     ('state', 'in', ['validate', 'confirm'])
    #                 ]).mapped('number_of_days'))
    #                 remaining_el = total_allocated - el_taken
    #
    #                 def create_and_validate_leave(leave_type, days, half_day=False, period=False):
    #                     vals = {
    #                         'name': f"Auto Leave {formatted_date}",
    #                         'employee_id': employee.id,
    #                         'holiday_status_id': leave_type.id,
    #                         'request_date_from': formatted_date,
    #                         'request_date_to': formatted_date,
    #                         'number_of_days': days,
    #                     }
    #                     if half_day:
    #                         vals.update({
    #                             'request_unit_half': True,
    #                             'request_date_from_period': period,
    #                         })
    #
    #                     leave_rec = leave_obj.create(vals)
    #
    #                     # --- Handle state transitions safely ---
    #                     if leave_rec.state == 'draft':
    #                         try:
    #                             leave_rec.action_confirm()
    #                         except Exception:
    #                             pass
    #
    #                     if leave_rec.state in ['confirm', 'validate1', 'validate']:
    #                         try:
    #                             leave_rec.action_validate()
    #                         except Exception:
    #                             pass
    #
    #                 # --- Apply EL or LOP based on remaining balance ---
    #                 if remaining_el >= leave_duration:
    #                     create_and_validate_leave(
    #                         el_type, leave_duration, half_day,
    #                         'am' if half_day and morning == 'Absent' else 'pm' if half_day else False
    #                     )
    #                 elif remaining_el > 0:
    #                     create_and_validate_leave(el_type, remaining_el)
    #                     create_and_validate_leave(lop_type, leave_duration - remaining_el)
    #                 else:
    #                     create_and_validate_leave(
    #                         lop_type, leave_duration, half_day,
    #                         'am' if half_day and morning == 'Absent' else 'pm' if half_day else False
    #                     )
    #
    #             # ===== Create Attendance =====
    #             if check_in and check_out:
    #                 def to_datetime_local(date_part, time_value):
    #                     if isinstance(time_value, str):
    #                         try:
    #                             t = datetime.strptime(time_value, "%H:%M").time()
    #                             local_dt = datetime.combine(date_part, t)
    #                         except Exception:
    #                             return False
    #                     else:
    #                         return False
    #                     local_dt = tz.localize(local_dt)
    #                     return local_dt.astimezone(pytz.UTC)
    #
    #                 check_in_dt = to_datetime_local(formatted_date, check_in)
    #                 check_out_dt = to_datetime_local(formatted_date, check_out)
    #
    #                 if not (check_in_dt and check_out_dt):
    #                     continue
    #
    #                 open_attendance = attendance_obj.search([
    #                     ('employee_id', '=', employee.id),
    #                     ('check_out', '=', False)
    #                 ], limit=1)
    #
    #                 if open_attendance:
    #                     open_attendance.write({'check_out': fields.Datetime.to_string(check_out_dt)})
    #                 else:
    #                     attendance_obj.create({
    #                         'employee_id': employee.id,
    #                         'date': formatted_date,
    #                         'check_in': fields.Datetime.to_string(check_in_dt),
    #                         'check_out': fields.Datetime.to_string(check_out_dt),
    #                     })
    #
    #     except Exception as e:
    #         raise UserError(_('Upload failed: %s') % e)







    # keerthana
    # def action_upload(self):
    #     """Upload attendance Excel file, create attendance, then create EL/LOP based on accrual plan levels and new rules."""
    #
    #     if not self.file:
    #         raise UserError(_("Please upload a file."))
    #
    #     try:
    #         file_data = base64.b64decode(self.file)
    #         excel_file = io.BytesIO(file_data)
    #         wb = openpyxl.load_workbook(excel_file, read_only=True)
    #         ws = wb.active
    #     except Exception as e:
    #         raise UserError(_("Invalid Excel file. Please upload a valid .xlsx file.\n\nError: %s") % e)
    #
    #     el_type = self.env['hr.leave.type'].search([('name', '=', 'EL')], limit=1)
    #     lop_type = self.env['hr.leave.type'].search([('name', '=', 'LOP')], limit=1)
    #     if not el_type or not lop_type:
    #         raise UserError(_("Please configure both 'EL' and 'LOP' Time Off Types."))
    #
    #     def combine_datetime(date_part, time_value):
    #         if not time_value:
    #             return False
    #         if isinstance(time_value, datetime):
    #             return time_value
    #         elif isinstance(time_value, (float, int)):
    #             from openpyxl.utils.datetime import from_excel
    #             dt = from_excel(time_value)
    #             return datetime.combine(date_part, dt.time())
    #         elif isinstance(time_value, str):
    #             for fmt in ("%H:%M:%S", "%H:%M"):
    #                 try:
    #                     t = datetime.strptime(time_value.strip(), fmt).time()
    #                     return datetime.combine(date_part, t)
    #                 except Exception:
    #                     continue
    #             return False
    #         else:
    #             return False
    #
    #     today = fields.Date.context_today(self)
    #
    #     row_number = 1
    #     for row in ws.iter_rows(min_row=2, values_only=True):
    #         row_number += 1
    #
    #         employee_code = row[0]
    #         employee_name = row[1]
    #         date_value = row[2]
    #         check_in = row[3]
    #         check_out = row[4]
    #
    #         if not employee_code or not employee_name:
    #             raise UserError(_("Row %d: Missing Employee details.") % row_number)
    #
    #         employee_name = str(employee_name).strip()
    #         try:
    #             employee_code_str = str(int(employee_code))
    #         except Exception:
    #             employee_code_str = str(employee_code).strip()
    #
    #         employee = self.env['hr.employee'].search([
    #             ('cmr_code', '=', employee_code_str),
    #             ('name', '=', employee_name)
    #         ], limit=1)
    #
    #         if not employee:
    #             raise UserError(_("Row %d: No employee found with code '%s' and name '%s'.") %
    #                             (row_number, employee_code_str, employee_name))
    #
    #         if isinstance(date_value, datetime):
    #             formatted_date = date_value.date()
    #         elif isinstance(date_value, str):
    #             formatted_date = False
    #             for fmt in ("%d-%b-%Y", "%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d/%m/%y"):
    #                 try:
    #                     formatted_date = datetime.strptime(date_value, fmt).date()
    #                     break
    #                 except Exception:
    #                     continue
    #             if not formatted_date:
    #                 raise UserError(_("Row %d: Invalid date format '%s'.") % (row_number, date_value))
    #         else:
    #             continue
    #
    #         check_in_dt = combine_datetime(formatted_date, check_in)
    #         check_out_dt = combine_datetime(formatted_date, check_out)
    #
    #         user_tz = self.env.user.tz or 'Asia/Kolkata'
    #         import pytz
    #         tz = pytz.timezone(user_tz)
    #         check_in_utc = tz.localize(check_in_dt).astimezone(pytz.UTC).replace(tzinfo=None) if check_in_dt else False
    #         check_out_utc = tz.localize(check_out_dt).astimezone(pytz.UTC).replace(
    #             tzinfo=None) if check_out_dt else False
    #
    #         upload_vals = {
    #             'employee_name': employee.id,
    #             'ctc_type': employee.ctc_type,
    #             'employee_code': employee_code_str,
    #             'date': formatted_date,
    #             'check_in_attendance': check_in,
    #             'check_out_attendance': check_out,
    #             'your_datetime': check_in_utc,
    #             'your_checkout_datetime': check_out_utc,
    #         }
    #         upload_rec = self.env['hr.upload'].create(upload_vals)
    #
    #         attendance_obj = self.env['hr.attendance']
    #         if check_in_utc and check_out_utc:
    #             open_att = attendance_obj.search([
    #                 ('employee_id', '=', employee.id),
    #                 ('check_out', '=', False)
    #             ], limit=1)
    #             if open_att:
    #                 open_att.write({
    #                     'check_out': fields.Datetime.to_string(check_out_utc),
    #                     'morning_session': upload_rec.morning_session,
    #                     'afternoon_session': upload_rec.afternoon_session,
    #                 })
    #             else:
    #                 attendance_obj.create({
    #                     'employee_id': employee.id,
    #                     'date': formatted_date,
    #                     'check_in': fields.Datetime.to_string(check_in_utc),
    #                     'check_out': fields.Datetime.to_string(check_out_utc),
    #                     'morning_session': upload_rec.morning_session,
    #                     'afternoon_session': upload_rec.afternoon_session,
    #                     'difference_check_in': upload_rec.difference_check_in,
    #                     'difference_check_out': upload_rec.difference_check_out,
    #                     'total_working_hours': upload_rec.total_working_hours,
    #                     'full_day_status': upload_rec.full_day_status,
    #                 })
    #
    #         morning = upload_rec.morning_session or ''
    #         afternoon = upload_rec.afternoon_session or ''
    #         leave_duration = 0.0
    #         half_day = False
    #         if morning == 'Absent' and afternoon == 'Absent':
    #             leave_duration = 1.0
    #         elif morning == 'Absent' or afternoon == 'Absent':
    #             leave_duration = 0.5
    #             half_day = True
    #
    #         if leave_duration <= 0:
    #             continue
    #
    #         allocation_obj = self.env['hr.leave.allocation']
    #         allocations = allocation_obj.search([
    #             ('employee_id', '=', employee.id),
    #             ('state', '=', 'validate'),
    #         ])
    #
    #         valid_alloc = False
    #         for alloc in allocations:
    #             start = alloc.date_from or date.min
    #             end = alloc.date_to or date.max
    #             if start <= formatted_date <= end:
    #                 valid_alloc = alloc
    #                 _logger.info("Matched Allocation: %s | %s -> %s | Days: %s", alloc.holiday_status_id.display_name,
    #                              start, end, alloc.number_of_days)
    #                 break
    #
    #         if not valid_alloc:
    #             def _create_and_validate(leave_type, days, half=False, period=False):
    #                 vals = {
    #                     'name': "Auto Leave %s" % formatted_date,
    #                     'employee_id': employee.id,
    #                     'holiday_status_id': leave_type.id,
    #                     'request_date_from': formatted_date,
    #                     'request_date_to': formatted_date,
    #                     'number_of_days': days,
    #                     'payslip_state': 'done',
    #                 }
    #                 if half:
    #                     vals.update({'request_unit_half': True, 'request_date_from_period': period})
    #                 l = self.env['hr.leave'].create(vals)
    #                 if l.state == 'draft':
    #                     try:
    #                         l.action_confirm()
    #                     except Exception as e:
    #                         _logger.warning("Could not confirm: %s", e)
    #                 if l.state in ['confirm', 'validate1']:
    #                     try:
    #                         l.action_validate()
    #                     except Exception as e:
    #                         _logger.warning("Could not validate: %s", e)
    #                 return l
    #
    #             _create_and_validate(lop_type, leave_duration, half_day,
    #                                  'am' if half_day and morning == 'Absent' else 'pm' if half_day else False)
    #             continue
    #
    #         plan = valid_alloc.accrual_plan_id
    #         total_earned_till_date = 0.0
    #
    #         if plan:
    #             Level = self.env['hr.leave.accrual.level']
    #             plan_levels = Level.search([('accrual_plan_id', '=', plan.id)], order='sequence asc')
    #             for level in plan_levels:
    #                 alloc_start = valid_alloc.date_from or formatted_date
    #                 if level.start_type == 'day':
    #                     start_date = alloc_start + timedelta(days=(level.start_count or 0))
    #                 elif level.start_type == 'month':
    #                     from dateutil.relativedelta import relativedelta
    #                     start_date = alloc_start + relativedelta(months=(level.start_count or 0))
    #                 elif level.start_type == 'year':
    #                     from dateutil.relativedelta import relativedelta
    #                     start_date = alloc_start + relativedelta(years=(level.start_count or 0))
    #                 else:
    #                     start_date = alloc_start
    #
    #                 if formatted_date < start_date:
    #                     continue
    #
    #                 freq = (level.frequency or 'daily')
    #                 added = float(level.added_value or 0.0)
    #                 if level.added_value_type == 'hour':
    #                     hours_per_day = 8.0
    #                     added_in_days = added / hours_per_day
    #                 else:
    #                     added_in_days = added
    #
    #                 if freq == 'daily':
    #                     days_elapsed = (formatted_date - start_date).days
    #                     intervals = days_elapsed + 1
    #                     earned = intervals * added_in_days
    #                 elif freq == 'weekly':
    #                     days_elapsed = (formatted_date - start_date).days
    #                     intervals = (days_elapsed // 7) + 1
    #                     earned = intervals * added_in_days
    #                 elif freq == 'monthly':
    #                     from dateutil.relativedelta import relativedelta
    #                     months_elapsed = (formatted_date.year - start_date.year) * 12 + (
    #                             formatted_date.month - start_date.month)
    #                     intervals = months_elapsed + 1
    #                     earned = intervals * added_in_days
    #                 elif freq == 'yearly':
    #                     years_elapsed = formatted_date.year - start_date.year
    #                     intervals = years_elapsed + 1
    #                     earned = intervals * added_in_days
    #                 else:
    #                     earned = 0.0
    #
    #                 if level.cap_accrued_time and level.maximum_leave:
    #                     earned = min(earned, float(level.maximum_leave or 0.0))
    #
    #                 total_earned_till_date += earned
    #
    #         if valid_alloc.number_of_days:
    #             total_earned_till_date = min(total_earned_till_date, float(valid_alloc.number_of_days))
    #
    #         taken_leaves = self.env['hr.leave'].search([
    #             ('employee_id', '=', employee.id),
    #             ('holiday_status_id', '=', valid_alloc.holiday_status_id.id),
    #             ('state', 'not in', ('refuse',)),
    #             ('request_date_from', '>=', valid_alloc.date_from) if valid_alloc.date_from else ('id', '!=', False)
    #         ])
    #         total_taken = sum(l.number_of_days for l in taken_leaves) if taken_leaves else 0.0
    #
    #         remaining_el = total_earned_till_date - total_taken
    #         if remaining_el < 0:
    #             remaining_el = 0.0
    #
    #         def create_and_validate_leave(leave_type, days, half=False, period=False):
    #             vals = {
    #                 'name': "Auto Leave %s" % formatted_date,
    #                 'employee_id': employee.id,
    #                 'holiday_status_id': leave_type.id,
    #                 'request_date_from': formatted_date,
    #                 'request_date_to': formatted_date,
    #                 'number_of_days': days,
    #                 'payslip_state': 'done',
    #             }
    #             if half:
    #                 vals.update({'request_unit_half': True, 'request_date_from_period': period})
    #             leave_rec = self.env['hr.leave'].create(vals)
    #             if leave_rec.state == 'draft':
    #                 try:
    #                     leave_rec.action_confirm()
    #                 except Exception as e:
    #                     _logger.warning("Could not confirm leave: %s", e)
    #             if leave_rec.state in ['confirm', 'validate1']:
    #                 try:
    #                     leave_rec.action_validate()
    #                 except Exception as e:
    #                     _logger.warning("Could not validate leave: %s", e)
    #             return leave_rec
    #
    #         # ======================================================
    #         # === NEW INTELLIGENT EL/LOP DECISION BLOCK ============
    #         # ======================================================
    #         from dateutil.relativedelta import relativedelta
    #
    #         total_allocated = sum(valid_alloc.mapped('number_of_days_display')) if valid_alloc else 0.0
    #         paid_leaves = self.env['hr.leave'].search([
    #             ('employee_id', '=', employee.id),
    #             ('state', 'not in', ('cancel', 'refuse')),
    #             ('holiday_status_id.time_type', '=', 'paid'),
    #         ])
    #         total_paid_taken = sum(paid_leaves.mapped('number_of_days')) if paid_leaves else 0.0
    #
    #         remaining_paid_leave = total_allocated - total_paid_taken
    #
    #         month_start = formatted_date.replace(day=1)
    #         next_month = (month_start + relativedelta(months=1))
    #         month_end = next_month - timedelta(days=1)
    #         monthly_paid = self.env['hr.leave'].search([
    #             ('employee_id', '=', employee.id),
    #             ('state', 'not in', ('cancel', 'refuse')),
    #             ('holiday_status_id.time_type', '=', 'paid'),
    #             ('request_date_from', '>=', month_start),
    #             ('request_date_to', '<=', month_end),
    #         ])
    #         total_monthly_paid = sum(monthly_paid.mapped('number_of_days'))
    #
    #         _logger.info(
    #             f"[EL/LOP Decision] {employee.name}: Allocated={total_allocated}, PaidTaken={total_paid_taken}, "
    #             f"Remaining={remaining_paid_leave}, MonthlyPaid={total_monthly_paid}, LeaveDuration={leave_duration}")
    #
    #         if (total_monthly_paid + leave_duration) > 3:
    #             _logger.info(f"[EL/LOP Decision] Employee {employee.name}: Monthly paid > 3, creating LOP")
    #             create_and_validate_leave(lop_type, leave_duration, half_day,
    #                                       'am' if half_day and morning == 'Absent' else 'pm' if half_day else False)
    #         else:
    #             _logger.info(f"[EL/LOP Decision] Employee {employee.name}: Monthly paid <= 3, creating EL")
    #             create_and_validate_leave(el_type, leave_duration, half_day,
    #                                       'am' if half_day and morning == 'Absent' else 'pm' if half_day else False)

    # 29-10-2025
    # def action_upload(self):
    #         """Uploads attendance Excel file and creates hr.upload, hr.attendance, and hr.leave records with EL/LOP rules.
    #         Integrates with hr.contract.leaves_available for non_ctc employees only.
    #         """
    #         if not self.file:
    #             raise UserError(_("Please upload a file."))
    #
    #         try:
    #             wb = openpyxl.load_workbook(
    #                 filename=BytesIO(base64.b64decode(self.file)),
    #                 read_only=True
    #             )
    #             ws = wb.active
    #
    #             hr_upload_obj = self.env['hr.upload']
    #             employee_obj = self.env['hr.employee']
    #             attendance_obj = self.env['hr.attendance']
    #             leave_obj = self.env['hr.leave']
    #
    #             el_type = self.env['hr.leave.type'].search([('name', '=', 'EL')], limit=1)
    #             lop_type = self.env['hr.leave.type'].search([('name', '=', 'LOP')], limit=1)
    #             if not el_type or not lop_type:
    #                 raise UserError(_("Please configure both 'EL' and 'LOP' in Time Off Types."))
    #
    #             user_tz = self.env.user.tz or 'Asia/Kolkata'
    #             tz = pytz.timezone(user_tz)
    #
    #             def combine_datetime(date_part, time_value):
    #                 if not time_value:
    #                     return False
    #                 if isinstance(time_value, datetime):
    #                     return time_value
    #                 elif isinstance(time_value, (float, int)):
    #                     from openpyxl.utils.datetime import from_excel
    #                     dt = from_excel(time_value)
    #                     return datetime.combine(date_part, dt.time())
    #                 elif isinstance(time_value, str):
    #                     for fmt in ("%H:%M:%S", "%H:%M"):
    #                         try:
    #                             t = datetime.strptime(time_value.strip(), fmt).time()
    #                             return datetime.combine(date_part, t)
    #                         except Exception:
    #                             continue
    #                     return False
    #                 else:
    #                     return False
    #
    #             row_number = 1
    #             for row in ws.iter_rows(min_row=2, values_only=True):
    #                 row_number += 1
    #
    #                 employee_code = row[1]
    #                 employee_name = row[2]
    #                 date_value = row[9]
    #                 check_in = row[12]
    #                 check_out = row[14]
    #
    #                 if not employee_code:
    #                     raise UserError(_("Row %d: Employee code is missing.") % row_number)
    #                 if not employee_name:
    #                     raise UserError(_("Row %d: Employee name is missing.") % row_number)
    #
    #                 employee_name = str(employee_name).strip()
    #                 try:
    #                     employee_code_str = str(int(employee_code))
    #                 except Exception:
    #                     employee_code_str = str(employee_code).strip()
    #
    #                 employee_exact = employee_obj.search([
    #                     ('cmr_code', '=', employee_code_str),
    #                     ('name', '=', employee_name)
    #                 ], limit=1)
    #
    #                 if not employee_exact:
    #                     raise UserError(_(
    #                         "Row %d: No employee found with code '%s' and name '%s'."
    #                     ) % (row_number, employee_code_str, employee_name))
    #
    #                 employee = employee_exact
    #
    #                 formatted_date = False
    #                 if isinstance(date_value, datetime):
    #                     formatted_date = date_value.date()
    #                 elif isinstance(date_value, str):
    #                     for fmt in ("%d-%b-%Y", "%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d/%m/%y"):
    #                         try:
    #                             formatted_date = datetime.strptime(date_value, fmt).date()
    #                             break
    #                         except Exception:
    #                             continue
    #                 if not formatted_date:
    #                     raise UserError(_("Row %d: Invalid date format '%s'.") % (row_number, date_value))
    #
    #                 check_in_dt = combine_datetime(formatted_date, check_in)
    #                 check_out_dt = combine_datetime(formatted_date, check_out)
    #
    #                 check_in_utc = check_out_utc = False
    #                 if check_in_dt:
    #                     check_in_utc = tz.localize(check_in_dt).astimezone(pytz.UTC).replace(tzinfo=None)
    #                 if check_out_dt:
    #                     check_out_utc = tz.localize(check_out_dt).astimezone(pytz.UTC).replace(tzinfo=None)
    #
    #                 upload_rec = hr_upload_obj.create({
    #                     'employee_name': employee.id,
    #                     'ctc_type': employee.ctc_type,
    #                     'employee_code': employee_code_str,
    #                     'date': formatted_date,
    #                     'check_in_attendance': check_in,
    #                     'check_out_attendance': check_out,
    #                     'your_datetime': check_in_utc,
    #                     'your_checkout_datetime': check_out_utc,
    #                 })
    #
    #                 morning = upload_rec.morning_session or ''
    #                 afternoon = upload_rec.afternoon_session or ''
    #                 leave_duration = 0.0
    #                 half_day = False
    #
    #                 if morning == 'Absent' and afternoon == 'Absent':
    #                     leave_duration = 1.0
    #                 elif morning == 'Absent' or afternoon == 'Absent':
    #                     leave_duration = 0.5
    #                     half_day = True
    #
    #                 # ✅ Only for Non-CTC employees
    #                 if leave_duration > 0 and employee.ctc_type == 'non_ctc':
    #                     existing_leave = leave_obj.search([
    #                         ('employee_id', '=', employee.id),
    #                         ('request_date_from', '=', formatted_date),
    #                         ('request_date_to', '=', formatted_date),
    #                         ('state', '!=', 'refuse')
    #                     ], limit=1)
    #                     if existing_leave:
    #                         continue
    #
    #                     # Fetch active contract
    #                     contract = self.env['hr.contract'].search([
    #                         ('employee_id', '=', employee.id),
    #                         ('state', '=', 'open')
    #                     ], limit=1)
    #
    #                     remaining_el = 0.0
    #                     if contract:
    #                         remaining_el = contract.leaves_available or 0.0
    #
    #                     def create_and_validate_leave(leave_type, days, half_day=False, period=False):
    #                         vals = {
    #                             'name': f"Auto Leave {formatted_date}",
    #                             'employee_id': employee.id,
    #                             'holiday_status_id': leave_type.id,
    #                             'request_date_from': formatted_date,
    #                             'request_date_to': formatted_date,
    #                             'number_of_days': days,
    #                             'payslip_state': 'done'
    #                         }
    #                         if half_day:
    #                             vals.update({
    #                                 'request_unit_half': True,
    #                                 'request_date_from_period': period,
    #                             })
    #                         leave_rec = leave_obj.create(vals)
    #                         if leave_rec.state == 'draft':
    #                             leave_rec.action_confirm()
    #                         if leave_rec.state in ['confirm', 'validate1']:
    #                             leave_rec.action_validate()
    #                         return leave_rec
    #
    #                     # Deduction Logic
    #                     if remaining_el >= leave_duration:
    #                         create_and_validate_leave(
    #                             el_type, leave_duration, half_day,
    #                             'am' if half_day and morning == 'Absent' else 'pm' if half_day else False
    #                         )
    #                         if contract:
    #                             contract.leaves_available -= leave_duration
    #
    #                     elif 0 < remaining_el < leave_duration:
    #                         create_and_validate_leave(el_type, remaining_el)
    #                         lop_days = leave_duration - remaining_el
    #                         create_and_validate_leave(lop_type, lop_days)
    #                         if contract:
    #                             contract.leaves_available = 0.0
    #
    #                     else:
    #                         create_and_validate_leave(
    #                             lop_type, leave_duration, half_day,
    #                             'am' if half_day and morning == 'Absent' else 'pm' if half_day else False
    #                         )
    #
    #                     # Ensure leaves_available never negative
    #                     if contract and contract.leaves_available < 0:
    #                         contract.leaves_available = 0.0
    #
    #                 # Create Attendance
    #                 if check_in_utc and check_out_utc:
    #                     open_attendance = attendance_obj.search([
    #                         ('employee_id', '=', employee.id),
    #                         ('check_out', '=', False)
    #                     ], limit=1)
    #                     if open_attendance:
    #                         open_attendance.write({'check_out': fields.Datetime.to_string(check_out_utc)})
    #                     else:
    #                         attendance_obj.create({
    #                             'employee_id': employee.id,
    #                             'date': formatted_date,
    #                             'check_in': fields.Datetime.to_string(check_in_utc),
    #                             'check_out': fields.Datetime.to_string(check_out_utc),
    #                         })
    #
    #         except Exception as e:
    #             raise UserError(_('Upload failed: %s') % e)

    # Latest Code 28-10-2025 - 06:11 pm
    # def action_upload(self):
    #     """Uploads attendance Excel file and creates hr.upload, hr.attendance, and hr.leave records with EL/LOP rules."""
    #     if not self.file:
    #         raise UserError(_("Please upload a file."))
    #
    #     try:
    #         wb = openpyxl.load_workbook(
    #             filename=BytesIO(base64.b64decode(self.file)),
    #             read_only=True
    #         )
    #         ws = wb.active
    #
    #         hr_upload_obj = self.env['hr.upload']
    #         employee_obj = self.env['hr.employee']
    #         attendance_obj = self.env['hr.attendance']
    #         leave_obj = self.env['hr.leave']
    #
    #         # === Verify Leave Types Exist ===
    #         el_type = self.env['hr.leave.type'].search([('name', '=', 'EL')], limit=1)
    #         lop_type = self.env['hr.leave.type'].search([('name', '=', 'LOP')], limit=1)
    #         if not el_type or not lop_type:
    #             raise UserError(_("Please configure both 'EL' and 'LOP' in Time Off Types."))
    #
    #         user_tz = self.env.user.tz or 'Asia/Kolkata'
    #         tz = pytz.timezone(user_tz)
    #
    #         # Helper: combine Excel date + time
    #         def combine_datetime(date_part, time_value):
    #             if not time_value:
    #                 return False
    #             if isinstance(time_value, datetime):
    #                 return time_value
    #             elif isinstance(time_value, (float, int)):
    #                 # Excel float to time
    #                 from openpyxl.utils.datetime import from_excel
    #                 dt = from_excel(time_value)
    #                 return datetime.combine(date_part, dt.time())
    #             elif isinstance(time_value, str):
    #                 for fmt in ("%H:%M:%S", "%H:%M"):
    #                     try:
    #                         t = datetime.strptime(time_value.strip(), fmt).time()
    #                         return datetime.combine(date_part, t)
    #                     except Exception:
    #                         continue
    #                 return False
    #             else:
    #                 return False
    #
    #         row_number = 1
    #         for row in ws.iter_rows(min_row=2, values_only=True):
    #             row_number += 1
    #
    #             # === Extract Excel values ===
    #             employee_code = row[1]
    #             employee_name = row[2]
    #             date_value = row[9]
    #             check_in = row[12]
    #             check_out = row[14]
    #
    #             # === Validation ===
    #             if not employee_code:
    #                 raise UserError(_("Row %d: Employee code is missing.") % row_number)
    #             if not employee_name:
    #                 raise UserError(_("Row %d: Employee name is missing.") % row_number)
    #
    #             employee_name = str(employee_name).strip()
    #             try:
    #                 employee_code_str = str(int(employee_code))
    #             except Exception:
    #                 employee_code_str = str(employee_code).strip()
    #
    #             employee_exact = employee_obj.search([
    #                 ('cmr_code', '=', employee_code_str),
    #                 ('name', '=', employee_name)
    #             ], limit=1)
    #
    #             if not employee_exact:
    #                 raise UserError(_(
    #                     "Row %d: No employee found with code '%s' and name '%s'."
    #                 ) % (row_number, employee_code_str, employee_name))
    #
    #             employee = employee_exact
    #
    #             # === Convert Date ===
    #             formatted_date = False
    #             if isinstance(date_value, datetime):
    #                 formatted_date = date_value.date()
    #             elif isinstance(date_value, str):
    #                 for fmt in ("%d-%b-%Y", "%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d/%m/%y"):
    #                     try:
    #                         formatted_date = datetime.strptime(date_value, fmt).date()
    #                         break
    #                     except Exception:
    #                         continue
    #             if not formatted_date:
    #                 raise UserError(_("Row %d: Invalid date format '%s'.") % (row_number, date_value))
    #
    #             # === Combine Date + Time ===
    #             check_in_dt = combine_datetime(formatted_date, check_in)
    #             check_out_dt = combine_datetime(formatted_date, check_out)
    #
    #             check_in_utc = check_out_utc = False
    #             if check_in_dt:
    #                 check_in_utc = tz.localize(check_in_dt).astimezone(pytz.UTC).replace(tzinfo=None)
    #             if check_out_dt:
    #                 check_out_utc = tz.localize(check_out_dt).astimezone(pytz.UTC).replace(tzinfo=None)
    #
    #             # === Create hr.upload Record ===
    #             upload_rec = hr_upload_obj.create({
    #                 'employee_name': employee.id,
    #                 'ctc_type': employee.ctc_type,
    #                 'employee_code': employee_code_str,
    #                 'date': formatted_date,
    #                 'check_in_attendance': check_in,
    #                 'check_out_attendance': check_out,
    #                 'your_datetime': check_in_utc,
    #                 'your_checkout_datetime': check_out_utc,
    #             })
    #
    #             # === Determine Leave Type (EL/LOP) ===
    #             morning = upload_rec.morning_session or ''
    #             afternoon = upload_rec.afternoon_session or ''
    #             leave_duration = 0.0
    #             half_day = False
    #
    #             if morning == 'Absent' and afternoon == 'Absent':
    #                 leave_duration = 1.0
    #             elif morning == 'Absent' or afternoon == 'Absent':
    #                 leave_duration = 0.5
    #                 half_day = True
    #
    #             if leave_duration > 0:
    #                 existing_leave = leave_obj.search([
    #                     ('employee_id', '=', employee.id),
    #                     ('request_date_from', '=', formatted_date),
    #                     ('request_date_to', '=', formatted_date),
    #                     ('state', '!=', 'refuse')
    #                 ], limit=1)
    #                 if existing_leave:
    #                     continue
    #
    #                 leave_type = lop_type
    #                 remaining_el = 0.0
    #
    #                 if employee.ctc_type == 'non_ctc':
    #                     first_day = formatted_date.replace(day=1)
    #                     last_day = (first_day + relativedelta(months=1)) - timedelta(days=1)
    #                     total_els = leave_obj.search_read([
    #                         ('employee_id', '=', employee.id),
    #                         ('holiday_status_id', '=', el_type.id),
    #                         ('request_date_from', '>=', first_day),
    #                         ('request_date_to', '<=', last_day),
    #                         ('state', '!=', 'refuse')
    #                     ], ['number_of_days'])
    #                     total_el_taken = sum(l['number_of_days'] for l in total_els)
    #                     remaining_el = max(3 - total_el_taken, 0.0)
    #
    #                 def create_and_validate_leave(leave_type, days, half_day=False, period=False):
    #                     vals = {
    #                         'name': f"Auto Leave {formatted_date}",
    #                         'employee_id': employee.id,
    #                         'holiday_status_id': leave_type.id,
    #                         'request_date_from': formatted_date,
    #                         'request_date_to': formatted_date,
    #                         'number_of_days': days,
    #                         'payslip_state': 'done'
    #                     }
    #                     if half_day:
    #                         vals.update({
    #                             'request_unit_half': True,
    #                             'request_date_from_period': period,
    #                         })
    #                     leave_rec = leave_obj.create(vals)
    #                     if leave_rec.state == 'draft':
    #                         leave_rec.action_confirm()
    #                     if leave_rec.state in ['confirm', 'validate1']:
    #                         leave_rec.action_validate()
    #
    #                 if remaining_el >= leave_duration:
    #                     create_and_validate_leave(
    #                         el_type, leave_duration, half_day,
    #                         'am' if half_day and morning == 'Absent' else 'pm' if half_day else False
    #                     )
    #                 elif remaining_el > 0 and remaining_el < leave_duration:
    #                     create_and_validate_leave(el_type, remaining_el)
    #                     lop_days = leave_duration - remaining_el
    #                     create_and_validate_leave(lop_type, lop_days)
    #                 else:
    #                     create_and_validate_leave(
    #                         lop_type, leave_duration, half_day,
    #                         'am' if half_day and morning == 'Absent' else 'pm' if half_day else False
    #                     )
    #
    #             # === Create Attendance ===
    #             if check_in_utc and check_out_utc:
    #                 open_attendance = attendance_obj.search([
    #                     ('employee_id', '=', employee.id),
    #                     ('check_out', '=', False)
    #                 ], limit=1)
    #                 if open_attendance:
    #                     open_attendance.write({'check_out': fields.Datetime.to_string(check_out_utc)})
    #                 else:
    #                     attendance_obj.create({
    #                         'employee_id': employee.id,
    #                         'date': formatted_date,
    #                         'check_in': fields.Datetime.to_string(check_in_utc),
    #                         'check_out': fields.Datetime.to_string(check_out_utc),
    #                     })
    #
    #     except Exception as e:
    #         raise UserError(_('Upload failed: %s') % e)


    # def action_upload(self):
    #     """Uploads attendance Excel file and creates hr.upload, hr.attendance, and hr.leave records with EL/LOP rules."""
    #     if not self.file:
    #         raise UserError(_("Please upload a file."))
    #
    #     try:
    #         wb = openpyxl.load_workbook(
    #             filename=BytesIO(base64.b64decode(self.file)),
    #             read_only=True
    #         )
    #         ws = wb.active
    #
    #         hr_upload_obj = self.env['hr.upload']
    #         employee_obj = self.env['hr.employee']
    #         attendance_obj = self.env['hr.attendance']
    #         leave_obj = self.env['hr.leave']
    #
    #         # === Verify Leave Types Exist ===
    #         el_type = self.env['hr.leave.type'].search([('name', '=', 'EL')], limit=1)
    #         lop_type = self.env['hr.leave.type'].search([('name', '=', 'LOP')], limit=1)
    #         if not el_type or not lop_type:
    #             raise UserError(_("Please configure both 'EL' and 'LOP' in Time Off Types."))
    #
    #         user_tz = self.env.user.tz or 'Asia/Kolkata'
    #         tz = pytz.timezone(user_tz)
    #
    #         row_number = 1
    #         for row in ws.iter_rows(min_row=2, values_only=True):
    #             row_number += 1
    #
    #             # Extract Excel values
    #             employee_code = row[1]
    #             employee_name = row[2]
    #             date_value = row[9]
    #             check_in = row[12]
    #             check_out = row[14]
    #
    #             # ===== Validation =====
    #             if not employee_code:
    #                 raise UserError(_("Row %d: Employee code is missing.") % row_number)
    #             if not employee_name:
    #                 raise UserError(_("Row %d: Employee name is missing.") % row_number)
    #
    #             employee_name = str(employee_name).strip()
    #             try:
    #                 employee_code_str = str(int(employee_code))
    #             except Exception:
    #                 employee_code_str = str(employee_code).strip()
    #
    #             employee_exact = employee_obj.search([
    #                 ('cmr_code', '=', employee_code_str),
    #                 ('name', '=', employee_name)
    #             ], limit=1)
    #
    #             if not employee_exact:
    #                 raise UserError(_(
    #                     "Row %d: No employee found with code '%s' and name '%s'."
    #                 ) % (row_number, employee_code_str, employee_name))
    #
    #             employee = employee_exact
    #
    #             # ===== Date Conversion =====
    #             formatted_date = False
    #             if isinstance(date_value, datetime):
    #                 formatted_date = date_value.date()
    #             elif isinstance(date_value, str):
    #                 for fmt in ("%d-%b-%Y", "%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d/%m/%y"):
    #                     try:
    #                         formatted_date = datetime.strptime(date_value, fmt).date()
    #                         break
    #                     except Exception:
    #                         continue
    #
    #             if not formatted_date:
    #                 raise UserError(_("Row %d: Invalid date format '%s'.") % (row_number, date_value))
    #
    #             # ===== Create hr.upload Record =====
    #             upload_rec = hr_upload_obj.create({
    #                 'employee_name': employee.id,
    #                 'ctc_type': employee.ctc_type,
    #                 'employee_code': employee_code_str,
    #                 'date': formatted_date,
    #                 'check_in_attendance': check_in,
    #                 'check_out_attendance': check_out,
    #             })
    #
    #             # ===== Determine Leave Type (EL/LOP) TIME OFF CREATION=====
    #             morning = upload_rec.morning_session or ''
    #             afternoon = upload_rec.afternoon_session or ''
    #             leave_duration = 0.0
    #             half_day = False
    #
    #             if morning == 'Absent' and afternoon == 'Absent':
    #                 leave_duration = 1.0
    #             elif morning == 'Absent' or afternoon == 'Absent':
    #                 leave_duration = 0.5
    #                 half_day = True
    #
    #             if leave_duration > 0:
    #                 # Prevent duplicate leave on same day
    #                 existing_leave = leave_obj.search([
    #                     ('employee_id', '=', employee.id),
    #                     ('request_date_from', '=', formatted_date),
    #                     ('request_date_to', '=', formatted_date),
    #                     ('state', '!=', 'refuse')
    #                 ], limit=1)
    #                 if existing_leave:
    #                     continue
    #
    #                 # Monthly EL/LOP rule
    #                 leave_type = lop_type
    #                 remaining_el = 0.0
    #
    #                 if employee.ctc_type == 'non_ctc':
    #                     first_day = formatted_date.replace(day=1)
    #                     last_day = (first_day + relativedelta(months=1)) - timedelta(days=1)
    #
    #                     total_els = leave_obj.search_read([
    #                         ('employee_id', '=', employee.id),
    #                         ('holiday_status_id', '=', el_type.id),
    #                         ('request_date_from', '>=', first_day),
    #                         ('request_date_to', '<=', last_day),
    #                         ('state', '!=', 'refuse')
    #                     ], ['number_of_days'])
    #
    #                     total_el_taken = sum(l['number_of_days'] for l in total_els)
    #                     remaining_el = max(3 - total_el_taken, 0.0)
    #
    #                 def create_and_validate_leave(leave_type, days, half_day=False, period=False):
    #                     vals = {
    #                         'name': f"Auto Leave {formatted_date}",
    #                         'employee_id': employee.id,
    #                         'holiday_status_id': leave_type.id,
    #                         'request_date_from': formatted_date,
    #                         'request_date_to': formatted_date,
    #                         'number_of_days': days,
    #                     }
    #                     if half_day:
    #                         vals.update({
    #                             'request_unit_half': True,
    #                             'request_date_from_period': period,
    #                         })
    #                     leave_rec = leave_obj.create(vals)
    #
    #                     # Auto-confirm + validate
    #                     if leave_rec.state == 'draft':
    #                         leave_rec.action_confirm()
    #                     if leave_rec.state in ['confirm', 'validate1']:
    #                         leave_rec.action_validate()
    #
    #                 # Case 1: all EL
    #                 if remaining_el >= leave_duration:
    #                     create_and_validate_leave(
    #                         el_type, leave_duration, half_day,
    #                         'am' if half_day and morning == 'Absent' else 'pm' if half_day else False
    #                     )
    #
    #                 # Case 2: split between EL and LOP
    #                 elif remaining_el > 0 and remaining_el < leave_duration:
    #                     # EL part
    #                     create_and_validate_leave(el_type, remaining_el)
    #                     # LOP part
    #                     lop_days = leave_duration - remaining_el
    #                     create_and_validate_leave(lop_type, lop_days)
    #
    #                 # Case 3: all LOP
    #                 else:
    #                     create_and_validate_leave(
    #                         lop_type, leave_duration, half_day,
    #                         'am' if half_day and morning == 'Absent' else 'pm' if half_day else False
    #                     )
    #
    #             # ===== Create Attendance =====
    #             if check_in and check_out:
    #
    #                 def to_datetime_local(date_part, time_value):
    #                     if isinstance(time_value, datetime):
    #                         local_dt = datetime.combine(date_part, time_value.time())
    #                     elif isinstance(time_value, str):
    #                         for fmt in ("%H:%M:%S", "%H:%M"):
    #                             try:
    #                                 t = datetime.strptime(time_value.strip(), fmt).time()
    #                                 local_dt = datetime.combine(date_part, t)
    #                                 break
    #                             except Exception:
    #                                 continue
    #                         else:
    #                             return False
    #                     elif isinstance(time_value, (float, int)):
    #                         from openpyxl.utils.datetime import from_excel
    #                         t = from_excel(time_value).time()
    #                         local_dt = datetime.combine(date_part, t)
    #                     else:
    #                         return False
    #
    #                     local_dt = tz.localize(local_dt)
    #                     return local_dt.astimezone(pytz.UTC)
    #
    #                 check_in_dt = to_datetime_local(formatted_date, check_in)
    #                 check_out_dt = to_datetime_local(formatted_date, check_out)
    #
    #                 if not (check_in_dt and check_out_dt):
    #                     continue
    #
    #                 open_attendance = attendance_obj.search([
    #                     ('employee_id', '=', employee.id),
    #                     ('check_out', '=', False)
    #                 ], limit=1)
    #
    #                 if open_attendance:
    #                     open_attendance.write({'check_out': fields.Datetime.to_string(check_out_dt)})
    #                 else:
    #                     attendance_obj.create({
    #                         'employee_id': employee.id,
    #                         'date': formatted_date,
    #                         'check_in': fields.Datetime.to_string(check_in_dt),
    #                         'check_out': fields.Datetime.to_string(check_out_dt),
    #                     })
    #
    #     except Exception as e:
    #         raise UserError(_('Upload failed: %s') % e)



class HrLateDeductionMaster(models.Model):
    _name = "hr.late.deduction.master"
    _description = "Late Deduction Master"
    _order = "start_time asc"

    name = fields.Char(string="Name", required=True)
    start_time = fields.Char(string="Start Time (HH:MM)", required=True)  # Example: 10:10
    end_time = fields.Char(string="End Time (HH:MM)", required=True)  # Example: 10:30
    deduction_amount = fields.Float(string="Deduction Amount", required=True)

class HrMonthlyLateDeduction(models.Model):
    _name = "hr.monthly.late.deduction"
    _description = "Monthly Late Deduction"

    employee_id = fields.Many2one('hr.employee', string="Employee")
    month = fields.Selection(
        [(str(i), calendar.month_name[i]) for i in range(1, 13)],
        string="Month",
        required=True,
        default=lambda self: str(datetime.now().month)
    )

    year = fields.Integer(string="Year", required=True, default=lambda self: datetime.now().year)


    line_ids = fields.One2many(
        'hr.late.deduction.line',
        'monthly_id',
        string="Late Deduction Lines"
    )



    # def action_generate_all_employees(self):
    #     self.ensure_one()
    #     month = self.month
    #     year = self.year
    #
    #     start_date = datetime(year, month, 1)
    #     last_day = calendar.monthrange(year, month)[1]
    #     end_date = datetime(year, month, last_day)
    #
    #     uploads = self.env['hr.upload'].search([
    #         ('date', '>=', start_date.date()),
    #         ('date', '<=', end_date.date()),
    #     ])
    #     employee_ids = uploads.mapped('employee_name.id')
    #
    #     if not employee_ids:
    #         return {
    #             'effect': {
    #                 'fadeout': 'slow',
    #                 'message': 'No uploaded employees found for this period!',
    #                 'type': 'rainbow_man',
    #             }
    #         }
    #
    #     existing = self.search([('month', '=', month), ('year', '=', year)])
    #     existing_emp_ids = existing.mapped('employee_id.id')
    #
    #     created_records = self.env['hr.monthly.late.deduction']
    #     for emp_id in employee_ids:
    #         if emp_id not in existing_emp_ids:
    #             rec = self.create({
    #                 'employee_id': emp_id,
    #                 'month': month,
    #                 'year': year,
    #             })
    #             created_records += rec
    #
    #     if created_records:
    #         return {
    #             'type': 'ir.actions.act_window',
    #             'res_model': 'hr.monthly.late.deduction',
    #             'res_id': created_records[0].id,
    #             'view_mode': 'form',
    #             'target': 'current',
    #         }
    #     else:
    #         return {
    #             'effect': {
    #                 'fadeout': 'slow',
    #                 'message': 'Records already exist for uploaded employees!',
    #                 'type': 'rainbow_man',
    #             }
    #         }

    def action_generate_all_employees(self):
        self.ensure_one()

        # ✅ Clean and convert string to integer
        month = int(self.month)
        year = int(str(self.year).replace(',', ''))

        start_date = datetime(year, month, 1)
        last_day = calendar.monthrange(year, month)[1]
        end_date = datetime(year, month, last_day)

        uploads = self.env['hr.upload'].search([
            ('date', '>=', start_date.date()),
            ('date', '<=', end_date.date()),
        ])
        employee_ids = uploads.mapped('employee_name.id')

        if not employee_ids:
            return {
                'effect': {
                    'fadeout': 'slow',
                    'message': 'No uploaded employees found for this period!',
                    'type': 'rainbow_man',
                }
            }

        # Create lines for each employee
        line_vals = []
        for emp_id in employee_ids:
            line_vals.append((0, 0, {
                'employee_id': emp_id,
                'late_days': 0,
                'late_hours': 0.0,
                'deduction_amount': 0.0,
            }))

        # Update current record with lines
        self.line_ids = [(5, 0, 0)] + line_vals  # clear old + add new

        return {
            'effect': {
                'fadeout': 'slow',
                'message': f'{len(employee_ids)} Employee Records Loaded!',
                'type': 'rainbow_man',
            }
        }

class HrLateDeductionLine(models.Model):
    _name = "hr.late.deduction.line"
    _description = "Late Deduction Line"

    monthly_id = fields.Many2one('hr.monthly.late.deduction', string="Monthly Record")
    employee_id = fields.Many2one('hr.employee', string="Employee")
    late_days = fields.Integer(string="Late Days")
    late_hours = fields.Float(string="Late Hours")
    deduction_amount = fields.Float(string="Deduction Amount (₹)", compute="_compute_total_late_deduction",store=True)

    month = fields.Selection(
        related='monthly_id.month',
        string="Month",
        store=True,
        readonly=True
    )
    year = fields.Integer(
        related='monthly_id.year',
        string="Year",
        store=True,
        readonly=True
    )

    @api.depends('employee_id', 'monthly_id.month', 'monthly_id.year')
    def _compute_total_late_deduction(self):
        """Compute total monthly deduction for each employee based on hr.upload."""
        for rec in self:
            rec.deduction_amount = 0.0  # default
            if not rec.employee_id or not rec.monthly_id:
                continue

            # Safely get month & year
            try:
                month = int(rec.month)
                year = int(str(rec.year).replace(',', ''))
            except Exception:
                continue

            start_date = datetime(year, month, 1)
            last_day = calendar.monthrange(year, month)[1]
            end_date = datetime(year, month, last_day)

            # ✅ Search all hr.upload entries for that employee in the month
            uploads = self.env['hr.upload'].search([
                ('employee_name', '=', rec.employee_id.name),
                ('date', '>=', start_date.date()),
                ('date', '<=', end_date.date()),
            ])

            # ✅ Compute total late deduction
            rec.deduction_amount = sum(u.late_deduction for u in uploads)

            # Optional: if you want to count number of days with deductions


class HrOvertimeMaster(models.Model):
    _name = "hr.overtime.master"
    _description = "Overtime Master"
    _order = "start_time asc"

    name = fields.Char(string="Name", required=True)
    start_time = fields.Char(string="Start Time (HH:MM)", required=True)  # e.g. 18:00
    end_time = fields.Char(string="End Time (HH:MM)", required=True)      # e.g. 20:00
    overtime_amount = fields.Float(string="Overtime Amount (₹)")



class HrMonthlyOvertime(models.Model):
    _name = "hr.monthly.overtime"
    _description = "Monthly Overtime"

    employee_id = fields.Many2one('hr.employee', string="Employee")
    month = fields.Selection(
        [(str(i), calendar.month_name[i]) for i in range(1, 13)],
        string="Month",
        required=True,
        default=lambda self: str(datetime.now().month)
    )
    year = fields.Integer(string="Year", required=True, default=lambda self: datetime.now().year)

    line_ids = fields.One2many('hr.overtime.line', 'monthly_id', string="Overtime Lines")

    def action_generate_all_employees(self):
        self.ensure_one()
        month = int(self.month)
        year = int(self.year)

        start_date = datetime(year, month, 1)
        last_day = calendar.monthrange(year, month)[1]
        end_date = datetime(year, month, last_day)

        uploads = self.env['hr.upload'].search([
            ('date', '>=', start_date.date()),
            ('date', '<=', end_date.date()),
        ])
        employee_ids = uploads.mapped('employee_name.id')

        if not employee_ids:
            return {
                'effect': {
                    'fadeout': 'slow',
                    'message': 'No uploaded employees found for this period!',
                    'type': 'rainbow_man',
                }
            }

        line_vals = []
        for emp_id in employee_ids:
            line_vals.append((0, 0, {
                'employee_id': emp_id,

                'overtime_amount': 0.0,
            }))

        self.line_ids = [(5, 0, 0)] + line_vals

        return {
            'effect': {
                'fadeout': 'slow',
                'message': f'{len(employee_ids)} Employee Overtime Records Loaded!',
                'type': 'rainbow_man',
            }
        }


class HrOvertimeLine(models.Model):
    _name = "hr.overtime.line"
    _description = "Overtime Line"

    monthly_id = fields.Many2one('hr.monthly.overtime', string="Monthly Record")
    employee_id = fields.Many2one('hr.employee', string="Employee")

    overtime_amount = fields.Float(string="Overtime Amount (₹)", compute="_compute_total_overtime", store=True)

    month = fields.Selection(related='monthly_id.month', store=True, readonly=True)
    year = fields.Integer(related='monthly_id.year', store=True, readonly=True)

    @api.depends('employee_id', 'monthly_id.month', 'monthly_id.year')
    def _compute_total_overtime(self):
        """Compute total monthly overtime for each employee based on hr.upload."""
        for rec in self:
            rec.overtime_amount = 0.0
            if not rec.employee_id or not rec.monthly_id:
                continue

            month = int(rec.month)
            year = int(rec.year)
            start_date = datetime(year, month, 1)
            last_day = calendar.monthrange(year, month)[1]
            end_date = datetime(year, month, last_day)

            uploads = self.env['hr.upload'].search([
                ('employee_name', '=', rec.employee_id.name),
                ('date', '>=', start_date.date()),
                ('date', '<=', end_date.date()),
            ])

            rec.overtime_amount = sum(u.overtime_amount for u in uploads)














