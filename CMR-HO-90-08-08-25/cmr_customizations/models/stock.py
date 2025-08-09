import csv
import socket
import textwrap

import openpyxl
import requests

from odoo import models, fields, api, _
from odoo.exceptions import UserError, ValidationError
from odoo.tests import Form
import re
from datetime import datetime, timedelta
import base64
import xlrd
import io
from odf.opendocument import load
from odf.table import Table, TableRow, TableCell
import logging
import pytz
from collections import defaultdict
_logger = logging.getLogger(__name__)

class ProductLabelLayout(models.TransientModel):
    _inherit = 'product.label.layout'

    def process(self):
        res = super(ProductLabelLayout, self).process()
        if self.move_ids.picking_id.is_landed_cost == 'yes':
            landed_cost = self.env['stock.landed.cost'].search(
                [('picking_ids.name', '=', self.move_ids.picking_id.name), ('state', '=', 'done')])
            if not landed_cost:
                raise ValidationError(f"Landed Cost was not generated for {self.move_ids.picking_id.name} GRC")
            self.move_ids.picking_id.label_click_count += 1
        return res


class Picking(models.Model):
    """Inherited stock.picking class to add fields and functions"""
    _inherit = "stock.picking"

    has_landed_cost = fields.Boolean(
        string="Has Landed Cost",
        compute="_compute_has_landed_cost",
        store=True
    )

    stock_picking_delivery_ids = fields.One2many("stock.picking.barcode", "stock_picking_delivery_id")

    @api.onchange('no_of_parcel')
    def _onchange_no_of_parcel(self):
        if self.no_of_parcel < 0:
            raise ValidationError(_("No of Parcels cannot be negative."))

        self.stock_picking_delivery_ids = [(5, 0, 0)]  # Clear existing lines

        for i in range(1, self.no_of_parcel + 1):
            self.stock_picking_delivery_ids += self.stock_picking_delivery_ids.new({
                'serial_no': i,
            })

    @api.depends('move_ids_without_package')
    def _compute_has_landed_cost(self):
        for picking in self:
            picking.has_landed_cost = bool(self.env['stock.landed.cost'].search([
                ('picking_ids', 'in', picking.ids)
            ], limit=1))

    def action_view_landed_costs(self):
        return {
            'name': 'Landed Costs',
            'type': 'ir.actions.act_window',
            'res_model': 'stock.landed.cost',
            'view_mode': 'tree,form',
            'domain': [('picking_ids', 'in', self.ids)],  # Show only relevant landed costs
            'context': {'default_picking_ids': [(6, 0, self.ids)]},
        }

    def print_barcodes_stock_picking(self):
        report_name = 'cmr_customizations.report_stock_picking_delivery_barcode'
        return {
            'type': 'ir.actions.report',
            'report_name': report_name,
            'report_type': 'qweb-pdf',
            'res_id': self.id,
            'res_model': 'stock.picking',
        }

    def action_import_excel(self):
        return {
            'type': 'ir.actions.act_window',
            'name': 'Import Excel',
            'res_model': 'stock.verification.import',
            'view_mode': 'form',
            'target': 'new',
            'context': {'default_stock_picking_id': self.id},
        }

    def action_open_landed_cost(self):
        self.ensure_one()  # Ensure only one picking is handled at a time

        # Check if there's already a Landed Cost linked to this picking
        existing_landed_cost = self.env['stock.landed.cost'].search([
            ('picking_ids', 'in', self.ids)
        ], limit=1)

        if existing_landed_cost:
            # If a Landed Cost exists, open it instead of creating a new one
            return {
                'name': 'Landed Costs',
                'type': 'ir.actions.act_window',
                'res_model': 'stock.landed.cost',
                'view_mode': 'form',
                'res_id': existing_landed_cost.id,  # Open existing record
                'target': 'current'
            }

        # If no Landed Cost exists, create a new one
        return {
            'name': 'Landed Costs',
            'type': 'ir.actions.act_window',
            'res_model': 'stock.landed.cost',
            'view_mode': 'form',
            'view_id': self.env.ref('stock_landed_costs.view_stock_landed_cost_form').id,
            'target': 'current',
            'context': {
                'default_picking_ids': [(6, 0, self.ids)]
            }
        }


    def button_import_receipts(self):
        return {
            "name": _("Import Receipts"),
            "type": "ir.actions.act_window",
            "res_model": "import.stock.receipt.wizard",
            "target": "new",
            "views": [[False, "form"]],
        }



    def dev_transport_entry_create(self, picking):
        data = {'partner_id': picking.partner_id and picking.partner_id.id or False,
                'picking_id': picking and picking.id or False,
                'lr_number': picking.lr_number or ' ',
                'transport_details_id': picking.transpoter_id and picking.transpoter_id.id or False,
                'contact_name': picking.transpoter_id and picking.transpoter_id.contact_name or ' ',
                'no_of_parcel': picking.no_of_parcel or 0,
                'name': picking.tracking_number or ' ',
                }
        tra_ent = self.env['dev.transport.entry'].create(data)
        if tra_ent and picking.transpoter_route_id:
            for line in picking.transpoter_route_id.location_details_ids:
                location_detail = {
                    'source_location_id': line.source_location_id and line.source_location_id.id or False,
                    'destination_location_id': line.destination_location_id and line.destination_location_id.id or False,
                    'distance': line.distance,
                    'transport_charges': line.transport_charges,
                    'time_hour': line.time_hour or ' ',
                    'tracking_number': picking.tracking_number or ' ',
                    'picking_id': picking.id,
                    'transport_entry_id': tra_ent.id,
                }
                self.env['transport.location.details'].create(location_detail)
        return tra_ent

    @api.model
    def create(self, vals):
        res = super(Picking, self).create(vals)
        res.tracking_number = self.env['ir.sequence'].next_by_code('stock.picking.tracking')
        res.lr_number = self.env['ir.sequence'].next_by_code('stock.picking.lr_number')
        if res.return_id:
            vals['stock_bill_reference'] = res.return_id.stock_bill_reference
        return res

    is_quality_done = fields.Boolean('Quality Done', copy=False)
    transpoter_id = fields.Many2one('dev.transport.details', string='Transport by')
    transpoter_route_id = fields.Many2one('dev.routes.details', string='Transporter Route')
    no_of_parcel = fields.Integer(string='No Of Parcels')
    tracking_number = fields.Char(string='Tracking Number')
    nhcl_tracking_number = fields.Char(string='Source Tracking Number')
    lr_number = fields.Char(string='LR Number')
    vehicle_number = fields.Char(string='Vehicle Number')
    driver_name = fields.Char(string='Driver Name')
    transport_location_line = fields.One2many('transport.location.details', 'picking_id', string='Transport Routes')
    transport_entry_ids = fields.One2many('dev.transport.entry', 'picking_id', string='Transport Entry')

    nhcl_last_serial_number = fields.Char('Last Serial Number', compute='_get_last_serial_number')
    is_confirm = fields.Boolean('Is Confirm', copy=False)
    stock_type = fields.Selection(
        [('advertisement', 'Advertisement'), ('ho_operation', 'HO Operation'), ('sub_contract', 'Sub Contracting'), ('data_import', 'Data Import'), ('inter_state', 'Inter State'),
         ('intra_state', 'Intra State'),
         ('pos_exchange', 'POS Exchange'), ('others', 'Others')], string='Stock Type', tracking=True)
    dummy_stock_type = fields.Selection(
        [('advertisement', 'Advertisement'), ('ho_operation', 'HO Operation'),
         ('others', 'Others')], string='Dummy Stock Type', compute='_compute_dummy_stock_type')
    stock_barcode = fields.Char(string='Barcode Scan')
    label_click_count = fields.Integer(string="Label Click Count", default=0)
    is_landed_cost = fields.Selection([('yes', 'Yes'), ('no', 'No')], string='Landed Cost', copy=False)
    verify_barcode = fields.Char(string='Verification Scan')
    stock_verification_ids = fields.One2many('stock.verification', 'stock_picking_id')
    owner_id = fields.Many2one(
        'res.partner', 'Assign Owner',
        check_company=True, index='btree_not_null',
        help="When validating the transfer, the products will be assigned to this owner.")
    is_landed_cost_confirm = fields.Boolean('Confirm', copy=False)
    is_outgoing = fields.Boolean(string="Is Outgoing", compute="_compute_is_outgoing")
    verification_success = fields.Selection([('matched', 'Matched')], string='Verification Status',
                                            tracking=True)
    stock_picking_type = fields.Selection(
        [('exchange', 'Customer-Return'), ('receipt', 'Receipt'), ('goods_return', 'Goods Return')
            , ('delivery', 'Delivery'), ('pos_order', 'POS Order'), ('regular', 'Regular'), ('damage', 'Damage'),
         ('return', 'Return'),
         ('damage_main', 'Damage-Main'), ('main_damage', 'Main-Damage'),
         ('return_main', 'Return-Main')], string='Type',
        tracking=True, related='picking_type_id.stock_picking_type',store=True)

    scan_or_import = fields.Selection([
        ('scan', 'Scan'),
        ('import', 'Import')
    ], string="Scan or Import", default='scan')
    nhcl_company_bool = fields.Boolean(related="company_id.nhcl_company_bool", readonly=False)
    stock_bill_reference = fields.Char(string="Bill Reference")
    allow_import_order = fields.Boolean('Allow Import', compute='_compute_import_move_lines')
    vendor_refund = fields.Many2one('account.move', copy=False, domain=[('move_type', '=', 'in_refund')])

    def _compute_import_move_lines(self):
        if self.env.user and self.env.user.import_stock_move_line == True:
            self.allow_import_order = True
        else:
            self.allow_import_order = False

    @api.onchange('partner_id')
    def _onchange_import_partner_id(self):
        self._compute_import_move_lines()


    @api.depends('picking_type_id')
    def _compute_is_outgoing(self):
        for picking in self:
            if picking.picking_type_id.code == 'outgoing' and picking.stock_picking_type == 'delivery':
                picking.is_outgoing = picking.picking_type_id.code == 'outgoing'
            else:
                picking.is_outgoing = False

    @api.depends('stock_type')
    def _compute_dummy_stock_type(self):
        for i in self:
            if i.stock_type == 'ho_operation':
                i.dummy_stock_type = 'ho_operation'
            elif i.stock_type == 'advertisement':
                i.dummy_stock_type = 'advertisement'
            elif i.stock_type == 'others':
                i.dummy_stock_type = 'others'
            elif i.stock_type == 'inter_state':
                i.dummy_stock_type = 'ho_operation'
            elif i.stock_type == 'intra_state':
                i.dummy_stock_type = 'ho_operation'
            elif i.stock_type == 'pos_exchange':
                i.dummy_stock_type = 'ho_operation'
            elif i.stock_type == 'sub_contract':
                i.dummy_stock_type = 'ho_operation'
            elif i.stock_type == 'data_import':
                i.dummy_stock_type = 'ho_operation'
            else:
                i.dummy_stock_type = ''

    @api.onchange('stock_verification_ids')
    def check_product_verification(self):
        for rec in self:
            operation_qty = sum(rec.move_ids_without_package.mapped('quantity'))
            verification_qty = sum(rec.stock_verification_ids.mapped('stock_qty'))
            verify_check = operation_qty == verification_qty
            if verify_check:
                rec.verification_success = 'matched'
            else:
                rec.verification_success = ''

    def _get_last_serial_number(self):
        auto_generate_seq_rec = self.env['nhcl.master.sequence'].search(
            [('nhcl_code', '=', 'Auto Serial Number'), ('nhcl_state', '=', 'activate')])
        if auto_generate_seq_rec and auto_generate_seq_rec.nhcl_next_number > 1:
            self.nhcl_last_serial_number = auto_generate_seq_rec.nhcl_prefix + str(
                auto_generate_seq_rec.nhcl_next_number - 1)
        else:
            self.nhcl_last_serial_number = '0'

    def validate_related_po_receipt(self):
        for picking in self:
            if picking.picking_type_id.code == 'outgoing' and picking.sale_id:
                sale_order = picking.sale_id
                # Find the related intercompany Purchase Order (assuming there's a link between the sale and purchase)
                related_po = self.env['purchase.order'].sudo().search([('origin', '=', sale_order.name)], limit=1)
                # If a related Purchase Order is found
                if related_po and sale_order.stock_type != 'return':
                    # Get the related receipt (incoming picking)
                    related_receipt_picking = self.env['stock.picking'].sudo().search([
                        ('purchase_id', '=', related_po.id),('picking_type_id.code', '=', 'incoming'),('state', '!=', 'done')], limit=1)
                    if related_receipt_picking:
                        related_receipt_picking._compute_check()
                        related_receipt_picking.action_confirm()  # Confirm the receipt
                        related_receipt_picking.button_validate()  # Validate the receipt
                        for move_line in related_receipt_picking.move_line_ids_without_package:
                            # move_line.lot_id.serial_type = 'return'
                            move_line.lot_id.is_uploaded = False
                        return True
                elif related_po and sale_order.stock_type == 'return':
                    # Get the related receipt (incoming picking)
                    related_receipt_picking = self.env['stock.picking'].sudo().search([
                        ('purchase_id', '=', related_po.id),('picking_type_id.code', '=', 'incoming'),('state', '!=', 'done')], limit=1)
                    if related_receipt_picking:
                        related_receipt_picking.sudo().location_id = self.env.ref('stock.stock_location_customers').id
                        related_receipt_picking._compute_check()
                        related_receipt_picking.action_confirm()  # Confirm the receipt
                        related_receipt_picking.button_validate()  # Validate the receipt
                        for move_line in related_receipt_picking.move_line_ids_without_package:
                            move_line.sudo().location_id = self.env.ref('stock.stock_location_customers').id
                            move_line.lot_id.write({
                                'serial_type':'return',
                                'is_uploaded':False,
                            })
                        return True
        return False


    def button_validate(self):
        res = None
        for picking in self:
            if picking.quality_check_todo == True and picking.stock_type == 'ho_operation':
                raise ValidationError("Please Check te Product Quality, Before valiating.")
            is_subcontract = picking.move_ids_without_package.filtered(lambda x: x.is_subcontract == True)
            if picking.picking_type_id.code == 'incoming' and (picking.stock_type in ['ho_operation','sub_contract','data_import'] or is_subcontract):
                for line in picking.move_ids_without_package:
                    if line.product_id.barcode == False:
                        raise ValidationError(f"Please Generate The Barcode Of This Product.'{line.product_id.name}'")
                    if line.type_product == False:
                        raise ValidationError("Please Select The Brand Type.")
                    if line.product_uom_qty < line.quantity:
                        raise ValidationError(_("You cannot receive more quantity (%s) than ordered (%s) for product %s." % (line.quantity, line.product_uom_qty, line.product_id.display_name)))
                for move_line in picking.move_line_ids_without_package:
                    if move_line.type_product == 'brand' and move_line.internal_ref_lot == False:
                        raise ValidationError("Please enter the Branded Barcode No in GRC.")
                if picking.is_landed_cost == False:
                    raise ValidationError(_('Please Select the Landed Cost Included in RSP'))
                elif picking.is_landed_cost_confirm == False:
                    return {
                        'name': _('Landed Cost Confirmation'),
                        'type': 'ir.actions.act_window',
                        'target': 'new',
                        'res_model': 'nhcl.landed.cost.confirmation.popup',
                        'view_mode': 'form',
                        'view_id': self.env.ref('cmr_customizations.crm_landed_cost_confirm_popup_view').id,
                        'context': {'default_nhcl_picking_id': picking.id},
                    }
                if picking.purchase_id:
                    picking.purchase_id.verify_draft_pi_to_po()
                if not self.env['grc.master'].sudo().search([('name', '=', picking.name)], limit=1):
                    self.env['grc.master'].sudo().create({'name': picking.name})
                branded_product_move_ids = picking.move_ids.filtered(
                    lambda x: x.product_id.nhcl_product_type == 'others')
                move_lines = branded_product_move_ids.move_line_ids
                if move_lines:
                    raise UserError(
                        _("Please assign a serial number for branded product percentages.") % (
                            move_lines[0].product_id.name))
                else:
                    unbranded_product_move_ids = picking.move_ids
                    un_move_lines = unbranded_product_move_ids.move_line_ids
                    no_serial_move_lines = un_move_lines.filtered(lambda x:x.lot_id != True)
                    if (not picking.is_confirm and un_move_lines) or (
                        picking.check_ids and no_serial_move_lines and picking.is_quality_done == False
                    ):
                        return {
                            'name': _('Auto Serial No Confirmation'),
                            'type': 'ir.actions.act_window',
                            'target': 'new',
                            'res_model': 'nhcl.serial.no.popup',
                            'view_mode': 'form',
                            'view_id': self.env.ref('cmr_customizations.crm_serial_no_confirm_popup_view').id,
                            'context': {'default_nhcl_picking_id': picking.id},
                        }
                    else:
                        res = super(Picking, picking).button_validate()
                        if res and picking.purchase_id and picking.is_landed_cost == 'no':
                            for move_line in picking.move_ids_without_package.move_line_ids:
                                move_line.lot_values_update()
                                lot_id = move_line.lot_id
                                purchase_line_id = move_line.move_id.purchase_line_id
                                if move_line.move_id.product_uom != purchase_line_id.product_uom:
                                    cost_price = purchase_line_id.price_unit / move_line.move_id.product_uom.ratio
                                else:
                                    cost_price = purchase_line_id.price_unit
                                lot_id.write({
                                    'actual_cp': cost_price,
                                    'cost_price': cost_price,
                                    'category_8': purchase_line_id.purchase_category_id.id,
                                    'ho_grc_no': picking.name,
                                    'nhcl_margin_lot': move_line.approval_margin,
                                })
                                lot_id.calculate_rsp_price()
                                lot_id.get_mrp_margin_from_category()
                                lot_id._get_product_aging()
                                move_line.write({
                                    'cost_price': cost_price,
                                    'categ_8': lot_id.category_8.id,
                                    'rs_price': lot_id.rs_price,
                                    'mr_price': lot_id.mr_price,
                                })
                        if res and picking.is_landed_cost == 'yes':
                            for move_line in picking.move_ids_without_package.move_line_ids:
                                move_line.lot_values_update()
                                lot_id = move_line.lot_id
                                lot_id._get_product_aging()
                                lot_id.ho_grc_no = picking.name

            elif picking.picking_type_id.code == 'incoming' and picking.stock_type == 'pos_exchange':
                for line in picking.move_line_ids:
                    if line.lot_id:
                        main_company = self.env['res.company'].sudo().search([('nhcl_company_bool', '=', True)])
                        cost_price = self.env['stock.lot'].sudo().search(
                            [('name', '=', line.lot_id.name), ('company_id', '=', main_company.id)])
                        line.lot_id.write({
                            'product_qty':1,
                            'cost_price': cost_price.cost_price,
                            'type_product': cost_price.type_product,
                        })
                res = super(Picking, self).button_validate()
            elif picking.picking_type_id.code == 'incoming' and picking.picking_type_id.name == 'Store Return':
                picking.sudo().location_id = picking.picking_type_id.default_location_src_id.id
                res = super(Picking, self).button_validate()

            else:
                if picking.picking_type_id.code != 'incoming':
                    for move in picking.move_ids_without_package:
                        if move.dummy_lot_ids:
                            # Update lot_ids with dummy_lot_ids
                            move.lot_ids = [(6, 0, move.dummy_lot_ids.ids)]
                res = super(Picking, self).button_validate()

            if picking.picking_type_id.code == 'incoming' and picking.stock_type in ['inter_state', 'intra_state']:
                sale_no = picking.purchase_id.partner_ref
                sale_delivery = self.env['sale.order'].sudo().search([('name', '=', sale_no)])
                delivery = sale_delivery.picking_ids
                picking.write({
                    'transpoter_id': delivery.transpoter_id.id,
                    'transpoter_route_id': delivery.transpoter_route_id.id,
                    'no_of_parcel': delivery.no_of_parcel,
                    'lr_number': delivery.lr_number,
                    'tracking_number': delivery.tracking_number,
                    'transport_location_line': delivery.transport_location_line,
                    'transport_entry_ids': delivery.transport_entry_ids,
                })
                for res in picking:
                    for line in res.move_ids:
                        corresponding_delivery_move = delivery.move_ids.filtered(
                            lambda move: move.product_id == line.product_id)
                        if corresponding_delivery_move:
                            line.write({
                                'type_product': corresponding_delivery_move[0].type_product
                            })
                    for move_line in res.move_line_ids:
                        corresponding_delivery_move_line = delivery.move_line_ids.filtered(
                            lambda dml: dml.product_id == move_line.product_id
                        )
                        if corresponding_delivery_move_line:
                            move_line.write({
                                'internal_ref_lot': corresponding_delivery_move_line[0].internal_ref_lot,
                                'type_product': corresponding_delivery_move_line[0].type_product,
                                'categ_1': corresponding_delivery_move_line[0].categ_1,
                                'categ_2': corresponding_delivery_move_line[0].categ_2,
                                'categ_3': corresponding_delivery_move_line[0].categ_3,
                                'categ_4': corresponding_delivery_move_line[0].categ_4,
                                'categ_5': corresponding_delivery_move_line[0].categ_5,
                                'categ_6': corresponding_delivery_move_line[0].categ_6,
                                'categ_7': corresponding_delivery_move_line[0].categ_7,
                                'categ_8': corresponding_delivery_move_line[0].categ_8,
                                'descrip_1': corresponding_delivery_move_line[0].descrip_1,
                                'descrip_2': corresponding_delivery_move_line[0].descrip_2,
                                'descrip_3': corresponding_delivery_move_line[0].descrip_3,
                                'descrip_4': corresponding_delivery_move_line[0].descrip_4,
                                'descrip_5': corresponding_delivery_move_line[0].descrip_5,
                                'descrip_6': corresponding_delivery_move_line[0].descrip_6,
                                'descrip_7': corresponding_delivery_move_line[0].descrip_7,
                                'descrip_8': corresponding_delivery_move_line[0].descrip_8,
                                'cost_price': corresponding_delivery_move_line[0].cost_price,
                                'mr_price': corresponding_delivery_move_line[0].mr_price,
                                'rs_price': corresponding_delivery_move_line[0].rs_price,
                            })

                            # if move_line.type_product == 'brand':
                            #     move_line.lot_id.sudo().write({
                            #         'ref': corresponding_delivery_move_line[0].internal_ref_lot,
                            #         'type_product': corresponding_delivery_move_line[0].type_product,
                            #     })
                            # elif move_line.type_product == 'un_brand':
                            #     move_line.lot_id.sudo().write({
                            #         'ref': move_line.product_id.barcode,
                            #         'type_product': move_line.type_product,
                            #     })
                            if move_line.type_product == 'brand':
                                move_line.lot_id.sudo().write({
                                    'ref': move_line.internal_ref_lot,
                                    'type_product': move_line.type_product,
                                })
                            elif move_line.type_product == 'un_brand':
                                move_line.lot_id.sudo().write({
                                    'ref': move_line.product_id.barcode,
                                    'type_product': move_line.type_product,
                                })
            if picking.state == 'done' and picking.transpoter_id and picking.stock_type in ['inter_state',
                                                                                            'intra_state']:
                picking.dev_transport_entry_create(picking)
            if picking.state == 'done' and picking.stock_type == 'data_import':
                picking.update_stock_valuation()
            if picking.state == 'done' and picking.stock_type == 'pos_exchange':
                picking.update_stock_adjust_valuation()
        return res

    @api.onchange('is_landed_cost')
    def _onchange_is_landed_cost(self):
        if self.is_landed_cost == 'yes':
            self.is_landed_cost_confirm = True
        else:
            self.is_landed_cost_confirm = False

    @api.onchange('stock_barcode')
    def _onchange_stock_barcode(self):
        if not self.stock_type:
            if self.stock_barcode:
                raise UserError('Prior to scanning a barcode, please select a Stock Type.')
        if self.stock_barcode:
            barcode = self.stock_barcode
            # Check for GS1 nomenclature
            gs1_pattern = r'01(\d{14})21([A-Za-z0-9]+)'
            gs1_match = re.match(gs1_pattern, barcode)
            # Check for EAN-13 format
            ean13_pattern = r'(\d{13})'
            ean13_match = re.match(ean13_pattern, barcode)
            custom_serial_pattern = r'^(R\d+)'

            def search_product(barcode_field, barcode_value):
                """Helper function to search in product.product and product.template"""
                product = self.env['product.product'].search([(barcode_field, '=', barcode_value)], limit=1)
                if not product:
                    # Search in product.template if not found in product.product
                    template = self.env['product.template'].search([(barcode_field, '=', barcode_value)], limit=1)
                    if template:
                        # Return the first product variant linked to the template
                        product = template.product_variant_id
                return product

            if re.match(gs1_pattern, barcode) and self.picking_type_code == 'delivery':
                product_barcode = re.match(gs1_pattern, barcode).group(1)
                serial_number = re.match(gs1_pattern, barcode).group(2)
                product = search_product('barcode', product_barcode)
                if product:
                    # Check if the serial number already exists in any line for the product
                    for line in self.move_ids_without_package:
                        if serial_number in line.dummy_lot_ids.mapped('name') or serial_number in line.lot_ids.mapped(
                                'name'):
                            raise UserError(f'The serial number {serial_number} is already assigned to a product line.')

                    # Check if the serial number exists in any stock picking with 'internal' or 'outgoing' type
                    pickings_with_serial = self.env['stock.picking'].search([
                        ('picking_type_code', 'in', [ 'outgoing']),
                        ('state', '!=', 'cancel'),('move_line_ids.lot_id.name', '=', serial_number),])
                    if pickings_with_serial:
                        picking_names = ', '.join(pickings_with_serial.mapped('name'))
                        raise UserError(
                            f'The serial number {serial_number} is already assigned in the following pickings: {picking_names}.')
                    # If validation passes, add or update the line
                    existing_line = self.move_ids_without_package.filtered(lambda l: l.product_id == product)
                    if existing_line:
                        # Add the serial number (stock.lot) to the dummy field if not already added
                        lot = self.env['stock.lot'].search(
                            [('product_id', '=', product.id), ('name', '=', serial_number)], limit=1)
                        if not lot:
                            raise UserError(f'No serial number found for {serial_number}.' )
                        if lot not in existing_line.dummy_lot_ids:
                            existing_line.dummy_lot_ids = [(4, lot.id)]
                        existing_line.product_uom_qty = len(existing_line.dummy_lot_ids)
                    else:
                        # Add new line for product and assign serial number
                        lot = self.env['stock.lot'].search(
                            [('product_id', '=', product.id), ('name', '=', serial_number)], limit=1)
                        if not lot:
                            raise UserError(f'No serial number found for {serial_number}.')
                        self.move_ids_without_package = [(0, 0, {
                            'product_id': product.id,
                            'product_uom_qty': 1,
                            'dummy_lot_ids': [(4, lot.id)],
                            'location_id': self.location_id.id,
                            'location_dest_id': self.location_dest_id.id,
                            'name': product.display_name,
                        })]
                else:
                    raise UserError(f'No products found with barcode {product_barcode}')
            elif re.match(gs1_pattern, barcode) and self.picking_type_code == 'internal':
                product_barcode = re.match(gs1_pattern, barcode).group(1)
                serial_number = re.match(gs1_pattern, barcode).group(2)
                product = search_product('barcode', product_barcode)
                if product:
                    # Check if the serial number already exists in any line for the product
                    for line in self.move_ids_without_package:
                        if serial_number in line.dummy_lot_ids.mapped('name') or serial_number in line.lot_ids.mapped(
                                'name'):
                            raise UserError(f'The serial number {serial_number} is already assigned to a product line.')

                    # Check if the serial number exists in any stock picking with 'internal' or 'outgoing' type
                    pickings_with_serial = self.env['stock.picking'].search([
                        ('picking_type_code', 'in', [ 'internal']),
                        ('state', '!=', 'cancel'),('move_line_ids.lot_id.name', '=', serial_number),])
                    if pickings_with_serial:
                        picking_names = ', '.join(pickings_with_serial.mapped('name'))
                        raise UserError(
                            f'The serial number {serial_number} is already assigned in the following pickings: {picking_names}.')
                    # If validation passes, add or update the line
                    existing_line = self.move_ids_without_package.filtered(lambda l: l.product_id == product)
                    if existing_line:
                        # Add the serial number (stock.lot) to the dummy field if not already added
                        lot = self.env['stock.lot'].search(
                            [('product_id', '=', product.id), ('name', '=', serial_number)], limit=1)
                        if not lot:
                            raise UserError(f'No serial number found for {serial_number}.' )
                        if lot not in existing_line.dummy_lot_ids:
                            existing_line.dummy_lot_ids = [(4, lot.id)]
                        existing_line.product_uom_qty = len(existing_line.dummy_lot_ids)
                    else:
                        # Add new line for product and assign serial number
                        lot = self.env['stock.lot'].search(
                            [('product_id', '=', product.id), ('name', '=', serial_number)], limit=1)
                        if not lot:
                            raise UserError(f'No serial number found for {serial_number}.')
                        self.move_ids_without_package = [(0, 0, {
                            'product_id': product.id,
                            'product_uom_qty': 1,
                            'dummy_lot_ids': [(4, lot.id)],
                            'location_id': self.location_id.id,
                            'location_dest_id': self.location_dest_id.id,
                            'name': product.display_name,
                        })]
                else:
                    raise UserError(f'No products found with barcode {product_barcode}')
            elif ean13_match and self.picking_type_code == 'outgoing':
                ean13_barcode = ean13_match.group(1)
                # Search for all lots that match the EAN-13 barcode in the ref field
                lots = self.env['stock.lot'].search([('ref', '=', ean13_barcode), ('product_qty', '>', 0)])

                if lots:
                    product = lots[0].product_id
                    if not product:
                        raise UserError( f"No product is associated with lots for barcode {ean13_barcode}.")
                    # Get serial numbers that are used in stock.picking with 'internal' or 'delivery' types, and not in 'cancel' state
                    used_serials = self.env['stock.picking'].search([
                        ('picking_type_code', 'in', ['outgoing']),
                        ('state', '!=', 'cancel'),
                        ('move_ids_without_package.dummy_lot_ids', 'in', lots.ids),
                    ]).mapped('move_ids_without_package.dummy_lot_ids.name')

                    # Filter out lots that are either assigned in the current picking or used in stock.picking
                    assigned_serial_numbers = self.move_ids_without_package.mapped('dummy_lot_ids.name')
                    available_lots = lots.filtered(
                        lambda l: l.name not in assigned_serial_numbers and l.name not in used_serials)

                    if not available_lots:
                        raise UserError(
                            f'The serial numbers for barcode {ean13_barcode} are already issued or used in stock.picking')

                    # Take the next available lot
                    next_lot = available_lots[0]

                    # Check if the product already exists in the picking lines
                    existing_line = self.move_ids_without_package.filtered(lambda l: l.product_id == product)

                    if existing_line:
                        # Add the next available lot to the existing line
                        existing_line.dummy_lot_ids = [(4, next_lot.id)]
                        # Increment the quantity
                        existing_line.product_uom_qty = len(existing_line.dummy_lot_ids)
                    else:
                        # Add a new line with the product, assign the lot, and set product_uom_qty to 1
                        self.move_ids_without_package = [(0, 0, {
                            'product_id': product.id,
                            'product_uom_qty': 1,
                            'location_id': self.location_id.id,
                            'location_dest_id': self.location_dest_id.id,
                            'name': product.display_name,  # Set the name field
                            'dummy_lot_ids': [(4, next_lot.id)],  # Assign the next available lot
                        })]
                else:
                    raise UserError( f"No lots found with EAN-13 barcode {ean13_barcode} or insufficient quantity")
            elif ean13_match and self.picking_type_code == 'internal':
                ean13_barcode = ean13_match.group(1)
                # Search for all lots that match the EAN-13 barcode in the ref field
                lots = self.env['stock.lot'].search([('ref', '=', ean13_barcode), ('product_qty', '>', 0)])

                if lots:
                    product = lots[0].product_id
                    if not product:
                        raise UserError( f"No product is associated with lots for barcode {ean13_barcode}.")
                    # Get serial numbers that are used in stock.picking with 'internal' or 'delivery' types, and not in 'cancel' state
                    used_serials = self.env['stock.picking'].search([
                        ('picking_type_code', 'in', ['internal']),
                        ('state', '!=', 'cancel'),
                        ('move_ids_without_package.dummy_lot_ids', 'in', lots.ids),
                    ]).mapped('move_ids_without_package.dummy_lot_ids.name')

                    # Filter out lots that are either assigned in the current picking or used in stock.picking
                    assigned_serial_numbers = self.move_ids_without_package.mapped('dummy_lot_ids.name')
                    available_lots = lots.filtered(
                        lambda l: l.name not in assigned_serial_numbers and l.name not in used_serials)

                    if not available_lots:
                        raise UserError(
                            f'The serial numbers for barcode {ean13_barcode} are already issued or used in stock.picking')

                    # Take the next available lot
                    next_lot = available_lots[0]

                    # Check if the product already exists in the picking lines
                    existing_line = self.move_ids_without_package.filtered(lambda l: l.product_id == product)

                    if existing_line:
                        # Add the next available lot to the existing line
                        existing_line.dummy_lot_ids = [(4, next_lot.id)]
                        # Increment the quantity
                        existing_line.product_uom_qty = len(existing_line.dummy_lot_ids)
                    else:
                        # Add a new line with the product, assign the lot, and set product_uom_qty to 1
                        self.move_ids_without_package = [(0, 0, {
                            'product_id': product.id,
                            'product_uom_qty': 1,
                            'location_id': self.location_id.id,
                            'location_dest_id': self.location_dest_id.id,
                            'name': product.display_name,  # Set the name field
                            'dummy_lot_ids': [(4, next_lot.id)],  # Assign the next available lot
                        })]
                else:
                    raise UserError( f"No lots found with EAN-13 barcode {ean13_barcode} or insufficient quantity")
            elif re.match(custom_serial_pattern, barcode) and self.picking_type_code == 'outgoing':
                # Handle custom serial numbers that start with R1, R2, R3, etc.
                prefix = re.match(custom_serial_pattern, barcode).group(1)
                # Search for a lot with this prefix in the stock.lot model
                lot = self.env['stock.lot'].search([('name', '=like', f'{prefix}%')], limit=1)
                if lot:
                    product = lot.product_id
                    # Validation: Check if the serial number is already used in the current order
                    for line in self.move_ids_without_package:
                        if lot.name in line.lot_ids.mapped('name'):
                            raise UserError(
                                f'Serial number {lot.name} is already assigned to a product line in the current order.')
                    # Validation: Check if the serial number is used in a 'stock.picking' with type 'internal' or 'outgoing'
                    pickings_with_serial = self.env['stock.picking'].search([
                        ('picking_type_id.code', 'in', ['outgoing']), ('state', '!=', 'cancel'),
                        ('move_line_ids.lot_id.name', '=', lot.name), ])
                    if pickings_with_serial:
                        picking_names = ', '.join(pickings_with_serial.mapped('name'))
                        raise UserError(
                            f'The serial number {lot.name} is already assigned in the pickings: {picking_names}.' )
                    if product:
                        # Check if the product already exists in the order lines
                        existing_line = self.move_ids_without_package.filtered(lambda l: l.product_id == product)
                        if existing_line:
                            # Add the new lot to the existing line
                            existing_line.dummy_lot_ids = [(4, lot.id)]
                            # Update product_uom_qty based on the number of lots
                            existing_line.product_uom_qty = len(existing_line.dummy_lot_ids)
                        else:
                            # Add a new line with the product and set product_uom_qty to 1
                            self.move_ids_without_package = [(0, 0, {
                                'product_id': product.id,
                                'product_uom_qty': 1,
                                'dummy_lot_ids': [(4, lot.id)],
                                'location_id': self.location_id.id,
                                'location_dest_id': self.location_dest_id.id,
                                'name': product.display_name,  # Set the name field
                            })]
                    else:
                        raise UserError(f'No product found for the lot with serial number prefix {prefix}')
                else:
                    raise UserError(f'No lot found with serial number prefix {prefix}')
            elif re.match(custom_serial_pattern, barcode) and self.picking_type_code == 'internal':
                # Handle custom serial numbers that start with R1, R2, R3, etc.
                prefix = re.match(custom_serial_pattern, barcode).group(1)
                # Search for a lot with this prefix in the stock.lot model
                lot = self.env['stock.lot'].search([('name', '=like', f'{prefix}%')], limit=1)
                if lot:
                    product = lot.product_id
                    # Validation: Check if the serial number is already used in the current order
                    for line in self.move_ids_without_package:
                        if lot.name in line.lot_ids.mapped('name'):
                            raise UserError(
                                f'Serial number {lot.name} is already assigned to a product line in the current order.')
                    # Validation: Check if the serial number is used in a 'stock.picking' with type 'internal' or 'outgoing'
                    pickings_with_serial = self.env['stock.picking'].search([
                        ('picking_type_id.code', 'in', ['internal']), ('state', '!=', 'cancel'),
                        ('move_line_ids.lot_id.name', '=', lot.name), ])
                    if pickings_with_serial:
                        picking_names = ', '.join(pickings_with_serial.mapped('name'))
                        raise UserError(
                            f'The serial number {lot.name} is already assigned in the pickings: {picking_names}.' )
                    if product:
                        # Check if the product already exists in the order lines
                        existing_line = self.move_ids_without_package.filtered(lambda l: l.product_id == product)
                        if existing_line:
                            # Add the new lot to the existing line
                            existing_line.dummy_lot_ids = [(4, lot.id)]
                            # Update product_uom_qty based on the number of lots
                            existing_line.product_uom_qty = len(existing_line.dummy_lot_ids)
                        else:
                            # Add a new line with the product and set product_uom_qty to 1
                            self.move_ids_without_package = [(0, 0, {
                                'product_id': product.id,
                                'product_uom_qty': 1,
                                'dummy_lot_ids': [(4, lot.id)],
                                'location_id': self.location_id.id,
                                'location_dest_id': self.location_dest_id.id,
                                'name': product.display_name,  # Set the name field
                            })]
                    else:
                        raise UserError(f'No product found for the lot with serial number prefix {prefix}')
                else:
                    raise UserError(f'No lot found with serial number prefix {prefix}')
            else:
                raise UserError('Invalid barcode format')

            # Clear the barcode field after processing
            self.stock_barcode = False

    @api.onchange('verify_barcode')
    def _onchange_verify_barcode(self):
        if self.verify_barcode:
            barcode = self.verify_barcode
            # Patterns for barcode formats
            gs1_pattern = r'01(\d{14})21([A-Za-z0-9]+)'
            ean13_pattern = r'(\d{13})'
            custom_serial_pattern = r'^(R\d+)'

            if re.match(gs1_pattern, barcode):
                # GS1 Barcode
                product_barcode = re.match(gs1_pattern, barcode).group(1)
                code = re.match(gs1_pattern, barcode).group(2)
                matched_line = self.move_line_ids_without_package.filtered(lambda x: x.lot_id.name == code)
                if not matched_line:
                    raise ValidationError('No matching product or serial/lot number found.')
                tracking_type = matched_line[0].product_id.tracking
                product_id = matched_line[0].product_id.id

                if tracking_type == 'serial':
                    # Serial-tracked: ensure uniqueness
                    existing_line = self.stock_verification_ids.filtered(
                        lambda x: x.stock_serial and code in x.stock_serial.split(',')
                    )
                    if existing_line:
                        raise ValidationError(f'Serial number {code} is already added.')
                    self.stock_verification_ids = [(0, 0, {
                        'stock_product_id': product_id,
                        'stock_serial': code,
                        'stock_qty': 1,
                    })]
                elif tracking_type == 'lot':
                    # Lot-tracked: check available quantity from move lines
                    product_qty = sum(matched_line.mapped('quantity'))
                    existing_product_line = self.stock_verification_ids.filtered(
                        lambda x: x.stock_product_id.id == product_id and x.stock_serial == code
                    )
                    if existing_product_line:
                        if existing_product_line.stock_qty + 1 > product_qty:
                            raise ValidationError(
                                f"Scanned quantity exceeds allowed quantity for product '{matched_line[0].product_id.display_name}'."
                            )
                        existing_product_line.stock_qty += 1
                    else:
                        if 1 > product_qty:
                            raise ValidationError(
                                f"Scanned quantity exceeds allowed quantity for product '{matched_line[0].product_id.display_name}'."
                            )
                        self.stock_verification_ids = [(0, 0, {
                            'stock_product_id': product_id,
                            'stock_serial': code,  # Optionally store the lot number
                            'stock_qty': 1,
                        })]

            elif re.match(ean13_pattern, barcode):
                # EAN-13 Barcode
                ean13_barcode = re.match(ean13_pattern, barcode).group(1)
                matched_line = self.move_line_ids_without_package.filtered(
                    lambda x: x.internal_ref_lot == ean13_barcode
                )
                if not matched_line:
                    raise ValidationError('No matching product found for the EAN-13 barcode.')
                tracking_type = matched_line[0].product_id.tracking
                product_id = matched_line[0].product_id.id
                product_qty = sum(self.move_line_ids_without_package.filtered(
                    lambda x: x.internal_ref_lot == ean13_barcode
                ).mapped('quantity'))

                if tracking_type == 'serial':
                    existing_product_qty = sum(self.stock_verification_ids.filtered(
                        lambda x: x.stock_product_id.id == product_id
                    ).mapped('stock_qty'))
                    if existing_product_qty + 1 > product_qty:
                        raise ValidationError(
                            f"For Branded Article '{matched_line[0].product_id.display_name}', the quantities are matched."
                        )
                    self.stock_verification_ids = [(0, 0, {
                        'stock_product_id': product_id,
                        'stock_serial': ean13_barcode,
                        'stock_qty': 1,
                    })]
                elif tracking_type == 'lot':
                    existing_product_line = self.stock_verification_ids.filtered(
                        lambda x: x.stock_product_id.id == product_id and x.stock_serial == ean13_barcode
                    )
                    if existing_product_line:
                        if existing_product_line.stock_qty + 1 > product_qty:
                            raise ValidationError(
                                f"Scanned quantity exceeds allowed quantity for product '{matched_line[0].product_id.display_name}'."
                            )
                        existing_product_line.stock_qty += 1
                    else:
                        if 1 > product_qty:
                            raise ValidationError(
                                f"Scanned quantity exceeds allowed quantity for product '{matched_line[0].product_id.display_name}'."
                            )
                        self.stock_verification_ids = [(0, 0, {
                            'stock_product_id': product_id,
                            'stock_serial': ean13_barcode,  # Optionally store the lot number
                            'stock_qty': 1,
                        })]

            elif re.match(custom_serial_pattern, barcode):
                # Custom Serial Barcode
                code = re.match(custom_serial_pattern, barcode).group(1)
                matched_line = self.move_line_ids_without_package.filtered(lambda x: x.lot_id.name == code and x.type_product == 'un_brand')
                if not matched_line:
                    # Fallback to internal_ref_lot check
                    matched_line = self.move_line_ids_without_package.filtered(lambda x: x.internal_ref_lot == code and x.type_product == 'brand')
                    if not matched_line:
                        raise ValidationError('No matching product or serial/lot number found.')
                    tracking_type = matched_line[0].product_id.tracking
                    product_id = matched_line[0].product_id.id
                    product_qty = sum(matched_line.mapped('quantity'))
                    # Same logic as EAN-13 fallback
                    existing_product_line = self.stock_verification_ids.filtered(
                        lambda x: x.stock_product_id.id == product_id and x.stock_serial == code)
                    if existing_product_line:
                        if existing_product_line.stock_qty + 1 > product_qty:
                            raise ValidationError(
                                f"Scanned quantity exceeds allowed quantity for product '{matched_line[0].product_id.display_name}'.")
                        existing_product_line.stock_qty += 1
                    else:
                        if 1 > product_qty:
                            raise ValidationError(
                                f"Scanned quantity exceeds allowed quantity for product '{matched_line[0].product_id.display_name}'.")
                        self.stock_verification_ids = [(0, 0, {
                            'stock_product_id': product_id,
                            'stock_serial': code,
                            'stock_qty': 1,
                        })]
                else:
                    # Existing logic for lot_name match
                    tracking_type = matched_line[0].product_id.tracking
                    product_id = matched_line[0].product_id.id
                    if tracking_type == 'serial':
                        existing_line = self.stock_verification_ids.filtered(
                            lambda x: x.stock_serial and code in x.stock_serial.split(','))
                        if existing_line:
                            raise ValidationError(f'Serial number {code} is already added.')
                        self.stock_verification_ids = [(0, 0, {
                            'stock_product_id': product_id,
                            'stock_serial': code,
                            'stock_qty': 1,
                        })]
                    elif tracking_type == 'lot':
                        product_qty = sum(matched_line.mapped('quantity'))
                        existing_product_line = self.stock_verification_ids.filtered(
                            lambda x: x.stock_product_id.id == product_id and x.stock_serial == code)
                        if existing_product_line:
                            if existing_product_line.stock_qty + 1 > product_qty:
                                raise ValidationError(
                                    f"Scanned quantity exceeds allowed quantity for product '{matched_line[0].product_id.display_name}'.")
                            existing_product_line.stock_qty += 1
                        else:
                            if 1 > product_qty:
                                raise ValidationError(
                                    f"Scanned quantity exceeds allowed quantity for product '{matched_line[0].product_id.display_name}'.")
                            self.stock_verification_ids = [(0, 0, {
                                'stock_product_id': product_id,
                                'stock_serial': code,
                                'stock_qty': 1,
                            })]
            else:
                raise ValidationError('Invalid barcode format.')
            self.verify_barcode = False

    def action_confirm(self):
        res = super(Picking, self).action_confirm()
        for picking in self:
            if picking.picking_type_id.code == 'internal':
                for move in picking.move_ids_without_package:
                    if move.dummy_lot_ids:
                        move.lot_ids = [(6, 0, move.dummy_lot_ids.ids)]
            if picking:
                for line in picking.move_ids_without_package:
                    if line.product_id.nhcl_product_type and line.product_id.nhcl_product_type == 'unbranded':
                        line.type_product = 'un_brand'
                    elif line.product_id.nhcl_product_type and line.product_id.nhcl_product_type == 'branded':
                        line.type_product = 'brand'
        return res

    def reset_product_lines(self):
        for rec in self:
            rec.move_ids_without_package.unlink()

    def _compute_check(self):
        for picking in self:
            if (picking.purchase_id and picking.purchase_id.auto_sale_order_id and picking.purchase_id.auto_sale_order_id.picking_ids.batch_id and
                    any(keyword in picking.purchase_id.auto_sale_order_id.picking_ids.batch_id.mapped('name')[0] for keyword in ['return', 'damage'])):
                picking.check_ids.sudo().unlink()
                picking.quality_check_fail = False
                picking.quality_check_todo = False
            else:
                super(Picking, picking)._compute_check()


    def update_stock_valuation(self):
        for pick in self:
            for move in pick.move_ids_without_package:
                valuaton = self.env['stock.valuation.layer'].search([('stock_move_id','=', move.id)])
                if valuaton:
                    valuaton.unit_cost = move.move_cp
                    valuaton.value = move.move_cp * move.quantity
                    valuaton.remaining_value = move.move_cp * valuaton.remaining_qty

    def update_stock_adjust_valuation(self):
        for pick in self:
            for move in pick.move_line_ids_without_package:
                valuaton = self.env['stock.valuation.layer'].search([('stock_move_id', '=', move.move_id.id)])
                if valuaton:
                    valuaton.unit_cost = move.cost_price
                    valuaton.value = move.cost_price * valuaton.quantity
                    valuaton.remaining_value = move.cost_price * valuaton.remaining_qty


    def _create_backorder(self):
        res = super(Picking,self)._create_backorder()
        if res.backorder_id:
            for main_move in res.backorder_id.move_ids_without_package:
                for back_move in res.move_ids_without_package:
                    if back_move.product_id.id == main_move.product_id.id:
                        if main_move.product_id.nhcl_product_type == 'unbranded':
                            back_move.type_product = 'un_brand'
                        elif main_move.product_id.nhcl_product_type == 'branded':
                            back_move.type_product = 'brand'
        return res


    def print_barcodes_preview(self):
        if self.is_landed_cost == 'yes':
            landed_costs = self.env['stock.landed.cost'].search([
                ('picking_ids', 'in', [self.id]),
                ('state', '=', 'done')
            ])

            if not landed_costs:
                raise UserError("Landed cost is not posted for this receipt.")
        for line in self.move_line_ids:
            line.nhcl_name = "01" + str(line.product_id.barcode or '') + "21" + str(line.lot_id.name or '')
            print("--------------------------------",line.nhcl_name)

        report_name = 'cmr_customizations.nhcl_report_productlabel_brand'
        return {
            'type': 'ir.actions.report',
            'report_name': report_name,
            'report_type': 'qweb-pdf',
        }

    def print_ready_made_barcodes_preview(self):
        if self.is_landed_cost == 'yes':
            landed_costs = self.env['stock.landed.cost'].search([
                ('picking_ids', 'in', [self.id]),
                ('state', '=', 'done')
            ])

            if not landed_costs:
                raise UserError("Landed cost is not posted for this receipt.")
        for line in self.move_line_ids:
            line.nhcl_name = "01" + str(line.product_id.barcode) + "21" + line.lot_id.name
        report_name = 'cmr_customizations.nhcl_report_product_template_label_ready_made'
        # return {
        #     'type': 'ir.actions.report',
        #     'report_name': report_name,
        #     'report_type': 'qweb-pdf',
        # }
        return self.env.ref('cmr_customizations.nhcl_report_product_template_label_ready_made').report_action(self)


    def print_barcodes(self):
        self.ensure_one()
        zpl_data = ""

        for line in self.move_line_ids:
            line.nhcl_name = "01" + str(line.product_id.barcode or '') + "21" + (line.lot_id.name or '')

            description_1 = line.descrip_1.name if line.descrip_1 else ''
            category_2 = line.categ_2.name if line.categ_2 else ''
            category_3 = line.categ_3.name if line.categ_3 else ''
            description_8 = line.descrip_8.name if line.descrip_8 else ''
            mr_price = int(line.mr_price)
            rs_price = int(line.rs_price)
            product_name = line.product_id.display_name or ''
            wrapped_name = textwrap.wrap(product_name, width=38)
            zpl_data += f"""^XA
       ^PW1000
       ^LL400
       ^FWN
       ^CI28
       ^PR3
       """
            # Product Name
            zpl_data += "^CF0,30,30\n"
            if wrapped_name:
                zpl_data += f"^FO150,20^FD{wrapped_name[0]}^FS\n"
            if len(wrapped_name) > 1:
                zpl_data += f"^FO150,50^FD{wrapped_name[1]}^FS\n"
            if len(wrapped_name) > 2:
                zpl_data += f"^FO150,80^FD{wrapped_name[2]}^FS\n"
            # Description (Aging)
            if description_1:
                zpl_data += "^CF0,22,22\n"
                zpl_data += f"^FO700,100^FD{description_1}^FS\n"
            # Barcode + Human-readable
            zpl_data += f"""^CF0,30,30
       ^FO150,150^BY2,2,300
       ^BCN,50,N,Y,N

       ^FD{line.nhcl_name}^FS
       ^FO200,230^FD{line.nhcl_name}^FS
       """

            # Brand, Offer, MRP
            zpl_data += "^CF0,35,35\n"
            if line.categ_2 and line.categ_2.name:
                zpl_data += f"^FO150,320^FD{line.categ_2.name}^FS\n"
            zpl_data += f"^FO150,100^FD{description_8}^FS\n"
            zpl_data += f"^FO500,340^FDMRP: ₹{mr_price}/-^FS\n"

            # Fit and RSP
            zpl_data += "^CF0,35,35\n"
            if line.categ_3 and line.categ_3.name:
                zpl_data += f"^FO150,380^FD{line.categ_3.name}^FS\n"

            zpl_data += f"^FO500,420^FDRSP: ₹{rs_price}/-^FS\n"

            # Fit and RSP
            zpl_data += "^CF0,35,35\n"
            if line.categ_4 and line.categ_4.name:
                zpl_data += f"^FO150,420^FD{line.categ_4.name}^FS\n"

            zpl_data += "^XZ\n"
            return zpl_data

        # print("✅ Final ZPL Data:\n", zpl_data)

        # Send to printer
        # printer_ip = "192.168.168.100"
        # printer_port = 9100
        # try:
        #     with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as s:
        #         s.connect((printer_ip, printer_port))
        #         s.sendall(zpl_data.encode("utf-8"))
        # except Exception as e:
        #     raise UserError(f"⚠️ Zebra Printer Connection Error:\n{str(e)}")
        #
        # return {
        #     'type': 'ir.actions.client',
        #     'tag': 'display_notification',
        #     'params': {
        #         'title': '✅ ZPL Sent',
        #         'message': 'DYMO Labels sent successfully to Zebra printer.',
        #         'type': 'success',
        #         'sticky': False,
        #     }
        # }

    def print_ready_made_barcodes(self):
        self.ensure_one()
        zpl_data = ""

        for line in self.move_line_ids:
            # Step 0: Generate encoded barcode GS1 format
            line.nhcl_name = line.internal_ref_lot

            # Debug Print Summary for the line
            print("🧾 GENERATING ZPL FOR LINE", line.id)
            print("🆔 Product Name       :", line.product_id.display_name)
            # print("📦 Product Barcode    :", line.product_id.barcode)
            print("🏷️  NHCL Name (GS1)   :", line.nhcl_name)
            print("📝 Description         :", line.descrip_1.name if line.descrip_1 else "None")
            print("🎨 Category 1          :", line.categ_1.name if line.categ_1 else "None")
            print("🎨 Category 2          :", line.categ_2.name if line.categ_2 else "None")
            print("🎨 Category 3          :", line.categ_3.name if line.categ_3 else "None")
            print("🎨 Category 4          :", line.categ_4.name if line.categ_4 else "None")
            print("📐 Size                :", line.categ_7.name if line.categ_7 else "None")
            print("📦 Quantity            :", line.quantity)
            print("📅 Mfd.Date            :", line.mfd_date() if line.lot_id else "None")
            print("💰 MRP                 :", line.mr_price)
            print("💰 RSP                 :", line.rs_price)
            print("🏷️  Lot Name           :", line.lot_id.name if line.lot_id else "None")
            print("🏭 Manufacturer        :", self.partner_id.name or '')
            print("📍 Address             :", self.partner_id.street or '')
            print("📞 Phone               :", self.partner_id.phone or '')
            print("✉️  Email              :", self.partner_id.email or '')
            print("-" * 60)

            # Wrap product name for display
            product_name = line.product_id.display_name or ''
            wrapped_lines = textwrap.wrap(product_name, width=40)
            street_name = self.partner_id.street or ''
            wrapped_lines_street = textwrap.wrap(street_name, width=40)
            manufacture_name = self.partner_id.name or ''
            wrapped_lines_manufacture = textwrap.wrap(manufacture_name, width=40)
            size_text = line.categ_7.name or ''
            wrapped_size = textwrap.wrap(size_text, width=40, break_long_words=True)

            zpl_data += f"""
               ^XA
       ^PW1100
       ^LL400
       ^FWB
       ^CI28

       ^PR3
       ^CF0,22
       ^FO0,460^GB1000,3,3,B,0^FS      ; Top border
        ^FO100,15^GB3,1000,3,B,0^FS       ; Left border
        ^FO900,15^GB3,1000,3,B,0^FS     ; Right border (1000 - 3)
        ^FO0,15^GB1000,3,3,B,0^FS    ; Bottom border (660 - 3)

       """

            # Size, Quantity, Color
            #      zpl_data += f"""^CF0,22,22
            # ^FO180,280^FDSize: {line.categ_7.name or ''}^FS
            # ^CF0,26,26
            # ^FO180,220^FDNETQ: {int(line.quantity)}N^FS
            # ^CF0,26,26
            # ^FO180,40^FDColor: {line.categ_1.name or ''}^FS
            # ^FO240,15^GB3,1000,3,B,0^FS
            # """
            zpl_data += "^CF0,26,26\n"

            # Print wrapped Size (max 2 lines), starting at Y = 280
            for j, part in enumerate(wrapped_size[:2]):
                y_position = 320  # More spacing to fit two lines (label + value)
                zpl_data += f"^FO120,{y_position}^FDSize:^FS\n"
                zpl_data += f"^FO150,{y_position}^FD{part}^FS\n"

            # Color field (label + value on next line)
            zpl_data += (
                f"^CF0,26,26\n"
                f"^FO120,100^FDColor:^FS\n"
                f"^FO150,100^FD{line.categ_1.name or ''}^FS\n"
            )

            # NETQ field (label + value on next line)
            zpl_data += (
                f"^CF0,26,26\n"
                f"^FO120,220^FDNETQ:^FS\n"
                f"^FO150,220^FD{int(line.quantity)}N^FS\n"
            )

            # Mfd Date and Description
            zpl_data += f"""^CF0,30,30
       ^FO210,240^FDMfd.Date: {line.mfd_date() if line.lot_id else ''}^FS
       ^CF0,33,33
       ^FO210,100^FD{line.descrip_1.name if line.descrip_1 else ''}^FS
       ^FO240,15^GB3,1000,3,B,0^FS 
       """

            # Product Name (wrapped lines)
            zpl_data += "^CF0,20,20\n"
            for i, part in enumerate(wrapped_lines[:2]):
                zpl_data += f"^FO{260 + i * 25},25^FD{part}^FS\n"

            # Barcode & Human-readable
            zpl_data += f"""
                          ^FO350,45
                          ^BY2,2,100
                          ^BCB,80,Y,N
                          ^FD{line.nhcl_name}^FS
                          ^FO320,20^GB3,1000,3,B,0^FS
                          ^FO480,20^GB3,1000,3,B,0^FS
                          ^FO550,20^GB3,1000,3,B,0^FS
                          ^FO610,20^GB3,1000,3,B,0^FS
                          ^FO650,20^GB3,1000,3,B,0^FS
                          ^FO690,20^GB3,1000,3,B,0^FS
                          ^FO180,15^GB3,1000,3,B,0^FS 
                          """

            # Manufacturer Details
            zpl_data += "^CF0,22,22\n"
            zpl_data += "^FO500,300^FDManufactured By:^FS\n"

            for i, part in enumerate(wrapped_lines_manufacture):
                y_offset = 40 + (i * 30)  # Adjust vertical spacing
                zpl_data += "^CF0,22,22\n"
                zpl_data += f"^FO530,{y_offset}^FD{part}^FS\n"

            zpl_data += "^CF0,22,22\n"
            for i, part in enumerate(wrapped_lines_street[:3]):
                zpl_data += f"^FO{565 + i * 25},25^FD{part}^FS\n"

            zpl_data += f"""

               ^FO630,100^FDCustomer Care: {self.partner_id.phone or ''}^FS
               ^FO670,100^FD{self.partner_id.email or ''}^FS
               """

            # MRP & RSP
            zpl_data += f"""^CF0,50,50
       ^FO760,340^FDMRP:^FS
       ^CF0,80,80
       ^FO750,30^FD₹{int(line.mr_price)} /-^FS
       ^FO820,20^GB3,1000,3,B,0^FS
       ^CF0,50,50
       ^FO840,340^FDRSP:^FS
       ^CF0,80,80
       ^FO830,30^FD₹{int(line.rs_price)} /-^FS

       ^XZ
       """

        return zpl_data

        # print("✅ Final ZPL Data:\n", zpl_data)

        # Send to Zebra printer
        # printer_ip = "192.168.168.100"  # Replace with your actual printer IP
        # printer_port = 9100
        #
        # try:
        #     with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as s:
        #         s.connect((printer_ip, printer_port))
        #         s.sendall(zpl_data.encode("utf-8"))
        # except Exception as e:
        #     raise UserError(f"⚠️ Zebra Printer Connection Error:\n{str(e)}")
        #
        # return {
        #     'type': 'ir.actions.client',
        #     'tag': 'display_notification',
        #     'params': {
        #         'title': '✅ ZPL Sent',
        #         'message': 'Ready Made Labels sent successfully to Zebra printer.',
        #         'type': 'success',
        #         'sticky': False,
        #     }
        # }

    def zpl_preview_barcodes(self):
        self.ensure_one()

        # Get ZPL data (barcode labels)
        zpl_data = self.print_barcodes()

        # Send to Labelary for PDF or image preview
        url = "http://api.labelary.com/v1/printers/8dpmm/labels/4x6/0/"
        headers = {"Accept": "application/pdf"}  # Change to application/pdf for PDF

        try:
            # Send ZPL data to Labelary
            response = requests.post(url, headers=headers, data=zpl_data.encode('utf-8'))
            if response.status_code != 200:
                raise UserError(f"Labelary Error: {response.text}")
            attachment = self.env['ir.attachment'].create({
                'name': 'ZPL Barcode Preview.pdf',
                'type': 'binary',
                'datas': base64.b64encode(response.content),
                'mimetype': 'application/pdf',
                'res_model': self._name,
                'res_id': self.id,
            })
            return {
                'type': 'ir.actions.act_url',
                'url': f'/web/content/{attachment.id}?raw=true',
                'target': 'new',
            }
        except Exception as e:
            raise UserError(f"Preview Failed:\n{str(e)}")

    def zpl_preview_ready_made(self):
        self.ensure_one()

        # Get ZPL data (barcode labels)
        zpl_data = self.print_ready_made_barcodes()

        # Send to Labelary for PDF or image preview
        url = "http://api.labelary.com/v1/printers/8dpmm/labels/4x6/0/"
        headers = {"Accept": "application/pdf"}  # Change to application/pdf for PDF

        try:
            # Send ZPL data to Labelary
            response = requests.post(url, headers=headers, data=zpl_data.encode('utf-8'))
            if response.status_code != 200:
                raise UserError(f"Labelary Error: {response.text}")
            attachment = self.env['ir.attachment'].create({
                'name': 'ZPL Ready Made Barcode Preview.pdf',
                'type': 'binary',
                'datas': base64.b64encode(response.content),
                'mimetype': 'application/pdf',
                'res_model': self._name,
                'res_id': self.id,
            })
            return {
                'type': 'ir.actions.act_url',
                'url': f'/web/content/{attachment.id}?raw=true',
                'target': 'new',
            }
        except Exception as e:
            raise UserError(f"Preview Failed:\n{str(e)}")

    def zpl_preview_dymo(self):
        self.ensure_one()

        # Get ZPL data (barcode labels)
        zpl_data = self.print_dymo()

        # Send to Labelary for PDF or image preview
        url = "http://api.labelary.com/v1/printers/8dpmm/labels/4x6/0/"
        headers = {"Accept": "application/pdf"}  # Change to application/pdf for PDF

        try:
            # Send ZPL data to Labelary
            response = requests.post(url, headers=headers, data=zpl_data.encode('utf-8'))
            if response.status_code != 200:
                raise UserError(f"Labelary Error: {response.text}")
            attachment = self.env['ir.attachment'].create({
                'name': 'ZPL General Preview.pdf',
                'type': 'binary',
                'datas': base64.b64encode(response.content),
                'mimetype': 'application/pdf',
                'res_model': self._name,
                'res_id': self.id,
            })
            return {
                'type': 'ir.actions.act_url',
                'url': f'/web/content/{attachment.id}?raw=true',
                'target': 'new',
            }
        except Exception as e:
            raise UserError(f"Preview Failed:\n{str(e)}")

    def print_dymo(self):
        self.ensure_one()
        zpl_data = ""

        for line in self.move_line_ids:
            line.nhcl_name = "01" + str(line.product_id.barcode or '') + "21" + (line.lot_id.name or '')

            description_1 = line.descrip_1.name if line.descrip_1 else ''
            category_2 = line.categ_2.name if line.categ_2 else ''
            category_3 = line.categ_3.name if line.categ_3 else ''
            description_8 = line.descrip_8.name if line.descrip_8 else ''
            mr_price = int(line.mr_price)
            rs_price = int(line.rs_price)
            product_name = line.product_id.display_name or ''
            wrapped_name = textwrap.wrap(product_name, width=38)

            print("🧾 LINE ID               :", line.id)
            print("📦 Product Name          :", product_name)
            print("📦 Barcode               :", line.product_id.barcode)
            print("🏷️  NHCL Name (GS1)      :", line.nhcl_name)
            print("📚 Aging (description_1) :", description_1)
            print("🏢 Brand (category_3)    :", category_3)
            print("💡 Offer (description_8) :", description_8)
            print("📐 Fit (category_2)      :", category_2)
            print("💰 MRP Price             :", mr_price)
            print("💰 RSP Price             :", rs_price)
            print("-" * 60)

            zpl_data += f"""^XA
       ^PW1000
       ^LL400
       ^FWN
       ^CI28
       ^PR3
       """

            # Product Name
            zpl_data += "^CF0,30,30\n"
            if wrapped_name:
                print("wrapped name[0]", wrapped_name[0])
                zpl_data += f"^FO150,20^FD{wrapped_name[0]}^FS\n"
            if len(wrapped_name) > 1:
                print("wrapped name[1]", wrapped_name[1])
                zpl_data += f"^FO150,50^FD{wrapped_name[1]}^FS\n"
            if len(wrapped_name) > 2:
                print("wrapped name[2]", wrapped_name[2])
                zpl_data += f"^FO150,80^FD{wrapped_name[2]}^FS\n"

            # Description (Aging)
            if description_1:
                zpl_data += "^CF0,22,22\n"
                zpl_data += f"^FO700,100^FD{description_1}^FS\n"
            # Barcode + Human-readable
            zpl_data += f"""^CF0,30,30
       ^FO150,150^BY1.5,2,300
       ^BCN,50,N,Y,N

       ^FD{line.nhcl_name}^FS
       ^FO200,230^FD{line.nhcl_name}^FS
       """

            # Brand, Offer, MRP
            zpl_data += "^CF0,35,35\n"
            zpl_data += f"^FO150,100^FD{category_3}^FS\n"
            zpl_data += f"^FO150,100^FD{description_8}^FS\n"
            zpl_data += f"^FO500,340^FDMRP: ₹{mr_price}/-^FS\n"

            # Fit and RSP
            zpl_data += "^CF0,35,35\n"
            zpl_data += f"^FO150,420^FD{category_2}^FS\n"
            zpl_data += f"^FO500,420^FDNET: ₹{rs_price}/-^FS\n"

            zpl_data += "^XZ\n"
            return zpl_data

        # print("✅ Final ZPL Data:\n", zpl_data)

        # Send to printer
        # printer_ip = "192.168.168.100"
        # printer_port = 9100
        # try:
        #     with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as s:
        #         s.connect((printer_ip, printer_port))
        #         s.sendall(zpl_data.encode("utf-8"))
        # except Exception as e:
        #     raise UserError(f"⚠️ Zebra Printer Connection Error:\n{str(e)}")
        #
        # return {
        #     'type': 'ir.actions.client',
        #     'tag': 'display_notification',
        #     'params': {
        #         'title': '✅ ZPL Sent',
        #         'message': 'DYMO Labels sent successfully to Zebra printer.',
        #         'type': 'success',
        #         'sticky': False,
        #     }
        # }

    def action_open_barcode_print_wizard(self):
        self.ensure_one()
        if self.is_landed_cost == 'yes':
            landed_costs = self.env['stock.landed.cost'].search([
                ('picking_ids', 'in', [self.id]),
                ('state', '=', 'done')
            ])

            if not landed_costs:
                raise UserError("Landed cost is not posted for this receipt.")
        return {
            'type': 'ir.actions.act_window',
            'name': 'Print Barcode Labels',
            'res_model': 'print.label',
            'view_mode': 'form',
            'target': 'new',
            'context': {
                'default_picking_id': self.id
            }
        }


class StockMove(models.Model):
    """Inherited stock.move class to add fields and functions"""
    _inherit = "stock.move"

    serial_no = fields.Char(string="Serial No")
    dummy_lot_ids = fields.Many2many('stock.lot', string='Dummy Lots')
    type_product = fields.Selection([('brand', 'Brand'), ('un_brand', 'UnBrand'), ('others', 'Others')],string='Brand Type', copy=False)
    move_brand_barcode = fields.Char(string="CMR Barcode", copy=False)
    move_cp = fields.Float(string="CMR CP", copy=False)
    move_mrp = fields.Float(string="CMR MRP", copy=False)
    move_rsp = fields.Float(string="CMR RSP", copy=False)
    move_design_id = fields.Many2one('product.attribute.value', string="Design",related='purchase_line_id.design_id')
    return_parent_qty = fields.Float('Return Original Qty', copy=False)
    return_pend_qty = fields.Float('Return Pending Qty', copy=False, compute='compute_nhcl_pending_qty', store=True)
    prod_barcode = fields.Char('Barcode', copy=False)
    zone_id = fields.Many2one('placement.master.data', string='Zone', copy=False)
    family = fields.Many2one('product.category', string="Family", domain="[('parent_id','=',False)]")
    category = fields.Many2one(
        'product.category',
        string="Category",
        domain="[('parent_id','=',family)]"
    )

    class_level_id = fields.Many2one(
        'product.category',
        string="Class",
        domain="[('parent_id','=',category)]"
    )

    brick = fields.Many2one(
        'product.category',
        string="Brick",
        domain="[('parent_id','=',class_level_id)]"
    )

    def _get_zone_id(self):
        """Get zone_id from the product's category or default to False."""
        for move in self:
            if move.product_id.categ_id:
                move.zone_id = move.product_id.categ_id.parent_id.parent_id.parent_id.zone_id.id if move.product_id.categ_id.parent_id.parent_id.parent_id.zone_id else False
                move.family = move.product_id.categ_id.parent_id.parent_id.parent_id.id
                move.category = move.product_id.categ_id.parent_id.parent_id.id
                move.class_level_id = move.product_id.categ_id.parent_id.id
                move.brick = move.product_id.categ_id.id
            else:
                move.zone_id = False

    @api.model
    def create(self, vals):
        move = super(StockMove, self).create(vals)
        move._get_zone_id()
        return move

    @api.depends('return_parent_qty','picking_id')
    def compute_nhcl_pending_qty(self):
        for move in self:
            if move.picking_id.stock_picking_type == 'goods_return' and move.return_parent_qty > 0:
                move.return_pend_qty =  move.return_parent_qty - move.product_uom_qty
            else:
                move.return_pend_qty = 0.0

    def _prepare_move_line_vals(self, *args, **kwargs):
        vals = super(StockMove, self)._prepare_move_line_vals(*args, **kwargs)
        # For serial products, you might split move lines per serial number (as handled elsewhere)
        sale_line = self.sale_line_id
        if sale_line:
            if self.product_id.tracking == 'lot' and sale_line.lot_ids:
                vals['lot_id'] = sale_line.lot_ids.id
            elif self.product_id.tracking == 'serial' and sale_line.lot_ids:
                vals['lot_id'] = sale_line.lot_ids[0].id
        return vals

    @api.model
    def _prepare_merge_moves_distinct_fields(self):
        distinct_fields = super(StockMove, self)._prepare_merge_moves_distinct_fields()
        distinct_fields.append('move_brand_barcode')
        return distinct_fields


    def action_import_excel(self):
        return {
            'type': 'ir.actions.act_window',
            'name': 'Import Excel',
            'res_model': 'manafacturing.import.wizard',
            'view_mode': 'form',
            'target': 'new',
            'context': {'default_manafacturing_id': self.id},
        }

    def _get_new_picking_values(self):
        res = super(StockMove, self)._get_new_picking_values()
        if res.get('origin'):
            sale_id = self.env['sale.order'].search([('name', '=', res.get('origin'))], limit=1)
            stock_operation_type = self.env['stock.picking.type'].search(
                [('stock_picking_type', '=', sale_id.stock_type), ('company_id', '=', sale_id.company_id.id)], limit=1)
            partner = sale_id.partner_id.name.split('-')
            customer_id = partner[-1]
            route = self.env['dev.routes.details'].search([('location_details_ids.destination_location_id.name', '=', customer_id), ('transpoter_id.name', '=', sale_id.transpoter_id.name)], limit=1)

            if sale_id:
                res.update({
                    'transpoter_id': sale_id.transpoter_id.id if sale_id.transpoter_id else False,
                    'stock_type': sale_id.so_type,
                    'transpoter_route_id': route.id,

                })
                if stock_operation_type:
                    res.update({
                        'picking_type_id': stock_operation_type.id,
                        'location_id': stock_operation_type.default_location_src_id.id

                    })
                for move in self:
                    for line in sale_id.order_line:
                        if move.product_id == line.product_id:
                            move.write({
                                'type_product': line.type_product,
                            })
                            if stock_operation_type:
                                move.write({
                                    'location_id': stock_operation_type.default_location_src_id.id

                                })
                        # for move_line in move.move_line_ids:
                        #     if move_line.product_id == line.product_id and move_line.lot_id.name == line.lot_ids.name:
                        #         move_line.sudo().write({
                        #             'internal_ref_lot': line.branded_barcode,
                        #         })
        return res

    @api.onchange('product_id')
    def _onchange_product_id_stock(self):
        if self.picking_id and not self.picking_id.stock_type:
            # Clear the product_id and raise an error if no stock_type is selected
            self.product_id = False
            raise UserError(
                "Before you can select a product, you must first choose a Stock Type."
            )


    def action_assign_serial(self):
        if self.picking_id.picking_type_id.code == 'incoming' and self.picking_id.stock_type in ['ho_operation','sub_contract','data_import']:
            raise UserError(
                _("You are not allowed to Assign the Serial Number For Product %s") % (self.product_id.name))
        else:
            return super(StockMove, self).action_assign_serial()

    def auto_generate_serial_numbers(self):
        for move in self:
            if move.product_id.tracking in ['serial','lot'] and move.picking_id.picking_type_id.code == 'incoming' and (move.picking_id.stock_type in ['ho_operation','sub_contract','data_import'] or move.is_subcontract):
                picking_receipt = move.picking_id
                auto_generate_seq = self.env['nhcl.master.sequence'].search(
                    [('nhcl_code', '=', 'Auto Serial Number'), ('nhcl_state', '=', 'activate')])
                lot_qty = move.quantity
                if auto_generate_seq:
                    if move.product_id.tracking == 'serial' and move.quantity > 0:
                        Form(self.env['stock.assign.serial'].with_context(
                            default_move_id=move.id,
                            default_next_serial_number=self.env['stock.lot']._get_next_serial(
                                picking_receipt.company_id,
                                move.product_id),
                            default_next_serial_count=move.quantity
                        )).save().generate_serial_numbers()
                    elif move.quantity > 0:
                        lot_qty = move.quantity
                        Form(self.env['stock.assign.serial'].with_context(
                            default_move_id=move.id,
                            default_next_serial_number=self.env['stock.lot']._get_next_serial(
                                picking_receipt.company_id,
                                move.product_id),
                            default_next_serial_count=1
                        )).save().generate_serial_numbers()
                    if move.move_line_ids:
                        if move.product_id.tracking == 'lot':
                            move.move_line_ids.quantity = lot_qty
                        if move.move_line_ids[-1].lot_id or move.move_line_ids[-1].lot_name:
                            lot_name = move.move_line_ids[-1].lot_name or move.move_line_ids[-1].lot_id.name
                            prefix = auto_generate_seq.nhcl_prefix
                            serial_no = lot_name.split(prefix)
                            if len(serial_no) > 1:
                                auto_generate_seq.nhcl_next_number = int(serial_no[1]) + 1
        return True

    def _action_done(self, cancel_backorder=False):
        for move in self:
            if move.product_id.tracking in ['serial', 'lot'] and move.picking_id.picking_type_id.code == 'incoming':
                if float(move.quantity) == sum(
                        move.move_line_ids.filtered(lambda x: x.lot_name == False).mapped('quantity')):
                    auto_generate = self.env['nhcl.master.sequence'].search(
                        [('nhcl_code', '=', 'Auto Serial Number'), ('nhcl_state', '=', 'activate')])
                    if auto_generate:
                        move.auto_generate_serial_numbers()
        return super()._action_done(cancel_backorder=cancel_backorder)

    def write(self, vals):
        res = super(StockMove, self).write(vals)
        if 'quantity' in vals or any(picking.is_confirm for picking in self.mapped('picking_id')):
            self.picking_id.is_confirm = False
        for move in self:
            for move_line in move.picking_id.move_line_ids_without_package:
                if move.picking_id.stock_type in ['ho_operation']:
                    if move.product_id == move_line.product_id:
                        if move.type_product == 'un_brand':
                            move_line.sudo().write({
                                'type_product': move.type_product,
                                'internal_ref_lot': move.product_id.barcode
                            })
                        elif move.type_product == 'brand':
                            move_line.write({
                                'type_product': move.type_product,
                            })

        return res


class StockMoveLine(models.Model):
    _inherit = "stock.move.line"

    nhcl_name = fields.Char(string="Name", copy=False)
    internal_ref_lot = fields.Char(string="Barcode", copy=False, tracking=True)
    type_product = fields.Selection([('brand', 'Brand'), ('un_brand', 'UnBrand'), ('others', 'Others')],string='Brand Type', copy=False)
    categ_1 = fields.Many2one('product.attribute.value', string='Color', copy=False,
                              domain=[('attribute_id.name', '=', 'Color')])
    categ_2 = fields.Many2one('product.attribute.value', string='Fit', copy=False,
                              domain=[('attribute_id.name', '=', 'Fit')])
    categ_3 = fields.Many2one('product.attribute.value', string='Brand', copy=False,
                              domain=[('attribute_id.name', '=', 'Brand')])
    categ_4 = fields.Many2one('product.attribute.value', string='Pattern', copy=False,
                              domain=[('attribute_id.name', '=', 'Pattern')])
    categ_5 = fields.Many2one('product.attribute.value', string='Border Type', copy=False,
                              domain=[('attribute_id.name', '=', 'Border Type')])
    categ_6 = fields.Many2one('product.attribute.value', string='Border Size', copy=False,
                              domain=[('attribute_id.name', '=', 'Border Size')])
    categ_7 = fields.Many2one('product.attribute.value', string='Size', copy=False,
                              domain=[('attribute_id.name', '=', 'Size')])
    categ_8 = fields.Many2one('product.attribute.value', string='Design', copy=False,
                              domain=[('attribute_id.name', '=', 'Design')])
    descrip_1 = fields.Many2one('product.aging.line', string="Product Aging", copy=False, related="lot_id.description_1")
    descrip_2 = fields.Many2one('product.attribute.value', string='Range', copy=False,
                                domain=[('attribute_id.name', '=', 'Range')])
    descrip_3 = fields.Many2one('product.attribute.value', string='Collection', copy=False,
                                domain=[('attribute_id.name', '=', 'Collection')])
    descrip_4 = fields.Many2one('product.attribute.value', string='Fabric', copy=False,
                                domain=[('attribute_id.name', '=', 'Fabric')])
    descrip_5 = fields.Many2one('product.attribute.value', string='Exclusive', copy=False,
                                domain=[('attribute_id.name', '=', 'Exclusive')])
    descrip_6 = fields.Many2one('product.attribute.value', string='Print', copy=False, domain=[('attribute_id.name', '=', 'Print')])
    descrip_7 = fields.Many2one('product.attribute.value', string='Days Ageing', copy=False, domain=[('attribute_id.name', '=', 'Days Ageing')])
    descrip_8 = fields.Many2one('product.attribute.value', string='Description 8', copy=False, domain=[('attribute_id.name', '=', 'Offer')])
    acutal_line_cp = fields.Float(string='Actual CP', related='lot_id.actual_cp' , copy=False)
    cost_price = fields.Float(string='CP', copy=False)
    mr_price = fields.Float(string='MRP', copy=False)
    rs_price = fields.Float(string='RSP', copy=False)
    approval_margin = fields.Integer(related='move_id.purchase_line_id.purchase_rsp_margin', string="App Margin",
                                     copy=False)
    segment = fields.Selection([('apparel', 'Apparel'), ('non_apparel', 'Non Apparel'), ('others', 'Others')],
                               string="segment", copy=False, store=True, related='product_id.segment')

    picking_type_id = fields.Many2one(
        'stock.picking.type', 'Operation type', compute='_compute_picking_type_id', search='_search_picking_type_id',
        store=True)
    zone_id = fields.Many2one('placement.master.data', string='Zone', copy=False)
    family = fields.Many2one('product.category', string="Family", domain="[('parent_id','=',False)]")
    category = fields.Many2one(
        'product.category',
        string="Category",
        domain="[('parent_id','=',family)]"
    )

    class_level_id = fields.Many2one(
        'product.category',
        string="Class",
        domain="[('parent_id','=',category)]"
    )

    brick = fields.Many2one(
        'product.category',
        string="Brick",
        domain="[('parent_id','=',class_level_id)]"
    )

    def _get_zone_id(self):
        """Get zone_id from the product's category or default to False."""
        for move_line in self:
            if move_line.product_id.categ_id:
                move_line.zone_id = move_line.product_id.categ_id.parent_id.parent_id.parent_id.zone_id.id if move_line.product_id.categ_id.parent_id.parent_id.parent_id.zone_id else False
                move_line.family = move_line.product_id.categ_id.parent_id.parent_id.parent_id.id
                move_line.category = move_line.product_id.categ_id.parent_id.parent_id.id
                move_line.class_level_id = move_line.product_id.categ_id.parent_id.id
                move_line.brick = move_line.product_id.categ_id.id
            else:
                move_line.zone_id = False

    def mfd_date(self):
        for rec in self:
            if rec.create_date:
                print("rec.create_date",rec.create_date)
                return rec.create_date.strftime('%m/%Y')
            return ''


    @api.depends('product_id','move_id','move_id.move_brand_barcode')
    def _update_type_product(self):
        for move in self:
            for move_line in move.picking_id.move_ids_without_package:
                if move.picking_id.stock_type in ['ho_operation','inter_state','intra_state']:
                    if move_line.type_product == 'un_brand':
                        if move.product_id == move_line.product_id:
                            move.type_product = move_line.type_product
                            move.internal_ref_lot = move_line.product_id.barcode
                    elif move_line.type_product == 'brand':
                        if move.product_id == move_line.product_id:
                            move.type_product = move_line.type_product
                    else:
                        move.type_product = ''
                        move.internal_ref_lot = ''
                elif move.picking_id.stock_type == 'data_import':
                    if move_line.type_product == 'un_brand':
                        if move.product_id == move_line.product_id:
                            move.type_product = move_line.type_product
                            move.internal_ref_lot = move_line.product_id.barcode
                    elif move_line.type_product == 'brand':
                        matched_line = move_line.filtered(lambda l: l.id == move.move_id.id)
                        if move.product_id == move_line.product_id and matched_line:
                            move.type_product = move_line.type_product
                            move.internal_ref_lot = move_line.move_brand_barcode
                            move.cost_price = move_line.move_cp
                            move.mr_price = move_line.move_mrp
                            move.rs_price = move_line.move_rsp
                    else:
                        move.type_product = ''
                        move.internal_ref_lot = ''
                else:
                    move.type_product = ''
                    move.internal_ref_lot = ''


    def get_product_attributes(self):
        for rec in self:
            val = rec.product_id.product_template_variant_value_ids
            for i in val:
                attribute = self.env['product.attribute.value'].search([('name', '=', i.name)])
                for j in attribute:
                    if j.attribute_id.name == i.attribute_id.name and i.attribute_id.name.startswith('Color'):
                        rec.categ_1 = j.id
                    if j.attribute_id.name == i.attribute_id.name and i.attribute_id.name.startswith('Fit'):
                        rec.categ_2 = j.id
                    if j.attribute_id.name == i.attribute_id.name and i.attribute_id.name.startswith('Brand'):
                        rec.categ_3 = j.id
                    if j.attribute_id.name == i.attribute_id.name and i.attribute_id.name.startswith('Pattern'):
                        rec.categ_4 = j.id
                    if j.attribute_id.name == i.attribute_id.name and i.attribute_id.name.startswith('Border Type'):
                        rec.categ_5 = j.id
                    if j.attribute_id.name == i.attribute_id.name and i.attribute_id.name.startswith('Border Size'):
                        rec.categ_6 = j.id
                    if j.attribute_id.name == i.attribute_id.name and i.attribute_id.name.startswith('Size'):
                        rec.categ_7 = j.id
                    if j.attribute_id.name == i.attribute_id.name and i.attribute_id.name.startswith('Range'):
                        rec.descrip_2 = j.id
                    if j.attribute_id.name == i.attribute_id.name and i.attribute_id.name.startswith('Collection'):
                        rec.descrip_3 = j.id
                    if j.attribute_id.name == i.attribute_id.name and i.attribute_id.name.startswith('Fabric'):
                        rec.descrip_4 = j.id
                    if j.attribute_id.name == i.attribute_id.name and i.attribute_id.name.startswith('Exclusive'):
                        rec.descrip_5 = j.id
                    if j.attribute_id.name == i.attribute_id.name and i.attribute_id.name.startswith('Print'):
                        rec.descrip_6 = j.id

    def get_branded_barcode_from_store(self):
        for rec in self:
            if rec.move_id and rec.move_id.sale_line_id and rec.move_id.sale_line_id.branded_barcode and rec.lot_id.name == rec.move_id.sale_line_id.lot_ids.name:
                rec.internal_ref_lot = rec.move_id.sale_line_id.branded_barcode



    @api.model
    def create(self, vals):
        record = super(StockMoveLine, self).create(vals)
        record.get_product_attributes()
        record._get_zone_id()
        if record.picking_id.stock_picking_type in ['return','damage']:
            record.get_branded_barcode_from_store()
        return record

    def write(self, vals):
        res = super(StockMoveLine, self).write(vals)
        for move_line in self:
            if move_line.lot_id:
                lot_values = {}
                for categ_field in ['categ_1', 'categ_2', 'categ_3', 'categ_4', 'categ_5', 'categ_6','categ_7','categ_8']:
                    if categ_field in vals:
                        lot_values[categ_field.replace('categ', 'category')] = vals[categ_field]
                for desc_field in ['descrip_1', 'descrip_2', 'descrip_3', 'descrip_4', 'descrip_5', 'descrip_6','descrip_7','descrip_8']:
                    if desc_field in vals:
                        lot_values[desc_field.replace('descrip', 'description')] = vals[desc_field]
                for price_field in ['cost_price', 'mr_price', 'rs_price']:
                    if price_field in vals:
                        lot_values[price_field] = vals[price_field]
                if 'internal_ref_lot' in vals:
                    lot_values['ref'] = vals['internal_ref_lot']
                if lot_values:
                    move_line.lot_id.write(lot_values)
        return res


    def lot_values_update(self):
        for move_line in self:
            if move_line.picking_id.picking_type_code == 'incoming':
                move_line.lot_id.write({
                    'type_product': move_line.type_product,
                    'picking_id': move_line.picking_id.id,
                    'ref': move_line.internal_ref_lot,
                    'category_1': move_line.categ_1,
                    'category_2': move_line.categ_2,
                    'category_3': move_line.categ_3,
                    'category_4': move_line.categ_4,
                    'category_5': move_line.categ_5,
                    'category_6': move_line.categ_6,
                    'category_7': move_line.categ_7,
                    'category_8': move_line.categ_8,
                    'description_1': move_line.descrip_1,
                    'description_2': move_line.descrip_2,
                    'description_3': move_line.descrip_3,
                    'description_4': move_line.descrip_4,
                    'description_5': move_line.descrip_5,
                    'description_6': move_line.descrip_6,
                    'description_7': move_line.descrip_7,
                    'description_8': move_line.descrip_8,
                    'mr_price': move_line.mr_price,
                    'rs_price': move_line.rs_price,
                })
                if move_line.type_product == 'brand':
                    # Search for existing barcode
                    existing_barcode = self.env['product.barcode'].sudo().search(
                        [('barcode', '=', move_line.internal_ref_lot)],
                        limit=1)

                    if move_line.picking_id.picking_type_code == 'incoming':  # If the picking code is for incoming transfers (receipts)
                        if existing_barcode:
                            # Increment nhcl_inward_qty for existing barcode
                            existing_barcode.sudo().write({'nhcl_inward_qty': existing_barcode.nhcl_inward_qty + 1})
                        else:
                            # Create a new barcode and set nhcl_inward_qty to 1
                            self.env['product.barcode'].create({
                                'barcode': move_line.internal_ref_lot,
                                'product_tmpl_id': move_line.product_id.product_tmpl_id.id,
                                'product_id': move_line.product_id.id,
                                'nhcl_inward_qty': 1,
                            })

                    elif move_line.picking_id.picking_type_id.code == 'outgoing':  # If the picking code is for outgoing transfers (deliveries)
                        if existing_barcode:
                            # Increment nhcl_outward_qty for existing barcode
                            existing_barcode.sudo().write({'nhcl_outward_qty': existing_barcode.nhcl_outward_qty + 1})
                        else:
                            # Create a new barcode and set nhcl_outward_qty to 1
                            self.env['product.barcode'].sudo().create({
                                'barcode': move_line.internal_ref_lot,
                                'product_tmpl_id': move_line.product_id.product_tmpl_id.id,
                                'product_id': move_line.product_id.id,
                                'nhcl_outward_qty': 1,
                            })



    @api.onchange('internal_ref_lot')
    def sending_no_to_lot(self):
        for rec in self:
            if rec.lot_id and rec.internal_ref_lot:
                # Set the reference on the lot
                # rec.lot_id.ref = rec.internal_ref_lot
                # Search for existing barcode
                existing_barcode = self.env['product.barcode'].sudo().search([('barcode', '=', rec.internal_ref_lot)],
                                                                             limit=1)

                if rec.picking_id.picking_type_id.code == 'incoming':  # If the picking code is for incoming transfers (receipts)
                    if existing_barcode:
                        # Increment nhcl_inward_qty for existing barcode
                        existing_barcode.sudo().write({'nhcl_inward_qty': existing_barcode.nhcl_inward_qty + 1})
                    else:
                        # Create a new barcode and set nhcl_inward_qty to 1
                        self.env['product.barcode'].sudo().create({
                            'barcode': rec.internal_ref_lot,
                            'product_id': rec.lot_id.product_id.id,
                            'nhcl_inward_qty': 1,
                        })

                elif rec.picking_id.picking_type_id.code == 'outgoing':  # If the picking code is for outgoing transfers (deliveries)
                    if existing_barcode:
                        # Increment nhcl_outward_qty for existing barcode
                        existing_barcode.sudo().write({'nhcl_outward_qty': existing_barcode.nhcl_outward_qty + 1})
                    else:
                        # Create a new barcode and set nhcl_outward_qty to 1
                        self.env['product.barcode'].sudo().create({
                            'barcode': rec.internal_ref_lot,
                            'product_id': rec.lot_id.product_id.id,
                            'nhcl_outward_qty': 1,
                        })


class StockLot(models.Model):
    """Inherited stock.lot class to add fields and functions"""
    _inherit = 'stock.lot'

    nhcl_done_qty = fields.Boolean(string="Done Qty", compute='done_qty')
    is_under_plan = fields.Boolean(string="Is Under Plan")
    category_1 = fields.Many2one('product.attribute.value', string='Color', copy=False,
                                 domain=[('attribute_id.name', '=', 'Color')])
    category_2 = fields.Many2one('product.attribute.value', string='Fit', copy=False,
                                 domain=[('attribute_id.name', '=', 'Fit')])
    category_3 = fields.Many2one('product.attribute.value', string='Brand', copy=False,
                                 domain=[('attribute_id.name', '=', 'Brand')])
    category_4 = fields.Many2one('product.attribute.value', string='Pattern', copy=False,
                                 domain=[('attribute_id.name', '=', 'Pattern')])
    category_5 = fields.Many2one('product.attribute.value', string='Border Type', copy=False,
                                 domain=[('attribute_id.name', '=', 'Border Type')])
    category_6 = fields.Many2one('product.attribute.value', string='Border Size', copy=False,
                                 domain=[('attribute_id.name', '=', 'Border Size')])
    category_7 = fields.Many2one('product.attribute.value', string='Size', copy=False,
                                 domain=[('attribute_id.name', '=', 'Size')])
    category_8 = fields.Many2one('product.attribute.value', string='Design', copy=False,
                                 domain=[('attribute_id.name', '=', 'Design')])

    description_1 = fields.Many2one('product.aging.line', string="Product Aging",
                                    copy=False)
    description_2 = fields.Many2one('product.attribute.value', string='Range', copy=False,
                                    domain=[('attribute_id.name', '=', 'Range')])
    description_3 = fields.Many2one('product.attribute.value', string='Collection', copy=False,
                                    domain=[('attribute_id.name', '=', 'Collection')])
    description_4 = fields.Many2one('product.attribute.value', string='Fabric', copy=False,
                                    domain=[('attribute_id.name', '=', 'Fabric')])
    description_5 = fields.Many2one('product.attribute.value', string='Exclusive', copy=False,
                                    domain=[('attribute_id.name', '=', 'Exclusive')])
    description_6 = fields.Many2one('product.attribute.value', string='Print', copy=False,
                                    domain=[('attribute_id.name', '=', 'Print')])
    description_7 = fields.Many2one('product.attribute.value', string='Days Ageing', copy=False,
                                    domain=[('attribute_id.name', '=', 'Days Ageing')])
    description_8 = fields.Many2one('product.attribute.value', string='Description 8', copy=False, domain=[('attribute_id.name', '=', 'Offer')])
    product_description = fields.Html(string="Product Description", copy=False)
    web_product = fields.Char(string="Website Product Name", copy=False)
    cost_price = fields.Float(string='CP', copy=False, tracking=True)
    actual_cp = fields.Float(string='Actual CP', copy=False, tracking=True)
    mr_price = fields.Float(string='MRP', copy=False, tracking=True)
    rs_price = fields.Float(string='RSP', copy=False, tracking=True)
    # = fields.Float(string='TRP', copy=False)
    type_product = fields.Selection([('brand', 'Brand'), ('un_brand', 'UnBrand'), ('others', 'Others')],string='Brand Type', copy=False)
    picking_id = fields.Many2one('stock.picking', string="GRC No", copy=False)
    grc_partner_id = fields.Many2one('res.partner', string="Vendor", related='picking_id.partner_id')
    nhcl_margin_lot = fields.Integer(string=" RSP Margin", copy=False, tracking=True)
    nhcl_mrp_margin_lot = fields.Integer(string="MRP Margin", copy=False, tracking=True)
    segment = fields.Selection([('apparel','Apparel'), ('non_apparel','Non Apparel'), ('others','Others')], string="Segment", copy=False, related='product_id.segment')
    ref = fields.Char('Barcode', help="Internal reference number in case it differs from the manufacturer's lot/serial number")

    aging_id = fields.Many2one('product.aging', string="Aging id")
    shortage = fields.Float(string='Excess/Shortage', copy=False)
    serial_type = fields.Selection([('regular', 'Regular'), ('return', 'Returned')],
                                   string='Serial Type', copy=False, tracking=True, default='regular')
    nhcl_updated_margin_lot = fields.Integer(string="Updated Margin", copy=False, tracking=True)
    ho_grc_no = fields.Char(string='HO GRC NO.')
    is_uploaded = fields.Boolean('Is Uploaded', copy=False)
    zone_id = fields.Many2one('placement.master.data', string='Zone', copy=False)
    family = fields.Many2one('product.category',
                             string="Family",
                             domain="[('parent_id','=',False)]")
    category = fields.Many2one(
        'product.category',
        string="Category",
        domain="[('parent_id','=',family)]"
    )
    class_level_id = fields.Many2one(
        'product.category',
        string="Class",
        domain="[('parent_id','=',category)]"
    )
    brick = fields.Many2one(
        'product.category',
        string="Brick",
        domain="[('parent_id','=',class_level_id)]"
    )

    def _get_zone_id(self):
        """Get zone_id from the product's category or default to False."""
        for lot in self:
            if lot.product_id.categ_id:
                lot.zone_id = lot.product_id.categ_id.parent_id.parent_id.parent_id.zone_id.id if lot.product_id.categ_id.parent_id.parent_id.parent_id.zone_id else False
                lot.family = lot.product_id.categ_id.parent_id.parent_id.parent_id.id
                lot.category = lot.product_id.categ_id.parent_id.parent_id.id
                lot.class_level_id = lot.product_id.categ_id.parent_id.id
                lot.brick = lot.product_id.categ_id.id
            else:
                lot.zone_id = False

    @api.depends('product_qty')
    def done_qty(self):
        company = self.env['res.company'].search([('nhcl_company_bool','=',True)])
        for rec in self:
            if rec.product_qty < 0 and rec.company_id == company:
                rec.nhcl_done_qty = True
            else:
                rec.nhcl_done_qty = False

    def _get_product_aging(self):
        for res in self:
            if res.nhcl_margin_lot != 0:
                # Calculate margin price and preliminary price
                margin_price = ((res.nhcl_margin_lot / 100) * res.cost_price)
                temp_price = round(res.cost_price + margin_price)
            else:
                temp_price = round(res.cost_price)
            # Convert price to nearest integer
            if temp_price != 0:
                temp_price = int(temp_price)
                # Adjust the price based on last two digits
                last_two_digits = temp_price % 100
                if last_two_digits < 50:
                    temp_price = temp_price - last_two_digits + 49
                else:
                    temp_price = temp_price - last_two_digits + 99
                res.rs_price = round(temp_price)
            product = self.env['product.aging.line'].search([])
            if isinstance(res.create_date, datetime):
                create_date = res.create_date.date()
            else:
                create_date = res.create_date
            for rec in product:
                if rec.from_date <= create_date <= rec.to_date:
                    res.description_1 = rec.id
                    break
            else:
                res.description_1 = False


    @api.depends('nhcl_margin_lot', 'cost_price')
    def calculate_rsp_price(self):
        for rec in self:
            if rec.nhcl_margin_lot != 0:
                # Calculate margin price and preliminary price
                margin_price = ((rec.nhcl_margin_lot / 100) * rec.cost_price)
                without_tax_temp_price = round(rec.cost_price + margin_price)
                tax_ids = rec.product_id.supplier_taxes_id
                if len(tax_ids) > 1:
                    for tax in tax_ids:
                        if tax.min_amount <= without_tax_temp_price <= tax.max_amount:
                            after_tax_temp_price = without_tax_temp_price *  (tax.amount / 100.0)
                            temp_price = without_tax_temp_price + after_tax_temp_price
                            break
                else:
                    after_tax_temp_price = without_tax_temp_price * (
                            tax_ids.amount / 100.0) if tax_ids else without_tax_temp_price
                    temp_price = without_tax_temp_price + after_tax_temp_price


            else:
                temp_price = round(rec.cost_price)
            # Convert price to nearest integer
            if temp_price != 0:
                temp_price = int(temp_price)
                # Adjust the price based on last two digits
                last_two_digits = temp_price % 100
                if last_two_digits < 50:
                    temp_price = temp_price - last_two_digits + 49
                else:
                    temp_price = temp_price - last_two_digits + 99
                rec.rs_price = round(temp_price)

    # naseer
    @api.depends('name')
    def get_mrp_margin_from_category(self):
        for rec in self:
            rec.segment = rec.product_id.segment
            # Fetch the MRP margin lines once if the nested parent exists
            mrp_lines = False
            categ = rec.product_id.product_tmpl_id.categ_id
            if categ and categ.parent_id and categ.parent_id.parent_id and categ.parent_id.parent_id.parent_id:
                mrp_lines = categ.parent_id.parent_id.parent_id.product_category_mrp_ids
            margin = 0
            if mrp_lines:
                margin = next((line.margin for line in mrp_lines if line.from_range <= rec.rs_price <= line.to_range),
                              0)
            rec.nhcl_mrp_margin_lot = margin
            rec.calculate_mrp_price()

    @api.depends('nhcl_mrp_margin_lot', 'rs_price')
    def calculate_mrp_price(self):
        for rec in self:
            if rec.nhcl_mrp_margin_lot != 0:
                # Calculate margin price and preliminary price
                margin_price = ((rec.nhcl_mrp_margin_lot / 100) * rec.rs_price)
                temp_price = round(rec.rs_price + margin_price)
            else:
                temp_price = round(rec.rs_price)
            # Convert price to nearest integer
            if temp_price != 0:
                temp_price = int(temp_price)
                # Adjust the price based on last two digits
                last_two_digits = temp_price % 100
                if last_two_digits < 50:
                    temp_price = temp_price - last_two_digits + 49
                else:
                    temp_price = temp_price - last_two_digits + 99
                rec.mr_price = round(temp_price)



    @api.onchange('category_1')
    def updating_line_to_lot_category_1(self):
        detail = self.env['stock.move.line'].search([('lot_id.name', '=', self.name)])
        detail.categ_1 = self.category_1

    @api.onchange('category_2')
    def updating_line_to_lot_category_2(self):
        detail = self.env['stock.move.line'].search([('lot_id.name', '=', self.name)])
        detail.categ_2 = self.category_2

    @api.onchange('category_3')
    def updating_line_to_lot_category_3(self):
        detail = self.env['stock.move.line'].search([('lot_id.name', '=', self.name)])
        detail.categ_3 = self.category_3

    @api.onchange('category_4')
    def updating_line_to_lot_category_4(self):
        detail = self.env['stock.move.line'].search([('lot_id.name', '=', self.name)])
        detail.categ_4 = self.category_4

    @api.onchange('category_5')
    def updating_line_to_lot_category_5(self):
        detail = self.env['stock.move.line'].search([('lot_id.name', '=', self.name)])
        detail.categ_5 = self.category_5

    @api.onchange('category_6')
    def updating_line_to_lot_category_6(self):
        detail = self.env['stock.move.line'].search([('lot_id.name', '=', self.name)])
        detail.categ_6 = self.category_6

    @api.onchange('category_7')
    def updating_line_to_lot_category_6(self):
        detail = self.env['stock.move.line'].search([('lot_id.name', '=', self.name)])
        detail.categ_7 = self.category_7

    @api.onchange('category_8')
    def updating_line_to_lot_category_6(self):
        detail = self.env['stock.move.line'].search([('lot_id.name', '=', self.name)])
        detail.categ_8 = self.category_8


    @api.onchange('description_2')
    def updating_line_to_lot_description_2(self):
        detail = self.env['stock.move.line'].search([('lot_id.name', '=', self.name)])
        detail.descrip_2 = self.description_2

    @api.onchange('description_3')
    def updating_line_to_lot_description_3(self):
        detail = self.env['stock.move.line'].search([('lot_id.name', '=', self.name)])
        detail.descrip_3 = self.description_3

    @api.onchange('description_4')
    def updating_line_to_lot_description_4(self):
        detail = self.env['stock.move.line'].search([('lot_id.name', '=', self.name)])
        detail.descrip_4 = self.description_4

    @api.onchange('description_5')
    def updating_line_to_lot_description_5(self):
        detail = self.env['stock.move.line'].search([('lot_id.name', '=', self.name)])
        detail.descrip_5 = self.description_5

    @api.onchange('description_6')
    def updating_line_to_lot_description_6(self):
        detail = self.env['stock.move.line'].search([('lot_id.name', '=', self.name)])
        detail.descrip_6 = self.description_6

    @api.onchange('description_7')
    def updating_line_to_lot_description_7(self):
        detail = self.env['stock.move.line'].search([('lot_id.name', '=', self.name)])
        detail.descrip_7 = self.description_7

    @api.onchange('description_8')
    def updating_line_to_lot_description_8(self):
        detail = self.env['stock.move.line'].search([('lot_id.name', '=', self.name)])
        detail.descrip_8 = self.description_8

    @api.onchange('mr_price')
    def updating_line_to_lot_mr_price(self):
        detail = self.env['stock.move.line'].search([('lot_id.name', '=', self.name)])
        detail.mr_price = self.mr_price

    @api.onchange('rs_price')
    def updating_line_to_lot_rs_price(self):
        detail = self.env['stock.move.line'].search([('lot_id.name', '=', self.name)])
        detail.rs_price = self.rs_price

    @api.onchange('ref')
    def updating_line_to_lot_ref(self):
        detail = self.env['stock.move.line'].search([('lot_id.name', '=', self.name)])
        detail.internal_ref_lot = self.ref

    def get_attributes(self):
        for rec in self:
            val = rec.product_id.product_template_variant_value_ids
            for i in val:
                attribute = self.env['product.attribute.value'].search([('name', '=', i.name)])
                for j in attribute:
                    if j and j.attribute_id.name == i.attribute_id.name and i.attribute_id.name.startswith('Color'):
                        rec.category_1 = j.id
                    if j and j.attribute_id.name == i.attribute_id.name and i.attribute_id.name.startswith('Fit'):
                        rec.category_2 = j.id
                    if j and j.attribute_id.name == i.attribute_id.name and i.attribute_id.name.startswith('Brand'):
                        rec.category_3 = j.id
                    if j and j.attribute_id.name == i.attribute_id.name and i.attribute_id.name.startswith('Pattern'):
                        rec.category_4 = j.id
                    if j and j.attribute_id.name == i.attribute_id.name and i.attribute_id.name.startswith(
                            'Border Type'):
                        rec.category_5 = j.id
                    if j and j.attribute_id.name == i.attribute_id.name and i.attribute_id.name.startswith(
                            'Border Size'):
                        rec.category_6 = j.id
                    if j and j.attribute_id.name == i.attribute_id.name and i.attribute_id.name.startswith(
                            'Size'):
                        rec.category_7 = j.id
                    if j and j.attribute_id.name == i.attribute_id.name and i.attribute_id.name.startswith('Range'):
                        rec.description_2 = j.id
                    if j and j.attribute_id.name == i.attribute_id.name and i.attribute_id.name.startswith(
                            'Collection'):
                        rec.description_3 = j.id
                    if j and j.attribute_id.name == i.attribute_id.name and i.attribute_id.name.startswith('Fabric'):
                        rec.description_4 = j.id
                    if j and j.attribute_id.name == i.attribute_id.name and i.attribute_id.name.startswith('Exclusive'):
                        rec.description_5 = j.id
                    if j and j.attribute_id.name == i.attribute_id.name and i.attribute_id.name.startswith('Print'):
                        rec.description_6 = j.id
                    if j and j.attribute_id.name == i.attribute_id.name and i.attribute_id.name.startswith('Days Ageing'):
                        rec.description_6 = j.id

    @api.model
    def create(self, vals):
        record = super(StockLot, self).create(vals)
        record.get_attributes()
        record._get_zone_id()

        return record

    def search_by_product_margin(self, product_margin):
        domain = []
        today = datetime.today()

        # Get the user's timezone
        user_tz = self.env.user.tz or pytz.utc
        local = pytz.timezone(user_tz)

        # Get the selected ageing slab range
        if product_margin.pm_day_ageing_slab:
            # Mapping for ageing slabs
            slab_mapping = {
                '1': (0, 30),
                '2': (30, 60),
                '3': (60, 90),
                '4': (90, 120),
                '5': (120, 150),
                '6': (150, 180),
                '7': (180, 210),
                '8': (210, 240),
                '9': (240, 270),
                '10': (270, 300),
                '11': (300, 330),
                '12': (330, 360)
            }
            # Get the start and end days for the slab
            slab_start, slab_end = slab_mapping.get(str(product_margin.pm_day_ageing_slab), (0, 360))
            # Calculate the lower and upper bounds for the matching date range
            ageing_date_start = today - timedelta(days=slab_end)
            ageing_date_end = today - timedelta(days=slab_start)
            # Ensure the start date is earlier than the end date
            if ageing_date_start > ageing_date_end:
                ageing_date_start, ageing_date_end = ageing_date_end, ageing_date_start
            # Localize the dates to the user's timezone
            from_date_local = ageing_date_start.replace(hour=0, minute=0, second=0, microsecond=0)
            to_date_local = ageing_date_end.replace(hour=23, minute=59, second=59, microsecond=999999)
            from_date_local = local.localize(from_date_local)
            to_date_local = local.localize(to_date_local)
            # Convert the localized dates to UTC
            from_date_utc = from_date_local.astimezone(pytz.utc)
            to_date_utc = to_date_local.astimezone(pytz.utc)
            # Format the dates into ISO 8601 format
            from_date_str = from_date_utc.strftime("%Y-%m-%dT%H:%M:%S")
            to_date_str = to_date_utc.strftime("%Y-%m-%dT%H:%M:%S")
            # Add create_date range condition to the domain
            domain.append(('create_date', '>=', from_date_str))
            domain.append(('create_date', '<=', to_date_str))
        # Loop through all category and description fields
        for i in range(1, 8):
            # Dynamically construct field names
            category_field = f'category_{i}'
            description_field = f'description_{i}'
            # Get the corresponding many2many fields in loyalty.rule
            category_pm_field = f'pm_category_{i}_ids'
            description_pm_field = f'pm_description_{i}_ids' if i != 7 else None
            # Add to domain if the loyalty rule fields have values
            if getattr(product_margin, category_pm_field):
                domain.append((category_field, 'in', getattr(product_margin, category_pm_field).ids))
            if description_pm_field and getattr(product_margin, description_pm_field):
                domain.append((description_field, 'in', getattr(product_margin, description_pm_field).ids))
        lots = self.env['stock.lot'].search(domain)
        return lots

    def search_by_loyalty_rule(self, loyalty_rule):
        domain = []
        today = datetime.today()

        # Get the user's timezone
        user_tz = self.env.user.tz or pytz.utc
        local = pytz.timezone(user_tz)

        # Get the selected ageing slab range
        if loyalty_rule.day_ageing_slab:
            # Mapping for ageing slabs
            slab_mapping = {
                '1': (0, 30),
                '2': (30, 60),
                '3': (60, 90),
                '4': (90, 120),
                '5': (120, 150),
                '6': (150, 180),
                '7': (180, 210),
                '8': (210, 240),
                '9': (240, 270),
                '10': (270, 300),
                '11': (300, 330),
                '12': (330, 360)
            }
            # Get the start and end days for the slab
            slab_start, slab_end = slab_mapping.get(str(loyalty_rule.day_ageing_slab), (0, 360))

            # Calculate the lower and upper bounds for the matching date range
            ageing_date_start = today - timedelta(days=slab_end)
            ageing_date_end = today - timedelta(days=slab_start)

            # Ensure the start date is earlier than the end date
            if ageing_date_start > ageing_date_end:
                ageing_date_start, ageing_date_end = ageing_date_end, ageing_date_start

            # Localize the dates to the user's timezone
            from_date_local = ageing_date_start.replace(hour=0, minute=0, second=0, microsecond=0)
            to_date_local = ageing_date_end.replace(hour=23, minute=59, second=59, microsecond=999999)

            from_date_local = local.localize(from_date_local)
            to_date_local = local.localize(to_date_local)

            # Convert the localized dates to UTC
            from_date_utc = from_date_local.astimezone(pytz.utc)
            to_date_utc = to_date_local.astimezone(pytz.utc)

            # Format the dates into ISO 8601 format
            from_date_str = from_date_utc.strftime("%Y-%m-%dT%H:%M:%S")
            to_date_str = to_date_utc.strftime("%Y-%m-%dT%H:%M:%S")

            # Add create_date range condition to the domain
            domain.append(('create_date', '>=', from_date_str))
            domain.append(('create_date', '<=', to_date_str))
        # Loop through all category and description fields
        for i in range(1, 8):
            # Dynamically construct field names
            category_field = f'category_{i}'
            description_field = f'description_{i}'
            # Get the corresponding many2many fields in loyalty.rule
            category_rule_field = f'category_{i}_ids'
            description_rule_field = f'description_{i}_ids' if i != 7 else None

            # Add to domain if the loyalty rule fields have values
            if getattr(loyalty_rule, category_rule_field):
                domain.append((category_field, 'in', getattr(loyalty_rule, category_rule_field).ids))
            if description_rule_field and getattr(loyalty_rule, description_rule_field):
                domain.append((description_field, 'in', getattr(loyalty_rule, description_rule_field).ids))

        # Add product filtering if ref_product_ids is defined in the loyalty rule
        if loyalty_rule.ref_product_ids:
            domain.append(('product_id', 'in', loyalty_rule.ref_product_ids.ids))
        # Add category filtering if product_category_id is defined in the loyalty rule
        if loyalty_rule.product_category_ids:
            selected_categories = loyalty_rule.product_category_id.ids
            for category in loyalty_rule.product_category_ids:
                selected_categories += category.search([('id', 'child_of', category.id)]).ids
            selected_categories = list(set(selected_categories))
            domain.append(('product_id.categ_id', 'in', selected_categories))
        # Add product tag filtering if product_tag_id is defined in the loyalty rule
        if loyalty_rule.product_tag_id:
            domain.append(('product_id.product_tag_ids', '=', loyalty_rule.product_tag_id.id))
        # Additional checks for stock.lot records
        domain.append(('product_qty', '>=', 1))
        domain.append(('product_id.item_type', '=', 'inventory'))
        domain.append(('company_id.nhcl_company_bool', '=', False))
        lots = self.env['stock.lot'].search(domain)
        return lots


    @api.model
    def _get_next_serial(self, company, product):
        """Return the next serial number to be attributed to the product."""
        if product.tracking in ['serial','lot']:
            auto_generate_sequence = self.env['nhcl.master.sequence'].search(
                [('nhcl_code', '=', 'Auto Serial Number'), ('nhcl_state', '=', 'activate')])
            if auto_generate_sequence:
                if auto_generate_sequence.nhcl_next_number == 1:
                    last_serial = auto_generate_sequence.nhcl_prefix + '0'
                else:
                    last_serial = auto_generate_sequence.nhcl_prefix + str(auto_generate_sequence.nhcl_next_number -1)
                if last_serial:
                    return self.env['stock.lot'].generate_lot_names(last_serial, 2)[1]['lot_name']
            else:
                raise ValidationError(
                    "Serial sequence is not configured in the Sequence Master. Please configure it.")
        else:
            return super(StockLot, self)._get_next_serial(company, product)


class StockBackorderConfirmation(models.TransientModel):
    """Inherited stock.backorder.confirmation class to override existing functions"""
    _inherit = 'stock.backorder.confirmation'

    def process(self):
        for pick_id in self.pick_ids:
            if pick_id.picking_type_id.code == 'incoming' and pick_id.stock_type == 'ho_operation' and pick_id.is_confirm == False:
                # pick_id.move_ids.auto_generate_serial_numbers()
                pick_id.is_confirm = True
        return super(StockBackorderConfirmation, self).process()

    def process_cancel_backorder(self):
        for pick_id in self.pick_ids:
            if pick_id.picking_type_id.code == 'incoming' and pick_id.stock_type == 'ho_operation' and pick_id.is_confirm == False:
                pick_id.move_ids.auto_generate_serial_numbers()
                pick_id.is_confirm = True
        return super(StockBackorderConfirmation, self).process_cancel_backorder()


class StockLandedCost(models.Model):
    _inherit = 'stock.landed.cost'

    def button_validate(self):
        for cost in self:
            for line in cost.cost_lines:
                if line.price_unit <= 0:
                    raise ValidationError(_("Unit Price should be greater than or equal to 1."))
        res = super(StockLandedCost, self).button_validate()
        for cost in self:
            for line in cost.valuation_adjustment_lines.filtered(lambda line: line.move_id):
                for move_line in line.move_id.move_line_ids:
                    if move_line.lot_id:
                        lot = move_line.lot_id
                        purchase_line_id = move_line.move_id.purchase_line_id
                        if line.move_id.product_uom != purchase_line_id.product_uom:
                            cost_price = purchase_line_id.price_unit / line.move_id.product_uom.ratio
                        else:
                            cost_price = line.nhcl_cost_price
                        lot.cost_price = cost_price
                        lot.actual_cp = line.actual_cp_lc
                        lot.nhcl_margin_lot = move_line.approval_margin
                        lot.category_8 = move_line.move_id.purchase_line_id.purchase_category_id
                        # lot.get_margin_from_category()
                        lot.calculate_rsp_price()
                        lot.get_mrp_margin_from_category()
                        # lot._get_product_aging()
                        move_line.cost_price = line.nhcl_cost_price
                        move_line.rs_price = lot.rs_price
                        move_line.mr_price = lot.mr_price
                        move_line.categ_8 = lot.category_8.id

        return res


class StockVerification(models.Model):
    _name = 'stock.verification'

    stock_picking_id = fields.Many2one('stock.picking', copy=False)
    stock_product_id = fields.Many2one('product.product', string='Product', copy=False)
    stock_serial = fields.Char(string="Serial's", copy=False)
    stock_qty = fields.Float(string='Qty', copy=False)

    @api.model
    def create(self, vals):
        product = self.env['product.product'].browse(vals.get('stock_product_id'))
        serial_or_barcode = vals.get('stock_serial')
        qty = vals.get('stock_qty', 0)

        # Only raise error if product is serial-tracked and input is a serial (not 13-digit barcode)
        if product and product.tracking == 'serial' and qty > 1:
            if not re.match(r'^\d{13}$', serial_or_barcode):  # not a 13-digit barcode
                raise ValidationError(
                    "You cannot add more than 1 quantity for a serial-tracked product (by serial number).")

        return super().create(vals)

    def write(self, vals):
        res = super().write(vals)
        for record in self:
            product = record.stock_product_id
            serial_or_barcode = record.stock_serial
            qty = record.stock_qty

            if product.tracking == 'serial' and qty > 1:
                if not re.match(r'^\d{13}$', serial_or_barcode):  # not a barcode
                    raise ValidationError(
                        "You cannot set quantity > 1 for a serial-tracked product (by serial number).")
        return res


class AdjustmentLines(models.Model):
    _inherit = 'stock.valuation.adjustment.lines'

    nhcl_cost_price = fields.Float('Cost Price',copy=False,compute='compute_cost_price_per_unit')
    actual_cp_lc = fields.Float('Actual Cost Price',copy=False)

    def compute_cost_price_per_unit(self):
        for rec in self:
            if rec.final_cost and rec.quantity:
                temp_nhcl_cost_price = rec.final_cost / rec.quantity
                rec.actual_cp_lc = rec.former_cost / rec.quantity
                rec.nhcl_cost_price = temp_nhcl_cost_price
            else:
                rec.nhcl_cost_price = 0
                rec.actual_cp_lc = 0

class StockPickingBatch(models.Model):
    _inherit = 'stock.picking.batch'

    nhcl_receipt_validate = fields.Selection([('yes', "YES"), ('no', 'NO')], string="Check", copy=False, default='no')

    warehouse_name = fields.Char(string="Warehouse")

    #In transfer page filter based on store
    @api.onchange('nhcl_company')
    def _onchange_nhcl_company(self):
        if self.nhcl_company and self.nhcl_company.nhcl_store_name:
            print(self.nhcl_company.nhcl_store_name.name)
            self.warehouse_name = self.nhcl_company.nhcl_store_name.name
        else:
            self.warehouse_name = False


    def action_done(self):
        for rec in self:
            company = self.env['res.company'].sudo().search([('nhcl_company_bool','=',True)])
            if rec.picking_type_id.company_id.name != company.name:
                for line in rec.picking_ids:
                    operation_qty = sum(line.move_ids_without_package.mapped('quantity'))
                    verification_qty = sum(line.stock_verification_ids.mapped('stock_qty'))
                    verify_check = operation_qty == verification_qty
                    if verify_check == False and line.picking_type_code == 'outgoing':
                        raise ValidationError("Verify Check is not done.")
                    if rec.nhcl_company.nhcl_store_name.name != line.partner_id.name and self.env.user.id != 1:
                        raise ValidationError("You should give same contact and to store.")
            res = super().action_done()
        return res

    def post_lot_serial_numbers(self):
        if self.nhcl_receipt_validate == 'no':
            all_validated = all(picking.validate_related_po_receipt() for picking in self.picking_ids)

            if all_validated:
                self.nhcl_receipt_validate = 'yes'
                return {
                    'effect': {
                        'fadeout': 'slow',
                        'type': 'rainbow_man',
                        'message': _("Successfully Validated.."),
                    }
                }
            else:
                raise ValidationError(_("Some transfers are not validated. Please check."))
        else:
            raise ValidationError(_("Already validated."))

    def action_confirm(self):
        for batch in self:
            for picking in batch.picking_ids:
                if picking.picking_type_code != 'outgoing':
                    continue

                if not all([
                    picking.transpoter_id,
                    picking.transpoter_route_id,
                    picking.lr_number,
                    picking.no_of_parcel,
                    picking.driver_name,
                    picking.vehicle_number,
                ]):
                    raise UserError(_(
                        "Please fill all required transporter details in Delivery Order %s."
                    ) % picking.name)

        return super().action_confirm()


class ManafacturingImportWizard(models.TransientModel):
    _name = 'manafacturing.import.wizard'

    manafacturing_id = fields.Many2one('stock.move', string="Manufacturing Move", required=True)  # Ensure it's required
    file_data = fields.Binary("File", required=True)
    file_name = fields.Char("Filename", required=True)
    file_type = fields.Selection([
        ('excel', 'Excel (.xlsx, .xls)'),
        ('csv', 'CSV (.csv, .txt)')
    ], string="File Type", required=True, default='excel')

    def is_valid_file_extension(self, file_name):
        valid_extensions = ['.xls', '.xlsx', '.ods', '.csv', '.txt']
        return any(file_name.lower().endswith(ext) for ext in valid_extensions)

    def action_import(self):
        if not self.file_name or isinstance(self.file_name, bool):
            raise UserError("File name is missing or invalid.")

        if not self.is_valid_file_extension(self.file_name):
            raise UserError("Invalid file format! Allowed: .xls, .xlsx, .ods, .csv, .txt")

        file_content = base64.b64decode(self.file_data)
        if len(file_content) == 0:
            raise UserError("File is empty!")

        # For serial tracked products we simply collect lot IDs.
        # For lot tracked products, we collect lot IDs with their quantity.
        serial_lot_ids = []
        lot_data = {}  # key: lot id, value: quantity (float)
        total_qty = 0.0
        demand_qty = self.manafacturing_id.product_uom_qty

        try:
            if self.file_type == 'excel':
                workbook = openpyxl.load_workbook(io.BytesIO(file_content))
                sheet = workbook.active
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    # Expecting four columns: serial_no, barcode, tracking, quantity
                    serial_no = row[0] or ""
                    barcode = row[1] or ""
                    tracking = row[2].strip().lower() if row[2] else ""
                    qty = row[3]  # May be None

                    if self.manafacturing_id.product_id.tracking == 'serial':
                        # For serial tracked items, quantity must not be provided
                        if qty:
                            raise UserError("Quantity is not allowed for serial tracked items.")
                        lot_id = self.process_lot(serial_no, barcode, tracking, return_id=True)
                        if lot_id:
                            serial_lot_ids.append(lot_id)
                    elif self.manafacturing_id.product_id.tracking == 'lot':
                        # For lot tracked items, quantity is required
                        if not qty:
                            raise UserError("Quantity is required for lot tracked items.")
                        try:
                            qty_val = float(qty)
                        except Exception:
                            raise UserError("Invalid quantity value provided.")
                        total_qty += qty_val
                        lot_id = self.process_lot(serial_no, barcode, tracking, return_id=True)
                        if not lot_id:
                            raise UserError("No lot found for row with Serial/Barcode '%s'" % (serial_no or barcode))
                        lot_data[lot_id] = lot_data.get(lot_id, 0.0) + qty_val
            elif self.file_type == 'csv':
                decoded_file = io.StringIO(file_content.decode("utf-8-sig"))
                reader = csv.reader(decoded_file)
                next(reader)  # Skip header row
                for row in reader:
                    if len(row) < 4:
                        _logger.warning(f"Skipping incomplete row: {row}")
                        continue
                    serial_no = row[0].strip() if row[0] else ""
                    barcode = row[1].strip() if row[1] else ""
                    tracking = row[2].strip().lower() if row[2] else ""
                    qty = row[3].strip() if row[3] else None

                    if self.manafacturing_id.product_id.tracking == 'serial':
                        if qty:
                            raise UserError("Quantity is not allowed for serial tracked items in CSV.")
                        lot_id = self.process_lot(serial_no, barcode, tracking, return_id=True)
                        if lot_id:
                            serial_lot_ids.append(lot_id)
                    elif self.manafacturing_id.product_id.tracking == 'lot':
                        if not qty:
                            raise UserError("Quantity is required for lot tracked items in CSV.")
                        try:
                            qty_val = float(qty)
                        except Exception:
                            raise UserError("Invalid quantity value provided in CSV.")
                        total_qty += qty_val
                        lot_id = self.process_lot(serial_no, barcode, tracking, return_id=True)
                        if not lot_id:
                            raise UserError("No lot found for row with Serial/Barcode '%s'" % (serial_no or barcode))
                        lot_data[lot_id] = lot_data.get(lot_id, 0.0) + qty_val

            # Validate against demand quantity
            if self.manafacturing_id.product_id.tracking == 'serial':
                if len(serial_lot_ids) > demand_qty:
                    raise UserError(
                        f"Too many serial numbers! Required: {demand_qty}, Provided: {len(serial_lot_ids)}"
                    )
            elif self.manafacturing_id.product_id.tracking == 'lot':
                if total_qty > demand_qty:
                    raise UserError(
                        f"Total quantity ({total_qty}) exceeds required quantity ({demand_qty})."
                    )
            # Assign Serial Numbers or create move lines based on product tracking
            if self.manafacturing_id.product_id.tracking == 'serial':
                self.manafacturing_id.lot_ids = [(6, 0, serial_lot_ids)]
            elif self.manafacturing_id.product_id.tracking == 'lot':
                for lot_id, qty in lot_data.items():
                    move_line_vals = {
                        'lot_id': lot_id,
                        'company_id': self.env['res.company'].sudo().search([('nhcl_company_bool', '=', True)]).id,
                        'product_uom_id': self.manafacturing_id.product_id.uom_po_id.id,
                        'product_id': self.manafacturing_id.product_id.id,
                        'location_id': self.env.ref('stock.stock_location_stock').id,
                        'location_dest_id': self.env['stock.location'].search([('usage','=','production'),('company_id.nhcl_company_bool','=',True)]).id,
                        'quantity': qty,
                        'move_id': self.manafacturing_id.id,
                    }
                    self.env['stock.move.line'].create(move_line_vals)
        except Exception as e:
            _logger.error(f"Error: {str(e)}")
            raise UserError(f"Error importing file: {str(e)}")
        return {'type': 'ir.actions.act_window_close'}

    def process_lot(self, serial_no, barcode, tracking, return_id=False):
        _logger.info(f"Processing Lot - Serial: '{serial_no}', Barcode: '{barcode}', Tracking: '{tracking}'")
        lot = None
        if tracking == 'serial' and serial_no:
            lot = self.env['stock.lot'].search([('name', '=', serial_no)], limit=1)
            if lot:
                _logger.info(f"Found Serial Lot: {lot.name}")
            else:
                _logger.warning(f"No lot found for serial number: {serial_no}")
        elif tracking == 'lot' and serial_no:
            lot = self.env['stock.lot'].search([('name', '=', serial_no)], limit=1)
            if lot:
                _logger.info(f"Found Lot: {lot.name}")
            else:
                _logger.warning(f"No lot found for serial number: {serial_no}")
        elif tracking == 'serial' and barcode:
            lot = self.env['stock.lot'].search([('ref', '=', barcode)], limit=1)
            if lot:
                _logger.info(f"Found Lot for Barcode: {lot.name}")
            else:
                _logger.warning(f"No lot found for barcode: {barcode}")
        elif tracking == 'lot' and barcode:
            lot = self.env['stock.lot'].search([('ref', '=', barcode)], limit=1)
            if lot:
                _logger.info(f"Found Lot for Barcode: {lot.name}")
            else:
                _logger.warning(f"No lot found for barcode: {barcode}")
        else:
            _logger.warning("Skipping row due to missing or invalid tracking information")
        if return_id:
            return lot.id if lot else None
        return


class PickingType(models.Model):
    _inherit = "stock.picking.type"

    stock_picking_type = fields.Selection(
        [('exchange', 'Customer-Return'), ('receipt', 'Receipt'), ('goods_return', 'Goods Return')
            , ('delivery', 'Delivery'), ('pos_order', 'POS Order'),
         ('regular', 'Regular'), ('damage', 'Damage'), ('return', 'Return'),
         ('damage_main', 'Damage-Main'), ('main_damage', 'Main-Damage'),
         ('return_main', 'Return-Main')], string='Type', tracking=True)

class StockVerificationImport(models.TransientModel):
    _name = 'stock.verification.import'
    _description = 'Import Stock Verification'

    stock_picking_id = fields.Many2one('stock.picking', string="Stock Picking", required=True)
    file_type = fields.Selection([
        ('excel', 'Excel'),
        ('csv', 'CSV')
    ], string="File Type", required=True, default='excel')
    file_data = fields.Binary(string="Upload File", required=True)
    file_name = fields.Char(string="File Name", required=True)

    def is_valid_file_extension(self, file_name):
        valid_extensions = ['.xls', '.xlsx', '.ods', '.csv', '.txt']
        return any(file_name.lower().endswith(ext) for ext in valid_extensions)

    def clean_string(self, value):
        if value is None:
            return ''
        if isinstance(value, float):
            return str(int(value))
        return str(value).strip().replace('.0', '')


    def action_import(self):
        _logger.info("Starting barcode verification import process.")

        if self.stock_picking_id.verification_success == 'matched':
            raise UserError("This picking is already verified and matched. Further imports are not allowed.")

        if not self.file_name or isinstance(self.file_name, bool):
            raise UserError("File name is missing or invalid.")

        if not self.is_valid_file_extension(self.file_name):
            raise UserError("Invalid file format! Allowed: .xls, .xlsx, .csv, .txt")

        file_content = base64.b64decode(self.file_data)
        if len(file_content) == 0:
            raise UserError("File is empty!")

        # ✅ Reset verification lines before starting fresh
        self.stock_picking_id.stock_verification_ids = [(5, 0, 0)]

        gs1_pattern = r'01(\d{14})21([A-Za-z0-9]+)'  # GS1 format
        ean13_pattern = r'^\d{13}$'

        verification_lines = []
        found_serials = set()
        barcode_qty_map = defaultdict(float)
        barcode_row_map = {}
        move_lines = self.stock_picking_id.move_line_ids_without_package

        try:
            if self.file_type == 'excel':
                workbook = openpyxl.load_workbook(io.BytesIO(file_content))
                sheet = workbook.active

                for idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                    raw_serial_input = self.clean_string(row[0])
                    barcode = self.clean_string(row[1])
                    qty = row[2] or 0

                    if not raw_serial_input and not barcode:
                        raise UserError(f"Row {idx}: Either Serial or Barcode must be provided.")
                    if raw_serial_input and barcode:
                        raise UserError(f"Row {idx}: Provide either Serial or Barcode, not both.")
                    if not isinstance(qty, (int, float)) or qty <= 0:
                        raise UserError(f"Row {idx}: Quantity must be a positive number.")

                    # GS1 Serial Parse
                    gs1_match = re.match(gs1_pattern, raw_serial_input)
                    serial_no = gs1_match.group(2) if gs1_match else raw_serial_input

                    # Unbranded: Use Serial/GS1
                    if serial_no:
                        matching_lines = move_lines.filtered(lambda ml: ml.lot_id.name == serial_no)
                        if not matching_lines:
                            raise UserError(f"Row {idx}: Serial/Lot '{serial_no}' not found in move lines.")

                        line = next((ml for ml in matching_lines if ml.product_id), False)
                        if not line:
                            raise UserError(f"Row {idx}: No product found for serial '{serial_no}'.")

                        if line.type_product == 'brand':
                            raise UserError(f"Row {idx}: Branded product '{serial_no}' must use barcode, not serial.")

                        tracking = line.product_id.tracking
                        if tracking == 'serial' and qty != 1:
                            raise UserError(f"Row {idx}: Serial-tracked product '{serial_no}' must have quantity 1.")

                        if serial_no in found_serials:
                            raise UserError(f"Row {idx}: Duplicate serial '{serial_no}' in file.")

                        found_serials.add(serial_no)
                        verification_lines.append((0, 0, {
                            'stock_product_id': line.product_id.id,
                            'stock_serial': serial_no,
                            'stock_qty': qty,
                        }))

                    # Branded: Use Barcode
                    elif barcode:
                        if not re.match(ean13_pattern, barcode):
                            raise UserError(f"Row {idx}: Invalid barcode format '{barcode}'. Must be 13 digits.")
                        barcode_qty_map[barcode] += qty
                        barcode_row_map[barcode] = idx

            # ✅ Second pass: process barcode lines
            for barcode, total_qty in barcode_qty_map.items():
                idx = barcode_row_map.get(barcode, 0)
                matching_lines = move_lines.filtered(lambda ml: ml.internal_ref_lot == barcode and ml.lot_id.name)
                if not matching_lines:
                    raise UserError(f"Row {idx}: Barcode '{barcode}' not found in move lines.")

                branded_lines = matching_lines.filtered(lambda ml: ml.type_product == 'brand')
                if not branded_lines:
                    raise UserError(f"Row {idx}: Barcode '{barcode}' refers to unbranded product. Use serial.")

                product_line = next((ml for ml in branded_lines if ml.product_id), False)
                if not product_line:
                    raise UserError(f"Row {idx}: Product not found for barcode '{barcode}'.")

                tracking = product_line.product_id.tracking

                if tracking == 'serial':
                    available_lines = [ml for ml in branded_lines if ml.lot_id.name not in found_serials]
                    if len(available_lines) < total_qty:
                        raise UserError(
                            f"Row {idx}: Qty = {total_qty}, but only {len(available_lines)} available for barcode '{barcode}'.")

                    for ml in available_lines[:int(total_qty)]:
                        found_serials.add(ml.lot_id.name)

                    verification_lines.append((0, 0, {
                        'stock_product_id': product_line.product_id.id,
                        'stock_serial': barcode,
                        'stock_qty': total_qty,
                    }))

                elif tracking == 'lot':
                    available_qty = sum(ml.quantity for ml in branded_lines if ml.lot_id.name not in found_serials)
                    if available_qty < total_qty:
                        raise UserError(
                            f"Row {idx}: Qty = {total_qty}, but only {available_qty} available for barcode '{barcode}'.")

                    for ml in branded_lines:
                        found_serials.add(ml.lot_id.name)

                    verification_lines.append((0, 0, {
                        'stock_product_id': product_line.product_id.id,
                        'stock_serial': barcode,
                        'stock_qty': total_qty,
                    }))

            # ✅ Push new verification lines
            self.stock_picking_id.stock_verification_ids = verification_lines

            # ✅ Check final quantity match
            operation_qty = sum(self.stock_picking_id.move_line_ids_without_package.mapped('quantity'))
            verification_qty = sum(line[2]['stock_qty'] for line in verification_lines)
            self.stock_picking_id.verification_success = 'matched' if operation_qty == verification_qty else ''

        except Exception as e:
            _logger.error(f"Import error: {str(e)}")
            raise UserError(f"Error importing file: {str(e)}")

        return {'type': 'ir.actions.act_window_close'}



class StockReturnPicking(models.TransientModel):
    _inherit = 'stock.return.picking'

    def update_return_parent_qty(self, new_picking):
        """ Prevent copy of the carrier and carrier price when generating return picking
        (we have no integration of returns for now).
        """
        picking = self.env['stock.picking'].browse(new_picking)
        for line in picking.move_ids:
            line.write({
                'return_parent_qty': sum(line.move_orig_ids.mapped('product_uom_qty')),
            })

    def _create_returns(self):
        # Prevent copy of the carrier and carrier price when generating return picking
        # (we have no integration of returns for now)
        new_picking, pick_type_id = super()._create_returns()
        self.update_return_parent_qty(new_picking)
        return new_picking, pick_type_id



class StockPickingBarcode(models.Model):
    _name = 'stock.picking.barcode'

    @api.model
    def create(self, vals):
        picking = self.env['stock.picking'].browse(vals.get('stock_picking_delivery_id'))
        serial_no = vals.get('serial_no')
        delivery_check_seq = self.env['nhcl.master.sequence'].search(
            [('nhcl_code', '=', 'cmr.delivery'), ('nhcl_state', '=', 'activate')], limit=1)
        if not delivery_check_seq:
            raise ValidationError(_('The Delivery Check Sequence is not specified in the sequence master. "Please configure it!.'))

        if not picking.lr_number:
            raise ValidationError(_('LR Number is missing in the Stock Picking.'))
        final_barcode = f"{picking.lr_number}-{serial_no}"
        barcode = f"{picking.name}-{serial_no}"
        vals['barcode'] = barcode

        if vals.get('sequence', 'New') == 'New':
            vals['sequence'] = final_barcode

        res = super(StockPickingBarcode, self).create(vals)
        return res

    barcode = fields.Char(string='Barcode')
    stock_picking_delivery_id = fields.Many2one('stock.picking', string="Delivery Number")
    lr_number = fields.Char(
        string="LR Number",
        related='stock_picking_delivery_id.lr_number',
        store=False,
        readonly=True
    )
    sequence = fields.Char(string="Sequence",copy=False, default=lambda self: _("New"))
    serial_no = fields.Integer(string='S.NO')
