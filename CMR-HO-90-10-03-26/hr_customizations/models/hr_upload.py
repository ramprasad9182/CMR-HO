import io
import logging
import re
import xlrd
from dateutil.parser import parse
from dateutil.relativedelta import relativedelta

from odoo import models, fields, api, _
import base64
from io import BytesIO
import openpyxl
import pytz
from datetime import datetime, timedelta, time, date
from odoo.exceptions import UserError, ValidationError, AccessError
from odoo.tools import DEFAULT_SERVER_DATETIME_FORMAT
import calendar
import traceback

_logger = logging.getLogger(__name__)


class Hrupload(models.Model):
    _name = "hr.upload"

    date = fields.Datetime(string='Date')
    employee_code = fields.Char(related='employee_name.barcode', string='Employee Code', store=True)
    # your_datetime = fields.Datetime(string='Check in', compute='_compute_your_datetime')
    # your_checkout_datetime = fields.Datetime(string='Check-Out', compute='_compute_your_checkout_datetime',
    #                                          store=True)
    employee_name = fields.Many2one('hr.employee', string="Employee Name", required=True)
    check_in_attendance = fields.Char(string="Check In")
    check_out_attendance = fields.Char(string="Check Out")
    difference_check_in = fields.Char(string="IN Difference (mins)", compute='_compute_attendance', store=True)
    difference_check_out = fields.Char(string="OUT Difference (mins)", compute='_compute_attendance', store=True)
    total_working_hours = fields.Char(string="Total Working Hours", compute='_compute_attendance', store=True)
    attendance_status = fields.Selection([
        ('grace', 'Grace'),
        ('early', 'Early'),
        ('late', 'Late'),
        ('others', 'Others')
    ], string="Attendance Status", compute='_compute_attendance', store=True)
    late_deduction = fields.Char(string="Late Deduction(Rupees)", compute="_compute_late_deduction", store=True)
    overtime_amount = fields.Char(string="Overtime Amount (₹)", compute="_compute_overtime_amount", store=True)
    morning_session = fields.Selection([
        ('Present', 'Present'),
        ('Absent', 'Absent')
    ], string="Morning Session", compute='_compute_full_day_status', store=True)
    afternoon_session = fields.Selection([
        ('Present', 'Present'),
        ('Absent', 'Absent')
    ], string="Afternoon Session", compute='_compute_full_day_status', store=True)

    full_day_status = fields.Char(string="Full Day Status", compute='_compute_full_day_status', store=True)

    ctc_type = fields.Selection([
        ('with_bonus', 'With Bonus'),
        ('without_bonus', 'Without Bonus'),
        ('non_ctc', 'Non Ctc'),
    ], string='CTC Type', related='employee_name.ctc_type', store=True, readonly=True)
    designation_id = fields.Many2one('hr.job', string="Designation", related='employee_name.job_id', store=True,
                                     readonly=True)
    department_id = fields.Many2one('hr.department', string="Department", related='employee_name.department_id',
                                    store=True, readonly=True)

    company_id = fields.Many2one('res.company', string="Company", related='employee_name.company_id', store=True,
                                 readonly=True)
    division_id = fields.Many2one(
        'product.category',
        string='Division',
        domain=[('parent_id', '=', False)],
        related='employee_name.division_id', store=True, readonly=True
    )

    overtime_hours = fields.Char(
        string="Overtime (Hours)",
        compute="_compute_overtime_hours",
        store=True,
        help="Shows extra hours worked after 22:00"
    )

    your_datetime = fields.Datetime(
        string='Check-In',
        compute='_compute_all_times',
        store=True
    )

    your_checkout_datetime = fields.Datetime(
        string='Check-Out',
        compute='_compute_all_times',
        store=True
    )

    session_check_in = fields.Datetime(
        string='Session Check-In',
        compute='_compute_all_times',
        store=True
    )

    session_check_out = fields.Datetime(
        string='Session Check-Out',
        compute='_compute_all_times',
        store=True
    )

    lunch_in = fields.Char("Lunch In")
    lunch_out = fields.Char("Lunch Out")

    break_in = fields.Char("Break In")
    break_out = fields.Char("Break Out")

    lunch_delay = fields.Char("Lunch Delay (HH:MM)", compute="_compute_lunch_break_delay", store=True)
    break_delay = fields.Char("Break Delay (HH:MM)", compute="_compute_lunch_break_delay", store=True)

    lunch_delay_amount = fields.Float("Lunch Deduction Amount", compute="_compute_lunch_break_amount", store=True)
    break_delay_amount = fields.Float("Break Deduction Amount", compute="_compute_lunch_break_amount", store=True)

    lunch_addition_amount = fields.Float(
        string="Lunch Addition Amount",
        compute="_compute_lunch_addition_amount",
        store=True
    )

    lunch_bonus_amount = fields.Float(
        string='Lunch Bonus',
        compute='_compute_lunch_addition_amount',
        store=True
    )

    @api.depends('lunch_delay', 'employee_name', 'date')
    def _compute_lunch_addition_amount(self):
        for rec in self:
            rec.lunch_addition_amount = 0

            if not rec.employee_name or not rec.date or not rec.lunch_delay:
                continue

            employee = rec.employee_name
            shift_id = employee.resource_calendar_id.id

            # --------------------------
            # Convert HH:MM / HH:MM:SS → minutes (DURATION)
            # --------------------------
            def to_minutes(val):
                if not val:
                    return 0
                if isinstance(val, str):
                    h, m, *_ = val.split(':')
                    return int(h) * 60 + int(m)
                if hasattr(val, 'hour'):
                    return val.hour * 60 + val.minute
                return 0

            lunch_minutes = to_minutes(rec.lunch_delay)
            if lunch_minutes <= 0:
                continue

            upload_month = rec.date.month
            upload_year = rec.date.year

            # use_bonus_master = employee.lunch_bonus_category in (
            #     'female_supervisor',
            #     'female_senior',
            #     'male_supervisor',
            #     'male_senior'
            # )

            # ==============================
            # FETCH MASTER
            # ==============================
            # if use_bonus_master:
            #     master = self.env['lunch.bonus.master'].search([
            #         ('shift_id', '=', shift_id),
            #         ('category', '=', employee.lunch_bonus_category),
            #         ('month', '=', upload_month),
            #         ('year', '=', upload_year),
            #     ], limit=1)
            # else:
            master = self.env['shift.addition.master'].search([
                    ('shift_id', '=', shift_id),
                    ('category', '=', employee.lunch_bonus_category),
                    ('addition_type', '=', 'lunch'),
                    ('month', '=', upload_month),
                    ('year', '=', upload_year),
                ], limit=1)

            if not master:
                continue

            # ==============================
            # MATCH DELAY SLAB (DURATION)
            # ==============================
            for slab in master.slab_ids:
                from_m = to_minutes(slab.from_time)
                to_m = to_minutes(slab.to_time)

                if from_m <= lunch_minutes <= to_m:
                    rec.lunch_addition_amount = slab.amount
                    break

    @api.depends('lunch_in', 'lunch_out', 'break_in', 'break_out')
    def _compute_lunch_break_delay(self):
        for rec in self:

            rec.lunch_delay = "0:00"
            rec.break_delay = "0:00"

            def to_minutes(value):
                if not value:
                    return 0
                if isinstance(value, float):
                    try:
                        total_seconds = int(value * 24 * 3600)
                        return total_seconds // 60
                    except:
                        return 0
                if isinstance(value, datetime):
                    return value.hour * 60 + value.minute
                if isinstance(value, time):
                    return value.hour * 60 + value.minute

                value = str(value).strip()
                if ":" in value:
                    try:
                        h, m = value.split(":")
                        return int(h) * 60 + int(m)
                    except:
                        return 0

                if value.isdigit():
                    return int(value)

                return 0

            # ---------------- LUNCH -----------------
            lunch_in_min = to_minutes(rec.lunch_in)
            lunch_out_min = to_minutes(rec.lunch_out)

            if lunch_in_min > 0:
                if lunch_out_min >= lunch_in_min:
                    # normal case
                    diff = lunch_out_min - lunch_in_min
                else:
                    # MISSING OUT → assign max slab
                    lunch_master = self.env['shift.deduction.master'].search([
                        ('shift_id', '=', rec.employee_name.resource_calendar_id.id),
                        ('category', '=', rec.employee_name.lunch_bonus_category),
                        ('break_type', '=', 'lunch')
                    ], limit=1)

                    if lunch_master and lunch_master.slab_ids:
                        last_slab = sorted(
                            lunch_master.slab_ids,
                            key=lambda s: to_minutes(s.to_time)
                        )[-1]
                        diff = to_minutes(last_slab.to_time)
                    else:
                        diff = 0

                hr = diff // 60
                mn = diff % 60
                rec.lunch_delay = f"{hr}:{mn:02d}"

            # ---------------- BREAK -----------------
            break_in_min = to_minutes(rec.break_in)
            break_out_min = to_minutes(rec.break_out)

            if break_in_min > 0:
                if break_out_min >= break_in_min:
                    diff = break_out_min - break_in_min
                else:
                    # MISSING OUT → assign max slab
                    break_master = self.env['shift.deduction.master'].search([
                        ('shift_id', '=', rec.employee_name.resource_calendar_id.id),
                        ('category', '=', rec.employee_name.lunch_bonus_category),
                        ('break_type', '=', 'evening')
                    ], limit=1)

                    if break_master and break_master.slab_ids:
                        last_slab = sorted(
                            break_master.slab_ids,
                            key=lambda s: to_minutes(s.to_time)
                        )[-1]
                        diff = to_minutes(last_slab.to_time)
                    else:
                        diff = 0

                hr = diff // 60
                mn = diff % 60
                rec.break_delay = f"{hr}:{mn:02d}"

    @api.depends('lunch_delay', 'break_delay', 'employee_name', 'break_in', 'break_out')
    def _compute_lunch_break_amount(self):
        for rec in self:

            rec.lunch_delay_amount = 0
            rec.break_delay_amount = 0

            employee = rec.employee_name
            if not employee or not employee.resource_calendar_id:
                continue

            shift_id = employee.resource_calendar_id.id
            category = employee.lunch_bonus_category


            # ------------------------------------------
            # Convert HH:MM → total minutes
            # ------------------------------------------
            def to_minutes(val):
                if not val:
                    return 0
                try:
                    h, m = str(val).split(":")
                    return int(h) * 60 + int(m)
                except:
                    return 0

            lunch_min = to_minutes(rec.lunch_delay)
            break_min = to_minutes(rec.break_delay)

            # ===========================================================
            #                LUNCH SLAB AMOUNT
            # ===========================================================
            lunch_master = self.env['shift.deduction.master'].search([
                ('shift_id', '=', shift_id),
                ('category', '=', category),
                ('break_type', '=', 'lunch')
            ], limit=1)

            if lunch_master and lunch_master.slab_ids:

                # sort slabs by from_time for safety
                slabs = sorted(lunch_master.slab_ids, key=lambda s: to_minutes(s.from_time))

                matched = False

                for slab in slabs:
                    fm = to_minutes(slab.from_time)
                    tm = to_minutes(slab.to_time)

                    if fm <= lunch_min <= tm:
                        rec.lunch_delay_amount = slab.amount
                        matched = True
                        break

                        # If no slab matched but delay is large → last slab amount
                if not matched and lunch_min > to_minutes(slabs[-1].to_time):
                    rec.lunch_delay_amount = slabs[-1].amount

                # ===========================================================
                #     IF BREAK-IN OR BREAK-OUT MISSING → DIRECT DEDUCTION
                # ===========================================================
            if not rec.break_in or not rec.break_out:
                rec.break_delay_amount = 25
                continue

            # ===========================================================
            #                EVENING BREAK SLAB AMOUNT
            # ===========================================================
            break_master = self.env['shift.deduction.master'].search([
                ('shift_id', '=', shift_id),
                ('category', '=', category),
                ('break_type', '=', 'evening')
            ], limit=1)

            if break_master and break_master.slab_ids:

                slabs = sorted(break_master.slab_ids, key=lambda s: to_minutes(s.from_time))

                matched = False

                for slab in slabs:
                    fm = to_minutes(slab.from_time)
                    tm = to_minutes(slab.to_time)

                    if fm <= break_min <= tm:
                        rec.break_delay_amount = slab.amount
                        matched = True
                        break

                if not matched and break_min > to_minutes(slabs[-1].to_time):
                    last_to_min = to_minutes(slabs[-1].to_time)

                    # Extra minutes after last slab
                    extra_minutes = break_min - last_to_min

                    # ₹5 for every 5 minutes
                    extra_blocks = extra_minutes // 5

                    rec.break_delay_amount = slabs[-1].amount + (extra_blocks * 5)

    @api.depends(
        'date',
        'check_in_attendance',
        'check_out_attendance',
        'morning_session',
        'afternoon_session',
        'employee_name',
    )
    def _compute_all_times(self):

        for rec in self:

            rec.your_datetime = False
            rec.your_checkout_datetime = False
            rec.session_check_in = False
            rec.session_check_out = False

            if not rec.date:
                continue

            # --------------------------------------------------------------
            # Setup timezone & base date
            # --------------------------------------------------------------
            user_tz = pytz.timezone(self.env.user.tz or "Asia/Kolkata")
            date_val = rec.date

            # --------------------------------------------------------------
            # Helper: Convert time string → UTC naive datetime
            # --------------------------------------------------------------
            def to_utc_naive(tstr):
                # Empty => 00:00:00
                if not tstr:
                    hh = mm = ss = 0
                else:
                    parts = tstr.split(':')
                    hh = int(parts[0])
                    mm = int(parts[1]) if len(parts) > 1 else 0
                    ss = int(parts[2]) if len(parts) > 2 else 0

                dt = datetime.combine(date_val, time(hh, mm, ss))
                local_dt = user_tz.localize(dt)
                return local_dt.astimezone(pytz.utc).replace(tzinfo=None)

            # --------------------------------------------------------------
            # 1) Actual check-in/out
            # --------------------------------------------------------------
            rec.your_datetime = to_utc_naive(rec.check_in_attendance)
            rec.your_checkout_datetime = to_utc_naive(rec.check_out_attendance)

            # --------------------------------------------------------------
            # 2) Need employee to continue
            # --------------------------------------------------------------
            if not rec.employee_name:
                continue

            # --------------------------------------------------------------
            # 3) Load working schedule
            # --------------------------------------------------------------
            calendar = rec.employee_name.resource_calendar_id
            weekday = date_val.strftime("%A")

            intervals = calendar.attendance_ids.filtered(
                lambda a: a.display_name.startswith(weekday)
            )

            morning = intervals.filtered(lambda a: a.day_period == "morning")
            afternoon = intervals.filtered(lambda a: a.day_period == "afternoon")

            morning_start = morning.hour_from if morning else None
            morning_end = morning.hour_to if morning else None
            afternoon_start = afternoon.hour_from if afternoon else None

            # Helper for shift times
            def shift_to_utc_naive(float_hour):
                if float_hour is None:
                    return False
                h = int(float_hour)
                m = int((float_hour - h) * 60)
                dt = datetime.combine(date_val, time(h, m))
                return user_tz.localize(dt).astimezone(pytz.utc).replace(tzinfo=None)

            # --------------------------------------------------------------
            # 4) Full-day absent => 00:00
            # --------------------------------------------------------------
            if rec.morning_session == "Absent" and rec.afternoon_session == "Absent":
                zero = to_utc_naive("00:00:00")
                rec.session_check_in = zero
                rec.session_check_out = zero
                continue

            # --------------------------------------------------------------
            # 5) Session check-in
            # --------------------------------------------------------------
            if rec.morning_session == "Present" and rec.check_in_attendance:
                rec.session_check_in = rec.your_datetime
            else:
                rec.session_check_in = shift_to_utc_naive(afternoon_start)

            # --------------------------------------------------------------
            # 6) Session check-out
            # --------------------------------------------------------------
            if rec.afternoon_session == "Present" and rec.check_out_attendance:
                rec.session_check_out = rec.your_checkout_datetime
            else:
                rec.session_check_out = shift_to_utc_naive(morning_end)

    @api.depends('check_out_attendance')
    def _compute_overtime_hours(self):
        for rec in self:
            rec.overtime_hours = "0:00"

            if not rec.check_out_attendance:
                continue

            try:
                # ✅ Get first overtime slab (lowest start_time)
                first_slab = self.env['hr.overtime.master'].search([], order="start_time asc", limit=1)
                if not first_slab:
                    continue

                # Convert string times (HH:MM) to datetime.time
                check_out_time = datetime.strptime(rec.check_out_attendance.strip(), "%H:%M").time()
                start_time = datetime.strptime(first_slab.start_time.strip(), "%H:%M").time()

                # Combine with today's date
                today = datetime.today()
                start_dt = datetime.combine(today, start_time)
                check_out_dt = datetime.combine(today, check_out_time)

                # ✅ If check-out is before start → no OT
                if check_out_time <= start_time:
                    rec.overtime_hours = "0:00"
                    continue

                # ✅ If check-out crosses midnight (e.g. 02:00 next day)
                if check_out_time < start_time:
                    check_out_dt += timedelta(days=1)

                # ✅ Calculate overtime
                diff = check_out_dt - start_dt
                total_minutes = int(diff.total_seconds() / 60)
                hours = total_minutes // 60
                minutes = total_minutes % 60
                rec.overtime_hours = f"{hours}:{minutes:02d}"

            except Exception:
                rec.overtime_hours = "0:00"

    @api.depends('check_out_attendance')
    def _compute_overtime_amount(self):
        for rec in self:
            rec.overtime_amount = 0.0

            if not rec.check_out_attendance:
                continue

            try:
                # Parse employee check-out time
                try:
                    check_out_time = datetime.strptime(rec.check_out_attendance, "%H:%M:%S").time()
                except ValueError:
                    check_out_time = datetime.strptime(rec.check_out_attendance, "%H:%M").time()

                # Fetch all overtime slabs and sort by start time
                masters = self.env['hr.overtime.master'].search([])
                if not masters:
                    continue

                # Sort slabs by start time safely
                def parse_t(t):
                    try:
                        return datetime.strptime(t, "%H:%M:%S").time()
                    except ValueError:
                        return datetime.strptime(t, "%H:%M").time()

                masters = masters.sorted(lambda m: parse_t(m.start_time))

                # Loop through slabs
                for i, master in enumerate(masters):
                    try:
                        start_time = parse_t(master.start_time)
                        end_time = parse_t(master.end_time)
                        last_slab = (i == len(masters) - 1)

                        # Normal condition: within slab range
                        if start_time <= check_out_time <= end_time:
                            rec.overtime_amount = master.overtime_amount
                            break

                        # Special rule for LAST SLAB — open ended
                        # If checkout is AFTER last slab → still apply last slab
                        if last_slab and check_out_time >= end_time:
                            rec.overtime_amount = master.overtime_amount
                            break

                    except Exception:
                        # If any error in this slab, skip it and continue
                        continue

            except Exception:
                rec.overtime_amount = 0.0

    # @api.depends('check_out_attendance')
    # def _compute_overtime_amount(self):
    #     for rec in self:
    #         rec.overtime_amount = 0.0
    #
    #         if rec.check_out_attendance:
    #             try:
    #                 # Convert check_out_attendance (HH:MM) to datetime.time
    #                 try:
    #                     check_out_time = datetime.strptime(rec.check_out_attendance, "%H:%M:%S").time()
    #                 except ValueError:
    #                     check_out_time = datetime.strptime(rec.check_out_attendance, "%H:%M").time()
    #
    #                 # Fetch all overtime rules
    #                 masters = self.env['hr.overtime.master'].search([])
    #
    #                 for master in masters:
    #                     try:
    #                         try:
    #                             start_time = datetime.strptime(master.start_time, "%H:%M:%S").time()
    #                             end_time = datetime.strptime(master.end_time, "%H:%M:%S").time()
    #                         except ValueError:
    #                             start_time = datetime.strptime(master.start_time, "%H:%M").time()
    #                             end_time = datetime.strptime(master.end_time, "%H:%M").time()
    #
    #                         # ✅ If check-out is between start and end → assign overtime amount
    #                         if start_time <= check_out_time <= end_time:
    #                             rec.overtime_amount = master.overtime_amount
    #                             break
    #                     except Exception:
    #                         continue
    #
    #             except Exception:
    #                 rec.overtime_amount = 0.0
    #         else:
    #             rec.overtime_amount = 0.0
    #
    # @api.depends(
    #     'check_in_attendance',
    #     'check_out_attendance',
    #     'date',
    #     'employee_name',
    #     'employee_name.resource_calendar_id',
    #     'employee_name.resource_calendar_id.attendance_ids',
    #     'employee_name.resource_calendar_id.attendance_ids.hour_from',
    #     'employee_name.resource_calendar_id.attendance_ids.hour_to',
    #     'employee_name.resource_calendar_id.attendance_ids.day_period',
    # )
    # def _compute_full_day_status(self):
    #     for rec in self:
    #         rec.full_day_status = 'Absent (Full Day)'
    #         rec.morning_session = 'Absent'
    #         rec.afternoon_session = 'Absent'
    #
    #         if not rec.employee_name or not rec.date:
    #             _logger.debug("hr.upload[%s]: missing employee or date", rec.id)
    #             continue
    #
    #         calendar = rec.employee_name.resource_calendar_id
    #         if not calendar:
    #             _logger.debug("hr.upload[%s]: employee %s has no calendar", rec.id, rec.employee_name.name)
    #             continue
    #
    #         # parse check in/out times robustly (allow 'HH:MM' or 'HH:MM:SS')
    #         def parse_time_str(tstr):
    #             if not tstr:
    #                 return None
    #             for fmt in ("%H:%M:%S", "%H:%M"):
    #                 try:
    #                     return datetime.strptime(tstr, fmt).time()
    #                 except Exception:
    #                     continue
    #             # if already a time/datetime object, handle gracefully
    #             if isinstance(tstr, datetime):
    #                 return tstr.time()
    #             if isinstance(tstr, time):
    #                 return tstr
    #             return None
    #
    #         tin = parse_time_str(rec.check_in_attendance)
    #         tout = parse_time_str(rec.check_out_attendance)
    #         if not tin or not tout:
    #             _logger.debug("hr.upload[%s]: could not parse times: in=%s out=%s", rec.id, rec.check_in_attendance,
    #                           rec.check_out_attendance)
    #             continue
    #
    #         # combine with rec.date so all datetimes share same date
    #         check_in_dt = datetime.combine(rec.date, tin)
    #         check_out_dt = datetime.combine(rec.date, tout)
    #
    #         weekday = rec.date.weekday()
    #         sessions = calendar.attendance_ids.filtered(lambda a: int(a.dayofweek) == weekday)
    #         if not sessions:
    #             _logger.debug("hr.upload[%s]: no sessions for weekday %s", rec.id, weekday)
    #             continue
    #
    #         # thresholds
    #         MORNING_MIN_HOURS = 1.5
    #         AFTERNOON_MIN_HOURS = 4.5
    #         EPS = 1e-6  # float tolerance
    #
    #         for session in sessions:
    #             # compute session start/end on the same rec.date
    #             start_time = time(int(session.hour_from), int((session.hour_from % 1) * 60))
    #             end_time = time(int(session.hour_to), int((session.hour_to % 1) * 60))
    #             session_start = datetime.combine(rec.date, start_time)
    #             session_end = datetime.combine(rec.date, end_time)
    #
    #             # If session_end <= session_start skip invalid slot
    #             if session_end <= session_start:
    #                 _logger.warning("hr.upload[%s]: invalid session times %s-%s", rec.id, session.hour_from,
    #                                 session.hour_to)
    #                 continue
    #
    #             # compute overlap (inclusive at boundaries)
    #             overlap_start = max(check_in_dt, session_start)
    #             overlap_end = min(check_out_dt, session_end)
    #             worked_seconds = (overlap_end - overlap_start).total_seconds() if overlap_end > overlap_start else 0.0
    #             worked_hours = worked_seconds / 3600.0
    #
    #             _logger.debug(
    #                 "hr.upload[%s] emp=%s period=%s session=%s-%s check=%s-%s worked=%.3f",
    #                 rec.id, rec.employee_name.name, session.day_period,
    #                 session_start.time(), session_end.time(),
    #                 check_in_dt.time(), check_out_dt.time(), worked_hours
    #             )
    #
    #             # apply threshold per session
    #             if session.day_period == 'morning':
    #                 if worked_hours + EPS >= MORNING_MIN_HOURS:
    #                     rec.morning_session = 'Present'
    #             elif session.day_period == 'afternoon':
    #                 if worked_hours + EPS >= AFTERNOON_MIN_HOURS:
    #                     rec.afternoon_session = 'Present'
    #
    #         # decide final full_day_status
    #         if rec.morning_session == 'Present' and rec.afternoon_session == 'Present':
    #             rec.full_day_status = 'Present (Full Day)'
    #         elif rec.morning_session == 'Present' and rec.afternoon_session == 'Absent':
    #             rec.full_day_status = 'First Session Present and Second Session Absent'
    #         elif rec.morning_session == 'Absent' and rec.afternoon_session == 'Present':
    #             rec.full_day_status = 'First Session Absent and Second Session Present'
    #         else:
    #             rec.full_day_status = 'Absent (Full Day)'

    # @api.depends(
    #     'check_in_attendance',
    #     'check_out_attendance',
    #     'date',
    #     'employee_name',
    #     'employee_name.resource_calendar_id',
    #     'employee_name.resource_calendar_id.attendance_ids',
    #     'employee_name.resource_calendar_id.attendance_ids.hour_from',
    #     'employee_name.resource_calendar_id.attendance_ids.hour_to',
    #     'employee_name.resource_calendar_id.attendance_ids.day_period',
    # )
    # def _compute_full_day_status(self):
    #     for rec in self:
    #         rec.full_day_status = 'Absent (Full Day)'
    #         rec.morning_session = 'Absent'
    #         rec.afternoon_session = 'Absent'
    #
    #         if not rec.employee_name or not rec.date:
    #             continue
    #
    #         calendar = rec.employee_name.resource_calendar_id
    #         if not calendar:
    #             continue
    #
    #         # ──────────────────────────────
    #         # SAFE TIME PARSER
    #         # ──────────────────────────────
    #         def parse_time_str(tstr):
    #             if not tstr:
    #                 return None
    #             for fmt in ("%H:%M:%S", "%H:%M"):
    #                 try:
    #                     return datetime.strptime(tstr.strip(), fmt).time()
    #                 except:
    #                     continue
    #             return None
    #
    #         tin = parse_time_str(rec.check_in_attendance)
    #         tout = parse_time_str(rec.check_out_attendance)
    #         if not tin or not tout:
    #             continue
    #
    #         check_in_dt = datetime.combine(rec.date, tin)
    #         check_out_dt = datetime.combine(rec.date, tout)
    #
    #         weekday = rec.date.weekday()
    #         sessions = calendar.attendance_ids.filtered(lambda a: int(a.dayofweek) == weekday)
    #         if not sessions:
    #             continue
    #
    #         # ──────────────────────────────
    #         # FIX: Correctly get the LAST late-deduction rule
    #         # ──────────────────────────────
    #         rules = self.env['hr.late.deduction.master'].search([])
    #
    #         def get_rule_end(r):
    #             return parse_time_str(r.end_time) or time(0, 0)
    #
    #         last_rule = max(rules, key=lambda r: get_rule_end(r)) if rules else None
    #
    #         morning_last_end_dt = None
    #         if last_rule:
    #             last_end = get_rule_end(last_rule)
    #             morning_last_end_dt = datetime.combine(rec.date, last_end)
    #
    #         EPS = 1e-6
    #         total_worked_hours = (
    #                 (check_out_dt - check_in_dt).total_seconds() / 3600.0
    #         )
    #
    #
    #         # ──────────────────────────────
    #         # SESSION CHECKING
    #         # ──────────────────────────────
    #         for session in sessions:
    #             start_time = time(int(session.hour_from), int((session.hour_from % 1) * 60))
    #             end_time = time(int(session.hour_to), int((session.hour_to % 1) * 60))
    #
    #             session_start = datetime.combine(rec.date, start_time)
    #             session_end = datetime.combine(rec.date, end_time)
    #
    #             if session_end <= session_start:
    #                 continue
    #
    #             # overlap
    #             overlap_start = max(check_in_dt, session_start)
    #             overlap_end = min(check_out_dt, session_end)
    #
    #             worked = max((overlap_end - overlap_start).total_seconds(), 0)
    #             worked_hours = worked / 3600.0
    #
    #             # ──────────────────────────────
    #             # MORNING SESSION (based on last-rule end-time)
    #             # ──────────────────────────────
    #             if session.day_period == 'morning':
    #                 if morning_last_end_dt:
    #                     rec.morning_session = 'Present' if check_in_dt <= morning_last_end_dt else 'Absent'
    #
    #             # ──────────────────────────────
    #             # AFTERNOON SESSION
    #             # ──────────────────────────────
    #             if total_worked_hours + EPS >= 4.5:
    #                 rec.afternoon_session = 'Present'
    #             else:
    #                 rec.afternoon_session = 'Absent'
    #         # ──────────────────────────────
    #         # FINAL FULL DAY STATUS
    #         # ──────────────────────────────
    #         if rec.morning_session == 'Present' and rec.afternoon_session == 'Present':
    #             rec.full_day_status = 'Present (Full Day)'
    #         elif rec.morning_session == 'Present' and rec.afternoon_session == 'Absent':
    #             rec.full_day_status = 'First Session Present and Second Session Absent'
    #         elif rec.morning_session == 'Absent' and rec.afternoon_session == 'Present':
    #             rec.full_day_status = 'First Session Absent and Second Session Present'
    #         else:
    #             rec.full_day_status = 'Absent (Full Day)'

    #
    # @api.depends(
    #     'check_in_attendance',
    #     'check_out_attendance',
    #     'date',
    #     'employee_name',
    #     'employee_name.resource_calendar_id',
    #     'employee_name.resource_calendar_id.attendance_ids',
    # )
    # def _compute_full_day_status(self):
    #
    #     def parse_time_str(tstr):
    #         if not tstr:
    #             return None
    #         for fmt in ("%H:%M:%S", "%H:%M"):
    #             try:
    #                 return datetime.strptime(tstr.strip(), fmt).time()
    #             except:
    #                 continue
    #         return None
    #
    #     for rec in self:
    #         # -------------------------------
    #         # DEFAULT VALUES
    #         # -------------------------------
    #         rec.full_day_status = 'Absent (Full Day)'
    #         rec.morning_session = 'Absent'
    #         rec.afternoon_session = 'Absent'
    #
    #         if not rec.employee_name or not rec.date:
    #             continue
    #
    #         calendar = rec.employee_name.resource_calendar_id
    #         if not calendar:
    #             continue
    #
    #         tin = parse_time_str(rec.check_in_attendance)
    #         tout = parse_time_str(rec.check_out_attendance)
    #         if not tin or not tout:
    #             continue
    #
    #         check_in_dt = datetime.combine(rec.date, tin)
    #         check_out_dt = datetime.combine(rec.date, tout)
    #
    #         if check_out_dt <= check_in_dt:
    #             continue
    #
    #         # -------------------------------
    #         # TOTAL WORKING HOURS (NO CHANGE)
    #         # -------------------------------
    #         total_worked_hours = (
    #                 (check_out_dt - check_in_dt).total_seconds() / 3600.0
    #         )
    #
    #         # -------------------------------
    #         # GET TODAY'S SCHEDULE
    #         # -------------------------------
    #         weekday = rec.date.weekday()
    #         sessions = calendar.attendance_ids.filtered(
    #             lambda a: int(a.dayofweek) == weekday
    #         )
    #
    #         if not sessions:
    #             continue
    #
    #         # -------------------------------
    #         # LATE DEDUCTION → MORNING LOGIC
    #         # -------------------------------
    #         rules = self.env['hr.late.deduction.master'].search([])
    #
    #         def rule_end_time(r):
    #             return parse_time_str(r.end_time) or time(0, 0)
    #
    #         last_rule = max(rules, key=lambda r: rule_end_time(r)) if rules else None
    #
    #         if last_rule:
    #             morning_last_end_dt = datetime.combine(
    #                 rec.date, rule_end_time(last_rule)
    #             )
    #             if check_in_dt <= morning_last_end_dt:
    #                 rec.morning_session = 'Present'
    #         else:
    #             rec.morning_session = 'Present'
    #
    #         # -------------------------------
    #         # SCHEDULE-BASED CALCULATION
    #         # (INCLUDING BREAK)
    #         # -------------------------------
    #         scheduled_hours = 0.0
    #         schedule_worked_hours = 0.0
    #
    #         for s in sessions:
    #             start_time = time(
    #                 int(s.hour_from),
    #                 int((s.hour_from % 1) * 60)
    #             )
    #             end_time = time(
    #                 int(s.hour_to),
    #                 int((s.hour_to % 1) * 60)
    #             )
    #
    #             session_start = datetime.combine(rec.date, start_time)
    #             session_end = datetime.combine(rec.date, end_time)
    #
    #             if session_end <= session_start:
    #                 continue
    #
    #             # Full scheduled span (including break)
    #             span_hours = (
    #                     (session_end - session_start).total_seconds() / 3600.0
    #             )
    #             scheduled_hours += span_hours
    #
    #             # Overlap with attendance
    #             overlap_start = max(check_in_dt, session_start)
    #             overlap_end = min(check_out_dt, session_end)
    #
    #             if overlap_end > overlap_start:
    #                 schedule_worked_hours += (
    #                         (overlap_end - overlap_start).total_seconds() / 3600.0
    #                 )
    #
    #         half_day_hours = scheduled_hours / 2.0
    #
    #         # -------------------------------
    #         # AFTERNOON SESSION (FINAL RULE)
    #         # -------------------------------
    #         if rec.morning_session == 'Present':
    #             # Needs 9 hours within working schedule
    #             rec.afternoon_session = (
    #                 'Present' if schedule_worked_hours >= 9.0 else 'Absent'
    #             )
    #         else:
    #             # Needs 4.5 hours within working schedule
    #             rec.afternoon_session = (
    #                 'Present' if schedule_worked_hours >= 4.5 else 'Absent'
    #             )
    #
    #         # -------------------------------
    #         # FINAL DAY STATUS
    #         # -------------------------------
    #         if rec.morning_session == 'Present' and rec.afternoon_session == 'Present':
    #             rec.full_day_status = 'Present (Full Day)'
    #         elif rec.morning_session == 'Present':
    #             rec.full_day_status = 'First Session Present and Second Session Absent'
    #         elif rec.afternoon_session == 'Present':
    #             rec.full_day_status = 'First Session Absent and Second Session Present'
    #         else:
    #             rec.full_day_status = 'Absent (Full Day)'
    @api.depends(
        'check_in_attendance',
        'check_out_attendance',
        'date',
        'employee_name',
        'employee_name.resource_calendar_id',
        'employee_name.resource_calendar_id.attendance_ids',
    )
    def _compute_full_day_status(self):

        def parse_time_str(tstr):
            if not tstr:
                return None
            for fmt in ("%H:%M:%S", "%H:%M"):
                try:
                    return datetime.strptime(tstr.strip(), fmt).time()
                except Exception:
                    continue
            return None

        for rec in self:
            # -------------------------------
            # DEFAULT VALUES
            # -------------------------------
            rec.full_day_status = 'Absent (Full Day)'
            rec.morning_session = 'Absent'
            rec.afternoon_session = 'Absent'

            if not rec.employee_name or not rec.date:
                continue

            calendar = rec.employee_name.resource_calendar_id
            if not calendar:
                continue

            tin = parse_time_str(rec.check_in_attendance)
            tout = parse_time_str(rec.check_out_attendance)
            if not tin or not tout:
                continue

            check_in_dt = datetime.combine(rec.date, tin)
            check_out_dt = datetime.combine(rec.date, tout)

            if check_out_dt <= check_in_dt:
                continue

            # -------------------------------
            # GET TODAY'S SCHEDULE
            # -------------------------------
            weekday = rec.date.weekday()
            sessions = calendar.attendance_ids.filtered(
                lambda a: int(a.dayofweek) == weekday
            )

            if not sessions:
                continue

            # -------------------------------
            # MORNING SESSION (LATE RULE)
            # -------------------------------
            rules = self.env['hr.late.deduction.master'].search([])

            def rule_end_time(r):
                return parse_time_str(r.end_time) or time(0, 0)

            last_rule = max(rules, key=lambda r: rule_end_time(r)) if rules else None

            if last_rule:
                morning_last_end_dt = datetime.combine(
                    rec.date, rule_end_time(last_rule)
                )
                if check_in_dt <= morning_last_end_dt:
                    rec.morning_session = 'Present'
            else:
                rec.morning_session = 'Present'

            # -------------------------------
            # SCHEDULE-BASED HOURS
            # -------------------------------
            scheduled_hours = 0.0
            schedule_worked_hours = 0.0

            for s in sessions:
                start_time = time(
                    int(s.hour_from),
                    int((s.hour_from % 1) * 60)
                )
                end_time = time(
                    int(s.hour_to),
                    int((s.hour_to % 1) * 60)
                )

                session_start = datetime.combine(rec.date, start_time)
                session_end = datetime.combine(rec.date, end_time)

                if session_end <= session_start:
                    continue

                # Total scheduled span
                scheduled_hours += (
                        (session_end - session_start).total_seconds() / 3600.0
                )

                # Overlap with attendance
                overlap_start = max(check_in_dt, session_start)
                overlap_end = min(check_out_dt, session_end)

                if overlap_end > overlap_start:
                    schedule_worked_hours += (
                            (overlap_end - overlap_start).total_seconds() / 3600.0
                    )

            # -------------------------------
            # HARD RULE: BELOW 4.5 HOURS
            # -------------------------------
            if schedule_worked_hours < 4.5:
                rec.morning_session = 'Absent'
                rec.afternoon_session = 'Absent'
            else:
                # -------------------------------
                # AFTERNOON SESSION (EXISTING LOGIC)
                # -------------------------------
                if rec.morning_session == 'Present':
                    # Needs 9 hours
                    rec.afternoon_session = (
                        'Present' if schedule_worked_hours >= 9.0 else 'Absent'
                    )
                else:
                    # Needs 4.5 hours
                    rec.afternoon_session = (
                        'Present' if schedule_worked_hours >= 4.5 else 'Absent'
                    )

            # -------------------------------
            # FINAL DAY STATUS
            # -------------------------------
            if rec.morning_session == 'Present' and rec.afternoon_session == 'Present':
                rec.full_day_status = 'Present (Full Day)'
            elif rec.morning_session == 'Present':
                rec.full_day_status = 'First Session Present and Second Session Absent'
            elif rec.afternoon_session == 'Present':
                rec.full_day_status = 'First Session Absent and Second Session Present'
            else:
                rec.full_day_status = 'Absent (Full Day)'

    @api.depends(
        'check_in_attendance',
        'check_out_attendance',
        'date',
        'employee_name',
        'employee_name.resource_calendar_id',
    )
    def _compute_attendance(self):
        for rec in self:
            rec.difference_check_in = "0:00"
            rec.difference_check_out = "0:00"
            rec.total_working_hours = "0:00"
            rec.attendance_status = 'others'

            emp = rec.employee_name or rec.employee_id
            if not emp or not emp.resource_calendar_id:
                continue

            base_date = rec.date or fields.Date.today()
            if isinstance(base_date, fields.Date):
                base_date = datetime.strptime(str(base_date), "%Y-%m-%d").date()

            weekday = base_date.weekday()

            shifts_today = emp.resource_calendar_id.attendance_ids.filtered(
                lambda s: int(s.dayofweek) == weekday and not (s.name.lower() in ['break', 'lunch'])
            )
            if not shifts_today:
                continue

            shifts_today = shifts_today.sorted(key=lambda l: l.hour_from)

            def parse_dt(val):
                if not val:
                    return None
                if isinstance(val, datetime):
                    return val
                if isinstance(val, (int, float)):
                    h = int(val)
                    m = int(round((val - h) * 60))
                    return datetime.combine(base_date, time(h, m))
                if isinstance(val, str):
                    m = re.match(r'^(\d{1,2}):(\d{1,2})(?::(\d{1,2}))?$', val.strip())
                    if m:
                        h, min_, sec = int(m.group(1)), int(m.group(2)), int(m.group(3) or 0)
                        return datetime.combine(base_date, time(h, min_, sec))
                return None

            check_in_dt = parse_dt(rec.check_in_attendance)
            check_out_dt = parse_dt(rec.check_out_attendance)

            if not check_in_dt or not check_out_dt:
                continue

            # ---------- FIND CORRECT SHIFT FOR CHECK-IN ----------
            shift_in = None
            for s in shifts_today:
                s_start = datetime.combine(base_date, time(int(s.hour_from), int(round((s.hour_from % 1) * 60))))
                s_end = datetime.combine(base_date, time(int(s.hour_to), int(round((s.hour_to % 1) * 60))))
                if s_start <= check_in_dt <= s_end:
                    shift_in = s
                    break
            if not shift_in:
                shift_in = shifts_today[0]

            # ---------- FIND CORRECT SHIFT FOR CHECK-OUT ----------
            shift_out = None
            for s in reversed(shifts_today):
                s_start = datetime.combine(base_date, time(int(s.hour_from), int(round((s.hour_from % 1) * 60))))
                s_end = datetime.combine(base_date, time(int(s.hour_to), int(round((s.hour_to % 1) * 60))))
                if s_start <= check_out_dt <= s_end:
                    shift_out = s
                    break
            if not shift_out:
                shift_out = shifts_today[-1]

            # ---------- CALCULATE DIFFERENCE CHECK-IN ----------
            shift_start_in = datetime.combine(base_date,
                                              time(int(shift_in.hour_from), int(round((shift_in.hour_from % 1) * 60))))
            grace_end_in = shift_start_in + timedelta(minutes=15)

            if check_in_dt <= shift_start_in:
                diff_min = int((shift_start_in - check_in_dt).total_seconds() / 60)
                rec.attendance_status = 'early'
            elif shift_start_in < check_in_dt <= grace_end_in:
                diff_min = 0
                rec.attendance_status = 'grace'
            else:
                diff_min = -int((check_in_dt - grace_end_in).total_seconds() / 60)
                rec.attendance_status = 'late'

            # ✅ Convert minutes to HH:MM
            sign = '-' if diff_min < 0 else ''
            abs_min = abs(diff_min)
            h, m = divmod(abs_min, 60)
            rec.difference_check_in = f"{sign}{h}:{m:02d}"

            # ---------- CALCULATE DIFFERENCE CHECK-OUT ----------
            shift_end_out = datetime.combine(base_date,
                                             time(int(shift_out.hour_to), int(round((shift_out.hour_to % 1) * 60))))
            diff_out_min = int((check_out_dt - shift_end_out).total_seconds() / 60)
            sign_out = '-' if diff_out_min < 0 else ''
            abs_min_out = abs(diff_out_min)
            h_out, m_out = divmod(abs_min_out, 60)
            rec.difference_check_out = f"{sign_out}{h_out}:{m_out:02d}"

            # ---------- TOTAL WORKING HOURS ----------
            total_secs = (check_out_dt - check_in_dt).total_seconds()
            if total_secs < 0:
                total_secs += 24 * 3600
            h = int(total_secs // 3600)
            m = int((total_secs % 3600) // 60)
            rec.total_working_hours = f"{h}:{m:02d}"

    # @api.depends('check_in_attendance')
    # def _compute_late_deduction(self):
    #     for rec in self:
    #         rec.late_deduction = 0.0
    #         if rec.check_in_attendance:
    #             try:
    #                 # Convert check_in_attendance (HH:MM) to datetime.time
    #                 check_in_time = datetime.strptime(rec.check_in_attendance, "%H:%M").time()
    #
    #                 employee_shift = rec.employee_name.resource_calendar_id.id
    #
    #                 masters = self.env['hr.late.deduction.master'].search([
    #                     ('shift_id', '=', employee_shift)
    #                 ])
    #                 for master in masters:
    #                     start_time = datetime.strptime(master.start_time, "%H:%M").time()
    #                     end_time = datetime.strptime(master.end_time, "%H:%M").time()
    #
    #                     if start_time <= check_in_time <= end_time:
    #                         rec.late_deduction = master.deduction_amount
    #                         break
    #             except Exception:
    #                 rec.late_deduction = 0.0
    #         else:
    #             rec.late_deduction = 0.0
    #
    # DEFAULT_SERVER_DATETIME_FORMAT = "%Y-%m-%d %H:%M:%S "

    @api.depends('check_in_attendance')
    def _compute_late_deduction(self):
        for rec in self:
            rec.late_deduction = 0.0
            if rec.check_in_attendance:
                try:
                    # Convert check_in_attendance (HH:MM) to datetime.time
                    check_in_time = datetime.strptime(rec.check_in_attendance, "%H:%M").time()

                    employee_shift = rec.employee_name.resource_calendar_id.id

                    masters = self.env['hr.late.deduction.master'].search([
                        ('shift_id', '=', employee_shift)
                    ])
                    for master in masters:
                        start_time = datetime.strptime(master.start_time, "%H:%M").time()
                        end_time = datetime.strptime(master.end_time, "%H:%M").time()

                        if start_time <= check_in_time <= end_time:
                            rec.late_deduction = master.deduction_amount
                            break
                except Exception:
                    rec.late_deduction = 0.0
            else:
                rec.late_deduction = 0.0

    DEFAULT_SERVER_DATETIME_FORMAT = "%Y-%m-%d %H:%M:%S "

    @api.depends('date', 'check_in_attendance')
    def _compute_your_datetime(self):
        for record in self:
            if record.date:
                try:
                    check_in_time = record.check_in_attendance or "00:00"

                    # Normalize both formats HH:MM and HH:MM:SS
                    parts = check_in_time.split(':')
                    if len(parts) == 2:  # Only hours and minutes
                        check_in_time = f"{check_in_time}:00"
                    elif len(parts) == 1:  # Just hour (edge case)
                        check_in_time = f"{check_in_time}:00:00"

                    datetime_str = f"{record.date} {check_in_time}"
                    naive_datetime = datetime.strptime(datetime_str, DEFAULT_SERVER_DATETIME_FORMAT)

                    # Convert to UTC based on user's timezone
                    user_tz = self.env.user.tz or 'UTC'
                    local_tz = pytz.timezone(user_tz)
                    local_dt = local_tz.localize(naive_datetime, is_dst=None)
                    utc_dt = local_dt.astimezone(pytz.UTC)

                    record.your_datetime = fields.Datetime.to_string(utc_dt)
                except Exception:
                    record.your_datetime = False
            else:
                record.your_datetime = False

    @api.depends('date', 'check_out_attendance')
    def _compute_your_checkout_datetime(self):
        for record in self:
            if record.date:
                try:
                    checkout_time = record.check_out_attendance or "00:00"

                    # Normalize time to HH:MM:SS
                    parts = checkout_time.split(':')
                    if len(parts) == 2:
                        checkout_time = f"{checkout_time}:00"
                    elif len(parts) == 1:
                        checkout_time = f"{checkout_time}:00:00"

                    datetime_str = f"{record.date} {checkout_time}"
                    naive_datetime = datetime.strptime(datetime_str, DEFAULT_SERVER_DATETIME_FORMAT)

                    # Convert to UTC based on user's timezone
                    user_tz = self.env.user.tz or 'UTC'
                    local_tz = pytz.timezone(user_tz)
                    local_dt = local_tz.localize(naive_datetime, is_dst=None)
                    utc_dt = local_dt.astimezone(pytz.UTC)

                    record.your_checkout_datetime = fields.Datetime.to_string(utc_dt)
                except Exception:
                    record.your_checkout_datetime = False
            else:
                record.your_checkout_datetime = False

    def open_upload_wizard(self):
        return {
            'name': 'Upload HR Data',
            'type': 'ir.actions.act_window',
            'res_model': 'hr.upload.wizard',
            'view_mode': 'form',
            'target': 'new',
        }


class HrUploadWizard(models.TransientModel):
    _name = "hr.upload.wizard"
    _description = "HR Upload Wizard"

    file = fields.Binary(string="File", required=True)
    filename = fields.Char(string="File Name")

    def action_upload(self):
        """Upload attendance from Excel and:
           - create hr.upload lines
           - create hr.attendance for worked days
           - create & auto-approve EL / LOP from absences
        """
        if not self.file:
            raise UserError(_("Please upload a file."))

        try:
            # --------------------------------------------------------------
            # 0. SETUP
            # --------------------------------------------------------------
            workbook = openpyxl.load_workbook(
                BytesIO(base64.b64decode(self.file)),
                read_only=True,
                data_only=True
            )
            sheet = workbook.active

            Employee = self.env["hr.employee"]
            Upload = self.env["hr.upload"]
            Attendance = self.env["hr.attendance"]
            Leave = self.env["hr.leave"]
            Allocation = self.env["hr.leave.allocation"]
            AccrualLevel = self.env["hr.leave.accrual.level"]
            LeaveType = self.env["hr.leave.type"]

            earned_leave = LeaveType.search([('work_entry_type_id.code', '=', 'LEAVE120')], limit=1)
            loss_of_pay = LeaveType.search([('work_entry_type_id.code', '=', 'LEAVE90')], limit=1)

            if not earned_leave or not loss_of_pay:
                raise UserError(_("Please configure LEAVE120 and LEAVE90 codes."))

            # --------------------------------------------------------------
            # Approve Leave Helper
            # --------------------------------------------------------------
            def _approve_leave(values, leave_label, work_date, session_period=None):

                formatted = work_date.strftime("%d-%m-%Y")

                if session_period == "am":
                    session_txt = "morning"
                elif session_period == "pm":
                    session_txt = "afternoon"
                else:
                    session_txt = False

                desc = (
                    f"{leave_label} taken on {formatted} ({session_txt})"
                    if session_txt
                    else f"{leave_label} taken on {formatted}"
                )

                values.update({
                    "name": desc,
                    "payslip_state": "done",
                })

                leave = Leave.create(values)
                try:
                    if leave.state == "draft":
                        leave.action_confirm()
                    if leave.state != "validate":
                        leave.action_validate()
                except:
                    pass

            # --------------------------------------------------------------
            # Normalizer for ALL time formats
            # --------------------------------------------------------------
            def normalize_time(val):
                if not val:
                    return False

                # Excel float (0.55678)
                if isinstance(val, float):
                    try:
                        total_seconds = int(val * 24 * 3600)
                        hh = total_seconds // 3600
                        mm = (total_seconds % 3600) // 60
                        return f"{hh:02d}:{mm:02d}"
                    except:
                        return False

                # datetime/time
                if isinstance(val, datetime):
                    return val.strftime("%H:%M")
                if isinstance(val, time):
                    return val.strftime("%H:%M")

                # String: HH:MM:SS
                if isinstance(val, str):
                    v = val.strip()

                    # HH:MM:SS
                    try:
                        return datetime.strptime(v, "%H:%M:%S").strftime("%H:%M")
                    except:
                        pass

                    # H:M / HH:MM / 01:5 / 1:05 etc.
                    if ":" in v:
                        try:
                            h, m = v.split(":")[0], v.split(":")[1]
                            if len(m) == 1:
                                m = "0" + m
                            return f"{int(h):02d}:{int(m):02d}"
                        except:
                            return False

                    return False

                return False

            # --------------------------------------------------------------
            # Duplicate prevention
            # --------------------------------------------------------------
            duplicate_checker = set()
            collected_absences = {}

            row_no = 1

            # --------------------------------------------------------------
            # MAIN EXCEL LOOP
            # --------------------------------------------------------------
            for row in sheet.iter_rows(min_row=2, values_only=True):
                row_no += 1

                emp_code_x = row[0]
                emp_name_x = row[1]
                excel_date = row[2]
                excel_in = row[3]
                excel_out = row[4]

                lunch_in = row[5]
                lunch_out = row[6]
                break_in = row[7]
                break_out = row[8]

                # ---------------- BASIC CHECKS ----------------
                if not emp_code_x and not emp_name_x and not excel_date:
                    continue
                if not emp_code_x or not emp_name_x:
                    raise UserError(_("Row %d: Employee code/name missing.") % row_no)
                if not excel_date:
                    raise UserError(_("Row %d: Date missing.") % row_no)

                if isinstance(excel_date, datetime):
                    work_date = excel_date.date()
                elif isinstance(excel_date, date):
                    work_date = excel_date
                else:
                    work_date = parse(str(excel_date), dayfirst=True).date()

                if work_date > fields.Date.today():
                    raise UserError(_("Row %d: Future date not allowed.") % row_no)

                emp_code = str(emp_code_x).strip()
                emp_name = str(emp_name_x).strip()

                # Duplicate check
                if (emp_code, work_date) in duplicate_checker:
                    raise UserError(_("Row %d: Duplicate entry for %s on %s") %
                                    (row_no, emp_code, work_date))
                duplicate_checker.add((emp_code, work_date))

                # ---------------- EMPLOYEE ----------------
                employee = Employee.search([
                    ('barcode', '=', emp_code),
                    ('name', '=', emp_name)
                ], limit=1)

                if not employee:
                    raise UserError(_("Row %d: Employee not found (%s - %s)") %
                                    (row_no, emp_code, emp_name))

                if (
                        not employee.contract_id or
                        not employee.resource_calendar_id or
                        employee.contract_id.resource_calendar_id != employee.resource_calendar_id
                ):
                    raise UserError(_("Row %d: Contract / Working Schedule mismatch.") % row_no)

                # ---------------- NORMALIZE TIMES ----------------
                check_in = normalize_time(excel_in)
                check_out = normalize_time(excel_out)

                lunch_in_t = normalize_time(lunch_in)
                lunch_out_t = normalize_time(lunch_out)
                break_in_t = normalize_time(break_in)
                break_out_t = normalize_time(break_out)

                # ---------------- CREATE hr.upload ----------------
                upload = Upload.create({
                    "employee_name": employee.id,
                    # "employee_code": emp_code,
                    "date": work_date,
                    "check_in_attendance": check_in,
                    "check_out_attendance": check_out,
                    "lunch_in": lunch_in_t,
                    "lunch_out": lunch_out_t,
                    "break_in": break_in_t,
                    "break_out": break_out_t,
                })

                # ---------------- ABSENCE CALC ----------------
                absence_qty = 0
                if upload.morning_session == "Absent":
                    absence_qty += 0.5
                if upload.afternoon_session == "Absent":
                    absence_qty += 0.5

                # ---------------- ATTENDANCE CREATION ----------------
                if absence_qty in (0.0, 0.5):
                    Attendance.create({
                        "employee_id": employee.id,
                        "date": work_date,
                        "check_in": upload.session_check_in,
                        # "check_in": upload.your_datetime,
                        "check_out": upload.session_check_out,
                        # "check_out": upload.your_checkout_datetime,
                        'check_in_attendance': upload.your_datetime,
                        'check_out_attendance': upload.your_checkout_datetime,
                        "morning_session": upload.morning_session,
                        "afternoon_session": upload.afternoon_session,
                        "full_day_status": upload.full_day_status,
                        "difference_check_in": upload.difference_check_in,
                        "difference_check_out": upload.difference_check_out,
                        "total_working_hours": upload.total_working_hours,
                        "lunch_in": upload.lunch_in,
                        "lunch_out": upload.lunch_out,
                        "break_in": upload.break_in,
                        "break_out": upload.break_out,
                        "lunch_delay": upload.lunch_delay,
                        "break_delay": upload.break_delay,
                        "lunch_delay_amount": upload.lunch_delay_amount,
                        "break_delay_amount": upload.break_delay_amount,
                        "lunch_addition_amount": upload.lunch_addition_amount,
                    })

                # ---------------- SAVE ABSENCES FOR EL/LOP ----------------
                if absence_qty in (0.5, 1.0):
                    month_key = work_date.strftime("%Y-%m")
                    ckey = (employee.id, month_key)
                    collected_absences.setdefault(ckey, [])
                    collected_absences[ckey].append(
                        (work_date, absence_qty,
                         upload.morning_session,
                         upload.afternoon_session)
                    )

            # --------------------------------------------------------------
            # NO ABSENCES → DONE
            # --------------------------------------------------------------
            if not collected_absences:
                return

            # --------------------------------------------------------------
            # EL / LOP PROCESSING
            # --------------------------------------------------------------
            for (emp_id, month_ym), absence_list in collected_absences.items():

                employee = Employee.browse(emp_id)
                absence_list.sort(key=lambda x: x[0])

                eligible_for_el = (
                        employee.leave_eligibility == "yes" and
                        employee.ctc_type == "non_ctc"
                )

                # Remaining EL
                remaining_el_days = 0.0

                if eligible_for_el:
                    allocation = Allocation.search([
                        ('employee_id', '=', emp_id),
                        ('holiday_status_id', '=', earned_leave.id),
                        ('state', '=', 'validate'),
                        ('accrual_plan_id', '!=', False)
                    ], limit=1)

                    if allocation:
                        level = AccrualLevel.search([
                            ('accrual_plan_id', '=', allocation.accrual_plan_id.id)
                        ], limit=1)
                        remaining_el_days = level.added_value if level else 0

                # ---- PROCESS EACH ABSENCE ----
                for work_date, qty, ms, asess in absence_list:

                    # CASE: 1.0 absent but only 0.5 EL left
                    if qty == 1.0 and remaining_el_days == 0.5:
                        el_period = "am" if ms == "Absent" else "pm"
                        lop_period = "pm" if el_period == "am" else "am"

                        _approve_leave({
                            "employee_id": emp_id,
                            "holiday_status_id": earned_leave.id,
                            "request_date_from": work_date,
                            "request_date_to": work_date,
                            "number_of_days": 0.5,
                            "request_unit_half": True,
                            "request_date_from_period": el_period,
                        }, "EL", work_date, el_period)

                        _approve_leave({
                            "employee_id": emp_id,
                            "holiday_status_id": loss_of_pay.id,
                            "request_date_from": work_date,
                            "request_date_to": work_date,
                            "number_of_days": 0.5,
                            "request_unit_half": True,
                            "request_date_from_period": lop_period,
                        }, "LOP", work_date, lop_period)

                        remaining_el_days = 0
                        continue

                    # CASE: Normal EL usage
                    if remaining_el_days > 0:
                        use_el = min(remaining_el_days, qty)
                        remaining_el_days -= use_el

                        # 0.5 EL
                        if use_el == 0.5:
                            period = "am" if ms == "Absent" else "pm"
                            _approve_leave({
                                "employee_id": emp_id,
                                "holiday_status_id": earned_leave.id,
                                "request_date_from": work_date,
                                "request_date_to": work_date,
                                "number_of_days": 0.5,
                                "request_unit_half": True,
                                "request_date_from_period": period,
                            }, "EL", work_date, period)

                        # 1.0 EL
                        elif use_el == 1.0:
                            _approve_leave({
                                "employee_id": emp_id,
                                "holiday_status_id": earned_leave.id,
                                "request_date_from": work_date,
                                "request_date_to": work_date,
                                "number_of_days": 1.0,
                            }, "EL", work_date, None)

                        qty -= use_el

                    # CASE: Remaining → LOP
                    if qty == 0.5:
                        period = "am" if ms == "Absent" else "pm"
                        _approve_leave({
                            "employee_id": emp_id,
                            "holiday_status_id": loss_of_pay.id,
                            "request_date_from": work_date,
                            "request_date_to": work_date,
                            "number_of_days": 0.5,
                            "request_unit_half": True,
                            "request_date_from_period": period,
                        }, "LOP", work_date, period)

                    elif qty == 1.0:
                        _approve_leave({
                            "employee_id": emp_id,
                            "holiday_status_id": loss_of_pay.id,
                            "request_date_from": work_date,
                            "request_date_to": work_date,
                            "number_of_days": 1.0,
                        }, "LOP", work_date, None)

        except Exception as e:
            raise UserError(_("Upload failed: %s") % e)

    # def action_upload(self):
    #     """Upload attendance from Excel and:
    #        - create hr.upload lines
    #        - create hr.attendance for worked days
    #        - create & auto-approve EL / LOP from absences
    #     """
    #     if not self.file:
    #         raise UserError(_("Please upload a file."))
    #
    #     try:
    #         # ==============================================================
    #         # 0. SETUP: MODELS, LEAVE TYPES, HELPERS
    #         # ==============================================================
    #         workbook = openpyxl.load_workbook(
    #             BytesIO(base64.b64decode(self.file)),
    #             read_only=True,
    #             data_only=True
    #         )
    #         sheet = workbook.active
    #
    #         Employee = self.env["hr.employee"]
    #         Upload = self.env["hr.upload"]
    #         Attendance = self.env["hr.attendance"]
    #         Leave = self.env["hr.leave"]
    #         Allocation = self.env["hr.leave.allocation"]
    #         AccrualLevel = self.env["hr.leave.accrual.level"]
    #         LeaveType = self.env["hr.leave.type"]
    #
    #         # ---- Time Off types based on Work Entry Code ----
    #         earned_leave = LeaveType.search([('work_entry_type_id.code', '=', 'LEAVE120')], limit=1)  # EL
    #         loss_of_pay = LeaveType.search([('work_entry_type_id.code', '=', 'LEAVE90')], limit=1)  # LOP
    #         if not earned_leave or not loss_of_pay:
    #             raise UserError(_("Please configure Time Off types LEAVE120 (EL) & LEAVE90 (LOP)."))
    #
    #         # ---- Helper: create a leave and try to confirm + validate it ----
    #         def _approve_leave(values, leave_label, work_date, session_period=None):
    #             """
    #             leave_label: 'EL' or 'LOP'
    #             work_date:   datetime.date used for description
    #             session_period: 'am' or 'pm' or None
    #             """
    #             formatted_date = work_date.strftime("%d-%m-%Y")
    #
    #             if session_period == "am":
    #                 session_text = "morning"
    #             elif session_period == "pm":
    #                 session_text = "afternoon"
    #             else:
    #                 session_text = False
    #
    #             if session_text:
    #                 description = f"{leave_label} taken on {formatted_date} ({session_text})"
    #             else:
    #                 description = f"{leave_label} taken on {formatted_date}"
    #
    #             values.update({
    #                 "name": description,
    #                 "payslip_state": "done",
    #             })
    #
    #             leave = Leave.create(values)
    #             try:
    #                 if leave.state == "draft":
    #                     leave.action_confirm()
    #                 if leave.state in ["confirm", "validate1", "validate"]:
    #                     leave.action_validate()
    #             except Exception:
    #                 # even if confirm/validate fails, we keep the leave record
    #                 pass
    #
    #         # ---- Internal memory structures ----
    #         duplicate_checker = set()  # prevent same (emp_code, work_date) twice in same file
    #         collected_absences = {}  # { (emp_id, 'YYYY-MM'): [(work_date, absence_qty, AM, PM)] }
    #
    #         # ==============================================================
    #         # 1. LOOP EXCEL ROWS: VALIDATE, CREATE UPLOAD, CREATE ATTENDANCE
    #         # ==============================================================
    #         row_no = 1
    #         for row in sheet.iter_rows(min_row=2, values_only=True):
    #             row_no += 1
    #
    #             excel_emp_code = row[0]
    #             excel_emp_name = row[1]
    #             excel_date = row[2]
    #             excel_in = row[3]
    #             excel_out = row[4]
    #
    #             # ---------- 1.1 BASIC FIELD VALIDATION ----------
    #             if not excel_emp_code or not excel_emp_name:
    #                 raise UserError(_("Row %d: Employee code/name missing.") % row_no)
    #             if not excel_date:
    #                 raise UserError(_("Row %d: Date missing.") % row_no)
    #
    #             # ---------- 1.2 PARSE DATE ----------
    #             if isinstance(excel_date, datetime):
    #                 work_date = excel_date.date()
    #             elif isinstance(excel_date, date):
    #                 work_date = excel_date
    #             else:
    #                 work_date = parse(str(excel_date), dayfirst=True).date()
    #
    #             # No future-dated uploads
    #             if work_date > fields.Date.today():
    #                 raise UserError(_("Row %d: Future date %s not allowed.") % (row_no, work_date))
    #
    #             # ---------- 1.3 PREVENT DUPLICATE (EMP + DATE) IN SAME FILE ----------
    #             emp_code = str(excel_emp_code).strip()
    #             duplicate_key = (emp_code, work_date)
    #             if duplicate_key in duplicate_checker:
    #                 raise UserError(_("Row %d: Duplicate entry for %s on %s.") %
    #                                 (row_no, emp_code, work_date))
    #             duplicate_checker.add(duplicate_key)
    #
    #             # ---------- 1.4 FIND EMPLOYEE ----------
    #             emp_name = str(excel_emp_name).strip()
    #             employee = Employee.search([
    #                 ('cmr_code', '=', emp_code),
    #                 ('name', '=', emp_name)
    #             ], limit=1)
    #             if not employee:
    #                 raise UserError(_("Row %d: Employee not found (%s - %s)") %
    #                                 (row_no, emp_code, emp_name))
    #
    #             # ---------- 1.5 CONTRACT + SCHEDULE VALIDATION ----------
    #             if (not employee.contract_id
    #                     or not employee.resource_calendar_id
    #                     or employee.contract_id.resource_calendar_id.id != employee.resource_calendar_id.id):
    #                 raise UserError(_("Row %d: Contract/Working schedule mismatch for %s.") %
    #                                 (row_no, employee.name))
    #
    #             # ---------- 1.6 NORMALIZE TIME TO 'HH:MM' OR False ----------
    #             def normalize_time(value):
    #                 """Convert cell value to 'HH:MM' or False (no time)."""
    #                 if value and isinstance(value, (datetime, time)):
    #                     return value.strftime("%H:%M")
    #                 if value and isinstance(value, str) and ":" in value:
    #                     return value[:5]  # cut seconds if present
    #                 return False  # means no in/out → considered Absent side
    #
    #             emp_in_time = normalize_time(excel_in)
    #             emp_out_time = normalize_time(excel_out)
    #
    #             # ---------- 1.7 CREATE UPLOAD LINE ----------
    #             upload_record = Upload.create({
    #                 "employee_name": employee.id,
    #                 "employee_code": emp_code,
    #                 "date": work_date,
    #                 "check_in_attendance": emp_in_time,
    #                 "check_out_attendance": emp_out_time,
    #             })
    #
    #             # ---------- 1.8 ABSENCE QUANTITY (0 / 0.5 / 1) ----------
    #             absence_qty = 0.0
    #             if upload_record.morning_session == "Absent":
    #                 absence_qty += 0.5
    #             if upload_record.afternoon_session == "Absent":
    #                 absence_qty += 0.5
    #
    #             # ---------- 1.9 ATTENDANCE CREATION ----------
    #             if absence_qty in (0.0, 0.5):
    #                 Attendance.create({
    #                     "employee_id": employee.id,
    #                     "date": work_date,
    #                     "check_in": upload_record.session_check_in,
    #                     "check_out": upload_record.session_check_out,
    #                     "morning_session": upload_record.morning_session,
    #                     "afternoon_session": upload_record.afternoon_session,
    #                     "full_day_status": upload_record.full_day_status,
    #                     'difference_check_in': upload_record.difference_check_in,
    #                     'difference_check_out': upload_record.difference_check_out,
    #                     'total_working_hours': upload_record.total_working_hours,
    #                 })
    #
    #             # ---------- 1.10 SAVE ABSENCE FOR EL/LOP CREATION ----------
    #             if absence_qty in (0.5, 1.0):
    #                 month_key = work_date.strftime("%Y-%m")
    #                 dict_key = (employee.id, month_key)
    #                 collected_absences.setdefault(dict_key, [])
    #                 collected_absences[dict_key].append(
    #                     (work_date, absence_qty,
    #                      upload_record.morning_session,
    #                      upload_record.afternoon_session)
    #                 )
    #
    #         # If there are no absences at all → nothing to do for time off
    #         if not collected_absences:
    #             return
    #
    #         # ==============================================================
    #         # 2. PER EMPLOYEE + MONTH: CREATE EL FIRST, THEN LOP
    #         # ==============================================================
    #         for (emp_id, month_ym), absence_list in collected_absences.items():
    #             employee = Employee.browse(emp_id)
    #
    #             # Sort by work_date so earliest absences use EL first
    #             absence_list.sort(key=lambda rec: rec[0])
    #
    #             # ---------- 2.1 ELIGIBILITY CHECK ----------
    #             eligible_for_el = (
    #                     employee.leave_eligibility == "yes"
    #                     and employee.ctc_type == "non_ctc"
    #             )
    #
    #             # ---------- 2.2 FETCH MONTHLY ENTITLEMENT (EL) ----------
    #             remaining_el_days = 0.0
    #             if eligible_for_el:
    #                 allocation = Allocation.search([
    #                     ('employee_id', '=', emp_id),
    #                     ('holiday_status_id', '=', earned_leave.id),
    #                     ('state', '=', 'validate'),
    #                     ('accrual_plan_id', '!=', False),
    #                 ], limit=1)
    #
    #                 if allocation:
    #                     level = AccrualLevel.search(
    #                         [('accrual_plan_id', '=', allocation.accrual_plan_id.id)],
    #                         limit=1
    #                     )
    #                     remaining_el_days = level.added_value if level else 0.0
    #
    #             # ---------- 2.3 PROCESS EACH ABSENT DATE ----------
    #             for work_date, absence_qty, morning_session, afternoon_session in absence_list:
    #
    #                 # CASE 1: Full-day absent, but only 0.5 EL remains → 0.5 EL + 0.5 LOP
    #                 if absence_qty == 1.0 and remaining_el_days == 0.5:
    #                     el_period = "am" if morning_session == "Absent" else "pm"
    #                     lop_period = "pm" if el_period == "am" else "am"
    #
    #                     # 0.5 EL (with session text)
    #                     _approve_leave({
    #                         "employee_id": emp_id,
    #                         "holiday_status_id": earned_leave.id,
    #                         "request_date_from": work_date,
    #                         "request_date_to": work_date,
    #                         "number_of_days": 0.5,
    #                         "request_unit_half": True,
    #                         "request_date_from_period": el_period,
    #                     }, "EL", work_date, el_period)
    #
    #                     # 0.5 LOP (opposite session)
    #                     _approve_leave({
    #                         "employee_id": emp_id,
    #                         "holiday_status_id": loss_of_pay.id,
    #                         "request_date_from": work_date,
    #                         "request_date_to": work_date,
    #                         "number_of_days": 0.5,
    #                         "request_unit_half": True,
    #                         "request_date_from_period": lop_period,
    #                     }, "LOP", work_date, lop_period)
    #
    #                     remaining_el_days = 0.0
    #                     continue  # go to next date
    #
    #                 # CASE 2: Use EL normally (as much as available)
    #                 if remaining_el_days > 0.0:
    #                     used_el = min(absence_qty, remaining_el_days)
    #                     remaining_el_days -= used_el
    #
    #                     if used_el == 0.5:
    #                         period = "am" if morning_session == "Absent" else "pm"
    #                         _approve_leave({
    #                             "employee_id": emp_id,
    #                             "holiday_status_id": earned_leave.id,
    #                             "request_date_from": work_date,
    #                             "request_date_to": work_date,
    #                             "number_of_days": 0.5,
    #                             "request_unit_half": True,
    #                             "request_date_from_period": period,
    #                         }, "EL", work_date, period)
    #
    #                     elif used_el == 1.0:
    #                         _approve_leave({
    #                             "employee_id": emp_id,
    #                             "holiday_status_id": earned_leave.id,
    #                             "request_date_from": work_date,
    #                             "request_date_to": work_date,
    #                             "number_of_days": 1.0,
    #                         }, "EL", work_date, None)
    #
    #                     # Remaining (if any) must go as LOP
    #                     absence_qty -= used_el
    #
    #                 # CASE 3: Remaining absence becomes LOP
    #                 if absence_qty == 0.5:
    #                     period = "am" if morning_session == "Absent" else "pm"
    #                     _approve_leave({
    #                         "employee_id": emp_id,
    #                         "holiday_status_id": loss_of_pay.id,
    #                         "request_date_from": work_date,
    #                         "request_date_to": work_date,
    #                         "number_of_days": 0.5,
    #                         "request_unit_half": True,
    #                         "request_date_from_period": period,
    #                     }, "LOP", work_date, period)
    #
    #                 elif absence_qty == 1.0:
    #                     _approve_leave({
    #                         "employee_id": emp_id,
    #                         "holiday_status_id": loss_of_pay.id,
    #                         "request_date_from": work_date,
    #                         "request_date_to": work_date,
    #                         "number_of_days": 1.0,
    #                     }, "LOP", work_date, None)
    #
    #     except Exception as e:
    #         raise UserError(_("Upload failed: %s") % e)

    # 18th nov
    # def action_upload(self):
    #     """Uploads attendance Excel file, with correct EL/LOP monthly logic and session-based splitting.
    #     Handles: multi-month import, proportional EL usage across months, half-day rules, AM/PM split.
    #     """
    #
    #     if not self.file:
    #         raise ValidationError("Please upload a file.")
    #
    #     try:
    #         import base64
    #         import openpyxl
    #         from io import BytesIO
    #         from datetime import datetime, date, time, timedelta
    #         import pytz
    #         from calendar import monthrange
    #
    #         from odoo import fields, _
    #         from odoo.exceptions import UserError
    #
    #         wb = openpyxl.load_workbook(
    #             filename=BytesIO(base64.b64decode(self.file)),
    #             read_only=True
    #         )
    #         ws = wb.active
    #
    #         HrUpload = self.env['hr.upload']
    #         Employee = self.env['hr.employee']
    #         Attendance = self.env['hr.attendance']
    #         Leave = self.env['hr.leave']
    #         Allocation = self.env['hr.leave.allocation']
    #         LeaveType = self.env['hr.leave.type']
    #
    #         # === Leave Type Lookup by Work Entry Code ===
    #         el_type = LeaveType.search([('work_entry_type_id.code', '=', 'LEAVE120')], limit=1)
    #         lop_type = LeaveType.search([('work_entry_type_id.code', '=', 'LEAVE90')], limit=1)
    #         if not el_type or not lop_type:
    #             raise UserError(
    #                 _("Configure Time Off types LEAVE120 (EL) and LEAVE90 (LOP)."))
    #
    #         # === Set timezone ===
    #         tz = pytz.timezone(self.env.user.tz or 'Asia/Kolkata')
    #
    #         def normalize_time(tv):
    #             if not tv:
    #                 return None
    #             try:
    #                 if isinstance(tv, (datetime, time)):
    #                     return tv.strftime("%H:%M")
    #                 if isinstance(tv, (int, float)):
    #                     from openpyxl.utils.datetime import from_excel
    #                     dt = from_excel(tv)
    #                     return dt.strftime("%H:%M")
    #                 s = str(tv).strip().replace(";", ":").replace(".", ":")
    #                 p = s.split(":")
    #                 return f"{p[0].zfill(2)}:{p[1].zfill(2)}" if len(p) >= 2 else None
    #             except:
    #                 return None
    #
    #         # Proportional EL usage across calendar days
    #         def proportion_of_leave_in_period(leave_rec, ps, pe):
    #             if not leave_rec.request_date_from:
    #                 return 0.0
    #             start = leave_rec.request_date_from
    #             end = leave_rec.request_date_to
    #             if isinstance(start, datetime): start = start.date()
    #             if isinstance(end, datetime): end = end.date()
    #
    #             os = max(start, ps)
    #             oe = min(end, pe)
    #             if os > oe:
    #                 return 0.0
    #
    #             overlap_days = (oe - os).days + 1
    #             total_days = (end - start).days + 1
    #
    #             # half-day single record
    #             if getattr(leave_rec, "request_unit_half", False) and start == end:
    #                 return 0.5 if (ps <= start <= pe) else 0.0
    #
    #             try:
    #                 nd = float(getattr(leave_rec, "number_of_days_display",
    #                                    leave_rec.number_of_days or 0.0))
    #             except:
    #                 nd = float(leave_rec.number_of_days or 0.0)
    #
    #             return (overlap_days / total_days) * nd if total_days > 0 else 0.0
    #
    #         # Shortcut create+validate
    #         def create_and_validate_leave(leave_type, date_val, days, period=None):
    #             """Always half-day if days=0.5, otherwise full-day."""
    #             if days <= 0:
    #                 return
    #
    #             vals = {
    #                 'name': f"Auto Leave {date_val}",
    #                 'employee_id': leave_employee.id,
    #                 'holiday_status_id': leave_type.id,
    #                 'request_date_from': date_val,
    #                 'request_date_to': date_val,
    #                 'number_of_days': days,
    #             }
    #
    #             if days == 0.5:
    #                 vals.update({
    #                     'request_unit_half': True,
    #                     'request_date_from_period': period,  # 'am' or 'pm'
    #                 })
    #
    #             try:
    #                 lr = Leave.create(vals)
    #             except:
    #                 return
    #
    #             try:
    #                 if lr.state == "draft":
    #                     lr.action_confirm()
    #             except:
    #                 pass
    #             try:
    #                 if lr.state in ["confirm", "validate1", "validate"]:
    #                     lr.action_validate()
    #             except:
    #                 pass
    #
    #         # === Process each row ===
    #         row_number = 1
    #         for row in ws.iter_rows(min_row=2, values_only=True):
    #             row_number += 1
    #             emp_code = row[0]
    #             emp_name = row[1]
    #             date_value = row[2]
    #             check_in_raw = row[3]
    #             check_out_raw = row[4]
    #
    #             if not emp_code or not emp_name:
    #                 raise UserError(f"Row {row_number}: Employee code/name missing")
    #
    #             emp_name = str(emp_name).strip()
    #             try:
    #                 emp_code_str = str(int(emp_code))
    #             except:
    #                 emp_code_str = str(emp_code).strip()
    #
    #             leave_employee = Employee.search([
    #                 ('cmr_code', '=', emp_code_str),
    #                 ('name', '=', emp_name)
    #             ], limit=1)
    #             if not leave_employee:
    #                 raise UserError(f"Row {row_number}: Employee not found: {emp_code_str} {emp_name}")
    #
    #             # === Parse date ===
    #             from openpyxl.utils.datetime import from_excel
    #             fdate = None
    #             if isinstance(date_value, datetime):
    #                 fdate = date_value.date()
    #             elif isinstance(date_value, (int, float)):
    #                 try:
    #                     fdate = from_excel(date_value).date()
    #                 except:
    #                     pass
    #             elif isinstance(date_value, str):
    #                 for fmt in ["%d-%m-%Y", "%d/%m/%Y", "%Y-%m-%d", "%d-%b-%Y",
    #                             "%d/%b/%Y", "%b-%d-%Y", "%d.%m.%Y", "%m/%d/%Y", "%d/%m/%y"]:
    #                     try:
    #                         fdate = datetime.strptime(date_value.strip(), fmt).date()
    #                         break
    #                     except:
    #                         continue
    #
    #             if not fdate:
    #                 raise UserError(f"Row {row_number}: Invalid date {date_value}")
    #
    #             check_in = normalize_time(check_in_raw)
    #             check_out = normalize_time(check_out_raw)
    #
    #             # === Upload record (sessions computed) ===
    #             up = HrUpload.create({
    #                 'employee_name': leave_employee.id,
    #                 'ctc_type': leave_employee.ctc_type,
    #                 'employee_code': emp_code_str,
    #                 'date': fdate,
    #                 'check_in_attendance': check_in,
    #                 'check_out_attendance': check_out,
    #             })
    #             if hasattr(up, "_compute_sessions"):
    #                 up._compute_sessions()
    #
    #             morning = up.morning_session
    #             afternoon = up.afternoon_session
    #
    #             # Absent duration
    #             if morning == "Absent" and afternoon == "Absent":
    #                 leave_duration = 1.0
    #             elif morning == "Absent" or afternoon == "Absent":
    #                 leave_duration = 0.5
    #             else:
    #                 leave_duration = 0
    #
    #             # === LEAVE PROCESSING ===
    #             if leave_duration > 0:
    #
    #                 # Non-CTC → Only LOP
    #                 if leave_employee.ctc_type != "non_ctc":
    #                     monthly_accrual = 0.0
    #                     remaining_el = 0.0
    #                 else:
    #                     # Get monthly accrual
    #                     alloc = Allocation.search([
    #                         ('employee_id', '=', leave_employee.id),
    #                         ('holiday_status_id', '=', el_type.id),
    #                         ('state', '=', 'validate')
    #                     ], limit=1)
    #
    #                     if alloc and alloc.accrual_plan_id:
    #                         lvl = self.env["hr.leave.accrual.level"].search([
    #                             ('accrual_plan_id', '=', alloc.accrual_plan_id.id)
    #                         ], limit=1)
    #                         monthly_accrual = lvl.added_value if lvl else 0.0
    #                     else:
    #                         monthly_accrual = 0.0
    #
    #                     # Compute EL used this month
    #                     y = fdate.year
    #                     m = fdate.month
    #                     first = date(y, m, 1)
    #                     last = date(y, m, monthrange(y, m)[1])
    #
    #                     used = 0.0
    #                     leaves = Leave.search([
    #                         ('employee_id', '=', leave_employee.id),
    #                         ('holiday_status_id', '=', el_type.id),
    #                         ('state', 'in', ['validate', 'confirm'])
    #                     ])
    #                     for l in leaves:
    #                         used += proportion_of_leave_in_period(l, first, last)
    #
    #                     remaining_el = max(monthly_accrual - used, 0.0)
    #
    #                 # === CASE 1: FULL DAY ABSENT & REMAINING EL = 0.5 → Split AM/PM ===
    #                 if leave_duration == 1.0 and 0 < remaining_el < 1.0:
    #
    #                     # If only 0.5 EL remains → assign EL to morning, LOP to afternoon.
    #                     create_and_validate_leave(el_type, fdate, 0.5, period="am")
    #                     create_and_validate_leave(lop_type, fdate, 0.5, period="pm")
    #                     # Skip normal flow
    #                     pass
    #
    #                 # === CASE 2: HALF DAY ABSENT ===
    #                 elif leave_duration == 0.5:
    #                     # If remaining EL available → assign EL
    #                     if remaining_el >= 0.5 and leave_employee.leave_eligibility == "yes":
    #                         period = "am" if morning == "Absent" else "pm"
    #                         create_and_validate_leave(el_type, fdate, 0.5, period=period)
    #                     else:
    #                         period = "am" if morning == "Absent" else "pm"
    #                         create_and_validate_leave(lop_type, fdate, 0.5, period=period)
    #
    #                 # === CASE 3: FULL DAY ABSENT NORMAL ===
    #                 else:
    #                     # enough EL available
    #                     if remaining_el >= 1.0 and leave_employee.leave_eligibility == "yes":
    #                         create_and_validate_leave(el_type, fdate, 1.0)
    #                     else:
    #                         create_and_validate_leave(lop_type, fdate, 1.0)
    #
    #             # ====================
    #             # ATTENDANCE LOGIC
    #             # ====================
    #             cal = leave_employee.resource_calendar_id
    #             if not cal:
    #                 continue
    #
    #             dow = fdate.strftime("%A")
    #             wh = cal.attendance_ids.filtered(
    #                 lambda a: a.display_name and dow.lower() in a.display_name.lower()
    #             )
    #             if not wh:
    #                 continue
    #
    #             morning_sched = wh.filtered(lambda a: "morning" in a.display_name.lower())
    #             afternoon_sched = wh.filtered(lambda a: "afternoon" in a.display_name.lower())
    #
    #             def to_utc(d, hf):
    #                 h = int(hf)
    #                 mi = int(round((hf - h) * 60))
    #                 dt = datetime.combine(d, time(h, mi))
    #                 dt = tz.localize(dt)
    #                 return dt.astimezone(pytz.UTC)
    #
    #             def to_dt_local(d, tstr):
    #                 try:
    #                     t = datetime.strptime(tstr, "%H:%M").time()
    #                     dt = datetime.combine(d, t)
    #                     return tz.localize(dt).astimezone(pytz.UTC)
    #                 except:
    #                     return None
    #
    #             check_in_dt = check_out_dt = None
    #
    #             # morning present
    #             if morning != "Absent" and morning_sched:
    #                 sched_out = to_utc(fdate, max(morning_sched.mapped("hour_to")))
    #                 actual_in = to_dt_local(fdate, check_in) if check_in else None
    #                 check_in_dt = actual_in or to_utc(fdate, min(morning_sched.mapped("hour_from")))
    #                 check_out_dt = sched_out
    #
    #             # afternoon present
    #             if afternoon != "Absent" and afternoon_sched:
    #                 sched_in = to_utc(fdate, min(afternoon_sched.mapped("hour_from")))
    #                 actual_out = to_dt_local(fdate, check_out) if check_out else None
    #
    #                 if check_in_dt:
    #                     check_out_dt = actual_out or to_utc(fdate, max(afternoon_sched.mapped("hour_to")))
    #                 else:
    #                     check_in_dt = sched_in
    #                     check_out_dt = actual_out or to_utc(fdate, max(afternoon_sched.mapped("hour_to")))
    #
    #             # create attendance
    #             if check_in_dt and check_out_dt and check_out_dt > check_in_dt:
    #                 open_att = Attendance.search([
    #                     ('employee_id', '=', leave_employee.id),
    #                     ('check_out', '=', False)
    #                 ], limit=1)
    #
    #                 if open_att:
    #                     open_att.write({'check_out': fields.Datetime.to_string(check_out_dt)})
    #                 else:
    #                     Attendance.create({
    #                         'employee_id': leave_employee.id,
    #                         'date': fdate,
    #                         'check_in': fields.Datetime.to_string(check_in_dt),
    #                         'check_out': fields.Datetime.to_string(check_out_dt),
    #                         'difference_check_in': up.difference_check_in,
    #                         'difference_check_out': up.difference_check_out,
    #                         'total_working_hours': up.total_working_hours,
    #                         'morning_session': up.morning_session,
    #                         'afternoon_session': up.afternoon_session,
    #                         'full_day_status': up.full_day_status,
    #                     })
    #
    #     except Exception as e:
    #         raise UserError(_("Upload failed: %s") % e)

    # 17th nov
    # def action_upload(self):
    #     """Uploads attendance Excel file and creates hr.upload, hr.attendance, and hr.leave records
    #     with EL/LOP rules and hybrid working schedule–based attendance logic."""
    #     if not self.file:
    #         raise UserError(_("Please upload a file."))
    #
    #     try:
    #         import base64
    #         import openpyxl
    #         from io import BytesIO
    #         from datetime import datetime, date, time, timedelta
    #         import pytz
    #
    #         wb = openpyxl.load_workbook(
    #             filename=BytesIO(base64.b64decode(self.file)),
    #             read_only=True
    #         )
    #         ws = wb.active
    #
    #         HrUpload = self.env['hr.upload']
    #         Employee = self.env['hr.employee']
    #         Attendance = self.env['hr.attendance']
    #         Leave = self.env['hr.leave']
    #         Allocation = self.env['hr.leave.allocation']
    #         LeaveType = self.env['hr.leave.type']
    #
    #         # === Leave Type Lookup by Work Entry Code ===
    #         el_type = LeaveType.search([('work_entry_type_id.code', '=', 'LEAVE120')], limit=1)
    #         lop_type = LeaveType.search([('work_entry_type_id.code', '=', 'LEAVE90')], limit=1)
    #         if not el_type or not lop_type:
    #             raise UserError(
    #                 _("Please configure Time Off types with work entry codes LEAVE120 (EL) and LEAVE90 (LOP)."))
    #
    #         # === Set timezone ===
    #         user_tz = self.env.user.tz or 'Asia/Kolkata'
    #         tz = pytz.timezone(user_tz)
    #
    #         def normalize_time(time_value):
    #             """Convert Excel time or float value to HH:MM format, ignoring seconds."""
    #             if not time_value:
    #                 return None
    #             try:
    #                 if isinstance(time_value, (datetime, time)):
    #                     return time_value.strftime("%H:%M")
    #                 if isinstance(time_value, (int, float)):
    #                     from openpyxl.utils.datetime import from_excel
    #                     dt = from_excel(time_value)
    #                     return dt.strftime("%H:%M")
    #                 time_str = str(time_value).strip().replace(';', ':').replace('.', ':')
    #                 parts = time_str.split(':')
    #                 if len(parts) >= 2:
    #                     return f"{parts[0].zfill(2)}:{parts[1].zfill(2)}"
    #                 return None
    #             except Exception:
    #                 return None
    #
    #
    #         # === Process each row ===
    #         row_number = 1
    #         for row in ws.iter_rows(min_row=2, values_only=True):
    #             row_number += 1
    #
    #             employee_code = row[0]
    #             employee_name = row[1]
    #             date_value = row[2]
    #             check_in = row[3]
    #             check_out = row[4]
    #
    #             if not employee_code:
    #                 raise UserError(_("Row %d: Employee code is missing.") % row_number)
    #             if not employee_name:
    #                 raise UserError(_("Row %d: Employee name is missing.") % row_number)
    #
    #             employee_name = str(employee_name).strip()
    #             try:
    #                 employee_code_str = str(int(employee_code))
    #             except Exception:
    #                 employee_code_str = str(employee_code).strip()
    #
    #             employee = Employee.search([
    #                 ('cmr_code', '=', employee_code_str),
    #                 ('name', '=', employee_name)
    #             ], limit=1)
    #             if not employee:
    #                 raise UserError(_("Row %d: No employee found with code '%s' and name '%s'.") %
    #                                 (row_number, employee_code_str, employee_name))
    #
    #             # === Date Parsing ===
    #             from openpyxl.utils.datetime import from_excel
    #
    #             formatted_date = False
    #             if isinstance(date_value, datetime):
    #                 formatted_date = date_value.date()
    #             elif isinstance(date_value, (int, float)):
    #                 try:
    #                     formatted_date = from_excel(date_value).date()
    #                 except Exception:
    #                     pass
    #             elif isinstance(date_value, str):
    #                 date_value = date_value.strip()
    #                 for fmt in [
    #                     "%d-%m-%Y", "%d/%m/%Y", "%Y-%m-%d", "%d-%b-%Y",
    #                     "%d/%b/%Y", "%b-%d-%Y", "%d.%m.%Y", "%m/%d/%Y", "%d/%m/%y"
    #                 ]:
    #                     try:
    #                         formatted_date = datetime.strptime(date_value, fmt).date()
    #                         break
    #                     except Exception:
    #                         continue
    #
    #             if not formatted_date:
    #                 raise UserError(_("Row %d: Invalid date format '%s'.") % (row_number, date_value))
    #
    #             check_in = normalize_time(check_in)
    #             check_out = normalize_time(check_out)
    #
    #             # === Create hr.upload Record ===
    #             upload_rec = HrUpload.create({
    #                 'employee_name': employee.id,
    #                 'ctc_type': employee.ctc_type,
    #                 'employee_code': employee_code_str,
    #                 'date': formatted_date,
    #                 'check_in_attendance': check_in,
    #                 'check_out_attendance': check_out,
    #             })
    #
    #             if hasattr(upload_rec, '_compute_sessions'):
    #                 upload_rec._compute_sessions()
    #
    #             morning = getattr(upload_rec, 'morning_session', '') or ''
    #             afternoon = getattr(upload_rec, 'afternoon_session', '') or ''
    #             leave_duration = 0.0
    #             half_day = False
    #
    #             if morning == 'Absent' and afternoon == 'Absent':
    #                 leave_duration = 1.0
    #             elif morning == 'Absent' or afternoon == 'Absent':
    #                 leave_duration = 0.5
    #                 half_day = True
    #
    #             # === Leave Logic ===
    #             if leave_duration > 0:
    #                 existing_leave = Leave.search([
    #                     ('employee_id', '=', employee.id),
    #                     ('request_date_from', '=', formatted_date),
    #                     ('request_date_to', '=', formatted_date),
    #                     ('state', '!=', 'refuse')
    #                 ], limit=1)
    #                 if existing_leave:
    #                     continue
    #
    #                 if employee.ctc_type != 'non_ctc':
    #                     leave_type_to_use = lop_type
    #                     remaining_el = 0.0
    #                 else:
    #                     el_allocations = Allocation.search([
    #                         ('employee_id', '=', employee.id),
    #                         ('holiday_status_id', '=', el_type.id),
    #                         ('state', '=', 'validate')
    #                     ], limit=1)
    #
    #                     remaining_el = 0.0
    #                     if el_allocations:
    #                         alloc = el_allocations[0]
    #                         start_date = alloc.date_from
    #                         accrual_per_month = 0.0
    #
    #                         if alloc.accrual_plan_id:
    #                             accrual_level = self.env['hr.leave.accrual.level'].search([
    #                                 ('accrual_plan_id', '=', alloc.accrual_plan_id.id)
    #                             ], limit=1)
    #                             if accrual_level:
    #                                 accrual_per_month = accrual_level.added_value
    #
    #                         if not accrual_per_month:
    #                             raise UserError(_(
    #                                 "Row %d: No accrual value defined in Accrual Plan for %s."
    #                             ) % (row_number, employee.name))
    #
    #                         if start_date and formatted_date >= start_date:
    #                             months_passed = (formatted_date.year - start_date.year) * 12 + (
    #                                     formatted_date.month - start_date.month) + 1
    #                             total_allocated = months_passed * accrual_per_month
    #                         else:
    #                             total_allocated = 0.0
    #
    #                         el_taken = sum(Leave.search([
    #                             ('employee_id', '=', employee.id),
    #                             ('holiday_status_id', '=', el_type.id),
    #                             ('state', 'in', ['validate', 'confirm']),
    #                             ('request_date_from', '<=', formatted_date)
    #                         ]).mapped('number_of_days'))
    #
    #                         remaining_el = max(total_allocated - el_taken, 0.0)
    #                     else:
    #                         raise UserError(_(
    #                             "Row %d: No validated EL allocation found for %s."
    #                         ) % (row_number, employee.name))
    #
    #                 def create_and_validate_leave(leave_type, days, half_day=False, period=False):
    #                     vals = {
    #                         'name': f"Auto Leave {formatted_date}",
    #                         'employee_id': employee.id,
    #                         'holiday_status_id': leave_type.id,
    #                         'request_date_from': formatted_date,
    #                         'request_date_to': formatted_date,
    #                         'number_of_days': days,
    #                     }
    #                     if half_day:
    #                         vals.update({
    #                             'request_unit_half': True,
    #                             'request_date_from_period': period,
    #                         })
    #                     leave_rec = Leave.create(vals)
    #                     if leave_rec.state == 'draft':
    #                         try:
    #                             leave_rec.action_confirm()
    #                         except Exception:
    #                             pass
    #                     if leave_rec.state in ['confirm', 'validate1', 'validate']:
    #                         try:
    #                             leave_rec.action_validate()
    #                         except Exception:
    #                             pass
    #
    #                 if employee.ctc_type == 'non_ctc':
    #                     if remaining_el >= leave_duration:
    #                         create_and_validate_leave(
    #                             el_type, leave_duration, half_day,
    #                             'am' if half_day and morning == 'Absent' else 'pm' if half_day else False
    #                         )
    #                     elif remaining_el > 0:
    #                         create_and_validate_leave(el_type, remaining_el)
    #                         create_and_validate_leave(lop_type, leave_duration - remaining_el)
    #                     else:
    #                         create_and_validate_leave(
    #                             lop_type, leave_duration, half_day,
    #                             'am' if half_day and morning == 'Absent' else 'pm' if half_day else False
    #                         )
    #                 else:
    #                     create_and_validate_leave(
    #                         lop_type, leave_duration, half_day,
    #                         'am' if half_day and morning == 'Absent' else 'pm' if half_day else False
    #                     )
    #
    #             # === Attendance creation (Hybrid logic) ===
    #             calendar = employee.resource_calendar_id
    #             if not calendar:
    #                 continue
    #
    #             day_of_week = formatted_date.strftime("%A")
    #             working_hours = calendar.attendance_ids.filtered(
    #                 lambda a: a.display_name and day_of_week.lower() in a.display_name.lower()
    #             )
    #             if not working_hours:
    #                 continue
    #
    #             morning_sched = working_hours.filtered(lambda a: 'morning' in a.display_name.lower())
    #             afternoon_sched = working_hours.filtered(lambda a: 'afternoon' in a.display_name.lower())
    #
    #             def to_utc_datetime(date_part, hour_float):
    #                 hours = int(hour_float)
    #                 minutes = int(round((hour_float - hours) * 60))
    #                 local_dt = datetime.combine(date_part, time(hours, minutes))
    #                 local_dt = tz.localize(local_dt)
    #                 return local_dt.astimezone(pytz.UTC)
    #
    #             def to_datetime_local(date_part, time_str):
    #                 try:
    #                     t = datetime.strptime(time_str, "%H:%M").time()
    #                     local_dt = datetime.combine(date_part, t)
    #                     local_dt = tz.localize(local_dt)
    #                     return local_dt.astimezone(pytz.UTC)
    #                 except Exception:
    #                     return None
    #
    #             check_in_dt = check_out_dt = None
    #
    #             # Morning present: actual in, schedule out
    #             if morning != 'Absent' and morning_sched:
    #                 sched_out = to_utc_datetime(formatted_date, max(morning_sched.mapped('hour_to')))
    #                 actual_in = to_datetime_local(formatted_date, check_in) if check_in else None
    #                 check_in_dt = actual_in or to_utc_datetime(formatted_date, min(morning_sched.mapped('hour_from')))
    #                 check_out_dt = sched_out
    #
    #             # Afternoon present: schedule in, actual out
    #             if afternoon != 'Absent' and afternoon_sched:
    #                 sched_in = to_utc_datetime(formatted_date, min(afternoon_sched.mapped('hour_from')))
    #                 actual_out = to_datetime_local(formatted_date, check_out) if check_out else None
    #
    #                 if check_in_dt and check_out_dt:
    #                     # merge morning+afternoon
    #                     check_out_dt = actual_out or to_utc_datetime(formatted_date,
    #                                                                  max(afternoon_sched.mapped('hour_to')))
    #                 else:
    #                     check_in_dt = sched_in
    #                     check_out_dt = actual_out or to_utc_datetime(formatted_date,
    #                                                                  max(afternoon_sched.mapped('hour_to')))
    #
    #             if check_in_dt and check_out_dt and check_out_dt > check_in_dt:
    #                 open_attendance = Attendance.search([
    #                     ('employee_id', '=', employee.id),
    #                     ('check_out', '=', False)
    #                 ], limit=1)
    #
    #                 if open_attendance:
    #                     open_attendance.write({'check_out': fields.Datetime.to_string(check_out_dt)})
    #                 else:
    #                     Attendance.create({
    #                         'employee_id': employee.id,
    #                         'date': formatted_date,
    #                         'check_in': fields.Datetime.to_string(check_in_dt),
    #                         'check_out': fields.Datetime.to_string(check_out_dt),
    #                         'difference_check_in': upload_rec.difference_check_in,
    #                         'difference_check_out': upload_rec.difference_check_out,
    #                         'total_working_hours': upload_rec.total_working_hours,
    #                         'morning_session': upload_rec.morning_session,
    #                         'afternoon_session': upload_rec.afternoon_session,
    #                         'full_day_status': upload_rec.full_day_status,
    #                     })
    #
    #     except Exception as e:
    #         raise UserError(_('Upload failed: %s') % e)


class HrLateDeductionMaster(models.Model):
    _name = "hr.late.deduction.master"
    _description = "Late Deduction Master"
    _order = "start_time asc"

    shift_id = fields.Many2one('resource.calendar', string="Shift")
    name = fields.Char(string="Name", required=True)
    start_time = fields.Char(string="Start Time (HH:MM)", required=True)  # Example: 10:10
    end_time = fields.Char(string="End Time (HH:MM)", required=True)  # Example: 10:30
    deduction_amount = fields.Float(string="Deduction Amount", required=True)


class HrMonthlyLateDeduction(models.Model):
    _name = "hr.monthly.late.deduction"
    _description = "Monthly Late Deduction"

    employee_id = fields.Many2one('hr.employee', string="Employee")
    month = fields.Selection(
        [(str(i), calendar.month_name[i]) for i in range(1, 13)],
        string="Month",
        required=True,
        default=lambda self: str(datetime.now().month)
    )

    year = fields.Integer(string="Year", required=True, default=lambda self: datetime.now().year)

    line_ids = fields.One2many(
        'hr.late.deduction.line',
        'monthly_id',
        string="Late Deduction Lines"
    )
    select_all = fields.Boolean("Select All", compute=False, store=False)

    @api.onchange('select_all')
    def _onchange_select_all(self):
        for rec in self:
            for line in rec.line_ids:
                line.is_selected = rec.select_all

    def delete_selected_lines(self):
        for rec in self:
            selected = rec.line_ids.filtered(lambda l: l.is_selected)
            selected.unlink()

    def action_generate_all_employees(self):
        """Generate or refresh monthly employee late deduction lines from hr.upload dynamically."""
        self.ensure_one()

        # Convert month & year safely
        month = int(self.month)
        year = int(str(self.year).replace(',', ''))

        # Compute date range
        start_date = datetime(year, month, 1)
        last_day = calendar.monthrange(year, month)[1]
        end_date = datetime(year, month, last_day)

        # Get all upload records for this month
        uploads = self.env['hr.upload'].search([
            ('date', '>=', start_date.date()),
            ('date', '<=', end_date.date()),
        ])

        if not uploads:
            return {
                'effect': {
                    'fadeout': 'slow',
                    'message': 'No upload data found for this period!',
                    'type': 'rainbow_man',
                }
            }

        # Get unique employees from uploads
        employee_ids = uploads.mapped('employee_name.id')

        line_vals = []
        for emp_id in employee_ids:
            emp_uploads = uploads.filtered(lambda u: u.employee_name.id == emp_id)

            # Safely handle any string values in late_deduction
            late_days = len(emp_uploads.filtered(lambda u: float(u.late_deduction or 0) > 0))
            late_hours = sum(
                [float(u.late_hours or 0) for u in emp_uploads]) if 'late_hours' in emp_uploads._fields else 0.0
            deduction_amount = sum(
                [float(u.late_deduction or 0) for u in emp_uploads]) if 'late_deduction' in emp_uploads._fields else 0.0

            line_vals.append((0, 0, {
                'employee_id': emp_id,
                'late_days': late_days,
                'late_hours': late_hours,
                'deduction_amount': deduction_amount,
            }))

        # Replace existing lines with fresh ones
        self.line_ids = [(5, 0, 0)] + line_vals

        return {
            'effect': {
                'fadeout': 'slow',
                'message': f'{len(employee_ids)} Employee Records Recalculated Successfully!',
                'type': 'rainbow_man',
            }
        }


class HrLateDeductionLine(models.Model):
    _name = "hr.late.deduction.line"
    _description = "Late Deduction Line"

    monthly_id = fields.Many2one('hr.monthly.late.deduction', string="Monthly Record")
    employee_id = fields.Many2one('hr.employee', string="Employee")
    late_days = fields.Integer(string="Late Days")
    late_hours = fields.Float(string="Late Hours")
    deduction_amount = fields.Float(string="Deduction Amount (₹)", compute="_compute_total_late_deduction", store=True)

    month = fields.Selection(
        related='monthly_id.month',
        string="Month",
        store=True,
        readonly=True
    )
    year = fields.Integer(
        related='monthly_id.year',
        string="Year",
        store=True,
        readonly=True
    )
    is_selected = fields.Boolean(string="Select", default=True)

    @api.depends('employee_id', 'monthly_id.month', 'monthly_id.year')
    def _compute_total_late_deduction(self):
        """Compute total monthly deduction for each employee based on hr.upload."""
        for rec in self:
            rec.deduction_amount = 0.0  # default
            if not rec.employee_id or not rec.monthly_id:
                continue

            # Safely get month & year
            try:
                month = int(rec.month)
                year = int(str(rec.year).replace(',', ''))
            except Exception:
                continue

            start_date = datetime(year, month, 1)
            last_day = calendar.monthrange(year, month)[1]
            end_date = datetime(year, month, last_day)

            # ✅ Search all hr.upload entries for that employee in the month
            uploads = self.env['hr.upload'].search([
                ('employee_name', '=', rec.employee_id.name),
                ('date', '>=', start_date.date()),
                ('date', '<=', end_date.date()),
            ])

            # ✅ Compute total late deduction
            rec.deduction_amount = sum(float(u.late_deduction or 0) for u in uploads)

            # Optional: if you want to count number of days with deductions


class HrOvertimeMaster(models.Model):
    _name = "hr.overtime.master"
    _description = "Overtime Master"
    _order = "start_time asc"

    name = fields.Char(string="Name", required=True)
    start_time = fields.Char(string="Start Time (HH:MM)", required=True)  # e.g. 18:00
    end_time = fields.Char(string="End Time (HH:MM)", required=True)  # e.g. 20:00
    overtime_amount = fields.Float(string="Overtime Amount (₹)")


class HrMonthlyOvertime(models.Model):
    _name = "hr.monthly.overtime"
    _description = "Monthly Overtime"

    employee_id = fields.Many2one('hr.employee', string="Employee")
    month = fields.Selection(
        [(str(i), calendar.month_name[i]) for i in range(1, 13)],
        string="Month",
        required=True,
        default=lambda self: str(datetime.now().month)
    )
    year = fields.Integer(string="Year", required=True, default=lambda self: datetime.now().year)

    line_ids = fields.One2many('hr.overtime.line', 'monthly_id', string="Overtime Lines")
    select_all = fields.Boolean(string="Select All")

    #  Onchange to select/unselect all lines
    @api.onchange('select_all')
    def _onchange_select_all(self):
        for rec in self:
            for line in rec.line_ids:
                line.is_selected = rec.select_all

    # Delete selected overtime lines
    def delete_selected_lines(self):
        for rec in self:
            selected_lines = rec.line_ids.filtered(lambda l: l.is_selected)
            if not selected_lines:
                raise UserError("Please select at least one overtime line to delete.")
            selected_lines.unlink()
            rec.select_all = False

    def action_generate_all_employees(self):
        """Generate or refresh monthly employee overtime summary dynamically from hr.upload."""
        self.ensure_one()

        # ✅ Convert month & year properly
        month = int(self.month)
        year = int(str(self.year).replace(',', ''))

        # ✅ Compute date range
        start_date = datetime(year, month, 1)
        last_day = calendar.monthrange(year, month)[1]
        end_date = datetime(year, month, last_day)

        # ✅ Fetch hr.upload records for this month
        uploads = self.env['hr.upload'].search([
            ('date', '>=', start_date.date()),
            ('date', '<=', end_date.date()),
        ])

        if not uploads:
            return {
                'effect': {
                    'fadeout': 'slow',
                    'message': 'No uploaded employees found for this period!',
                    'type': 'rainbow_man',
                }
            }

        employee_ids = uploads.mapped('employee_name.id')
        line_vals = []

        for emp_id in employee_ids:
            emp_uploads = uploads.filtered(lambda u: u.employee_name.id == emp_id)

            # ✅ Safely calculate total overtime (handles str/None types)

            overtime_amount = sum(float(u.overtime_amount or 0) for u in emp_uploads if hasattr(u, 'overtime_amount'))

            line_vals.append((0, 0, {
                'employee_id': emp_id,

                'overtime_amount': overtime_amount,
            }))

        # ✅ Clear existing lines and add new ones
        self.line_ids = [(5, 0, 0)] + line_vals

        return {
            'effect': {
                'fadeout': 'slow',
                'message': f'{len(employee_ids)} Employee Overtime Records Calculated!',
                'type': 'rainbow_man',
            }
        }


class HrOvertimeLine(models.Model):
    _name = "hr.overtime.line"
    _description = "Overtime Line"

    monthly_id = fields.Many2one('hr.monthly.overtime', string="Monthly Record")
    employee_id = fields.Many2one('hr.employee', string="Employee")

    overtime_amount = fields.Float(string="Overtime Amount (₹)", compute="_compute_total_overtime", store=True)

    month = fields.Selection(related='monthly_id.month', store=True, readonly=True)
    year = fields.Integer(related='monthly_id.year', store=True, readonly=True)
    is_selected = fields.Boolean(string="Select")

    @api.depends('employee_id', 'monthly_id.month', 'monthly_id.year')
    def _compute_total_overtime(self):
        """Compute total monthly overtime for each employee based on hr.upload."""
        for rec in self:
            rec.overtime_amount = 0.0
            if not rec.employee_id or not rec.monthly_id:
                continue

            try:
                # ✅ Always take month/year from monthly_id (not from rec)
                month = int(rec.monthly_id.month)
                year = int(str(rec.monthly_id.year).replace(',', ''))
            except Exception:
                continue

            # ✅ Compute date range safely
            start_date = datetime(year, month, 1)
            last_day = calendar.monthrange(year, month)[1]
            end_date = datetime(year, month, last_day)

            # ✅ Search all hr.upload entries for that employee
            uploads = self.env['hr.upload'].search([
                ('employee_name', '=', rec.employee_id.name),
                ('date', '>=', start_date.date()),
                ('date', '<=', end_date.date()),
            ])

            # ✅ Safely sum overtime amount (handles str/None)
            rec.overtime_amount = sum(float(u.overtime_amount or 0) for u in uploads)


class StateTaxMaster(models.Model):
    _name = "state.tax.master"

    state_id = fields.Many2one(
        'res.country.state',
        string="State",
        required=True
    )

    slab_ids = fields.One2many(
        'state.tax.slab',
        'tax_master_id',
        string="CTC Slabs"
    )

    class StateTaxSlab(models.Model):
        _name = "state.tax.slab"
        _description = "State Tax Slab"
        _order = "ctc_from"

        tax_master_id = fields.Many2one(
            'state.tax.master',
            string="Tax Master",
            ondelete="cascade",
            required=True
        )

        ctc_from = fields.Float(
            string="CTC From",
            required=True
        )

        ctc_to = fields.Float(
            string="CTC To",
            required=True
        )

        tax_amount = fields.Float(
            string="Tax Amount",
            required=True
        )

        @api.constrains('ctc_from', 'ctc_to')
        def _check_ctc_range(self):
            for rec in self:
                if rec.ctc_from >= rec.ctc_to:
                    raise ValidationError("CTC From must be less than CTC To.")


class ShiftMaster(models.Model):
    _name = "shift.master"
    _description = "Shift Master"

    name = fields.Char(string="Name", required=True)

    shift_from_id = fields.Many2one(
        'resource.calendar',
        string="From Shift",
        required=True
    )
    shift_to_id = fields.Many2one(
        'resource.calendar',
        string="To Shift",
        required=True
    )

    employee_line_ids = fields.One2many(
        'shift.master.line',
        'master_id',
        string="Employees"
    )

    @api.onchange('shift_from_id')
    def _onchange_shift_from_id(self):
        """Load all employees belonging to selected FROM shift"""
        if not self.shift_from_id:
            self.employee_line_ids = [(5, 0, 0)]
            return

        employees = self.env['hr.employee'].search([
            ('resource_calendar_id', '=', self.shift_from_id.id)
        ])

        lines = []
        for emp in employees:
            lines.append((0, 0, {
                'employee_id': emp.id,
            }))

        self.employee_line_ids = [(5, 0, 0)]  # clear previous
        self.employee_line_ids = lines

    def action_move_shift(self):
        if self.shift_from_id.id == self.shift_to_id.id:
            raise UserError("From Shift and To Shift cannot be same.")

        employees = self.employee_line_ids.mapped('employee_id')

        if not employees:
            raise UserError("No employees found to update.")

        # 1️⃣ Update Employee shift
        employees.write({'resource_calendar_id': self.shift_to_id.id})

        # 2️⃣ Update Contract shift (ALL STATES)
        contracts = self.env['hr.contract'].search([
            ('employee_id', 'in', employees.ids)
        ])

        print("🔍 Found Contracts:", contracts.ids)

        if contracts:
            contracts.write({'resource_calendar_id': self.shift_to_id.id})

        return {
            'effect': {
                'fadeout': 'slow',
                'message': 'Shift updated in Employee & ALL Contracts!',
                'type': 'rainbow_man',
            }
        }


class ShiftMasterLine(models.Model):
    _name = "shift.master.line"
    _description = "Shift Master Line"

    master_id = fields.Many2one('shift.master')
    employee_id = fields.Many2one('hr.employee', string="Employee", required=True)


class HrEmployee(models.Model):
    _inherit = 'hr.employee'

    lunch_bonus_category = fields.Selection([
        ('general_employee', 'General Employee'),
        ('female_supervisor', 'Female Supervisor'),
        ('female_senior', 'Female Senior'),
        ('male_supervisor', 'Male Supervisor'),
        ('male_senior', 'Male Senior'),
    ], string='Category', required=True)


class LunchBonusMaster(models.Model):
    _name = 'lunch.bonus.master'
    _description = 'Lunch Bonus Master'

    shift_id = fields.Many2one('resource.calendar', required=True)
    category = fields.Selection([
        ('female_supervisor', 'Female Supervisor'),
        ('female_senior', 'Female Senior'),
        ('male_supervisor', 'Male Supervisor'),
        ('male_senior', 'Male Senior'),
    ], string='Category', required=True)

    month = fields.Selection([
        ('1', 'January'),
        ('2', 'February'),
        ('3', 'March'),
        ('4', 'April'),
        ('5', 'May'),
        ('6', 'June'),
        ('7', 'July'),
        ('8', 'August'),
        ('9', 'September'),
        ('10', 'October'),
        ('11', 'November'),
        ('12', 'December'),
    ], string="Applicable Month", required=True)

    year = fields.Integer(string="Year", required=True, store=True)

    slab_ids = fields.One2many(
        'lunch.bonus.slab',
        'master_id',
        string='Bonus Slabs'
    )


class LunchBonusSlab(models.Model):
    _name = 'lunch.bonus.slab'
    _description = 'Lunch Bonus Slab'

    master_id = fields.Many2one('lunch.bonus.master', required=True)
    from_time = fields.Char(required=True)  # HH:MM
    to_time = fields.Char(required=True)  # HH:MM
    amount = fields.Float(required=True)


class HrSalaryAttachment(models.Model):
    _inherit = 'hr.salary.attachment'


    state = fields.Selection(
        selection=[
            ('draft', 'Draft'),
            ('waiting', 'Waiting for Approval'),
            ('approved', 'Approved'),
            ('open', 'Running'),
            ('close', 'Completed'),
            ('cancel', 'Cancelled'),
        ],
        string='Status',
        default='draft',
        required=True,
        tracking=True,
        copy=False,
    )

    def action_send_to_approval(self):
        for rec in self:
            if rec.state == 'draft':
                rec.state = 'waiting'

    def action_approve(self):
        if not self.env.user.has_group('hr.group_hr_manager'):
            raise AccessError("You are not allowed to approve.")
        for rec in self:
            if rec.state == 'waiting':
                rec.state = 'approved'

    def action_run(self):
        """Move from approved → open (running)"""
        for rec in self:
            if rec.state != 'approved':
                raise UserError("Record must be approved before running.")
            rec.state = 'open'  # normal running state




