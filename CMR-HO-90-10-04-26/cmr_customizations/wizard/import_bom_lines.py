import openpyxl
from io import BytesIO
from odoo import api, fields, models, _
from odoo.exceptions import ValidationError
import logging

_logger = logging.getLogger(__name__)

try:
    import xlrd
except ImportError:
    _logger.debug('Oops! Cannot `import xlrd`.')
try:
    import csv
except ImportError:
    _logger.debug('Oops! Cannot `import csv`.')
try:
    import base64
except ImportError:
    _logger.debug('Oops! Cannot `import base64`.')


class bom_line_ids_wizard(models.TransientModel):
    _name = 'bom.line.wizard'
    _description = "BoM Line Wizard"

    bom_file = fields.Binary(string="Select File")
    import_option = fields.Selection([('csv', 'CSV File'), ('xls', 'XLS File')], string='Select', default='csv')
    import_prod_option = fields.Selection([('barcode', 'Barcode'), ('code', 'Code'), ('name', 'Name')],
                                          string='Import Product By ', default='name')
    product_details_option = fields.Selection(
        [('from_product', 'Take Details From The Product'), ('from_xls', 'Take Details From The XLS File'),
         ], default='from_xls')

    def import_bom_lines(self):
        res = False
        counter = 0
        if self.import_option == 'csv':
            keys = ['product', 'quantity']
            try:
                wb = openpyxl.load_workbook(
                    filename=BytesIO(base64.b64decode(self.bom_file)), read_only=True
                )
                ws = wb.active
            except Exception:
                raise ValidationError(_("Please select any file or You have selected invalid file"))

            for row_no in ws.iter_rows(min_row=2, max_row=None, min_col=None,
                                       max_col=None, values_only=True):
                counter += 1
                field = list(map(str, row_no))
                values = dict(zip(keys, field))
                if values:
                    if len(row_no) == 0:
                        continue
                    else:
                        product_barcode = ''
                        lot_name = ''
                        if row_no[0] == None:
                            continue
                        else:
                            if len(row_no[0]) > 13:
                                if row_no[0][0] == '0' and row_no[0][1] == '1' and row_no[0][16] == '2' and row_no[0][
                                    17] == '1':
                                    for i in range(0, len(row_no[0])):
                                        if i > 1 and i < 16:
                                            product_barcode += row_no[0][i]
                                        elif i > 17 and i < len(row_no[0]):
                                            lot_name += row_no[0][i]
                                            continue
                            else:
                                product_barcode = row_no[0]
                        if row_no[1] == None or row_no[1] <= 0:
                            raise ValidationError(_('%s Quantity must be greater than zero.') % (row_no[0]))
                        if self.product_details_option == 'from_product':
                            values.update({
                                'product': product_barcode,
                                'serial_no': lot_name,
                                'quantity': row_no[1]
                            })
                        elif self.product_details_option == 'from_xls':
                            values.update({'product': product_barcode,
                                           'serial_no': lot_name,
                                           'quantity': row_no[1],
                                           })
                        else:
                            values.update({
                                'product': product_barcode,
                                'serial_no': lot_name,
                                'quantity': row_no[1],
                            })
                        res = self.create_bom_line(values)
        else:
            try:
                wb = openpyxl.load_workbook(
                    filename=BytesIO(base64.b64decode(self.bom_file)), read_only=True
                )
                ws = wb.active
                values = {}
            except Exception:
                raise ValidationError(_("Please select any file or You have selected invalid file"))
            for row_no in ws.iter_rows(min_row=2, max_row=None, min_col=None,
                                       max_col=None, values_only=True):
                counter += 1
                product_barcode = ''
                lot_name = ''
                if row_no[0] == None:
                    continue
                else:
                    if len(row_no[0]) > 13:
                        if row_no[0][0] == '0' and row_no[0][1] == '1' and row_no[0][16] == '2' and row_no[0][
                            17] == '1':
                            for i in range(0, len(row_no[0])):
                                if i > 1 and i < 16:
                                    product_barcode += row_no[0][i]
                                elif i > 17 and i < len(row_no[0]):
                                    lot_name += row_no[0][i]
                                    continue
                    else:
                        product_barcode = row_no[0]
                if row_no[1] == None or row_no[1] <= 0:
                    raise ValidationError(_('%s Quantity must be greater than zero.') % (row_no[0]))
                if self.product_details_option == 'from_product':
                    values.update({
                        'product': product_barcode,
                        'quantity': row_no[1],
                        'serial_no': lot_name
                    })
                elif self.product_details_option == 'from_xls':
                    values.update({'product': product_barcode,
                                   'quantity': row_no[1],
                                   'serial_no': lot_name,
                                   })
                else:
                    values.update({
                        'product': product_barcode,
                        'quantity': row_no[1],
                        'serial_no': lot_name,
                    })
                res = self.create_bom_line(values)
        view_id = self.env.ref('bi_import_all_orders_lines.message_wizard_popup')
        context = dict(self._context or {})
        dict_msg = str(counter) + " Records Imported Successfully."
        context['message'] = dict_msg
        return {
            'name': _('Success'),
            'type': 'ir.actions.act_window',
            'view_mode': 'form',
            'res_model': 'message.wizard',
            'views': [(view_id.id, 'form')],
            'view_id': view_id.id,
            'target': 'new',
            'context': context,
        }

    def create_bom_line(self, values):
        bom_brw = self.env['mrp.bom'].browse(self._context.get('active_id'))
        serial_no = self.env['stock.lot']
        product_obj_search = self.env['product.product']
        if self.product_details_option == 'from_product':
            if self.import_prod_option == 'barcode':
                if len(values['product']) > 13:
                    product_obj_search = self.env['product.product'].search([('barcode', '=', values['product'])])
                    serial_no = self.env['stock.lot'].search([('name', '=', values['serial_no'])])
                    if not serial_no:
                        raise ValidationError(
                            _('The serial number for this is not found in the database.'))
                else:
                    product_barcodes = self.env['product.barcode'].search([('barcode', '=', values['product'])])
                    if len(product_barcodes) > 0:
                        product_obj_search = product_barcodes[0].product_id
            elif self.import_prod_option == 'code':
                raise ValidationError(_('Please set the import Option to Barcode.'))
            else:
                raise ValidationError(_('Please set the import Option to Barcode.'))
            if product_obj_search:
                product_id = product_obj_search
            else:
                raise ValidationError(_('%s Product was not found in the Database.') % values.get('product'))
            if product_id.nhcl_product_type == 'unbranded':
                existing_product_line = bom_brw.bom_line_ids.filtered(
                    lambda x: x.product_id == product_id)
                existing_line = existing_product_line.prod_serial_no.filtered(lambda x:x.id == serial_no.id)
            else:
                existing_line = bom_brw.bom_line_ids.filtered(
                    lambda x: x.product_id == product_id and x.prod_barcode == values.get('product'))
            if existing_line:
                raise ValidationError(_('%s The product already exists.') % values.get('product'))
            existing_bom_line_ids = bom_brw.bom_line_ids.filtered(
                lambda x: x.product_id == product_id)
            if existing_bom_line_ids:
                existing_bom_line_ids.product_qty += values.get('quantity')
                if product_id.nhcl_product_type == 'unbranded':
                    existing_bom_line_ids.prod_serial_no = [(4, serial_no.id)]
            else:
                bom_line_ids = self.env['mrp.bom.line'].create({
                    'bom_id': bom_brw.id,
                    'product_id': product_id.id,
                    'prod_serial_no': serial_no.ids,
                    'prod_barcode': values.get('product'),
                    'product_qty': values.get('quantity'),
                    'product_uom_id': product_id.uom_id.id,
                })
        elif self.product_details_option == 'from_xls':
            if self.import_prod_option == 'barcode':
                barcode = values.get('product')
                if len(barcode) > 13:
                    product_obj_search = self.env['product.product'].search([('barcode', '=', barcode)])
                    serial_no = self.env['stock.lot'].search([('name', '=', values['serial_no'])])
                    if not serial_no:
                        raise ValidationError(
                            _('The serial number for this is not found in the database.'))
                else:
                    product_barcodes = self.env['product.barcode'].search([('barcode', '=', values['product'])])
                    if len(product_barcodes) > 0:
                        product_obj_search = product_barcodes[0].product_id
            elif self.import_prod_option == 'code':
                raise ValidationError(_('Please set the import Option to Barcode.'))
            else:
                raise ValidationError(_('Please set the import Option to Barcode.'))
            if product_obj_search:
                product_id = product_obj_search
            else:
                if self.import_prod_option == 'name':
                    raise ValidationError(_('Please set the import Option to Barcode.'))
                else:
                    raise ValidationError(
                        _('%s Product was not found in the Database') % values.get(
                            'product'))
            if product_id.nhcl_product_type == 'unbranded':
                existing_product_line = bom_brw.bom_line_ids.filtered(
                    lambda x: x.product_id == product_id)
                existing_line = existing_product_line.prod_serial_no.filtered(lambda x: x.id == serial_no.id)
            else:
                existing_line = bom_brw.bom_line_ids.filtered(
                    lambda x: x.product_id == product_id and x.prod_barcode == values.get('product'))
            if existing_line:
                raise ValidationError(_('%s The product already exists.') % values.get('product'))
            existing_bom_line_ids = bom_brw.bom_line_ids.filtered(
                lambda x: x.product_id == product_id)
            if existing_bom_line_ids:
                existing_bom_line_ids.product_qty += values.get('quantity')
                if product_id.nhcl_product_type == 'unbranded':
                    existing_bom_line_ids.prod_serial_no = [(4, serial_no.id)]
            else:
                bom_line_ids = self.env['mrp.bom.line'].create({
                    'bom_id': bom_brw.id,
                    'product_id': product_id.id,
                    'prod_serial_no': serial_no.ids,
                    'prod_barcode': values.get('product'),
                    'product_qty': values.get('quantity'),
                    'product_uom_id': product_id.uom_id.id,
                })
        else:
            if self.import_prod_option == 'barcode':
                if len(values['product']) > 13:
                    product_obj_search = self.env['product.product'].search([('barcode', '=', values['product'])])
                    serial_no = self.env['stock.lot'].search([('name', '=', values['serial_no'])])
                    if not serial_no:
                        raise ValidationError(
                            _('The serial number for this is not found in the database.'))
                else:
                    product_barcodes = self.env['product.barcode'].search([('barcode', '=', values['product'])])
                    if len(product_barcodes) > 0:
                        product_obj_search = product_barcodes[0].product_id
            elif self.import_prod_option == 'code':
                raise ValidationError(_('Please set the import Option to Barcode.'))
            else:
                raise ValidationError(_('Please set the import Option to Barcode.'))
            if product_obj_search:
                product_id = product_obj_search
            else:
                if self.import_prod_option == 'name':
                    raise ValidationError(_('Please set the import Option to Barcode.'))
                else:
                    raise ValidationError(
                        _('%s Product was not found in the Database.') % values.get(
                            'product'))
            if product_id.nhcl_product_type == 'unbranded':
                existing_product_line = bom_brw.bom_line_ids.filtered(
                    lambda x: x.product_id == product_id)
                existing_line = existing_product_line.prod_serial_no.filtered(lambda x: x.id == serial_no.id)
            else:
                existing_line = bom_brw.bom_line_ids.filtered(
                    lambda x: x.product_id == product_id and x.prod_barcode == values.get('product'))
            if existing_line:
                raise ValidationError(_('%s The product already exists.') % values.get('product'))
            existing_bom_line_ids = bom_brw.bom_line_ids.filtered(
                lambda x: x.product_id == product_id)
            if existing_bom_line_ids:
                existing_bom_line_ids.product_qty += values.get('quantity')
                if product_id.nhcl_product_type == 'unbranded':
                    existing_bom_line_ids.prod_serial_no = [(4, serial_no.id)]
            else:
                bom_line_ids = self.env['mrp.bom.line'].create({
                    'bom_id': bom_brw.id,
                    'product_id': product_id.id,
                    'prod_barcode': values.get('product'),
                    'prod_serial_no': serial_no.ids,
                    'product_qty': values.get('quantity'),
                })
        return True


class BOMLotWizard(models.TransientModel):
    _name = 'nhcl.bom.lot.wizard'
    _description = "BoM Lot Wizard"


    production_ids = fields.Many2many('mrp.production', string='Production')
    nom_categ_1 = fields.Many2one('product.attribute.value', string='Color', copy=False,
                              domain=[('attribute_id.name', '=', 'Color')])
    nom_categ_2 = fields.Many2one('product.attribute.value', string='Fit', copy=False,
                              domain=[('attribute_id.name', '=', 'Fit')])
    nom_categ_3 = fields.Many2one('product.attribute.value', string='Brand', copy=False,
                              domain=[('attribute_id.name', '=', 'Brand')])
    nom_categ_4 = fields.Many2one('product.attribute.value', string='Pattern', copy=False,
                              domain=[('attribute_id.name', '=', 'Pattern')])
    nom_categ_5 = fields.Many2one('product.attribute.value', string='Border Type', copy=False,
                              domain=[('attribute_id.name', '=', 'Border Type')])
    nom_categ_6 = fields.Many2one('product.attribute.value', string='Border Size', copy=False,
                              domain=[('attribute_id.name', '=', 'Border Size')])
    nom_categ_7 = fields.Many2one('product.attribute.value', string='Size', copy=False,
                              domain=[('attribute_id.name', '=', 'Size')])
    nom_categ_8 = fields.Many2one('product.attribute.value', string='Design', copy=False,
                              domain=[('attribute_id.name', '=', 'Design')])
    nom_descrip_1 = fields.Many2one('product.aging.line', string="Product Aging", copy=False,
                                )
    nom_descrip_2 = fields.Many2one('product.attribute.value', string='Range', copy=False,
                                domain=[('attribute_id.name', '=', 'Range')])
    nom_descrip_3 = fields.Many2one('product.attribute.value', string='Collection', copy=False,
                                domain=[('attribute_id.name', '=', 'Collection')])
    nom_descrip_4 = fields.Many2one('product.attribute.value', string='Fabric', copy=False,
                                domain=[('attribute_id.name', '=', 'Fabric')])
    nom_descrip_5 = fields.Many2one('product.attribute.value', string='Exclusive', copy=False,
                                domain=[('attribute_id.name', '=', 'Exclusive')])
    nom_descrip_6 = fields.Many2one('product.attribute.value', string='Print', copy=False,
                                domain=[('attribute_id.name', '=', 'Print')])
    nom_descrip_7 = fields.Many2one('product.attribute.value', string='Days Ageing', copy=False,
                                domain=[('attribute_id.name', '=', 'Days Ageing')])
    nom_descrip_8 = fields.Many2one('product.attribute.value', string='Description 8', copy=False,
                                domain=[('attribute_id.name', '=', 'Offer')])
    nom_descrip_9 = fields.Many2one('product.attribute.value', string='Discount', copy=False,
                                domain=[('attribute_id.name', '=', 'Discount')])


    def manf_update_cat_and_desp(self):
        if self.production_ids:
            for lot in self.production_ids:
                if lot.lot_producing_id and lot.state == 'done':
                    lot.lot_producing_id.write({
                        'category_1': self.nom_categ_1.id,
                        'category_2': self.nom_categ_2.id,
                        'category_3': self.nom_categ_3.id,
                        'category_4': self.nom_categ_4.id,
                        'category_5': self.nom_categ_5.id,
                        'category_6': self.nom_categ_6.id,
                        'category_7': self.nom_categ_7.id,
                        'category_8': self.nom_categ_8.id,
                        'description_1': self.nom_descrip_1.id,
                        'description_2': self.nom_descrip_2.id,
                        'description_3': self.nom_descrip_3.id,
                        'description_4': self.nom_descrip_4.id,
                        'description_5': self.nom_descrip_5.id,
                        'description_6': self.nom_descrip_6.id,
                        'description_7': self.nom_descrip_7.id,
                        'description_8': self.nom_descrip_8.id,
                        'description_9': self.nom_descrip_9.id,
                    })
                else:
                    raise ValidationError("All Manufacturing order is not in Done state to update the attributes.")


