import openpyxl
from io import BytesIO
from openpyxl import load_workbook
from odoo import fields, models, _
from odoo.exceptions import ValidationError
import logging

_logger = logging.getLogger(__name__)

try:
    import xlrd
except ImportError:
    _logger.debug('Oops! Cannot `import xlrd`.')
try:
    import csv
except ImportError:
    _logger.debug('Oops! Cannot `import csv`.')
try:
    import base64
except ImportError:
    _logger.debug('Oops! Cannot `import base64`.')


class order_line_wizard(models.TransientModel):
    _name = 'order.line.wizard'
    _description = "Order Line Wizard"

    sale_order_id = fields.Many2one('sale.order', string="Sale Order", required=True)
    file = fields.Binary(string="Excel File", required=True)
    filename = fields.Char(string="Filename")

    def action_import_barcodes(self):
        """Read Excel and create/merge sale order lines based on GS1 or EAN-13 barcode format.
        - Merges quantities for the same lot in the same sale order (lot-tracked products).
        - Validates that the total requested qty for a barcode does not exceed actual available qty.
        - When merging/updating qty, always update the line's price_unit from the lot (cost_price)
          falling back to product.lst_price.
        - For serial-tracked products behavior:
            * Each serial (lot that represents a serial) is handled as a separate 1-qty line.
            * If the same serial is already used in this order or another active order it will be skipped.
        """
        self.ensure_one()
        if not self.file:
            raise ValidationError(_("Please upload an Excel file."))

        try:
            data = base64.b64decode(self.file)
            wb = load_workbook(filename=BytesIO(data), data_only=True)
            sheet = wb.active
        except Exception as e:
            raise ValidationError(_("Invalid Excel file: %s") % str(e))

        SaleOrderLine = self.env['sale.order.line']
        StockQuant = self.env['stock.quant']
        StockMoveLine = self.env['stock.move.line']
        location_id = self.env.ref('stock.stock_location_stock').id
        skipped_lines = []

        # Helper: find existing sale order line for same product + lot + serial type in this sale order
        def _find_existing_line(product, lot_id, serial_type):
            return SaleOrderLine.search([
                ('order_id', '=', self.sale_order_id.id),
                ('product_id', '=', product.id),
                ('lot_ids', 'in', [lot_id]),
                ('sale_serial_type', '=', serial_type),
                ('order_id.state', 'not in', ['cancel'])
            ], limit=1)

        # Helper: compute actual available qty for a lot (qty received in receipts - qty already allocated in active sale lines)
        def _actual_available_qty_for_lot(lot):
            # qty that has come in for this lot via receipts
            in_qty = sum(StockMoveLine.search([
                ('picking_id.picking_type_id.code', '=', 'incoming'),  # safer than string 'receipt'
                ('company_id.nhcl_company_bool', '=', True),
                ('lot_id', '=', lot.id)
            ]).mapped('qty_done')) or 0.0

            # qty already used in active sale order lines that reference this lot
            used_qty = sum(SaleOrderLine.search([
                ('order_id.state', 'not in', ['cancel']),
                ('lot_ids', 'in', [lot.id])
            ]).mapped('product_uom_qty')) or 0.0

            return float(in_qty) - float(used_qty)

        # Iterate rows
        for row in sheet.iter_rows(min_row=2, values_only=True):
            barcode = str(row[0]).strip() if row[0] else ""
            qty = row[1] or 0

            if not barcode or not qty:
                skipped_lines.append((barcode, "Missing barcode or quantity"))
                continue

            # Prepare variables reused across blocks
            product = False
            quants = False

            # ------------------ GS1 (identified by presence of 'R' in the barcode) ------------------
            if 'R' in barcode:
                # GS1 -> lot code starts from 'R'
                lot_code = barcode[barcode.find('R'):]
                quant_domain = [
                    ('lot_id.name', '=', lot_code),
                    ('company_id.nhcl_company_bool', '=', True),
                    ('quantity', '>', 0),
                    ('location_id', '=', location_id),
                    # product type filter used in original code
                    ('lot_id.type_product', '=', 'un_brand')
                ]
                quants = StockQuant.search(quant_domain, order='id asc')
                product = quants[:1].product_id if quants else False

                if not product:
                    skipped_lines.append((barcode, "No matching product for GS1 barcode"))
                    continue

                # Determine sale_serial_type already present in this order for this product (keep first if exists)
                existing_types = SaleOrderLine.search([
                    ('order_id', '=', self.sale_order_id.id),
                    ('order_id.state', 'not in', ['cancel']),
                    ('product_id', '=', product.id)
                ]).mapped('sale_serial_type')
                sale_serial_type = existing_types[0] if existing_types else 'regular'

                # SERIAL TRACKING (each serial is a unique lot representing a serial)
                if product.tracking == 'serial':
                    # For serials, treat each available quant (lot) as a single item.
                    available_quants = quants.filtered(lambda q: q.quantity > 0)
                    if not available_quants:
                        skipped_lines.append((barcode, "No stock available for GS1 serials"))
                        continue

                    # Validate overall availability: total available serials must be >= qty
                    if qty > len(available_quants):
                        skipped_lines.append(
                            (barcode, f"Requested qty {qty} exceeds available serials {len(available_quants)}"))
                        continue

                    created_count = 0
                    for q in available_quants:
                        lot = q.lot_id

                        # Check if this serial is already used in any active sale order
                        existing_lines = SaleOrderLine.search([
                            ('lot_ids', 'in', [lot.id]),
                            ('order_id.state', 'not in', ['cancel'])
                        ])
                        same_order_line = existing_lines.filtered(lambda l: l.order_id.id == self.sale_order_id.id)
                        if same_order_line:
                            skipped_lines.append((barcode, f"Serial {lot.name} already used in this sale order"))
                            continue

                        other_orders = existing_lines.filtered(lambda l: l.order_id.id != self.sale_order_id.id)
                        if other_orders:
                            skipped_lines.append((barcode, f"Serial {lot.name} already used in another sale order"))
                            continue

                        price_unit = getattr(lot, 'cost_price', product.lst_price)

                        # Create sale order line for this serial (qty = 1)
                        SaleOrderLine.create({
                            'order_id': self.sale_order_id.id,
                            'product_id': product.id,
                            'lot_ids': [(6, 0, [lot.id])],
                            'branded_barcode': lot.ref,
                            'name': product.display_name,
                            'product_uom_qty': 1,
                            'product_uom': product.uom_id.id,
                            'price_unit': price_unit,
                            'type_product': getattr(lot, 'type_product', False),
                            'sale_serial_type': sale_serial_type,
                        })
                        created_count += 1
                        if created_count >= qty:
                            break

                # LOT TRACKING (this is where merging is applied)
                else:
                    # Build list of lots (quants) with their actual available qty
                    available_quants = quants.filtered(lambda q: q.quantity > 0)
                    if not available_quants:
                        skipped_lines.append((barcode, "No stock available for GS1 lot items"))
                        continue

                    # Validate total available across all matching lots
                    total_available = sum(max(0.0, _actual_available_qty_for_lot(q.lot_id)) for q in available_quants)
                    if qty > total_available:
                        skipped_lines.append(
                            (barcode, f"Requested qty {qty} exceeds available stock {int(total_available)}"))
                        continue

                    remaining_qty = qty
                    for q in available_quants:
                        lot = q.lot_id
                        actual_available_qty = max(0.0, _actual_available_qty_for_lot(lot))
                        if actual_available_qty <= 0:
                            continue

                        allocate_qty = min(remaining_qty, actual_available_qty)
                        if allocate_qty <= 0:
                            continue

                        # Merge or create
                        existing_line = _find_existing_line(product, lot.id, sale_serial_type)
                        price_unit = getattr(lot, 'cost_price', product.lst_price)

                        if existing_line:
                            # Update qty and ALWAYS update price from lot (option A)
                            new_qty = (existing_line.product_uom_qty or 0.0) + allocate_qty
                            existing_line.write({
                                'product_uom_qty': new_qty,
                                'price_unit': price_unit,
                            })
                        else:
                            SaleOrderLine.create({
                                'order_id': self.sale_order_id.id,
                                'product_id': product.id,
                                'lot_ids': [(6, 0, [lot.id])],
                                'branded_barcode': lot.ref,
                                'name': product.display_name,
                                'product_uom_qty': allocate_qty,
                                'product_uom': product.uom_id.id,
                                'price_unit': price_unit,
                                'type_product': getattr(lot, 'type_product', False),
                                'sale_serial_type': sale_serial_type,
                            })

                        remaining_qty -= allocate_qty
                        if remaining_qty <= 0:
                            break

                    if remaining_qty > 0:
                        skipped_lines.append(
                            (barcode, f"Requested qty exceeds available stock. Short by {int(remaining_qty)}"))

            # ------------------ EAN-13 (length 13) ------------------
            elif len(barcode) == 13:
                lot_code = barcode
                matched_quant = StockQuant.search([
                    ('lot_id.ref', '=', lot_code),
                    ('company_id.nhcl_company_bool', '=', True),
                    ('quantity', '>', 0),
                    ('location_id', '=', location_id),
                    ('lot_id.type_product', '=', 'brand')
                ], limit=1)

                if not matched_quant:
                    skipped_lines.append((barcode, "No stock found for EAN-13 barcode"))
                    continue

                product = matched_quant.product_id

                # Domain to gather all quants for that lot_code + product
                quant_domain = [
                    ('product_id', '=', product.id),
                    ('lot_id.ref', '=', lot_code),
                    ('company_id.nhcl_company_bool', '=', True),
                    ('quantity', '>', 0),
                    ('location_id', '=', location_id),
                    ('lot_id.type_product', '=', 'brand')
                ]
                quants = StockQuant.search(quant_domain, order='id asc')
                available_quants = quants.filtered(lambda q: q.quantity > 0)

                existing_types = SaleOrderLine.search([
                    ('order_id', '=', self.sale_order_id.id),
                    ('order_id.state', 'not in', ['cancel']),
                    ('product_id', '=', product.id)
                ]).mapped('sale_serial_type')
                sale_serial_type = existing_types[0] if existing_types else 'regular'

                # SERIAL TRACKING (Enhanced: include regular + return serials)
                if product.tracking == 'serial':
                    # Build regular & return quants (prioritize regular then return)
                    regular_quants = StockQuant.search([
                        ('product_id', '=', product.id),
                        ('lot_id.ref', '=', lot_code),
                        ('company_id.nhcl_company_bool', '=', True),
                        ('quantity', '>', 0),
                        ('location_id', '=', location_id),
                        ('lot_id.type_product', '=', 'brand'),
                        '|', ('lot_id.serial_type', '=', False), ('lot_id.serial_type', '=', 'regular'),
                    ], order='id asc')

                    return_quants = StockQuant.search([
                        ('product_id', '=', product.id),
                        ('lot_id.ref', '=', lot_code),
                        ('company_id.nhcl_company_bool', '=', True),
                        ('quantity', '>', 0),
                        ('location_id', '=', location_id),
                        ('lot_id.type_product', '=', 'brand'),
                        ('lot_id.serial_type', '=', 'return'),
                    ], order='id asc')

                    combined_quants = regular_quants + return_quants
                    if not combined_quants:
                        skipped_lines.append((barcode, "No serial stock available for EAN-13"))
                        continue

                    # Validate total available serials
                    if qty > len(combined_quants):
                        skipped_lines.append(
                            (barcode, f"Requested qty {qty} exceeds available serials {len(combined_quants)}"))
                        continue

                    created_count = 0
                    for q in combined_quants:
                        lot = q.lot_id

                        # Skip if serial already used in any active sale order with same type
                        existing_lines = SaleOrderLine.search([
                            ('lot_ids', 'in', [lot.id]),
                            ('order_id.state', 'not in', ['cancel'])
                        ])
                        if existing_lines:
                            same_type_line = existing_lines.filtered(
                                lambda l: l.sale_serial_type == (lot.serial_type or 'regular'))
                            if same_type_line:
                                # Already used with same serial type → skip this serial
                                continue

                        current_serial_type = lot.serial_type or 'regular'
                        price_unit = getattr(lot, 'cost_price', product.lst_price)

                        # Create a 1-qty line for this serial
                        SaleOrderLine.create({
                            'order_id': self.sale_order_id.id,
                            'product_id': product.id,
                            'lot_ids': [(6, 0, [lot.id])],
                            'branded_barcode': lot.ref,
                            'name': product.display_name,
                            'product_uom_qty': 1,
                            'product_uom': product.uom_id.id,
                            'price_unit': price_unit,
                            'type_product': getattr(lot, 'type_product', False),
                            'sale_serial_type': current_serial_type,
                        })
                        created_count += 1
                        if created_count >= qty:
                            break

                    if created_count < qty:
                        short = qty - created_count
                        skipped_lines.append((barcode, f"Requested qty exceeds available stock. Short by {short}"))

                # LOT TRACKING -> merge/update qty and price from lot
                else:
                    available_quants = quants.filtered(lambda q: q.quantity > 0)
                    if not available_quants:
                        skipped_lines.append((barcode, "No lot stock available for EAN-13"))
                        continue

                    total_available = sum(max(0.0, _actual_available_qty_for_lot(q.lot_id)) for q in available_quants)
                    if qty > total_available:
                        skipped_lines.append(
                            (barcode, f"Requested qty {qty} exceeds available stock {int(total_available)}"))
                        continue

                    remaining_qty = qty
                    for q in available_quants:
                        lot = q.lot_id
                        actual_available_qty = max(0.0, _actual_available_qty_for_lot(lot))
                        if actual_available_qty <= 0:
                            continue

                        allocate_qty = min(remaining_qty, actual_available_qty)
                        if allocate_qty <= 0:
                            continue

                        existing_line = _find_existing_line(product, lot.id, sale_serial_type)
                        price_unit = getattr(lot, 'cost_price', product.lst_price)

                        if existing_line:
                            # Update qty and ALWAYS update price from lot (option A)
                            new_qty = (existing_line.product_uom_qty or 0.0) + allocate_qty
                            existing_line.write({
                                'product_uom_qty': new_qty,
                                'price_unit': price_unit,
                            })
                        else:
                            SaleOrderLine.create({
                                'order_id': self.sale_order_id.id,
                                'product_id': product.id,
                                'lot_ids': [(6, 0, [lot.id])],
                                'branded_barcode': lot.ref,
                                'name': product.display_name,
                                'product_uom_qty': allocate_qty,
                                'product_uom': product.uom_id.id,
                                'price_unit': price_unit,
                                'type_product': getattr(lot, 'type_product', False),
                                'sale_serial_type': sale_serial_type,
                            })

                        remaining_qty -= allocate_qty
                        if remaining_qty <= 0:
                            break

                    if remaining_qty > 0:
                        skipped_lines.append(
                            (barcode, f"Requested qty exceeds available stock. Short by {int(remaining_qty)}"))
            # ------------------ UNKNOWN BARCODE ------------------
            else:
                skipped_lines.append((barcode, "Unknown barcode format"))
                continue
        # After processing all rows, raise ValidationError if anything was skipped
        if skipped_lines:
            msg = "\n".join([f"{b or '<empty>'}: {reason}" for b, reason in skipped_lines])
            raise ValidationError(_("Some lines were skipped:\n%s") % msg)

    # def action_import_barcodes(self):
    #     """Read Excel and create sale order lines based on GS1 or EAN-13 barcode format"""
    #     self.ensure_one()
    #     if not self.file:
    #         raise ValidationError(_("Please upload an Excel file."))
    #
    #     try:
    #         data = base64.b64decode(self.file)
    #         wb = load_workbook(filename=BytesIO(data), data_only=True)
    #         sheet = wb.active
    #     except Exception as e:
    #         raise ValidationError(_("Invalid Excel file: %s") % str(e))
    #
    #     skipped_lines = []
    #     location_id = self.env.ref('stock.stock_location_stock').id
    #     SaleOrderLine = self.env['sale.order.line']
    #     StockQuant = self.env['stock.quant']
    #
    #     # Cache existing used lot_ids in this sale order for performance
    #     used_lot_ids = set(SaleOrderLine.search([
    #         ('order_id', '=', self.sale_order_id.id),
    #         ('order_id.state', 'not in', ['cancel'])
    #     ]).mapped('lot_ids.id'))
    #
    #     for row in sheet.iter_rows(min_row=2, values_only=True):
    #         barcode = str(row[0]).strip() if row[0] else ""
    #         qty = row[1] or 0
    #
    #         if not barcode or not qty:
    #             skipped_lines.append((barcode, "Missing barcode or quantity"))
    #             continue
    #
    #         product = False
    #         quants = False
    #
    #         # ===================== 🟩 GS1 BARCODE =====================
    #         if 'R' in barcode:
    #             lot_code = barcode[barcode.find('R'):]
    #             quant_domain = [
    #                 ('lot_id.name', '=', lot_code),
    #                 ('company_id.nhcl_company_bool', '=', True),
    #                 ('quantity', '>', 0),
    #                 ('location_id', '=', location_id),
    #                 ('lot_id.type_product', '=', 'un_brand')
    #             ]
    #             quants = StockQuant.search(quant_domain, order='id asc')
    #             product = quants[:1].product_id
    #
    #             if not product:
    #                 skipped_lines.append((barcode, "No matching product for GS1 barcode"))
    #                 continue
    #
    #             # Determine sale_serial_type
    #             existing_types = SaleOrderLine.search([
    #                 ('order_id', '=', self.sale_order_id.id),
    #                 ('order_id.state', 'not in', ['cancel']),
    #                 ('product_id', '=', product.id)
    #             ]).mapped('sale_serial_type')
    #             sale_serial_type = existing_types[0] if existing_types else 'regular'
    #
    #             # ---------- SERIAL TRACKING ----------
    #             if product.tracking == 'serial':
    #                 available_quants = quants.filtered(lambda q: q.quantity > 0)
    #                 if not available_quants:
    #                     skipped_lines.append((barcode, "No stock available for GS1 serials"))
    #                     continue
    #
    #                 if qty > len(available_quants):
    #                     skipped_lines.append(
    #                         (barcode, f"Requested qty {qty} exceeds available serials {len(available_quants)}"))
    #                     continue
    #
    #                 created_count = 0
    #                 for q in available_quants:
    #                     lot = q.lot_id
    #
    #                     # Already used anywhere?
    #                     existing_lines = SaleOrderLine.search([
    #                         ('lot_ids', 'in', [lot.id]),
    #                         ('order_id.state', 'not in', ['cancel'])
    #                     ])
    #                     same_order_line = existing_lines.filtered(lambda l: l.order_id.id == self.sale_order_id.id)
    #                     if same_order_line:
    #                         skipped_lines.append((barcode, f"Serial {lot.name} already used in this sale order"))
    #                         continue
    #
    #                     other_orders = existing_lines.filtered(lambda l: l.order_id.id != self.sale_order_id.id)
    #                     if other_orders:
    #                         skipped_lines.append((barcode, f"Serial {lot.name} already used in another sale order"))
    #                         continue
    #
    #                     SaleOrderLine.create({
    #                         'order_id': self.sale_order_id.id,
    #                         'product_id': product.id,
    #                         'lot_ids': [(6, 0, [lot.id])],
    #                         'branded_barcode': lot.ref,
    #                         'name': product.display_name,
    #                         'product_uom_qty': 1,
    #                         'product_uom': product.uom_id.id,
    #                         'price_unit': getattr(lot, 'cost_price', product.lst_price),
    #                         'type_product': getattr(lot, 'type_product', False),
    #                         'sale_serial_type': sale_serial_type,
    #                     })
    #
    #                     created_count += 1
    #                     if created_count >= qty:
    #                         break
    #
    #             # ---------- LOT TRACKING ----------
    #             else:
    #                 remaining_qty = qty
    #                 available_quants = quants.filtered(lambda q: q.quantity > 0)
    #                 for q in available_quants:
    #                     lot = q.lot_id
    #                     used_qty = sum(SaleOrderLine.search([
    #                         ('order_id.state', 'not in', ['cancel']),
    #                         ('lot_ids', 'in', [lot.id])
    #                     ]).mapped('product_uom_qty'))
    #                     in_qty = sum(self.env['stock.move.line'].search([
    #                         ('picking_id.stock_picking_type', '=', 'receipt'),
    #                         ('company_id.nhcl_company_bool', '=', True),
    #                         ('lot_id', '=', q.lot_id.id)
    #                     ]).mapped('qty_done'))
    #                     actual_available_qty = in_qty - used_qty
    #
    #                     if actual_available_qty <= 0:
    #                         continue
    #
    #                     allocate_qty = min(remaining_qty, actual_available_qty)
    #                     if allocate_qty <= 0:
    #                         continue
    #
    #                     SaleOrderLine.create({
    #                         'order_id': self.sale_order_id.id,
    #                         'product_id': product.id,
    #                         'lot_ids': [(6, 0, [lot.id])],
    #                         'branded_barcode': lot.ref,
    #                         'name': product.display_name,
    #                         'product_uom_qty': allocate_qty,
    #                         'product_uom': product.uom_id.id,
    #                         'price_unit': getattr(lot, 'cost_price', product.lst_price),
    #                         'type_product': getattr(lot, 'type_product', False),
    #                         'sale_serial_type': sale_serial_type,
    #                     })
    #
    #                     remaining_qty -= allocate_qty
    #                     if remaining_qty <= 0:
    #                         break
    #
    #                 if remaining_qty > 0:
    #                     skipped_lines.append(
    #                         (barcode, f"Requested qty exceeds available stock. Short by {remaining_qty}"))
    #
    #         # ===================== 🟦 EAN-13 BARCODE =====================
    #         elif len(barcode) == 13:
    #             lot_code = barcode
    #             matched_quant = StockQuant.search([
    #                 ('lot_id.ref', '=', lot_code),
    #                 ('company_id.nhcl_company_bool', '=', True),
    #                 ('quantity', '>', 0),
    #                 ('location_id', '=', location_id),
    #                 ('lot_id.type_product', '=', 'brand')
    #             ], limit=1)
    #
    #             if not matched_quant:
    #                 skipped_lines.append((barcode, "No stock found for EAN-13 barcode"))
    #                 continue
    #
    #             product = matched_quant.product_id
    #             quant_domain = [
    #                 ('product_id', '=', product.id),
    #                 ('lot_id.ref', '=', lot_code),
    #                 ('company_id.nhcl_company_bool', '=', True),
    #                 ('quantity', '>', 0),
    #                 ('location_id', '=', location_id),
    #                 ('lot_id.type_product', '=', 'brand')
    #             ]
    #             quants = StockQuant.search(quant_domain, order='id asc')
    #             available_quants = quants.filtered(lambda q: q.quantity > 0)
    #
    #             existing_types = SaleOrderLine.search([
    #                 ('order_id', '=', self.sale_order_id.id),
    #                 ('order_id.state', 'not in', ['cancel']),
    #                 ('product_id', '=', product.id)
    #             ]).mapped('sale_serial_type')
    #             sale_serial_type = existing_types[0] if existing_types else 'regular'
    #
    #             # ---------- SERIAL TRACKING (Enhanced for Regular + Return Lots) ----------
    #             if product.tracking == 'serial':
    #                 created_count = 0
    #
    #                 # Regular serials (normal outgoing ones)
    #                 regular_quants = StockQuant.search([
    #                     ('product_id', '=', product.id),
    #                     ('lot_id.ref', '=', lot_code),
    #                     ('company_id.nhcl_company_bool', '=', True),
    #                     ('quantity', '>', 0),
    #                     ('location_id', '=', location_id),
    #                     ('lot_id.type_product', '=', 'brand'),
    #                     '|', ('lot_id.serial_type', '=', False), ('lot_id.serial_type', '=', 'regular'),
    #                 ], order='id asc')
    #
    #                 # Return serials (came back from customers)
    #                 return_quants = StockQuant.search([
    #                     ('product_id', '=', product.id),
    #                     ('lot_id.ref', '=', lot_code),
    #                     ('company_id.nhcl_company_bool', '=', True),
    #                     ('quantity', '>', 0),
    #                     ('location_id', '=', location_id),
    #                     ('lot_id.type_product', '=', 'brand'),
    #                     ('lot_id.serial_type', '=', 'return'),
    #                 ], order='id asc')
    #
    #                 combined_quants = regular_quants + return_quants
    #
    #                 for q in combined_quants:
    #                     lot = q.lot_id
    #
    #                     # Skip if serial already used in any active sale order
    #                     existing_lines = SaleOrderLine.search([
    #                         ('lot_ids', 'in', [lot.id]),
    #                         ('order_id.state', 'not in', ['cancel'])
    #                     ])
    #                     if existing_lines:
    #                         same_type_line = existing_lines.filtered(
    #                             lambda l: l.sale_serial_type == (lot.serial_type or 'regular'))
    #                         if same_type_line:
    #                             # If already used with same serial type → skip
    #                             continue
    #
    #                     # Determine which type this serial belongs to
    #                     current_serial_type = lot.serial_type or 'regular'
    #
    #                     SaleOrderLine.create({
    #                         'order_id': self.sale_order_id.id,
    #                         'product_id': product.id,
    #                         'lot_ids': [(6, 0, [lot.id])],
    #                         'branded_barcode': lot.ref,
    #                         'name': product.display_name,
    #                         'product_uom_qty': 1,
    #                         'product_uom': product.uom_id.id,
    #                         'price_unit': getattr(lot, 'cost_price', product.lst_price),
    #                         'type_product': getattr(lot, 'type_product', False),
    #                         'sale_serial_type': current_serial_type,
    #                     })
    #
    #                     created_count += 1
    #                     if created_count >= qty:
    #                         break
    #
    #                 if created_count < qty:
    #                     short = qty - created_count
    #                     skipped_lines.append((barcode, f"Requested qty exceeds available stock. Short by {short}"))
    #
    #
    #
    #             # ---------- LOT TRACKING ----------
    #             else:
    #                 remaining_qty = qty
    #                 for q in available_quants:
    #                     lot = q.lot_id
    #                     used_qty = sum(SaleOrderLine.search([
    #                         ('order_id.state', 'not in', ['cancel']),
    #                         ('lot_ids', 'in', [lot.id])
    #                     ]).mapped('product_uom_qty'))
    #                     in_qty = sum(self.env['stock.move.line'].search([
    #                         ('picking_id.stock_picking_type', '=', 'receipt'),
    #                         ('company_id.nhcl_company_bool', '=', True),
    #                         ('lot_id', '=', q.lot_id.id)
    #                     ]).mapped('qty_done'))
    #                     actual_available_qty = in_qty - used_qty
    #
    #                     if actual_available_qty <= 0:
    #                         continue
    #
    #                     allocate_qty = min(remaining_qty, actual_available_qty)
    #                     if allocate_qty <= 0:
    #                         continue
    #
    #                     SaleOrderLine.create({
    #                         'order_id': self.sale_order_id.id,
    #                         'product_id': product.id,
    #                         'lot_ids': [(6, 0, [lot.id])],
    #                         'branded_barcode': lot.ref,
    #                         'name': product.display_name,
    #                         'product_uom_qty': allocate_qty,
    #                         'product_uom': product.uom_id.id,
    #                         'price_unit': getattr(lot, 'cost_price', product.lst_price),
    #                         'type_product': getattr(lot, 'type_product', False),
    #                         'sale_serial_type': sale_serial_type,
    #                     })
    #
    #                     remaining_qty -= allocate_qty
    #                     if remaining_qty <= 0:
    #                         break
    #                 if remaining_qty > 0:
    #                     skipped_lines.append(
    #                         (barcode, f"Requested qty exceeds available stock. Short by {remaining_qty}"))
    #         # ===================== ❌ UNKNOWN BARCODE =====================
    #         else:
    #             skipped_lines.append((barcode, "Unknown barcode format"))
    #             continue
    #
    #     if skipped_lines:
    #         msg = "\n".join([f"{b}: {reason}" for b, reason in skipped_lines])
    #         raise ValidationError(_("Some lines were skipped:\n%s") % msg)

    # sale_order_file = fields.Binary(string="Select File")
    # import_option = fields.Selection([('csv', 'CSV File'), ('xls', 'XLS File')], string='Select', default='xls')
    # import_prod_option = fields.Selection([('barcode', 'Barcode'), ('code', 'Code'), ('name', 'Name')],
    #                                       string='Import Product By ', default='barcode')
    # product_details_option = fields.Selection(
    #     [('from_product', 'Take Details From The Product'), ('from_xls', 'Take Details From The XLS File'),
    #      ('from_pricelist', 'Take Details With Adapted Pricelist')], default='from_product')
    #
    # sample_option = fields.Selection([('csv', 'CSV'), ('xls', 'XLS')], string='Sample Type', default='xls')
    # down_samp_file = fields.Boolean(string='Download Sample Files')
    #
    # def import_sol(self):
    #     res = False
    #     counter = 0
    #     if self.import_option == 'csv':
    #         keys = ['product', 'quantity', 'price']
    #         try:
    #             wb = openpyxl.load_workbook(
    #                 filename=BytesIO(base64.b64decode(self.sale_order_file)), read_only=True
    #             )
    #             ws = wb.active
    #         except Exception:
    #             raise ValidationError(_("Please select any file or You have selected invalid file"))
    #
    #         for row_no in ws.iter_rows(min_row=2, max_row=None, min_col=None,
    #                                    max_col=None, values_only=True):
    #
    #             field = list(map(str, row_no))
    #             values = dict(zip(keys, field))
    #             if values:
    #                 if len(row_no) == 0:
    #                     continue
    #                 else:
    #                     product_barcode = ''
    #                     lot_name = ''
    #                     if row_no[0] == None:
    #                         continue
    #                     else:
    #                         counter += 1
    #                         if len(row_no[0]) > 13:
    #                             if row_no[0][0] == '0' and row_no[0][1] == '1' and row_no[0][16] == '2' and row_no[0][
    #                                 17] == '1':
    #                                 for i in range(0, len(row_no[0])):
    #                                     if i > 1 and i < 16:
    #                                         product_barcode += row_no[0][i]
    #                                     elif i > 17 and i < len(row_no[0]):
    #                                         lot_name += row_no[0][i]
    #                                         continue
    #                         else:
    #                             product_barcode = row_no[0]
    #                     if row_no[1] == None or row_no[1] <= 0:
    #                         raise ValidationError(_('%s Quantity must be greater than zero.') % (row_no[0]))
    #                     if self.product_details_option == 'from_product':
    #                         values.update({
    #                             'product': product_barcode,
    #                             'serial_no': lot_name,
    #                             'quantity': row_no[1]
    #                         })
    #                     elif self.product_details_option == 'from_xls':
    #                         values.update({'product': product_barcode,
    #                                        'serial_no': lot_name,
    #                                        'quantity': row_no[1],
    #                                        'price': row_no[2],
    #                                        })
    #                     else:
    #                         values.update({
    #                             'product': product_barcode,
    #                             'serial_no': lot_name,
    #                             'quantity': row_no[1],
    #                         })
    #                     res = self.create_order_line(values)
    #     else:
    #         try:
    #             wb = openpyxl.load_workbook(
    #                 filename=BytesIO(base64.b64decode(self.sale_order_file)), read_only=True
    #             )
    #             ws = wb.active
    #             values = {}
    #         except Exception:
    #             raise ValidationError(_("Please select any file or You have selected invalid file"))
    #         for row_no in ws.iter_rows(min_row=2, max_row=None, min_col=None,
    #                                    max_col=None, values_only=True):
    #
    #             product_barcode = ''
    #             lot_name = ''
    #             if row_no[0] == None:
    #                 continue
    #             else:
    #                 counter += 1
    #                 if len(row_no[0]) > 13:
    #                     if row_no[0][0] == '0' and row_no[0][1] == '1' and row_no[0][16] == '2' and row_no[0][
    #                         17] == '1':
    #                         for i in range(0, len(row_no[0])):
    #                             if i > 1 and i < 16:
    #                                 product_barcode += row_no[0][i]
    #                             elif i > 17 and i < len(row_no[0]):
    #                                 lot_name += row_no[0][i]
    #                                 continue
    #                 else:
    #                     product_barcode = row_no[0]
    #             if row_no[1] == None or row_no[1] <= 0:
    #                 raise ValidationError(_('%s Quantity must be greater than zero.') % (row_no[0]))
    #             if self.product_details_option == 'from_product':
    #                 values.update({
    #                     'product': product_barcode,
    #                     'quantity': row_no[1],
    #                     'serial_no': lot_name
    #                 })
    #             elif self.product_details_option == 'from_xls':
    #                 values.update({'product': product_barcode,
    #                                'quantity': row_no[1],
    #                                'serial_no': lot_name,
    #                                'price': row_no[2],
    #                                })
    #             else:
    #                 values.update({
    #                     'product': product_barcode,
    #                     'quantity': row_no[1],
    #                     'serial_no': lot_name,
    #                 })
    #             res = self.create_order_line(values)
    #     view_id = self.env.ref('nhcl_customizations.message_wizard_popup')
    #     context = dict(self._context or {})
    #     dict_msg = str(counter) + " Records Imported Successfully."
    #     context['message'] = dict_msg
    #     return {
    #         'name': _('Success'),
    #         'type': 'ir.actions.act_window',
    #         'view_mode': 'form',
    #         'res_model': 'message.wizard',
    #         'views': [(view_id.id, 'form')],
    #         'view_id': view_id.id,
    #         'target': 'new',
    #         'context': context,
    #     }
    #
    # def create_order_line(self, values):
    #     sale_order_brw = self.env['sale.order'].browse(self._context.get('active_id'))
    #     serial_no = self.env['stock.lot']
    #     product_obj_search = self.env['product.product']
    #     main_company = self.env['res.company'].search([('nhcl_company_bool', '=', False)])
    #     if self.product_details_option == 'from_product':
    #         if self.import_prod_option == 'barcode':
    #             if len(values['product']) > 13 and values['serial_no']:
    #                 s_no = values['serial_no']
    #                 print(main_company,"ytdytfku")
    #                 product_obj_search = self.env['product.product'].search([('barcode', '=', values['product'])])
    #                 location = self.env.ref('stock.stock_location_stock').id
    #                 serial_nos = self.env['stock.quant'].search([('location_id.id','=',location),
    #                     ('lot_id.name', '=', s_no), ('quantity', '>', 0), ('company_id.id', '=', main_company.id)])
    #                 serial_no = serial_nos.lot_id
    #                 if not serial_no:
    #                     raise ValidationError(
    #                         _('The serial number for this is not found in the database.'))
    #                 serial_id_list = []
    #
    #                 for serial in serial_no:
    #                     if serial.is_under_plan == True:
    #                         raise ValidationError(_('Lot Number is Under Planned'))
    #                     sale_lines = self.env['sale.order.line'].search([
    #                         ('lot_ids.name', '=', serial.name), ('company_id', '=', main_company.id)
    #                         ,('order_id.state', 'not in', ['cancel','sale'])
    #                     ])
    #                     sale_qty = sum(sale_lines.mapped('product_uom_qty'))
    #                     if serial.id in serial_id_list:
    #                         raise ValidationError(
    #                             _('Duplicate serial number detected: ID %s (Name: %s)') % (serial.id, serial.name))
    #                     if sale_qty < serial.product_qty:
    #                         serial_id_list.append(serial.id)
    #                     else:
    #                         raise ValidationError(
    #                             _('Serial Number already used in Another Sale Order Line: (Lot Name: %s)') % (
    #                                 serial.name))
    #             else:
    #                 product_barcodes = self.env['product.barcode'].sudo().search([('barcode', '=', values['product'])])
    #
    #                 if len(product_barcodes) > 0:
    #                     product_obj_search = product_barcodes[0].product_id
    #                 if values.get('product'):
    #                     location = self.env.ref('stock.stock_location_stock').id
    #                     product = self.env['product.product'].search([('default_code', '=', values['product'])],
    #                                                                  limit=1)
    #                     domain = [('lot_id.ref', '=', values['product']),('quantity', '>', 0),('location_id', '=', location),
    #                         ('company_id', '=', main_company.id),('lot_id.type_product', '=', 'brand'),]
    #                     # Add 'is_uploaded' filter only if tracking is 'serial'
    #                     if product.tracking == 'serial':
    #                         domain.append(('lot_id.is_uploaded', '=', False))
    #                     serial_nos = self.env['stock.quant'].search(domain, order='id asc')
    #                     serial_no = serial_nos.lot_id
    #                     if not serial_no:
    #                         raise ValidationError(
    #                             _('The serial number for this is not found in the databaseee.'))
    #         elif self.import_prod_option == 'code':
    #             raise ValidationError(_('Please set the import Option to Barcode.'))
    #         else:
    #             raise ValidationError(_('Please set the import Option to Barcode.'))
    #         product_id = self.env['product.product']
    #         if product_obj_search:
    #             product_id = product_obj_search
    #         elif values['product']:
    #             raise ValidationError(_('%s Product was not found in the Database.') % values.get('product'))
    #
    #         if sale_order_brw.state == 'draft':
    #             # Checking if the product barcode is longer than 13 characters
    #             if len(values['product']) > 13:
    #                 # Filter order line by product ID
    #                 existing_product_line = sale_order_brw.order_line.filtered(
    #                     lambda x: x.product_id == product_id)
    #
    #                 # Check if the serial number exists in the lot_ids of the filtered order line
    #                 existing_line = existing_product_line.lot_ids.filtered(lambda x: x.id in serial_no.ids)
    #             else:
    #                 # Split the 'branded_barcode' field by commas and check if the incoming product barcode exists in any of them
    #                 existing_line = sale_order_brw.order_line.filtered(
    #                     lambda x: x.product_id == product_id and
    #                               values.get('product') in (x.branded_barcode or '').split(','))
    #
    #             # Raise error if duplicate found
    #
    #             # Check if there is a branded_barcode field in the values
    #             if len(values['product']) == 13 and product_obj_search.tracking == 'serial':
    #                 # Iterate for the number of lines you need
    #                 for i in range(values['quantity']):
    #                     lot = serial_no[
    #                         i % len(serial_no)]
    #                     if lot.is_under_plan == True:
    #                         raise ValidationError(_('Lot Number is Under Planned'))
    #                     if values['product']:
    #                         sale_lines = self.env['sale.order.line'].search([
    #                             ('lot_ids.name', '=', lot.name), ('company_id', '=', main_company.id)
    #                             ,('order_id.state', 'not in', ['cancel','sale'])
    #                         ])
    #                         brand_used_qty =  self.env['sale.order.line'].search([
    #                             ('lot_ids.name', '=', lot.name), ('company_id', '=', main_company.id)
    #                             ,('order_id.state', 'not in', ['cancel','sale'])
    #                         ]).mapped('product_uom_qty')
    #                         brand_total_qty = self.env['stock.quant'].search([('lot_id.name', '=', lot.name), ('company_id', '=', main_company.id)]).mapped('quantity')
    #                         if brand_total_qty < brand_used_qty and lot.serial_type == 'regular':
    #                             raise ValidationError(
    #                                 _('Given qty is exceeded than available qty')
    #                                 # % (
    #                                 #     lot.lot_id.name, values['serial_no'],
    #                                 #     sale_lines[0].order_id)
    #                             )
    #                         else:
    #                             self.env['sale.order.line'].create({
    #                                 'order_id': sale_order_brw.id,
    #                                 'product_id': product_id.id,
    #                                 'lot_ids': [(6, 0, [lot.id])],
    #                                 'branded_barcode': values['product'],
    #                                 'name': product_id.display_name,
    #                                 'product_uom_qty': 1,
    #                                 'product_uom': product_id.uom_id.id,
    #                                 'price_unit': lot.cost_price,  # Assuming you want the cost price from serial_no
    #                                 'type_product': lot.type_product,
    #                                 'sale_serial_type': lot.serial_type,
    #                             })
    #                         sale_lines = self.env['sale.order.line'].search([
    #                             ('lot_ids.name', '=', lot.name), ('company_id', '=', main_company.id)
    #                             ,('order_id.state', 'not in', ['cancel'])
    #                         ])
    #                         sale_line_qty = sum(sale_lines.mapped('product_uom_qty'))
    #                         lot_product_qty = lot.product_qty - sale_line_qty
    #                         if lot_product_qty <= 0 and lot.product_id.tracking == 'serial':
    #                             lot.write({'is_uploaded': True})
    #             else:
    #                 qty = 0
    #                 pending_qty = 0
    #
    #                 if 'serial_no' in values and values['serial_no'] != '':
    #                     location = self.env.ref('stock.stock_location_stock').id
    #                     total_serail_no = self.env['stock.quant'].sudo().search([
    #                         ('company_id.nhcl_company_bool', '=', False),('lot_id.name', '=', values['serial_no']),
    #                         ('lot_id.ref', '=', values['product']),('lot_id.type_product', '!=', 'brand'),
    #                         ('location_id.id', '=', location)])
    #                     # total_serail_no = self.env['stock.quant'].search(
    #                     #     [('lot_id.ref', '=', values['product']), ('quantity', '>', 0),
    #                     #      ('lot_id.name', '=', values['serial_no']),('location_id','=',location),
    #                     #      ('company_id', '=', main_company.id), ('lot_id.type_product', '!=', 'brand')], order='id asc'
    #                     # )
    #                     existing_sale_lines = self.env['sale.order.line'].search([
    #                         ('branded_barcode', '=', values['product']), ('lot_ids.name', '=', values['serial_no']),
    #                         ('order_id.state', 'not in', ['cancel']), ('company_id', '=', main_company.id)
    #                     ])
    #                 else:
    #                     location = self.env.ref('stock.stock_location_stock').id
    #                     # total_serail_no = self.env['stock.quant'].search(
    #                     #     [('lot_id.ref', '=', values['product']), ('quantity', '>', 0),('location_id','=',location),
    #                     #      ('company_id', '=', main_company.id), ('lot_id.type_product', '=', 'brand')], order='id asc'
    #                     # )
    #                     total_serail_no = self.env['stock.quant'].sudo().search([
    #                         ('company_id.nhcl_company_bool', '=', False),
    #                         ('lot_id.ref', '=', values['product']), ('lot_id.type_product', '=', 'brand'),
    #                         ('location_id.id', '=', location)], order='id asc')
    #                     existing_sale_lines = self.env['sale.order.line'].search([
    #                         ('branded_barcode', '=', values['product']), ('company_id', '=', main_company.id)
    #                         ,('order_id.state', 'not in', ['cancel','sale'])
    #                     ])
    #                 existing_sale_lines_qty = sum(existing_sale_lines.mapped('product_uom_qty'))
    #                 total_lot_qty = sum(total_serail_no.mapped('quantity'))
    #                 available_lot_product_qty = total_lot_qty - existing_sale_lines_qty
    #                 if total_serail_no and total_serail_no[0].lot_id.serial_type == 'regular' and len(total_serail_no) > 0 and values[
    #                     'quantity'] > available_lot_product_qty and 'serial_no' in values and values['serial_no'] != '':
    #                     raise ValidationError(
    #                         _('%s Qty is exceeded Serial No:%s, available qty:%s. Sale Order No:%s') % (
    #                             total_serail_no[0].lot_id.ref, values['serial_no'], available_lot_product_qty,
    #                             existing_sale_lines[:1].order_id.id))
    #                 if total_serail_no and total_serail_no[0].lot_id.serial_type == 'regular' and len(total_serail_no) > 0 and values[
    #                     'quantity'] > available_lot_product_qty and 'serial_no' in values and values['serial_no'] == '':
    #                     raise ValidationError(
    #                         _('%s Qty is exceeded Serial No:%s, available qty:%s.Sale Order No:%s') % (
    #                             total_serail_no[0].lot_id.ref, values['serial_no'], available_lot_product_qty,
    #                             existing_sale_lines[:1].order_id.id))
    #                 total_qty = 0
    #                 for lot in serial_no:
    #                     if lot.is_under_plan == True:
    #                         raise ValidationError(_('Lot Number is Under Planned'))
    #                     existing_line = self.env['sale.order.line'].search([
    #                         ('order_id', '=', sale_order_brw.id), ('company_id', '=', main_company.id),
    #                         ('lot_ids', 'in', lot.id)
    #                     ], limit=1)
    #
    #                     if existing_line and existing_line.product_id.tracking != 'lot':
    #                         raise ValidationError(
    #                             _('The serial number "%s" (ID: %s) is already used in this Sale Order.') % (
    #                                 lot.name, lot.id)
    #                         )
    #
    #                     if lot.ref:
    #                         sale_lines = self.env['sale.order.line'].search([
    #                             ('lot_ids.name', '=', lot.name), ('company_id', '=', main_company.id),
    #                             ('order_id.state', 'not in', ['cancel','sale'])
    #                         ])
    #                         sale_line_qty = sum(sale_lines.mapped('product_uom_qty'))
    #                         lot_product_qty = lot.product_qty - sale_line_qty
    #                         if (values['quantity'] > lot_product_qty) or (pending_qty > 0):
    #                             if pending_qty > 0:
    #                                 if pending_qty < lot_product_qty:
    #                                     qty = pending_qty
    #                                 else:
    #                                     qty = lot_product_qty
    #                             else:
    #                                 qty = lot_product_qty
    #                             if qty > 0:
    #                                 if existing_line:
    #                                     existing_line.product_uom_qty += qty
    #                                     existing_line.price_unit = lot.cost_price
    #                                 else:
    #                                     self.env['sale.order.line'].create({
    #                                         'order_id': sale_order_brw.id,
    #                                         'product_id': product_id.id,
    #                                         'lot_ids': [(6, 0, [lot.lot_id.id])],
    #                                         'branded_barcode': lot.ref,
    #                                         'name': product_id.display_name,
    #                                         'product_uom_qty': qty,
    #                                         'product_uom': product_id.uom_id.id,
    #                                         'price_unit': lot.lot_id.cost_price,
    #                                         'type_product': lot.lot_id.type_product,
    #                                         'sale_serial_type': lot.lot_id.serial_type,
    #                                     })
    #                                 total_qty += qty
    #                                 pending_qty = values['quantity'] - total_qty
    #                                 if values['quantity'] == total_qty:
    #                                     break
    #                         else:
    #                             if existing_line:
    #                                 existing_line.product_uom_qty += values['quantity']
    #                                 existing_line.price_unit = lot.cost_price
    #                             else:
    #                                 self.env['sale.order.line'].create({
    #                                     'order_id': sale_order_brw.id,
    #                                     'product_id': product_id.id,
    #                                     'lot_ids': [(6, 0, [lot.id])],
    #                                     'branded_barcode': lot.ref,
    #                                     'name': product_id.display_name,
    #                                     'product_uom_qty': values['quantity'],
    #                                     'product_uom': product_id.uom_id.id,
    #                                     'price_unit': lot.cost_price,
    #                                     'type_product': lot.type_product,
    #                                     'sale_serial_type': lot.serial_type,
    #                                 })
    #                             break
    #                     sale_lines = self.env['sale.order.line'].search([
    #                         ('lot_ids.name', '=', lot.name), ('company_id', '=', main_company.id),
    #                         ('order_id.state', 'not in', ['cancel'])
    #                     ])
    #                     sale_line_qty = sum(sale_lines.mapped('product_uom_qty'))
    #                     lot_product_qty = lot.quantity - sale_line_qty
    #                     if lot_product_qty == 0:
    #                         lot.lot_id.write({'is_uploaded': True})
    #
    #         elif sale_order_brw.state == 'sent':
    #             # Checking if the product barcode is longer than 13 characters
    #             if len(values['product']) > 13:
    #                 # Filter order line by product ID
    #                 existing_product_line = sale_order_brw.order_line.filtered(
    #                     lambda x: x.product_id == product_id)
    #
    #                 # Check if the serial number exists in the lot_ids of the filtered order line
    #                 existing_line = existing_product_line.lot_ids.filtered(lambda x: x.id in serial_no.ids)
    #
    #             else:
    #                 # Split the 'branded_barcode' field by commas and check if the incoming product barcode exists in any of them
    #                 existing_line = sale_order_brw.order_line.filtered(
    #                     lambda x: x.product_id == product_id and
    #                               values.get('product') in (x.branded_barcode or '').split(','))
    #
    #             # Raise an error if the product already exists
    #             if existing_line:
    #                 raise ValidationError(_('%s Product is already existing.') % values.get('product'))
    #
    #             # Filter order line by product ID to update existing lines
    #             existing_order_line = sale_order_brw.order_line.filtered(
    #                 lambda x: x.product_id == product_id)
    #
    #             if existing_order_line:
    #                 # Update the quantity of the existing order line
    #                 existing_order_line.product_uom_qty += values.get('quantity')
    #
    #                 # If serial numbers are present, link them to the existing order line
    #                 if serial_no:
    #                     if serial_no.is_under_plan == True:
    #                         raise ValidationError(_('Lot Number is Under Planned'))
    #                     # Add the serial numbers to lot_ids without replacing the existing ones
    #                     existing_order_line.lot_ids = [(4, lot.id) for lot in serial_no]
    #
    #                     # Collect all the barcode references from the lot_ids (serial numbers)
    #                     barcodes = existing_order_line.lot_ids.mapped('ref')
    #
    #                     # Filter out any None or empty values and ensure uniqueness
    #                     unique_barcodes = list(set([barcode for barcode in barcodes if barcode]))
    #
    #                     # Join the unique barcodes into a comma-separated string
    #                     branded_barcode_value = ', '.join(unique_barcodes)
    #
    #                     # Update the branded_barcode field
    #                     existing_order_line.branded_barcode = branded_barcode_value
    #
    #                     # Mark all the serial numbers as uploaded
    #                     sale_lines = self.env['sale.order.line'].search([
    #                         ('lot_ids.name', '=', serial_no.name), ('company_id', '=', main_company.id),
    #                         ('order_id.state', 'not in', ['cancel'])
    #                     ])
    #                     sale_line_qty = sum(sale_lines.mapped('product_uom_qty'))
    #                     lot_product_qty = serial_no.product_qty - sale_line_qty
    #                     if lot_product_qty == 0:
    #                         serial_no.write({'is_uploaded': True})
    #
    #             elif values.get('product'):
    #                 # Create a new sale order line with the product and quantity
    #                 order_lines = self.env['sale.order.line'].create({
    #                     'order_id': sale_order_brw.id,
    #                     'product_id': product_id.id,
    #                     'lot_ids': [(6, 0, serial_no.ids)],
    #                     'branded_barcode': values.get('product'),
    #                     'name': product_id.display_name,
    #                     'product_uom_qty': values.get('quantity'),
    #                     'product_uom': product_id.uom_id.id,
    #                     'price_unit': product_id.lst_price,
    #                     'type_product': serial_no.type_product,
    #                     'sale_serial_type': serial_no.serial_type,
    #
    #                 })
    #                 # If the order line was created successfully, mark all the serial numbers as uploaded
    #                 if order_lines:
    #                     sale_lines = self.env['sale.order.line'].search([
    #                         ('lot_ids.name', '=', serial_no.name), ('company_id', '=', main_company.id),
    #                         ('order_id.state', 'not in', ['cancel'])
    #                     ])
    #                     sale_line_qty = sum(sale_lines.mapped('product_uom_qty'))
    #                     lot_product_qty = serial_no.product_qty - sale_line_qty
    #                     if lot_product_qty == 0:
    #                         serial_no.write({'is_uploaded': True})
    #
    #         elif sale_order_brw.state != 'sent' or sale_order_brw.state != 'draft':
    #             raise ValidationError(_('We cannot import data in validated or confirmed order!.'))
    #     elif self.product_details_option == 'from_xls':
    #         if self.import_prod_option == 'barcode':
    #             barcode = values.get('product')
    #             if len(barcode) > 13:
    #                 product_obj_search = self.env['product.product'].search([('barcode', '=', barcode)])
    #                 location = self.env.ref('stock.stock_location_stock').id
    #                 serial_nos = self.env['stock.quant'].search([('lot_id.name', '=', values['serial_no']),('location_id','=',location), ('company_id', '=', main_company.id)])
    #                 serial_no = serial_nos.lot_id
    #                 if not serial_no:
    #                     raise ValidationError(
    #                         _('The serial number for this is not found in the database.'))
    #             else:
    #                 product_barcodes = self.env['product.barcode'].search([('barcode', '=', values['product'])])
    #                 if len(product_barcodes) > 0:
    #                     product_obj_search = product_barcodes[0].product_id
    #                 location = self.env.ref('stock.stock_location_stock').id
    #                 serial_nos = self.env['stock.quant'].search(
    #                     [('lot_id.ref', '=', values['product']), ('location_id','=',location),('company_id', '=', main_company.id),('quantity', '>', 0), ('lot_id.is_uploaded', '=', False)],
    #                     limit=1)
    #                 serial_no = serial_nos.lot_id
    #                 if not serial_no:
    #                     raise ValidationError(
    #                         _('The serial number for this is not found in the database.'))
    #         elif self.import_prod_option == 'code':
    #             raise ValidationError(_('Please set the import Option to Barcode.'))
    #         else:
    #             raise ValidationError(_('Please set the import Option to Barcode.'))
    #         if product_obj_search:
    #             product_id = product_obj_search
    #         else:
    #             if self.import_prod_option == 'name':
    #                 raise ValidationError(_('Please set the import Option to Barcode.'))
    #             else:
    #                 raise ValidationError(
    #                     _('%s Product was not found in the Database') % values.get(
    #                         'product'))
    #
    #         if sale_order_brw.state == 'draft':
    #             # Checking if the product barcode is longer than 13 characters
    #             if len(values['product']) > 13:
    #                 # Filter order line by product ID
    #                 existing_product_line = sale_order_brw.order_line.filtered(
    #                     lambda x: x.product_id == product_id)
    #
    #                 # Check if the serial number exists in the lot_ids of the filtered order line
    #                 existing_line = existing_product_line.lot_ids.filtered(lambda x: x.id in serial_no.ids)
    #
    #             else:
    #                 # Split the 'branded_barcode' field by commas and check if the incoming product barcode exists in any of them
    #                 existing_line = sale_order_brw.order_line.filtered(
    #                     lambda x: x.product_id == product_id and
    #                               values.get('product') in (x.branded_barcode or '').split(','))
    #
    #             # Raise an error if the product already exists
    #             if existing_line:
    #                 raise ValidationError(_('%s Product is already existing.') % values.get('product'))
    #
    #             # Filter order line by product ID to update existing lines
    #             existing_order_line = sale_order_brw.order_line.filtered(
    #                 lambda x: x.product_id == product_id)
    #
    #             if existing_order_line:
    #                 # Update the quantity of the existing order line
    #                 existing_order_line.product_uom_qty += values.get('quantity')
    #
    #                 # If serial numbers are present, link them to the existing order line
    #                 if serial_no:
    #                     if serial_no.is_under_plan == True:
    #                         raise ValidationError(_('Lot Number is Under Planned'))
    #                     # Add the serial numbers to lot_ids without replacing the existing ones
    #                     existing_order_line.lot_ids = [(4, lot.id) for lot in serial_no]
    #
    #                     # Collect all the barcode references from the lot_ids (serial numbers)
    #                     barcodes = existing_order_line.lot_ids.mapped('ref')
    #
    #                     # Filter out any None or empty values and ensure uniqueness
    #                     unique_barcodes = list(set([barcode for barcode in barcodes if barcode]))
    #
    #                     # Join the unique barcodes into a comma-separated string
    #                     branded_barcode_value = ', '.join(unique_barcodes)
    #
    #                     # Update the branded_barcode field
    #                     existing_order_line.branded_barcode = branded_barcode_value
    #
    #                     # Mark all the serial numbers as uploaded
    #                     sale_lines = self.env['sale.order.line'].search([
    #                         ('lot_ids.name', '=', serial_no.name),('company_id', '=', main_company.id),
    #                         ('order_id.state', 'not in', ['cancel'])
    #                     ])
    #                     sale_line_qty = sum(sale_lines.mapped('product_uom_qty'))
    #                     lot_product_qty = serial_no.product_qty - sale_line_qty
    #                     if lot_product_qty == 0:
    #                         serial_no.write({'is_uploaded': True})
    #
    #             elif values.get('product'):
    #                 # Create a new sale order line with the product and quantity
    #                 order_lines = self.env['sale.order.line'].create({
    #                     'order_id': sale_order_brw.id,
    #                     'product_id': product_id.id,
    #                     'lot_ids': [(6, 0, serial_no.ids)],  # Set the lot_ids with all serial numbers
    #                     'branded_barcode': values.get('product'),
    #                     'name': product_id.display_name,
    #                     'product_uom_qty': values.get('quantity'),
    #                     'product_uom': product_id.uom_id.id,
    #                     'price_unit': product_id.lst_price,
    #                     'type_product': serial_no.type_product,
    #                     'sale_serial_type': serial_no.serial_type,
    #
    #                 })
    #                 # If the order line was created successfully, mark all the serial numbers as uploaded
    #                 if order_lines:
    #                     sale_lines = self.env['sale.order.line'].search([
    #                         ('lot_ids.name', '=', serial_no.name),('company_id', '=', main_company.id),
    #                         ('order_id.state', 'not in', ['cancel'])
    #                     ])
    #                     sale_line_qty = sum(sale_lines.mapped('product_uom_qty'))
    #                     lot_product_qty = serial_no.product_qty - sale_line_qty
    #                     if lot_product_qty == 0:
    #                         serial_no.write({'is_uploaded': True})
    #
    #         elif sale_order_brw.state == 'sent':
    #             # Checking if the product barcode is longer than 13 characters
    #             if len(values['product']) > 13:
    #                 # Filter order line by product ID
    #                 existing_product_line = sale_order_brw.order_line.filtered(
    #                     lambda x: x.product_id == product_id)
    #
    #                 # Check if the serial number exists in the lot_ids of the filtered order line
    #                 existing_line = existing_product_line.lot_ids.filtered(lambda x: x.id in serial_no.ids)
    #
    #             else:
    #                 # Split the 'branded_barcode' field by commas and check if the incoming product barcode exists in any of them
    #                 existing_line = sale_order_brw.order_line.filtered(
    #                     lambda x: x.product_id == product_id and
    #                               values.get('product') in (x.branded_barcode or '').split(','))
    #
    #             # Raise an error if the product already exists
    #             if existing_line:
    #                 raise ValidationError(_('%s Product is already existing.') % values.get('product'))
    #
    #             # Filter order line by product ID to update existing lines
    #             existing_order_line = sale_order_brw.order_line.filtered(
    #                 lambda x: x.product_id == product_id)
    #
    #             if existing_order_line:
    #                 # Update the quantity of the existing order line
    #                 existing_order_line.product_uom_qty += values.get('quantity')
    #
    #                 # If serial numbers are present, link them to the existing order line
    #                 if serial_no:
    #                     if serial_no.is_under_plan == True:
    #                         raise ValidationError(_('Lot Number is Under Planned'))
    #                     # Add the serial numbers to lot_ids without replacing the existing ones
    #                     existing_order_line.lot_ids = [(4, lot.id) for lot in serial_no]
    #
    #                     # Collect all the barcode references from the lot_ids (serial numbers)
    #                     barcodes = existing_order_line.lot_ids.mapped('ref')
    #
    #                     # Filter out any None or empty values and ensure uniqueness
    #                     unique_barcodes = list(set([barcode for barcode in barcodes if barcode]))
    #
    #                     # Join the unique barcodes into a comma-separated string
    #                     branded_barcode_value = ', '.join(unique_barcodes)
    #
    #                     # Update the branded_barcode field
    #                     existing_order_line.branded_barcode = branded_barcode_value
    #
    #                     # Mark all the serial numbers as uploaded
    #                     sale_lines = self.env['sale.order.line'].search([
    #                         ('lot_ids.name', '=', serial_no.name),('company_id', '=', main_company.id),
    #                         ('order_id.state', 'not in', ['cancel'])
    #                     ])
    #                     sale_line_qty = sum(sale_lines.mapped('product_uom_qty'))
    #                     lot_product_qty = serial_no.product_qty - sale_line_qty
    #                     if lot_product_qty == 0:
    #                         serial_no.write({'is_uploaded': True})
    #
    #             elif values.get('product'):
    #                 # Create a new sale order line with the product and quantity
    #                 order_lines = self.env['sale.order.line'].create({
    #                     'order_id': sale_order_brw.id,
    #                     'product_id': product_id.id,
    #                     'lot_ids': [(6, 0, serial_no.ids)],  # Set the lot_ids with all serial numbers
    #                     'branded_barcode': values.get('product'),
    #                     'name': product_id.display_name,
    #                     'product_uom_qty': values.get('quantity'),
    #                     'product_uom': product_id.uom_id.id,
    #                     'price_unit': product_id.lst_price,
    #                     'type_product': serial_no.type_product,
    #                     'sale_serial_type': serial_no.serial_type,
    #
    #                 })
    #                 # If the order line was created successfully, mark all the serial numbers as uploaded
    #                 if order_lines:
    #                     sale_lines = self.env['sale.order.line'].search([
    #                         ('lot_ids.name', '=', serial_no.name),('company_id', '=', main_company.id),
    #                         ('order_id.state', 'not in', ['cancel'])
    #                     ])
    #                     sale_line_qty = sum(sale_lines.mapped('product_uom_qty'))
    #                     lot_product_qty = serial_no.product_qty - sale_line_qty
    #                     if lot_product_qty == 0:
    #                         serial_no.write({'is_uploaded': True})
    #
    #         elif sale_order_brw.state != 'sent' or sale_order_brw.state != 'draft':
    #             raise ValidationError(_('We cannot import data in validated or confirmed order!.'))
    #     else:
    #         if self.import_prod_option == 'barcode':
    #             if len(values['product']) > 13:
    #                 product_obj_search = self.env['product.product'].search([('barcode', '=', values['product'])])
    #                 location = self.env.ref('stock.stock_location_stock').id
    #                 serial_nos = self.env['stock.quant'].search([('lot_id.name', '=', values['serial_no']),('location_id','=',location),
    #                                                           ('company_id', '=', main_company.id),])
    #                 serial_no = serial_nos.lot_id
    #                 if not serial_no:
    #                     raise ValidationError(
    #                         _('The serial number for this is not found in the database.'))
    #             else:
    #                 product_barcodes = self.env['product.barcode'].search([('barcode', '=', values['product'])])
    #                 if len(product_barcodes) > 0:
    #                     product_obj_search = product_barcodes[0].product_id
    #                 location = self.env.ref('stock.stock_location_stock').id
    #                 serial_nos = self.env['stock.quant'].search(
    #                     [('lot_id.ref', '=', values['product']),('location_id','=',location), ('company_id', '=', main_company.id),('quantity', '>', 0), ('lot_id.is_uploaded', '=', False)],
    #                     limit=1)
    #                 serial_no = serial_nos.lot_id
    #                 if not serial_no:
    #                     raise ValidationError(
    #                         _('The serial number for this is not found in the database.'))
    #         elif self.import_prod_option == 'code':
    #             raise ValidationError(_('Please set the import Option to Barcode.'))
    #         else:
    #             raise ValidationError(_('Please set the import Option to Barcode.'))
    #         if product_obj_search:
    #             product_id = product_obj_search
    #         else:
    #             if self.import_prod_option == 'name':
    #                 raise ValidationError(_('Please set the import Option to Barcode.'))
    #             else:
    #                 raise ValidationError(
    #                     _('%s Product was not found in the Database.') % values.get(
    #                         'product'))
    #         if sale_order_brw.state == 'draft':
    #             # Checking if the product barcode is longer than 13 characters
    #             if len(values['product']) > 13:
    #                 # Filter order line by product ID
    #                 existing_product_line = sale_order_brw.order_line.filtered(
    #                     lambda x: x.product_id == product_id)
    #
    #                 # Check if the serial number exists in the lot_ids of the filtered order line
    #                 existing_line = existing_product_line.lot_ids.filtered(lambda x: x.id in serial_no.ids)
    #
    #             else:
    #                 # Split the 'branded_barcode' field by commas and check if the incoming product barcode exists in any of them
    #                 existing_line = sale_order_brw.order_line.filtered(
    #                     lambda x: x.product_id == product_id and
    #                               values.get('product') in (x.branded_barcode or '').split(','))
    #
    #             # Raise an error if the product already exists
    #             if existing_line:
    #                 raise ValidationError(_('%s Product is already existing.') % values.get('product'))
    #
    #             # Filter order line by product ID to update existing lines
    #             existing_order_line = sale_order_brw.order_line.filtered(
    #                 lambda x: x.product_id == product_id)
    #
    #             if existing_order_line:
    #                 # Update the quantity of the existing order line
    #                 existing_order_line.product_uom_qty += values.get('quantity')
    #
    #                 # If serial numbers are present, link them to the existing order line
    #                 if serial_no:
    #                     if serial_no.is_under_plan == True:
    #                         raise ValidationError(_('Lot Number is Under Planned'))
    #                     # Add the serial numbers to lot_ids without replacing the existing ones
    #                     existing_order_line.lot_ids = [(4, lot.id) for lot in serial_no]
    #
    #                     # Collect all the barcode references from the lot_ids (serial numbers)
    #                     barcodes = existing_order_line.lot_ids.mapped('ref')
    #
    #                     # Filter out any None or empty values and ensure uniqueness
    #                     unique_barcodes = list(set([barcode for barcode in barcodes if barcode]))
    #
    #                     # Join the unique barcodes into a comma-separated string
    #                     branded_barcode_value = ', '.join(unique_barcodes)
    #
    #                     # Update the branded_barcode field
    #                     existing_order_line.branded_barcode = branded_barcode_value
    #
    #                     # Mark all the serial numbers as uploaded
    #                     sale_lines = self.env['sale.order.line'].search([
    #                         ('lot_ids.name', '=', serial_no.name),('company_id', '=', main_company.id),
    #                         ('order_id.state', 'not in', ['cancel'])
    #                     ])
    #                     sale_line_qty = sum(sale_lines.mapped('product_uom_qty'))
    #                     lot_product_qty = serial_no.product_qty - sale_line_qty
    #                     if lot_product_qty == 0:
    #                         serial_no.write({'is_uploaded': True})
    #
    #             elif values.get('product'):
    #                 # Create a new sale order line with the product and quantity
    #                 order_lines = self.env['sale.order.line'].create({
    #                     'order_id': sale_order_brw.id,
    #                     'product_id': product_id.id,
    #                     'lot_ids': [(6, 0, serial_no.ids)],  # Set the lot_ids with all serial numbers
    #                     'branded_barcode': values.get('product'),
    #                     'product_uom_qty': values.get('quantity'),
    #                     'type_product': serial_no.type_product,
    #                     'sale_serial_type': serial_no.serial_type,
    #
    #                 })
    #                 # If the order line was created successfully, mark all the serial numbers as uploaded
    #                 if order_lines:
    #                     sale_lines = self.env['sale.order.line'].search([
    #                         ('lot_ids.name', '=', serial_no.name),('company_id', '=', main_company.id),
    #                         ('order_id.state', 'not in', ['cancel'])
    #                     ])
    #                     sale_line_qty = sum(sale_lines.mapped('product_uom_qty'))
    #                     lot_product_qty = serial_no.product_qty - sale_line_qty
    #                     if lot_product_qty == 0:
    #                         serial_no.write({'is_uploaded': True})
    #
    #         elif sale_order_brw.state == 'sent':
    #             # Checking if the product barcode is longer than 13 characters
    #             if len(values['product']) > 13:
    #                 # Filter order line by product ID
    #                 existing_product_line = sale_order_brw.order_line.filtered(
    #                     lambda x: x.product_id == product_id)
    #
    #                 # Check if the serial number exists in the lot_ids of the filtered order line
    #                 existing_line = existing_product_line.lot_ids.filtered(lambda x: x.id in serial_no.ids)
    #
    #             else:
    #                 # Split the 'branded_barcode' field by commas and check if the incoming product barcode exists in any of them
    #                 existing_line = sale_order_brw.order_line.filtered(
    #                     lambda x: x.product_id == product_id and
    #                               values.get('product') in (x.branded_barcode or '').split(','))
    #
    #             # Raise an error if the product already exists
    #             if existing_line:
    #                 raise ValidationError(_('%s Product is already existing.') % values.get('product'))
    #
    #             # Filter order line by product ID to update existing lines
    #             existing_order_line = sale_order_brw.order_line.filtered(
    #                 lambda x: x.product_id == product_id)
    #
    #             if existing_order_line:
    #                 # Update the quantity of the existing order line
    #                 existing_order_line.product_uom_qty += values.get('quantity')
    #
    #                 # If serial numbers are present, link them to the existing order line
    #                 if serial_no:
    #                     if serial_no.is_under_plan == True:
    #                         raise ValidationError(_('Lot Number is Under Planned'))
    #                     # Add the serial numbers to lot_ids without replacing the existing ones
    #                     existing_order_line.lot_ids = [(4, lot.id) for lot in serial_no]
    #
    #                     # Collect all the barcode references from the lot_ids (serial numbers)
    #                     barcodes = existing_order_line.lot_ids.mapped('ref')
    #
    #                     # Filter out any None or empty values and ensure uniqueness
    #                     unique_barcodes = list(set([barcode for barcode in barcodes if barcode]))
    #
    #                     # Join the unique barcodes into a comma-separated string
    #                     branded_barcode_value = ', '.join(unique_barcodes)
    #
    #                     # Update the branded_barcode field
    #                     existing_order_line.branded_barcode = branded_barcode_value
    #
    #                     # Mark all the serial numbers as uploaded
    #                     sale_lines = self.env['sale.order.line'].search([
    #                         ('lot_ids.name', '=', serial_no.name),('company_id', '=', main_company.id),
    #                         ('order_id.state', 'not in', ['cancel'])
    #                     ])
    #                     sale_line_qty = sum(sale_lines.mapped('product_uom_qty'))
    #                     lot_product_qty = serial_no.product_qty - sale_line_qty
    #                     if lot_product_qty == 0:
    #                         serial_no.write({'is_uploaded': True})
    #
    #             elif values.get('product'):
    #                 # Create a new sale order line with the product and quantity
    #                 order_lines = self.env['sale.order.line'].create({
    #                     'order_id': sale_order_brw.id,
    #                     'product_id': product_id.id,
    #                     'lot_ids': [(6, 0, serial_no.ids)],  # Set the lot_ids with all serial numbers
    #                     'branded_barcode': values.get('product'),
    #                     'product_uom_qty': values.get('quantity'),
    #                     'type_product': serial_no.type_product,
    #                     'sale_serial_type': serial_no.serial_type,
    #
    #                 })
    #                 # If the order line was created successfully, mark all the serial numbers as uploaded
    #                 if order_lines:
    #                     sale_lines = self.env['sale.order.line'].search([
    #                         ('lot_ids.name', '=', serial_no.name),('company_id', '=', main_company.id),
    #                         ('order_id.state', 'not in', ['cancel'])
    #                     ])
    #                     sale_line_qty = sum(sale_lines.mapped('product_uom_qty'))
    #                     lot_product_qty = serial_no.product_qty - sale_line_qty
    #                     if lot_product_qty == 0:
    #                         serial_no.write({'is_uploaded': True})
    #
    #         elif sale_order_brw.state != 'sent' or sale_order_brw.state != 'draft':
    #             raise ValidationError(_('We cannot import data in validated or confirmed order!.'))
    #     return True
