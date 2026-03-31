import openpyxl
from io import BytesIO
from openpyxl import load_workbook
from odoo import fields, models, _
from odoo.exceptions import ValidationError
import logging

_logger = logging.getLogger(__name__)

try:
    import xlrd
except ImportError:
    _logger.debug('Oops! Cannot `import xlrd`.')
try:
    import csv
except ImportError:
    _logger.debug('Oops! Cannot `import csv`.')
try:
    import base64
except ImportError:
    _logger.debug('Oops! Cannot `import base64`.')


class order_line_wizard(models.TransientModel):
    _name = 'order.line.wizard'
    _description = "Order Line Wizard"

    sale_order_id = fields.Many2one('sale.order', string="Sale Order", required=True)
    file = fields.Binary(string="Excel File", required=True)
    filename = fields.Char(string="Filename")

    def _check_serial_availability(self, lot, location_id):
        StockMoveLine = self.env['stock.move.line']
        SaleOrderLine = self.env['sale.order.line']

        # IN qty
        in_qty = sum(StockMoveLine.search([
            ('company_id.nhcl_company_bool', '=', True),
            ('lot_id', '=', lot.id),
            ('location_dest_id', '=', location_id),
        ]).mapped('quantity'))

        # OUT qty (returns)
        out_qty = sum(StockMoveLine.search([
            ('company_id.nhcl_company_bool', '=', True),
            ('lot_id', '=', lot.id),
            ('picking_id.stock_picking_type', '=', 'goods_return'),
        ]).mapped('quantity'))

        # Already used in other sale orders
        used_qty = sum(SaleOrderLine.search([
            ('lot_ids', 'in', lot.id),
            ('order_id.state', 'not in', ['cancel']),
        ]).mapped('product_uom_qty'))

        available = in_qty - out_qty - used_qty
        return available

    def action_import_barcodes(self):
        """Read Excel and create/merge sale order lines based on GS1 or EAN-13 barcode format.
        - Merges quantities for the same lot in the same sale order (lot-tracked products).
        - Validates that the total requested qty for a barcode does not exceed actual available qty.
        - When merging/updating qty, always update the line's price_unit from the lot (cost_price)
          falling back to product.lst_price.
        - For serial-tracked products behavior:
            * Each serial (lot that represents a serial) is handled as a separate 1-qty line.
            * If the same serial is already used in this order or another active order it will be skipped.
        """
        self.ensure_one()
        if not self.file:
            raise ValidationError(_("Please upload an Excel file."))

        try:
            data = base64.b64decode(self.file)
            wb = load_workbook(filename=BytesIO(data), data_only=True)
            sheet = wb.active
        except Exception as e:
            raise ValidationError(_("Invalid Excel file: %s") % str(e))

        SaleOrderLine = self.env['sale.order.line']
        StockQuant = self.env['stock.quant']
        StockMoveLine = self.env['stock.move.line']
        location_id = self.env.ref('stock.stock_location_stock').id
        skipped_lines = []

        # Helper: find existing sale order line for same product + lot + serial type in this sale order
        def _find_existing_line(product, lot_id, serial_type):
            return SaleOrderLine.search([
                ('order_id', '=', self.sale_order_id.id),
                ('product_id', '=', product.id),
                ('lot_ids', 'in', [lot_id]),
                ('sale_serial_type', '=', serial_type),
                ('order_id.state', 'not in', ['cancel'])
            ], limit=1)

        # Helper: compute actual available qty for a lot (qty received in receipts - qty already allocated in active sale lines)
        def _actual_available_qty_for_lot(lot):
            # qty that has come in for this lot via receipts
            in_qty = sum(StockMoveLine.search([
                ('picking_id.stock_picking_type', 'in', ['receipt']),  # safer than string 'receipt'
                ('company_id.nhcl_company_bool', '=', True),
                ('lot_id', '=', lot.id)
            ]).mapped('qty_done')) or 0.0
            out_qty = sum(StockMoveLine.search([
                ('picking_id.stock_picking_type', 'in', ['goods_return']),  # safer than string 'receipt'
                ('company_id.nhcl_company_bool', '=', True),
                ('lot_id', '=', lot.id)
            ]).mapped('qty_done')) or 0.0

            # qty already used in active sale order lines that reference this lot
            used_qty = sum(SaleOrderLine.search([
                ('order_id.state', 'not in', ['cancel']),
                ('lot_ids', 'in', [lot.id])
            ]).mapped('product_uom_qty')) or 0.0

            return float(in_qty) - float(used_qty) - float(out_qty)

        # Iterate rows
        for row in sheet.iter_rows(min_row=2, values_only=True):
            barcode = str(row[0]).strip() if row[0] else ""
            qty = row[1] or 0

            if not barcode or not qty:
                skipped_lines.append((barcode, "Missing barcode or quantity"))
                continue

            # Prepare variables reused across blocks
            product = False
            quants = False

            # ------------------ BRAND + 'R' BARCODE (NEW BLOCK) ------------------
            if barcode.startswith('R'):

                matched_quant = StockQuant.search([
                    ('lot_id.ref', '=', barcode),
                    ('company_id.nhcl_company_bool', '=', True),
                    ('quantity', '>', 0),
                    ('location_id', '=', location_id),
                    ('lot_id.type_product', '=', 'brand')
                ], limit=1)

                # ONLY handle if BRAND FOUND
                if matched_quant:
                    lot_code = barcode

                    quant_domain = [
                        ('lot_id.ref', '=', lot_code),
                        ('company_id.nhcl_company_bool', '=', True),
                        ('quantity', '>', 0),
                        ('location_id', '=', location_id),
                        ('lot_id.type_product', '=', 'brand')
                    ]

                    quants = StockQuant.search(quant_domain, order='id asc')
                    product = quants[:1].product_id if quants else False

                    if not product:
                        skipped_lines.append((barcode, "No matching product for R-type brand barcode"))
                        continue

                    existing_types = SaleOrderLine.search([
                        ('order_id', '=', self.sale_order_id.id),
                        ('order_id.state', 'not in', ['cancel']),
                        ('product_id', '=', product.id)
                    ]).mapped('sale_serial_type')
                    sale_serial_type = existing_types[0] if existing_types else 'regular'

                    # -------- SERIAL TRACKING --------
                    if product.tracking == 'serial':

                        created_count = 0
                        for q in quants:
                            lot = q.lot_id
                            if not lot:
                                continue

                            available_qty = self._check_serial_availability(lot, location_id)
                            if available_qty < 1:
                                continue

                            # skip already used
                            existing_lines = SaleOrderLine.search([
                                ('lot_ids', 'in', [lot.id]),
                                ('order_id.state', 'not in', ['cancel'])
                            ])
                            if existing_lines:
                                continue

                            SaleOrderLine.create({
                                'order_id': self.sale_order_id.id,
                                'product_id': product.id,
                                'lot_ids': [(6, 0, [lot.id])],
                                'branded_barcode': lot.ref,
                                'name': product.display_name,
                                'product_uom_qty': 1,
                                'product_uom': product.uom_id.id,
                                'price_unit': lot.cost_price,
                                'type_product': lot.type_product,
                                'sale_serial_type': lot.serial_type or 'regular',
                            })

                            created_count += 1
                            if created_count >= qty:
                                break

                        if created_count < qty:
                            skipped_lines.append(
                                (barcode, f"Only {created_count} serials available out of requested {qty}")
                            )

                        continue


                    # -------- LOT TRACKING --------
                    else:
                        available_quants = quants.filtered(lambda q: q.quantity > 0)

                        total_available = sum(
                            max(0.0, _actual_available_qty_for_lot(q.lot_id)) for q in available_quants)
                        if qty > total_available:
                            skipped_lines.append(
                                (barcode, f"Requested qty {qty} exceeds available stock {int(total_available)}"))
                            continue

                        remaining_qty = qty
                        for q in available_quants:
                            lot = q.lot_id
                            actual_available_qty = max(0.0, _actual_available_qty_for_lot(lot))
                            if actual_available_qty <= 0:
                                continue

                            allocate_qty = min(remaining_qty, actual_available_qty)

                            existing_line = _find_existing_line(product, lot.id, sale_serial_type)
                            price_unit = getattr(lot, 'cost_price', product.lst_price)

                            if existing_line:
                                existing_line.write({
                                    'product_uom_qty': existing_line.product_uom_qty + allocate_qty,
                                    'price_unit': price_unit,
                                })
                            else:
                                SaleOrderLine.create({
                                    'order_id': self.sale_order_id.id,
                                    'product_id': product.id,
                                    'lot_ids': [(6, 0, [lot.id])],
                                    'branded_barcode': lot.ref,
                                    'name': product.display_name,
                                    'product_uom_qty': allocate_qty,
                                    'product_uom': product.uom_id.id,
                                    'price_unit': price_unit,
                                    'type_product': lot.type_product,
                                    'sale_serial_type': sale_serial_type,
                                })

                            remaining_qty -= allocate_qty
                            if remaining_qty <= 0:
                                break

                        if remaining_qty > 0:
                            skipped_lines.append(
                                (barcode, f"Requested qty exceeds available stock. Short by {int(remaining_qty)}"))

                        continue

            # ------------------ GS1 (identified by presence of 'R' in the barcode) ------------------
            if 'R' in barcode:
                # GS1 -> lot code starts from 'R'
                lot_code = barcode[barcode.find('R'):]
                quant_domain = [
                    ('lot_id.name', '=', lot_code),
                    ('company_id.nhcl_company_bool', '=', True),
                    ('quantity', '>', 0),
                    ('location_id', '=', location_id),
                    # product type filter used in original code
                    ('lot_id.type_product', '=', 'un_brand')
                ]
                quants = StockQuant.search(quant_domain, order='id asc')
                product = quants[:1].product_id if quants else False

                if not product:
                    skipped_lines.append((barcode, "No matching product for GS1 barcode"))
                    continue

                # Determine sale_serial_type already present in this order for this product (keep first if exists)
                existing_types = SaleOrderLine.search([
                    ('order_id', '=', self.sale_order_id.id),
                    ('order_id.state', 'not in', ['cancel']),
                    ('product_id', '=', product.id)
                ]).mapped('sale_serial_type')
                sale_serial_type = existing_types[0] if existing_types else 'regular'

                if product.tracking == 'serial':

                    if qty != 1:
                        skipped_lines.append((barcode, "Serial product: Qty must be 1"))
                        continue

                    for q in quants:
                        lot = q.lot_id
                        if not lot:
                            continue

                        # -------- LANDED COST & PRICE VALIDATION --------
                        if lot.rs_price <= 0.0:
                            skipped_lines.append((barcode, f"{lot.name} not done with landed cost"))
                            continue

                        landed_cost = lot.picking_id.has_landed_cost_status
                        if not landed_cost and lot.picking_id.is_landed_cost == 'yes':
                            skipped_lines.append((barcode, f"Landed cost pending for {lot.name}"))
                            continue

                        # -------- AVAILABILITY CHECK (MOVEMENT BASED) --------
                        available_qty = self._check_serial_availability(lot, location_id)
                        if available_qty < 1:
                            skipped_lines.append((barcode, f"Serial {lot.name} not available for sale"))
                            continue

                        # -------- CREATE SALE LINE --------
                        SaleOrderLine.create({
                            'order_id': self.sale_order_id.id,
                            'product_id': product.id,
                            'lot_ids': [(6, 0, [lot.id])],
                            'branded_barcode': lot.ref,
                            'name': product.display_name,
                            'product_uom_qty': 1,
                            'product_uom': product.uom_id.id,
                            'price_unit': lot.cost_price,
                            'type_product': lot.type_product,
                            'sale_serial_type': lot.serial_type or 'regular',
                        })
                        break


                # LOT TRACKING (this is where merging is applied)
                else:
                    # Build list of lots (quants) with their actual available qty
                    available_quants = quants.filtered(lambda q: q.quantity > 0)
                    if not available_quants:
                        skipped_lines.append((barcode, "No stock available for GS1 lot items"))
                        continue

                    # Validate total available across all matching lots
                    total_available = sum(max(0.0, _actual_available_qty_for_lot(q.lot_id)) for q in available_quants)
                    if qty > total_available:
                        skipped_lines.append(
                            (barcode, f"Requested qty {qty} exceeds available stock {int(total_available)}"))
                        continue

                    remaining_qty = qty
                    for q in available_quants:
                        lot = q.lot_id
                        actual_available_qty = max(0.0, _actual_available_qty_for_lot(lot))
                        if actual_available_qty <= 0:
                            continue

                        allocate_qty = min(remaining_qty, actual_available_qty)
                        if allocate_qty <= 0:
                            continue

                        # Merge or create
                        existing_line = _find_existing_line(product, lot.id, sale_serial_type)
                        price_unit = getattr(lot, 'cost_price', product.lst_price)

                        if existing_line:
                            # Update qty and ALWAYS update price from lot (option A)
                            new_qty = (existing_line.product_uom_qty or 0.0) + allocate_qty
                            existing_line.write({
                                'product_uom_qty': new_qty,
                                'price_unit': price_unit,
                            })
                        else:
                            SaleOrderLine.create({
                                'order_id': self.sale_order_id.id,
                                'product_id': product.id,
                                'lot_ids': [(6, 0, [lot.id])],
                                'branded_barcode': lot.ref,
                                'name': product.display_name,
                                'product_uom_qty': allocate_qty,
                                'product_uom': product.uom_id.id,
                                'price_unit': price_unit,
                                'type_product': getattr(lot, 'type_product', False),
                                'sale_serial_type': sale_serial_type,
                            })

                        remaining_qty -= allocate_qty
                        if remaining_qty <= 0:
                            break

                    if remaining_qty > 0:
                        skipped_lines.append(
                            (barcode, f"Requested qty exceeds available stock. Short by {int(remaining_qty)}"))

            # ------------------ EAN-13 (length 13) ------------------
            elif len(barcode) == 13:
                lot_code = barcode
                matched_quant = StockQuant.search([
                    ('lot_id.ref', '=', lot_code),
                    ('company_id.nhcl_company_bool', '=', True),
                    ('quantity', '>', 0),
                    ('location_id', '=', location_id),
                    ('lot_id.type_product', '=', 'brand')
                ], limit=1)

                if not matched_quant:
                    skipped_lines.append((barcode, "No stock found for EAN-13 barcode"))
                    continue

                product = matched_quant.product_id

                # Domain to gather all quants for that lot_code + product
                quant_domain = [
                    ('product_id', '=', product.id),
                    ('lot_id.ref', '=', lot_code),
                    ('company_id.nhcl_company_bool', '=', True),
                    ('quantity', '>', 0),
                    ('location_id', '=', location_id),
                    ('lot_id.type_product', '=', 'brand')
                ]
                quants = StockQuant.search(quant_domain, order='id asc')
                available_quants = quants.filtered(lambda q: q.quantity > 0)

                existing_types = SaleOrderLine.search([
                    ('order_id', '=', self.sale_order_id.id),
                    ('order_id.state', 'not in', ['cancel']),
                    ('product_id', '=', product.id)
                ]).mapped('sale_serial_type')
                sale_serial_type = existing_types[0] if existing_types else 'regular'

                # SERIAL TRACKING (Enhanced: include regular + return serials)
                if product.tracking == 'serial':
                    # Build regular & return quants (prioritize regular then return)
                    regular_quants = StockQuant.search([
                        ('product_id', '=', product.id),
                        ('lot_id.ref', '=', lot_code),
                        ('company_id.nhcl_company_bool', '=', True),
                        ('quantity', '>', 0),
                        ('location_id', '=', location_id),
                        ('lot_id.type_product', '=', 'brand'),
                        '|', ('lot_id.serial_type', '=', False), ('lot_id.serial_type', '=', 'regular'),
                    ], order='id asc')

                    return_quants = StockQuant.search([
                        ('product_id', '=', product.id),
                        ('lot_id.ref', '=', lot_code),
                        ('company_id.nhcl_company_bool', '=', True),
                        ('quantity', '>', 0),
                        ('location_id', '=', location_id),
                        ('lot_id.type_product', '=', 'brand'),
                        ('lot_id.serial_type', '=', 'return'),
                    ], order='id asc')

                    combined_quants = regular_quants + return_quants
                    if not combined_quants:
                        skipped_lines.append((barcode, "No serial stock available for EAN-13"))
                        continue

                    # Validate total available serials
                    if qty > len(combined_quants):
                        skipped_lines.append(
                            (barcode, f"Requested qty {qty} exceeds available serials {len(combined_quants)}"))
                        continue

                    created_count = 0
                    for q in combined_quants:
                        lot = q.lot_id

                        # Skip if serial already used in any active sale order with same type
                        existing_lines = SaleOrderLine.search([
                            ('lot_ids', 'in', [lot.id]),
                            ('order_id.state', 'not in', ['cancel'])
                        ])
                        if existing_lines:
                            same_type_line = existing_lines.filtered(
                                lambda l: l.sale_serial_type == (lot.serial_type or 'regular'))
                            if same_type_line:
                                # Already used with same serial type → skip this serial
                                continue

                        current_serial_type = lot.serial_type or 'regular'
                        price_unit = getattr(lot, 'cost_price', product.lst_price)
                        landed_cost = lot.picking_id.has_landed_cost_status
                        if not landed_cost and lot.picking_id.is_landed_cost == 'yes':
                            raise ValidationError(
                                f"You are not allowed to add {lot.name} serial not done with landed cost.")

                        # Create a 1-qty line for this serial
                        SaleOrderLine.create({
                            'order_id': self.sale_order_id.id,
                            'product_id': product.id,
                            'lot_ids': [(6, 0, [lot.id])],
                            'branded_barcode': lot.ref,
                            'name': product.display_name,
                            'product_uom_qty': 1,
                            'product_uom': product.uom_id.id,
                            'price_unit': price_unit,
                            'type_product': getattr(lot, 'type_product', False),
                            'sale_serial_type': current_serial_type,
                        })
                        created_count += 1
                        if created_count >= qty:
                            break

                    if created_count < qty:
                        short = qty - created_count
                        skipped_lines.append((barcode, f"Requested qty exceeds available stock. Short by {short}"))

                # LOT TRACKING -> merge/update qty and price from lot
                else:
                    available_quants = quants.filtered(lambda q: q.quantity > 0)
                    if not available_quants:
                        skipped_lines.append((barcode, "No lot stock available for EAN-13"))
                        continue

                    total_available = sum(max(0.0, _actual_available_qty_for_lot(q.lot_id)) for q in available_quants)
                    if qty > total_available:
                        skipped_lines.append(
                            (barcode, f"Requested qty {qty} exceeds available stock {int(total_available)}"))
                        continue

                    remaining_qty = qty
                    for q in available_quants:
                        lot = q.lot_id
                        actual_available_qty = max(0.0, _actual_available_qty_for_lot(lot))
                        if actual_available_qty <= 0:
                            continue

                        allocate_qty = min(remaining_qty, actual_available_qty)
                        if allocate_qty <= 0:
                            continue

                        existing_line = _find_existing_line(product, lot.id, sale_serial_type)
                        price_unit = getattr(lot, 'cost_price', product.lst_price)

                        if existing_line:
                            # Update qty and ALWAYS update price from lot (option A)
                            new_qty = (existing_line.product_uom_qty or 0.0) + allocate_qty
                            existing_line.write({
                                'product_uom_qty': new_qty,
                                'price_unit': price_unit,
                            })
                        else:
                            landed_cost = lot.picking_id.has_landed_cost_status
                            if not landed_cost and lot.picking_id.is_landed_cost == 'yes':
                                raise ValidationError(f"You are not allowed to add {lot.name} serial not done with landed cost.")
                            SaleOrderLine.create({
                                'order_id': self.sale_order_id.id,
                                'product_id': product.id,
                                'lot_ids': [(6, 0, [lot.id])],
                                'branded_barcode': lot.ref,
                                'name': product.display_name,
                                'product_uom_qty': allocate_qty,
                                'product_uom': product.uom_id.id,
                                'price_unit': price_unit,
                                'type_product': getattr(lot, 'type_product', False),
                                'sale_serial_type': sale_serial_type,
                            })

                        remaining_qty -= allocate_qty
                        if remaining_qty <= 0:
                            break

                    if remaining_qty > 0:
                        skipped_lines.append(
                            (barcode, f"Requested qty exceeds available stock. Short by {int(remaining_qty)}"))
            # ------------------ UNKNOWN BARCODE ------------------
            else:
                skipped_lines.append((barcode, "Unknown barcode format"))
                continue
        # After processing all rows, raise ValidationError if anything was skipped
        if skipped_lines:
            msg = "\n".join([f"{b or '<empty>'}: {reason}" for b, reason in skipped_lines])
            raise ValidationError(_("Some lines were skipped:\n%s") % msg)
