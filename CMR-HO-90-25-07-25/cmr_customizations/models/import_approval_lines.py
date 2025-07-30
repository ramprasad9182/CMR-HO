import openpyxl
from odoo import fields, models, _, exceptions
from datetime import datetime
from odoo.exceptions import UserError, ValidationError
import logging
_logger = logging.getLogger(__name__)
import io

try:
    import xlrd
except ImportError:
    _logger.debug('Oops! Cannot `import xlrd`.')
try:
    import csv
except ImportError:
    _logger.debug('Oops! Cannot `import csv`.')
try:
    import base64
except ImportError:
    _logger.debug('Oops! Cannot `import base64`.')


class import_approval_line_wizard(models.TransientModel):
    _name = 'import.approval.line.wizard'
    _description = "Import Approval Request Lines"

    approval_line_file = fields.Binary(string="Select File")
    import_option = fields.Selection([('csv', 'CSV File'), ('xls', 'XLS File')], string='Select', default='xls')
    import_prod_option = fields.Selection([('barcode', 'Barcode'), ('code', 'Code'), ('name', 'Name')],
                                          string='Import Product By ', default='barcode')
    product_details_option = fields.Selection(
        [('from_product', 'Take Details From The Product'), ('from_xls', 'Take Details From The XLS/CSV File'),
         ], default='from_product')

    def import_approval_line(self):
        counter = 0
        if self.import_option == 'csv':
            keys = ['product', 'quantity']
            try:
                wb = openpyxl.load_workbook(
                    filename=io.BytesIO(base64.b64decode(self.approval_line_file)), read_only=True
                )
                ws = wb.active
            except Exception:
                raise exceptions.ValidationError(_("Invalid file!"))
            values = {}
            for row_no in ws.iter_rows(min_row=2, max_row=None, min_col=None,
                                       max_col=None, values_only=True):
                counter+=1
                # for i in range(len(ws)):
                field = list(map(str, row_no))
                values = dict(zip(keys, field))
                if values:
                    if len(row_no) == 0:
                        continue
                    else:
                        product_barcode = ''
                        lot_name = ''
                        if row_no[0] == None:
                            continue
                        else:
                            # line = list(map(lambda row:isinstance(row.value, bytes) and row.value.encode('utf-8') or ustr(row.value), sheet.row(row_no)))
                            if len(row_no[0]) > 13:
                                product_barcode = row_no[0]
                                # if row_no[0][0] == '0' and row_no[0][1] == '1' and row_no[0][16] == '2' and row_no[0][
                                #     17] == '1':
                                #     for i in range(0, len(row_no[0])):
                                #         if i > 1 and i < 16:
                                #             product_barcode += row_no[0][i]
                                #         elif i > 17 and i < len(row_no[0]):
                                #             lot_name += row_no[0][i]
                                #             continue
                            else:
                                product_barcode = row_no[0]
                            if row_no[1] == None or row_no[1] <= 0:
                                raise ValidationError(_('%s Quantity must be greater than zero.') % (row_no[0]))
                            if self.product_details_option == 'from_product':
                                values.update({
                                    'code': product_barcode,
                                    'serial_no': lot_name,
                                    'quantity': row_no[1]
                                })
                            elif self.product_details_option == 'from_xls':
                                values.update({
                                    'code': product_barcode,
                                    'serial_no': lot_name,
                                    'quantity': row_no[1],
                                })
                            else:
                                values.update({
                                    'code': product_barcode,
                                    'quantity': row_no[1],
                                    'serial_no': lot_name,
                                })
                        res = self.create_approval_line(values)
        else:
            counter = 0
            try:
                wb = openpyxl.load_workbook(
                    filename=io.BytesIO(base64.b64decode(self.approval_line_file)), read_only=True
                )
                ws = wb.active
                values = {}
            except Exception:
                raise exceptions.ValidationError(_("Invalid file!"))

            for row_no in ws.iter_rows(min_row=2, max_row=None, min_col=None,
                                       max_col=None, values_only=True):
                counter+=1
                # for row_no in range(sheet.nrows):
                val = {}
                product_barcode = ''
                lot_name = ''
                if row_no[0] == None:
                    continue
                else:
                    # line = list(map(lambda row:isinstance(row.value, bytes) and row.value.encode('utf-8') or ustr(row.value), sheet.row(row_no)))
                    if len(row_no[0]) > 13:
                        product_barcode = row_no[0]
                        # if row_no[0][0] == '0' and row_no[0][1] == '1' and row_no[0][16] == '2' and row_no[0][
                        #     17] == '1':
                        #     for i in range(0, len(row_no[0])):
                        #         if i > 1 and i < 16:
                        #             product_barcode += row_no[0][i]
                        #         elif i > 17 and i < len(row_no[0]):
                        #             lot_name += row_no[0][i]
                        #             continue
                    else:
                        product_barcode = row_no[0]
                    if row_no[1] == None or row_no[1] <= 0:
                        raise ValidationError(_('%s Quantity must be greater than zero.') % (row_no[0]))
                    if self.product_details_option == 'from_product':
                        values.update({
                            'code': product_barcode,
                            'serial_no': lot_name,
                            'quantity': row_no[1]
                        })
                    elif self.product_details_option == 'from_xls':
                        values.update({
                            'code': product_barcode,
                            'serial_no': lot_name,
                            'quantity': row_no[1],
                        })
                    else:
                        values.update({
                            'code': product_barcode,
                            'quantity': row_no[1],
                            'serial_no': lot_name,
                        })
                    res = self.create_approval_line(values)
        view_id = self.env.ref('cmr_customizations.message_wizard_popup')
        context = dict(self._context or {})
        print("counter",counter)
        dict_msg = str(counter) + " Records Imported Successfully."
        context['message'] = dict_msg
        return {
            'name': _('Success'),
            'type': 'ir.actions.act_window',
            'view_mode': 'form',
            'res_model': 'message.wizard',
            'views': [(view_id.id, 'form')],
            # pass the id
            'view_id': view_id.id,
            'target': 'new',
            'context': context,
        }

    def create_approval_line(self, values):
        approval_request_brw = self.env['approval.request'].browse(self._context.get('active_id'))
        product = values.get('code')
        serial_no = self.env['stock.lot']
        product_obj_search = self.env['product.product']
        if self.product_details_option == 'from_product':
            if self.import_prod_option == 'barcode':
                if len(values['code']) > 13:
                    product_obj_search = self.env['product.product'].search([('barcode', '=', values['code'])])
                    # serial_no = self.env['stock.lot'].search([('name', '=', values['serial_no'])])
                    # if not serial_no:
                    #     raise ValidationError(_('The serial number for this is not found in the database.'))
                    print('product_obj_search',product_obj_search)
                else:
                    product_barcodes = self.env['product.barcode'].search([('barcode', '=', values['code'])])
                    if len(product_barcodes) > 0:
                        product_obj_search = product_barcodes[0].product_id
            elif self.import_prod_option == 'code':
                # raise ValidationError(_('Please set the import Option to Barcode.'))
                product_obj_search = self.env['product.product'].search([('default_code', '=', values['code'])])

            else:
                raise ValidationError(_('Please set the import Option to Barcode.'))
            if product_obj_search:
                product_id = product_obj_search[0]
            else:
                raise ValidationError(_('%s Product was not found".') % values.get('code'))

            if approval_request_brw.request_status == 'new':
                # if product_id.nhcl_product_type == 'unbranded':
                existing_line = approval_request_brw.product_line_ids.filtered(
                    lambda x: x.product_id == product_id and x.prod_serial_no == serial_no)
                # else:
                #     existing_line = approval_request_brw.product_line_ids.filtered(
                #         lambda x: x.product_id == product_id and x.prod_barcode == values.get('code'))
                if existing_line:
                    raise ValidationError(_('%s The product already exists.') % values.get('code'))
                approval_lines = self.env['approval.product.line'].create({
                    'approval_request_id': approval_request_brw.id,
                    # 'section': product_id.categ_id.parent_id.id,
                    # 'division': product_id.categ_id.parent_id.parent_id.id,
                    # 'department': product_id.categ_id.id,
                    'product_id': product_id.id,
                    'prod_serial_no':serial_no.ids,
                    'prod_barcode': values.get('code'),
                    'description': product_id.display_name,
                    'quantity': values.get('quantity'),
                    'product_uom_id': product_id.uom_po_id.id,
                })

            elif approval_request_brw.request_status == 'pending':
                # if product_id.nhcl_product_type == 'unbranded':
                existing_line = approval_request_brw.product_line_ids.filtered(
                    lambda x: x.product_id == product_id and x.prod_serial_no == serial_no)
                # else:
                #     existing_line = approval_request_brw.product_line_ids.filtered(
                #         lambda x: x.product_id == product_id and x.prod_barcode == values.get('code'))
                if existing_line:
                    raise ValidationError(_('%s The product already exists.') % values.get('code'))
                approval_lines = self.env['approval.product.line'].create({
                    'approval_request_id': approval_request_brw.id,
                    # 'section': product_id.categ_id.parent_id.id,
                    # 'division': product_id.categ_id.parent_id.parent_id.id,
                    # 'department': product_id.categ_id.id,
                    'product_id': product_id.id,
                    'prod_serial_no': serial_no.ids,
                    'prod_barcode': values.get('code'),
                    'description': product_id.display_name,
                    'quantity': values.get('quantity'),
                    'product_uom_id': product_id.uom_po_id.id,
                })

            elif approval_request_brw.request_status != 'pending' or approval_request_brw.request_status != 'new':
                raise UserError(_('We cannot import data in validated or confirmed order!.'))

        elif self.product_details_option == 'from_xls':
            if self.import_prod_option == 'barcode':
                barcode = values.get('code')
                if len(barcode) > 13:
                    product_obj_search = self.env['product.product'].search([('barcode', '=', values['code'])])
                    # serial_no = self.env['stock.lot'].search([('name', '=', values['serial_no'])])
                    # if not serial_no:
                        # raise ValidationError(_('The serial number for this is not found in the database.'))
                else:
                    product_barcodes = self.env['product.barcode'].search([('barcode', '=', values['code'])])
                    if len(product_barcodes) > 0:
                        product_obj_search = product_barcodes[0].product_id
            elif self.import_prod_option == 'code':
                # raise ValidationError(_('Please set the import Option to Barcode.'))
                product_obj_search = self.env['product.product'].search([('default_code', '=', values['code'])])

            else:
                raise ValidationError(_('Please set the import Option to Barcode.'))
            # tax_id_lst = []
            # if values.get('tax'):
            #     if ';' in values.get('tax'):
            #         tax_names = values.get('tax').split(';')
            #         for name in tax_names:
            #             tax = self.env['account.tax'].search([('name', '=', name), ('type_tax_use', '=', 'purchase')])
            #             if not tax:
            #                 raise ValidationError(_('"%s" Tax is not in your system') % name)
            #             tax_id_lst.append(tax.id)
            #
            #     elif ',' in values.get('tax'):
            #         tax_names = values.get('tax').split(',')
            #         for name in tax_names:
            #             tax = self.env['account.tax'].search([('name', '=', name), ('type_tax_use', '=', 'purchase')])
            #             if not tax:
            #                 raise ValidationError(_('"%s" Tax is not in your system') % name)
            #             tax_id_lst.append(tax.id)
            #     else:
            #         tax_names = values.get('tax').split(',')
            #         tax = self.env['account.tax'].search([('name', '=', tax_names), ('type_tax_use', '=', 'purchase')])
            #         if not tax:
            #             raise ValidationError(_('"%s" Tax is not in your system') % tax_names)
            #         tax_id_lst.append(tax.id)
            if product_obj_search:
                product_id = product_obj_search[0]
            else:
                if self.import_prod_option == 'name':
                    if values.get('price'):
                        standard_price = float(values.get('price'))
                    else:
                        standard_price = False

                    product_id = self.env['product.product'].create({'name': product, 'standard_price': standard_price})
                else:
                    raise ValidationError(
                        _('%s Product was not found in the Database.') % values.get(
                            'code'))

            if values.get('quantity'):
                quantity = float(values.get('quantity'))
            else:
                quantity = False

            if approval_request_brw.request_status == 'new':
                # if product_id.nhcl_product_type == 'unbranded':
                existing_product_line = approval_request_brw.product_line_ids.filtered(
                    lambda x: x.product_id == product_id)
                existing_line = existing_product_line.prod_serial_no.filtered(lambda x: x.id == serial_no.id)
                # else:
                #     existing_line = approval_request_brw.product_line_ids.filtered(
                #         lambda x: x.product_id == product_id and x.prod_barcode == values.get('code'))
                if existing_line:
                    raise ValidationError(_('%s The product already exists.') % values.get('code'))
                existing_approval_line = approval_request_brw.product_line_ids.filtered(
                    lambda x: x.product_id == product_id)
                if existing_approval_line:
                    existing_approval_line.quantity += quantity
                    # if product_id.nhcl_product_type == 'unbranded':
                    #     existing_approval_line.prod_serial_no = [(4, serial_no.id)]
                else:
                    approval_lines = self.env['approval.product.line'].create({
                        'approval_request_id': approval_request_brw.id,
                        # 'section': product_id.categ_id.parent_id.id,
                        # 'division': product_id.categ_id.parent_id.parent_id.id,
                        # 'department': product_id.categ_id.id,
                        'product_id': product_id.id,
                        'prod_serial_no': serial_no.ids,
                        'prod_barcode': values.get('code'),
                        'description': product_id.display_name,
                        'quantity': quantity,
                        'product_uom_id': product_id.uom_po_id.id or False,
                    })
            elif approval_request_brw.request_status == 'pending':
                # if product_id.nhcl_product_type == 'unbranded':
                existing_product_line = approval_request_brw.product_line_ids.filtered(
                    lambda x: x.product_id == product_id)
                existing_line = existing_product_line.prod_serial_no.filtered(lambda x:x.id == serial_no.id)
                # else:
                #     existing_line = approval_request_brw.product_line_ids.filtered(
                #         lambda x: x.product_id == product_id and x.prod_barcode == values.get('code'))
                if existing_line:
                    raise ValidationError(_('%s The product already exists.') % values.get('code'))
                existing_approval_line = approval_request_brw.product_line_ids.filtered(
                    lambda x: x.product_id == product_id)
                if existing_approval_line:
                    existing_approval_line.quantity += quantity
                    # if product_id.nhcl_product_type == 'unbranded':
                    #     existing_approval_line.prod_serial_no = [(4, serial_no.id)]
                else:
                    approval_lines = self.env['approval.product.line'].create({
                        'approval_request_id': approval_request_brw.id,
                        # 'section': product_id.categ_id.parent_id.id,
                        # 'division': product_id.categ_id.parent_id.parent_id.id,
                        # 'department': product_id.categ_id.id,
                        'product_id': product_id.id,
                        'prod_serial_no': serial_no.ids,
                        'prod_barcode': values.get('code'),
                        'description': product_id.display_name,
                        'quantity': quantity,
                        'product_uom_id': product_id.uom_po_id.id or False,
                    })

            elif approval_request_brw.request_status != 'pending' or approval_request_brw.request_status != 'new':
                raise UserError(_('We cannot import data in validated or confirmed order!.'))
        else:
            if self.import_prod_option == 'barcode':
                if len(values['code']) > 13:
                    product_obj_search = self.env['product.product'].search([('barcode', '=', values['code'])])
                    # serial_no = self.env['stock.lot'].search([('name', '=', values['serial_no'])])
                    # if not serial_no:
                    #     raise ValidationError(_('The serial number for this is not found in the database.'))
                else:
                    product_barcodes = self.env['product.barcode'].search([('barcode', '=', values['code'])])
                    if len(product_barcodes) > 0:
                        product_obj_search = product_barcodes[0].product_id
            elif self.import_prod_option == 'code':
                # raise ValidationError(_('Please set the import Option to Barcode.'))
                product_obj_search = self.env['product.product'].search([('default_code', '=', values['code'])])

            else:
                raise ValidationError(_('Please set the import Option to Barcode.'))
            if product_obj_search:
                product_id = product_obj_search[0]
            else:
                if self.import_prod_option == 'name':
                    if values.get('price'):
                        standard_price = float(values.get('price'))
                    else:
                        standard_price = False
                    product_id = self.env['product.product'].create({'name': product, 'standard_price': standard_price})
                else:
                    raise ValidationError(
                        _('%s Product was not found in the Database.') % values.get(
                            'code'))

            if values.get('quantity'):
                quantity = float(values.get('quantity'))
            else:
                quantity = False

            if approval_request_brw.request_status == 'new':
                # if product_id.nhcl_product_type == 'unbranded':
                existing_product_line = approval_request_brw.product_line_ids.filtered(
                    lambda x: x.product_id == product_id)
                existing_line = existing_product_line.prod_serial_no.filtered(lambda x:x.id == serial_no.id)
                # else:
                #     existing_line = approval_request_brw.product_line_ids.filtered(
                #         lambda x: x.product_id == product_id and x.prod_barcode == values.get('code'))
                if existing_line:
                    raise ValidationError(_('%s The product already exists.') % values.get('code'))
                existing_approval_line = approval_request_brw.product_line_ids.filtered(
                    lambda x: x.product_id == product_id)
                if existing_approval_line:
                    existing_approval_line.quantity += quantity
                    # if product_id.nhcl_product_type == 'unbranded':
                    #     existing_approval_line.prod_serial_no = [(4, serial_no.id)]
                else:
                    approval_lines = self.env['approval.product.line'].create({
                        'approval_request_id': approval_request_brw.id,
                        # 'section': product_id.categ_id.parent_id.id,
                        # 'division': product_id.categ_id.parent_id.parent_id.id,
                        # 'department': product_id.categ_id.id,
                        'product_id': product_id.id,
                        'prod_serial_no': serial_no.ids,
                        'prod_barcode': values.get('code'),
                        'description': product_id.display_name,
                        'quantity': quantity,
                        'product_uom_id': product_id.uom_po_id.id,
                    })
                    approval_lines.update({
                        'quantity': values.get('quantity'),
                    })

            elif approval_request_brw.request_status == 'pending':
                # if product_id.nhcl_product_type == 'unbranded':
                existing_product_line = approval_request_brw.product_line_ids.filtered(
                    lambda x: x.product_id == product_id)
                existing_line = existing_product_line.prod_serial_no.filtered(lambda x:x.id == serial_no.id)
                # else:
                #     existing_line = approval_request_brw.product_line_ids.filtered(
                #         lambda x: x.product_id == product_id and x.prod_barcode == values.get('code'))
                if existing_line:
                    raise ValidationError(_('%s The product already exists.') % values.get('code'))
                existing_approval_line = approval_request_brw.product_line_ids.filtered(
                    lambda x: x.product_id == product_id)
                if existing_approval_line:
                    existing_approval_line.quantity += quantity
                    # if product_id.nhcl_product_type == 'unbranded':
                    #     existing_approval_line.prod_serial_no = [(4, serial_no.id)]
                else:
                    approval_lines = self.env['approval.product.line'].create({
                        'approval_request_id': approval_request_brw.id,
                        # 'section': product_id.categ_id.parent_id.id,
                        # 'division': product_id.categ_id.parent_id.parent_id.id,
                        # 'department': product_id.categ_id.id,
                        'product_id': product_id.id,
                        'prod_serial_no': serial_no.ids,
                        'prod_barcode': values.get('code'),
                        'description': product_id.display_name,
                        'quantity': quantity,
                        'product_uom_id': product_id.uom_po_id.id,
                    })
                    approval_lines.update({
                        'quantity': values.get('quantity'),
                    })

            elif approval_request_brw.request_status != 'pending' or approval_request_brw.request_status != 'new':
                raise UserError(_('We cannot import data in validated or confirmed order!.'))
        return True


