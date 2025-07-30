# -*- coding: utf-8 -*-
# Part of BrowseInfo. See LICENSE file for full copyright and licensing details.
import openpyxl

from odoo import api, fields, models, _, exceptions
from datetime import datetime
import binascii
import tempfile
from tempfile import TemporaryFile
from odoo.exceptions import UserError, ValidationError
import logging
from odoo.tools import ustr

_logger = logging.getLogger(__name__)
import io

try:
    import xlrd
except ImportError:
    _logger.debug('Oops! Cannot `import xlrd`.')
try:
    import csv
except ImportError:
    _logger.debug('Oops! Cannot `import csv`.')
try:
    import base64
except ImportError:
    _logger.debug('Oops! Cannot `import base64`.')


class import_po_line_wizard(models.TransientModel):
    _name = 'import.po.line.wizard'
    _description = "Import Purchase Order Line"

    purchase_order_file = fields.Binary(string="Select File")
    import_option = fields.Selection([('csv', 'CSV File'), ('xls', 'XLS File')], string='Select', default='xls')
    import_prod_option = fields.Selection([('barcode', 'Barcode'), ('code', 'Code'), ('name', 'Name')],
                                          string='Import Product By ', default='barcode')
    product_details_option = fields.Selection(
        [('from_product', 'Take Details From The Product'), ('from_xls', 'Take Details From The XLS/CSV File'),
         ('from_pricelist', 'Take Details With Adapted Pricelist')], default='from_product')

    sample_option = fields.Selection([('csv', 'CSV'), ('xls', 'XLS')], string='Sample Type', default='xls')
    down_samp_file = fields.Boolean(string='Download Sample Files')

    def import_pol(self):
        counter = 0
        if self.import_option == 'csv':
            # keys = ['code', 'quantity', 'uom','description', 'price', 'tax']
            keys = ['product', 'quantity', 'price']
            try:
                wb = openpyxl.load_workbook(
                    filename=io.BytesIO(base64.b64decode(self.purchase_order_file)), read_only=True
                )
                ws = wb.active
            except Exception:
                raise exceptions.ValidationError(_("Invalid file!"))
            values = {}
            for row_no in ws.iter_rows(min_row=2, max_row=None, min_col=None,
                                       max_col=None, values_only=True):
                counter+=1
                # for i in range(len(ws)):
                field = list(map(str, row_no))
                values = dict(zip(keys, field))
                if values:
                    if len(row_no) == 0:
                        continue
                    else:
                        product_barcode = ''
                        lot_name = ''
                        if row_no[0] == None:
                            continue
                        else:
                            # line = list(map(lambda row:isinstance(row.value, bytes) and row.value.encode('utf-8') or ustr(row.value), sheet.row(row_no)))
                            if len(row_no[0]) > 13:
                                product_barcode = row_no[0]
                                # if row_no[0][0] == '0' and row_no[0][1] == '1' and row_no[0][16] == '2' and row_no[0][
                                #     17] == '1':
                                #     for i in range(0, len(row_no[0])):
                                #         if i > 1 and i < 16:
                                #             product_barcode += row_no[0][i]
                                #         elif i > 17 and i < len(row_no[0]):
                                #             lot_name += row_no[0][i]
                                #             continue
                            else:
                                product_barcode = row_no[0]
                            if row_no[1] == None or row_no[1] <= 0:
                                raise ValidationError(_('%s Quantity must be greater than zero.') % (row_no[0]))
                            # if row_no[2] == None or row_no[2] <= 0:
                            #     raise ValidationError(_('%s Price Must be Greater than Zero.') % (row_no[0]))
                            if self.product_details_option == 'from_product':
                                values.update({
                                    'code': product_barcode,
                                    'serial_no': lot_name,
                                    'quantity': row_no[1]
                                })
                            elif self.product_details_option == 'from_xls':
                                values.update({
                                    'code': product_barcode,
                                    'serial_no': lot_name,
                                    'quantity': row_no[1],
                                    # 'uom':line[2],
                                    # 'description':line[3],
                                    'price': row_no[2],
                                    # 'tax':line[5],
                                })
                            else:
                                values.update({
                                    'code': product_barcode,
                                    'quantity': row_no[1],
                                    'serial_no': lot_name,
                                })
                        res = self.create_po_line(values)
        else:
            try:
                wb = openpyxl.load_workbook(
                    filename=io.BytesIO(base64.b64decode(self.purchase_order_file)), read_only=True
                )
                ws = wb.active
                values = {}
            except Exception:
                raise exceptions.ValidationError(_("Invalid file!"))

            for row_no in ws.iter_rows(min_row=2, max_row=None, min_col=None,
                                       max_col=None, values_only=True):
                counter+=1
                # for row_no in range(sheet.nrows):
                val = {}
                product_barcode = ''
                lot_name = ''
                if row_no[0] == None:
                    continue
                else:
                    # line = list(map(lambda row:isinstance(row.value, bytes) and row.value.encode('utf-8') or ustr(row.value), sheet.row(row_no)))
                    if len(row_no[0]) > 13:
                        product_barcode = row_no[0]
                        # if row_no[0][0] == '0' and row_no[0][1] == '1' and row_no[0][16] == '2' and row_no[0][
                        #     17] == '1':
                        #     for i in range(0, len(row_no[0])):
                        #         if i > 1 and i < 16:
                        #             product_barcode += row_no[0][i]
                        #         elif i > 17 and i < len(row_no[0]):
                        #             lot_name += row_no[0][i]
                        #             continue
                    else:
                        product_barcode = row_no[0]
                    if row_no[1] == None or row_no[1] <= 0:
                        raise ValidationError(_('%s Quantity must be greater than zero.') % (row_no[0]))
                    # if row_no[2] == None or row_no[2] <= 0:
                    #     raise ValidationError(_('%s Price Must be Greater than Zero.') % (row_no[0]))
                    if self.product_details_option == 'from_product':
                        values.update({
                            'code': product_barcode,
                            'serial_no': lot_name,
                            'quantity': row_no[1]
                        })
                    elif self.product_details_option == 'from_xls':
                        values.update({
                            'code': product_barcode,
                            'serial_no': lot_name,
                            'quantity': row_no[1],
                            # 'uom':line[2],
                            # 'description':line[3],
                            'price': row_no[2],
                            # 'tax':line[5],
                        })
                    else:
                        values.update({
                            'code': product_barcode,
                            'quantity': row_no[1],
                            'serial_no': lot_name,
                        })
                    res = self.create_po_line(values)
        view_id = self.env.ref('bi_import_all_orders_lines.message_wizard_popup')
        context = dict(self._context or {})
        dict_msg = str(counter) + " Records Imported Successfully."
        context['message'] = dict_msg
        return {
            'name': _('Success'),
            'type': 'ir.actions.act_window',
            'view_mode': 'form',
            'res_model': 'message.wizard',
            'views': [(view_id.id, 'form')],
            # pass the id
            'view_id': view_id.id,
            'target': 'new',
            'context': context,
        }

    def create_po_line(self, values):
        purchase_order_brw = self.env['purchase.order'].browse(self._context.get('active_id'))
        current_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        product = values.get('code')
        serial_no = self.env['stock.lot']
        product_obj_search = self.env['product.product']
        if self.product_details_option == 'from_product':
            if self.import_prod_option == 'barcode':
                if len(values['code']) > 13:
                    product_obj_search = self.env['product.product'].search([('barcode', '=', values['code'])])
                    serial_no = self.env['stock.lot'].search([('name', '=', values['serial_no'])])
                    if not serial_no:
                        raise ValidationError(_('The serial number for this is not found in the database.'))
                else:
                    product_barcodes = self.env['product.barcode'].search([('barcode', '=', values['code'])])
                    if len(product_barcodes) > 0:
                        product_obj_search = product_barcodes[0].product_id
            elif self.import_prod_option == 'code':
                raise ValidationError(_('Please set the import Option to Barcode.'))
                # product_obj_search = self.env['product.product'].search([('default_code', '=', values['code'])])
            else:
                raise ValidationError(_('Please set the import Option to Barcode.'))
                # product_obj_search = self.env['product.product'].search([('name', '=', values['code'])])

            if product_obj_search:
                product_id = product_obj_search[0]
            else:
                raise ValidationError(_('%s Product was not found".') % values.get('code'))

            if purchase_order_brw.state == 'draft':
                # if product_id.nhcl_product_type == 'unbranded':
                existing_po_line = purchase_order_brw.order_line.filtered(
                    lambda x: x.product_id == product_id)
                existing_line = existing_po_line.prod_serial_no.filtered(lambda x:x.id == serial_no.id)
                # else:
                #     existing_line = purchase_order_brw.order_line.filtered(
                #         lambda x: x.product_id == product_id and x.prod_barcode == values.get('code'))
                if existing_line:
                    raise ValidationError(_('%s The product already exists.') % values.get('code'))
                existing_order_line = purchase_order_brw.order_line.filtered(
                    lambda x: x.product_id == product_id)
                if existing_order_line:
                    existing_order_line.product_qty += values.get('quantity')
                    if product_id.nhcl_product_type == 'unbranded':
                        existing_order_line.prod_serial_no = [(4, serial_no.id)]
                else:
                    po_order_lines = self.env['purchase.order.line'].create({
                        'order_id': purchase_order_brw.id,
                        'product_id': product_id.id,
                        'prod_serial_no':serial_no.ids,
                        'prod_barcode': values.get('code'),
                        'name': product_id.display_name,
                        'date_planned': current_time,
                        'product_qty': values.get('quantity'),
                        'product_uom': product_id.uom_po_id.id,
                        'price_unit': product_id.standard_price
                    })

            elif purchase_order_brw.state == 'sent':
                # if product_id.nhcl_product_type == 'unbranded':
                existing_po_line = purchase_order_brw.order_line.filtered(
                    lambda x: x.product_id == product_id)
                existing_line = existing_po_line.prod_serial_no.filtered(lambda x:x.id == serial_no.id)
                # else:
                #     existing_line = purchase_order_brw.order_line.filtered(
                #         lambda x: x.product_id == product_id and x.prod_barcode == values.get('code'))
                if existing_line:
                    raise ValidationError(_('%s The product already exists.') % values.get('code'))
                existing_order_line = purchase_order_brw.order_line.filtered(
                    lambda x: x.product_id == product_id)
                if existing_order_line:
                    existing_order_line.product_qty += values.get('quantity')
                    if product_id.nhcl_product_type == 'unbranded':
                        existing_order_line.prod_serial_no = [(4, serial_no.id)]
                else:
                    po_order_lines = self.env['purchase.order.line'].create({
                        'order_id': purchase_order_brw.id,
                        'product_id': product_id.id,
                        'prod_serial_no': serial_no.ids,
                        'prod_barcode': values.get('code'),
                        'name': product_id.display_name,
                        'date_planned': current_time,
                        'product_qty': values.get('quantity'),
                        'product_uom': product_id.uom_po_id.id,
                        'price_unit': product_id.standard_price
                    })

            elif purchase_order_brw.state != 'sent' or purchase_order_brw.state != 'draft':
                raise UserError(_('We cannot import data in validated or confirmed order!.'))

        elif self.product_details_option == 'from_xls':
            # uom = values.get('uom')
            if self.import_prod_option == 'barcode':
                barcode = values.get('code')
                if len(barcode) > 13:
                    product_obj_search = self.env['product.product'].search([('barcode', '=', values['code'])])
                    serial_no = self.env['stock.lot'].search([('name', '=', values['serial_no'])])
                    if not serial_no:
                        raise ValidationError(_('The serial number for this is not found in the database.'))
                else:
                    product_barcodes = self.env['product.barcode'].search([('barcode', '=', values['code'])])
                    if len(product_barcodes) > 0:
                        product_obj_search = product_barcodes[0].product_id
            elif self.import_prod_option == 'code':
                raise ValidationError(_('Please set the import Option to Barcode.'))
                # product_obj_search = self.env['product.product'].search([('default_code', '=', values['code'])])
            else:
                raise ValidationError(_('Please set the import Option to Barcode.'))
                # product_obj_search = self.env['product.product'].search([('name', '=', values['code'])])

            # uom_obj_search = self.env['uom.uom'].search([('name', '=', uom)])
            # tax_id_lst = []
            # if values.get('tax'):
            #     if ';' in values.get('tax'):
            #         tax_names = values.get('tax').split(';')
            #         for name in tax_names:
            #             tax = self.env['account.tax'].search([('name', '=', name), ('type_tax_use', '=', 'purchase')])
            #             if not tax:
            #                 raise ValidationError(_('"%s" Tax is not in your system') % name)
            #             tax_id_lst.append(tax.id)
            #
            #     elif ',' in values.get('tax'):
            #         tax_names = values.get('tax').split(',')
            #         for name in tax_names:
            #             tax = self.env['account.tax'].search([('name', '=', name), ('type_tax_use', '=', 'purchase')])
            #             if not tax:
            #                 raise ValidationError(_('"%s" Tax is not in your system') % name)
            #             tax_id_lst.append(tax.id)
            #     else:
            #         tax_names = values.get('tax').split(',')
            #         tax = self.env['account.tax'].search([('name', '=', tax_names), ('type_tax_use', '=', 'purchase')])
            #         if not tax:
            #             raise ValidationError(_('"%s" Tax is not in your system') % tax_names)
            #         tax_id_lst.append(tax.id)
            # if not uom_obj_search:
            #     raise ValidationError(_('UOM "%s" is Not Available') % uom)

            if product_obj_search:
                product_id = product_obj_search[0]
            else:
                if self.import_prod_option == 'name':
                    if values.get('price'):
                        standard_price = float(values.get('price'))
                    else:
                        standard_price = False

                    product_id = self.env['product.product'].create({'name': product, 'standard_price': standard_price})
                else:
                    raise ValidationError(
                        _('%s Product was not found in the Database.') % values.get(
                            'code'))

            if values.get('quantity'):
                quantity = float(values.get('quantity'))
            else:
                quantity = False

            if purchase_order_brw.state == 'draft':
                # if product_id.nhcl_product_type == 'unbranded':
                existing_po_line = purchase_order_brw.order_line.filtered(
                    lambda x: x.product_id == product_id)
                existing_line = existing_po_line.prod_serial_no.filtered(lambda x:x.id == serial_no.id)
                # else:
                #     existing_line = purchase_order_brw.order_line.filtered(
                #         lambda x: x.product_id == product_id and x.prod_barcode == values.get('code'))
                if existing_line:
                    raise ValidationError(_('%s The product already exists.') % values.get('code'))
                if values.get('price'):
                    standard_price = float(values.get('price'))
                else:
                    standard_price = False
                existing_order_line = purchase_order_brw.order_line.filtered(
                    lambda x: x.product_id == product_id)
                if existing_order_line:
                    existing_order_line.product_qty += quantity
                    if product_id.nhcl_product_type == 'unbranded':
                        existing_order_line.prod_serial_no = [(4, serial_no.id)]
                else:
                    po_order_lines = self.env['purchase.order.line'].create({
                        'order_id': purchase_order_brw.id,
                        'product_id': product_id.id,
                        'prod_serial_no': serial_no.ids,
                        'prod_barcode': values.get('code'),
                        'name': product_id.display_name,
                        'date_planned': current_time,
                        'product_qty': quantity,
                        'product_uom': product_id.uom_po_id.id or False,
                        'price_unit': product_id.standard_price
                    })
            elif purchase_order_brw.state == 'sent':
                # if product_id.nhcl_product_type == 'unbranded':
                existing_po_line = purchase_order_brw.order_line.filtered(
                    lambda x: x.product_id == product_id)
                existing_line = existing_po_line.prod_serial_no.filtered(lambda x:x.id == serial_no.id)
                # else:
                #     existing_line = purchase_order_brw.order_line.filtered(
                #         lambda x: x.product_id == product_id and x.prod_barcode == values.get('code'))
                if existing_line:
                    raise ValidationError(_('%s The product already exists.') % values.get('code'))
                existing_order_line = purchase_order_brw.order_line.filtered(
                    lambda x: x.product_id == product_id)
                if existing_order_line:
                    existing_order_line.product_qty += quantity
                    # if product_id.nhcl_product_type == 'unbranded':
                    #     existing_order_line.prod_serial_no = [(4, serial_no.id)]
                else:
                    po_order_lines = self.env['purchase.order.line'].create({
                        'order_id': purchase_order_brw.id,
                        'product_id': product_id.id,
                        'prod_serial_no': serial_no.ids,
                        'prod_barcode': values.get('code'),
                        'name': product_id.display_name,
                        'date_planned': current_time,
                        'product_qty': quantity,
                        'product_uom': product_id.uom_po_id.id or False,
                        'price_unit': product_id.standard_price
                    })

            elif purchase_order_brw.state != 'sent' or purchase_order_brw.state != 'draft':
                raise UserError(_('We cannot import data in validated or confirmed order!.'))
            # if tax_id_lst:
            #     po_order_lines.write({'taxes_id': ([(6, 0, tax_id_lst)])})
        else:
            if self.import_prod_option == 'barcode':
                if len(values['code']) > 13:
                    product_obj_search = self.env['product.product'].search([('barcode', '=', values['code'])])
                    serial_no = self.env['stock.lot'].search([('name', '=', values['serial_no'])])
                    if not serial_no:
                        raise ValidationError(_('The serial number for this is not found in the database.'))
                else:
                    product_barcodes = self.env['product.barcode'].search([('barcode', '=', values['code'])])
                    if len(product_barcodes) > 0:
                        product_obj_search = product_barcodes[0].product_id
            elif self.import_prod_option == 'code':
                raise ValidationError(_('Please set the import Option to Barcode.'))
                # product_obj_search = self.env['product.product'].search([('default_code', '=', values['code'])])
            else:
                raise ValidationError(_('Please set the import Option to Barcode.'))
                # product_obj_search = self.env['product.product'].search([('name', '=', values['code'])])

            if product_obj_search:
                product_id = product_obj_search[0]
            else:
                if self.import_prod_option == 'name':
                    if values.get('price'):
                        standard_price = float(values.get('price'))
                    else:
                        standard_price = False
                    product_id = self.env['product.product'].create({'name': product, 'standard_price': standard_price})
                else:
                    raise ValidationError(
                        _('%s Product was not found in the Database.') % values.get(
                            'code'))

            if values.get('quantity'):
                quantity = float(values.get('quantity'))
            else:
                quantity = False

            if purchase_order_brw.state == 'draft':
                # if product_id.nhcl_product_type == 'unbranded':
                existing_po_line = purchase_order_brw.order_line.filtered(
                    lambda x: x.product_id == product_id)
                existing_line = existing_po_line.prod_serial_no.filtered(lambda x:x.id == serial_no.id)
                # else:
                #     existing_line = purchase_order_brw.order_line.filtered(
                #         lambda x: x.product_id == product_id and x.prod_barcode == values.get('code'))
                if existing_line:
                    raise ValidationError(_('%s The product already exists.') % values.get('code'))
                existing_order_line = purchase_order_brw.order_line.filtered(
                    lambda x: x.product_id == product_id)
                if existing_order_line:
                    existing_order_line.product_qty += quantity
                    # if product_id.nhcl_product_type == 'unbranded':
                    #     existing_order_line.prod_serial_no = [(4, serial_no.id)]
                else:
                    po_order_lines = self.env['purchase.order.line'].create({
                        'order_id': purchase_order_brw.id,
                        'product_id': product_id.id,
                        'prod_serial_no': serial_no.ids,
                        'prod_barcode': values.get('code'),
                        'name': product_id.display_name,
                        'date_planned': current_time,
                        'product_qty': quantity,
                        'product_uom': product_id.uom_po_id.id,
                        'price_unit': product_id.standard_price
                    })
                    po_order_lines.onchange_product_id()
                    po_order_lines.update({
                        'product_qty': values.get('quantity'),
                    })

            elif purchase_order_brw.state == 'sent':
                # if product_id.nhcl_product_type == 'unbranded':
                existing_po_line = purchase_order_brw.order_line.filtered(
                    lambda x: x.product_id == product_id)
                existing_line = existing_po_line.prod_serial_no.filtered(lambda x:x.id == serial_no.id)
                # else:
                #     existing_line = purchase_order_brw.order_line.filtered(
                #         lambda x: x.product_id == product_id and x.prod_barcode == values.get('code'))
                if existing_line:
                    raise ValidationError(_('%s The product already exists.') % values.get('code'))
                existing_order_line = purchase_order_brw.order_line.filtered(
                    lambda x: x.product_id == product_id)
                if existing_order_line:
                    existing_order_line.product_qty += values.get('quantity')
                    # if product_id.nhcl_product_type == 'unbranded':
                    #     existing_order_line.prod_serial_no = [(4, serial_no.id)]
                else:
                    po_order_lines = self.env['purchase.order.line'].create({
                        'order_id': purchase_order_brw.id,
                        'product_id': product_id.id,
                        'prod_serial_no': serial_no.ids,
                        'prod_barcode': values.get('code'),
                        'name': product_id.display_name,
                        'date_planned': current_time,
                        'product_qty': quantity,
                        'product_uom': product_id.uom_po_id.id,
                        'price_unit': product_id.standard_price
                    })
                    po_order_lines.onchange_product_id()
                    po_order_lines.update({
                        'product_qty': values.get('quantity'),
                    })

            elif purchase_order_brw.state != 'sent' or purchase_order_brw.state != 'draft':
                raise UserError(_('We cannot import data in validated or confirmed order!.'))
        return True

    # def download_auto(self):
    #     return {
    #         'type': 'ir.actions.act_url',
    #         'url': '/web/binary/download_document?model=import.po.line.wizard&id=%s' % (self.id),
    #         'target': 'new',
    #     }

