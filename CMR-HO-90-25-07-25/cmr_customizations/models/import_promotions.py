import openpyxl
from odoo import fields, models, _, exceptions
from datetime import datetime
from odoo.exceptions import UserError, ValidationError
import logging

_logger = logging.getLogger(__name__)
import io

try:
    import csv
except ImportError:
    _logger.debug('Oops! Cannot `import csv`.')
try:
    import base64
except ImportError:
    _logger.debug('Oops! Cannot `import base64`.')


class import_promotions_wizard(models.TransientModel):
    _name = 'import.promotions.wizard'
    _description = "Import Promotions"

    promotion_file = fields.Binary(string="Select File")
    import_option = fields.Selection([('csv', 'CSV File'), ('xls', 'XLS File')], string='Select', default='csv')

    def import_promotions(self):
        counter = 0
        if self.import_option == 'csv':
            keys = ['product', 'quantity']
            try:
                wb = openpyxl.load_workbook(
                    filename=io.BytesIO(base64.b64decode(self.promotion_file)), read_only=True
                )
                ws = wb.active
            except Exception:
                raise exceptions.ValidationError(_("Invalid file!"))
            values = {}
            for row_no in ws.iter_rows(min_row=2, max_row=None, min_col=None,
                                       max_col=None, values_only=True):
                counter += 1
                # for i in range(len(ws)):
                field = list(map(str, row_no))
                values = dict(zip(keys, field))
                if values:
                    if len(row_no) == 0:
                        continue
                    else:
                        if row_no[0] == None:
                            continue
                        else:
                            values.update({
                                'name': row_no[0],
                                'program_type': row_no[1],
                                'currency': row_no[2],
                                'portal_point_name': row_no[3],
                                'date_from': row_no[4],
                                'date_to': row_no[5],
                                'min_qty': row_no[6],
                                'min_purchase': row_no[7],
                                'product_barcode': row_no[8],
                                # 'lot_name': lot_name
                            })
                        res = self.create_promotions(values)
        else:
            try:
                wb = openpyxl.load_workbook(
                    filename=io.BytesIO(base64.b64decode(self.promotion_file)), read_only=True
                )
                ws = wb.active
                values = {}
            except Exception:
                raise exceptions.ValidationError(_("Invalid file!"))

            for row_no in ws.iter_rows(min_row=2, max_row=None, min_col=None,
                                       max_col=None, values_only=True):
                counter += 1
                if row_no[0] == None:
                    continue
                else:
                    values.update({
                        'name': row_no[0],
                        'program_type': row_no[1],
                        'currency': row_no[2],
                        'portal_point_name': row_no[3],
                        'date_from': row_no[4],
                        'date_to': row_no[5],
                        'min_qty': row_no[6],
                        'min_purchase': row_no[7],
                        'product_barcode': row_no[8],
                    })
                    res = self.create_promotions(values)
        view_id = self.env.ref('bi_import_all_orders_lines.message_wizard_popup')
        context = dict(self._context or {})
        dict_msg = str(counter) + " Records Imported Successfully."
        context['message'] = dict_msg
        return {
            'name': _('Success'),
            'type': 'ir.actions.act_window',
            'view_mode': 'form',
            'res_model': 'message.wizard',
            'views': [(view_id.id, 'form')],
            # pass the id
            'view_id': view_id.id,
            'target': 'new',
            'context': context,
        }

    def create_promotions(self, values):
        currency = self.env['res.currency'].search([('name', '=', values['currency'])])
        program_id = self.env['loyalty.program'].create({
            'name': values['name'],
            'program_type': values['program_type'],
            'portal_point_name': values['portal_point_name'],
            'currency_id': currency.id,
            'date_from': values['date_from'],
            'date_to': values['date_to'],
        })
        product_obj_search = []
        for barcode in values['product_barcode'].split(','):
            # product_barcode = ''
            # lot_name = ''
            # if len(barcode) > 13:
                # if barcode[0] == '0' and barcode[1] == '1' and barcode[16] == '2' and barcode[17] == '1':
                #     for i in range(0, len(barcode[0])):
                #         if i > 1 and i < 16:
                #             product_barcode += barcode[i]
                #         elif i > 17 and i < len(barcode):
                #             lot_name += barcode[i]
                #             continue
            # else:
            product_barcode = barcode
            if len(product_barcode) > 13:
                product_obj = self.env['product.product'].search([('barcode', '=', product_barcode)])
                product_obj_search.append((4, product_obj.id))
                # serial_no = self.env['stock.lot'].search([('name', '=', lot_name)])
                # if not serial_no:
                #     raise ValidationError(
                #         _('The serial number for this is not found in the database.'))
            else:
                product_barcodes = self.env['product.barcode'].search([('barcode', '=', product_barcode)])
                if len(product_barcodes) > 0:
                    product_obj_search.append((4, product_barcodes[0].product_id.id))
        promotion_lines = {
            'minimum_qty': values['min_qty'],
            'minimum_amount': values['min_purchase'],
            'program_id': program_id.id,
            'product_ids': product_obj_search

        }
        self.env['loyalty.rule'].create(promotion_lines)
        return True


class LoyaltyProgram(models.Model):
    _inherit = 'loyalty.program'

    def button_import_promotion_action(self):
        return {
            "name": _("Import Promotions"),
            "type": "ir.actions.act_window",
            "res_model": "import.promotions.wizard",
            "target": "new",
            "views": [[False, "form"]],
        }
