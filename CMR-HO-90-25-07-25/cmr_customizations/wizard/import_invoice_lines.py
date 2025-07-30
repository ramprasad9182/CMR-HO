# -*- coding: utf-8 -*-
# Part of BrowseInfo. See LICENSE file for full copyright and licensing details.
import openpyxl

from odoo import api, fields, models, _, exceptions
from datetime import datetime
import binascii
import tempfile
from tempfile import TemporaryFile
from odoo.exceptions import UserError, ValidationError
import logging

_logger = logging.getLogger(__name__)
import io

try:
    import xlrd
except ImportError:
    _logger.debug('Oops! Cannot `import xlrd`.')
try:
    import csv
except ImportError:
    _logger.debug('Oops! Cannot `import csv`.')
try:
    import base64
except ImportError:
    _logger.debug('Oops! Cannot `import base64`.')


class import_invoice_wizard(models.TransientModel):
    _name = 'import.invoice.wizard'
    _description = 'import invoice wizard'

    invoice_file = fields.Binary(string="Select File")
    import_option = fields.Selection([('csv', 'CSV File'), ('xls', 'XLS File')], string='File Format', default='csv')
    import_prod_option = fields.Selection([('barcode', 'Barcode'), ('code', 'Code'), ('name', 'Name')],
                                          string='Import Product By ', default='name')
    product_details_option = fields.Selection(
        [('from_product', 'Take Details From The Product'), ('from_xls', 'Take Details From The XLS File')],
        default='from_xls')
    import_analytic_account_tags = fields.Boolean("Import Analytic Account & Tags")

    sample_option = fields.Selection([('csv', 'CSV'), ('xls', 'XLS')], string='Sample Type', default='csv')
    down_samp_file = fields.Boolean(string='Download Sample Files')

    def import_inv(self):
        counter = 0
        if self.import_option == 'csv':
            if self.import_analytic_account_tags == True:
                # keys = ['product', 'quantity', 'uom','description', 'price', 'tax','analytic_account','analytic_tags']
                keys = ['product', 'quantity', 'price', 'analytic_account', 'analytic_tags']
                try:
                    wb = openpyxl.load_workbook(
                        filename=io.BytesIO(base64.b64decode(self.invoice_file)), read_only=True
                    )
                    ws = wb.active
                except Exception:
                    raise exceptions.ValidationError(_("Invalid file!"))
                values = {}
                for row_no in ws.iter_rows(min_row=2, max_row=None, min_col=None,
                                           max_col=None, values_only=True):
                    counter+=1
                    # for i in range(len(file_reader)):
                    field = list(map(str, row_no))
                    values = dict(zip(keys, field))
                    if values:
                        if len(row_no) == 0:
                            continue
                        else:
                            product_barcode = ''
                            lot_name = ''
                            if row_no[0] == None:
                                continue
                            else:
                                # line = list(map(lambda row:isinstance(row.value, bytes) and row.value.encode('utf-8') or ustr(row.value), sheet.row(row_no)))
                                if len(row_no[0]) > 13:
                                    if row_no[0][0] == '0' and row_no[0][1] == '1' and row_no[0][16] == '2' and \
                                            row_no[0][
                                                17] == '1':
                                        for i in range(0, len(row_no[0])):
                                            if i > 1 and i < 16:
                                                product_barcode += row_no[0][i]
                                            elif i > 17 and i < len(row_no[0]):
                                                lot_name += row_no[0][i]
                                                continue
                                else:
                                    product_barcode = row_no[0]
                                if row_no[1] == None or row_no[1] <= 0:
                                    raise ValidationError(_('%s Quantity must be greater than zero.') % (row_no[0]))
                                # if row_no[2] == None or row_no[2] <= 0:
                                #     raise ValidationError(_('%s Price Must be Greater than Zero.') % (row_no[0]))
                                if self.product_details_option == 'from_product':
                                    values.update({
                                        'code': product_barcode,
                                        'serial_no': lot_name,
                                        'quantity': row_no[1]
                                    })
                                elif self.product_details_option == 'from_xls':
                                    values.update({
                                        'code': product_barcode,
                                        'serial_no': lot_name,
                                        'quantity': row_no[1],
                                        # 'uom':line[2],
                                        # 'description':line[3],
                                        'price': row_no[2],
                                        # 'tax':line[5],
                                    })
                                else:
                                    values.update({
                                        'code': product_barcode,
                                        'quantity': row_no[1],
                                        'serial_no': lot_name,
                                    })
                            res = self.create_inv_line(values)
            else:
                # keys = ['product', 'quantity', 'uom','description', 'price', 'tax']
                keys = ['product', 'quantity', 'price']
                try:
                    wb = openpyxl.load_workbook(
                        filename=io.BytesIO(base64.b64decode(self.invoice_file)), read_only=True
                    )
                    ws = wb.active
                except Exception:
                    raise exceptions.ValidationError(_("Invalid file!"))
                values = {}
                for row_no in ws.iter_rows(min_row=2, max_row=None, min_col=None,
                                           max_col=None, values_only=True):
                    counter+=1
                    # for i in range(len(file_reader)):
                    field = list(map(str, row_no))
                    values = dict(zip(keys, field))
                    product_barcode = ''
                    lot_name = ''
                    if values:
                        if len(row_no[0]) > 13:
                            if row_no[0][0] == '0' and row_no[0][1] == '1' and row_no[0][16] == '2' and row_no[0][
                                17] == '1':
                                for i in range(0, len(row_no[0])):
                                    if i > 1 and i < 16:
                                        product_barcode += row_no[0][i]
                                    elif i > 17 and i < len(row_no[0]):
                                        lot_name += row_no[0][i]
                                        continue
                        else:
                            product_barcode = row_no[0]
                        if row_no[1] == None or row_no[1] <= 0:
                            raise ValidationError(_('%s Quantity must be greater than zero.') % (row_no[0]))
                        # if row_no[2] == None or row_no[2] <= 0:
                        #     raise ValidationError(_('%s Price Must be Greater than Zero.') % (row_no[0]))
                        if self.product_details_option == 'from_product':
                            values.update({
                                'code': product_barcode,
                                'serial_no': lot_name,
                                'quantity': row_no[1]
                            })
                        elif self.product_details_option == 'from_xls':
                            values.update({
                                'code': product_barcode,
                                'serial_no': lot_name,
                                'quantity': row_no[1],
                                # 'uom':line[2],
                                # 'description':line[3],
                                'price': row_no[2],
                                # 'tax':line[5],
                            })
                        else:
                            values.update({
                                'code': product_barcode,
                                'quantity': row_no[1],
                                'serial_no': lot_name,
                            })
                        res = self.create_inv_line(values)
        else:
            if self.import_analytic_account_tags == True:
                try:
                    wb = openpyxl.load_workbook(
                        filename=io.BytesIO(base64.b64decode(self.purchase_order_file)), read_only=True
                    )
                    ws = wb.active
                    values = {}
                except Exception:
                    raise exceptions.ValidationError(_("Invalid file!"))
                for row_no in ws.iter_rows(min_row=2, max_row=None, min_col=None,
                                           max_col=None, values_only=True):
                    counter+=1
                    # for row_no in range(sheet.nrows):
                    val = {}
                    product_barcode = ''
                    lot_name = ''
                    if row_no[0] == None:
                        continue
                    else:
                        if len(row_no[0]) > 13:
                            if row_no[0][0] == '0' and row_no[0][1] == '1' and row_no[0][16] == '2' and row_no[0][
                                17] == '1':
                                for i in range(0, len(row_no[0])):
                                    if i > 1 and i < 16:
                                        product_barcode += row_no[0][i]
                                    elif i > 17 and i < len(row_no[0]):
                                        lot_name += row_no[0][i]
                                        continue
                        else:
                            product_barcode = row_no[0]
                        if row_no[1] == None or row_no[1] <= 0:
                            raise ValidationError(_('%s Quantity must be greater than zero.') % (row_no[0]))
                        # if row_no[2] == None or row_no[2] <= 0:
                        #     raise ValidationError(_('%s Price Must be Greater than Zero.') % (row_no[0]))
                        if self.product_details_option == 'from_product':
                            values.update({
                                'code': product_barcode,
                                'serial_no': lot_name,
                                'quantity': row_no[1]
                            })
                        elif self.product_details_option == 'from_xls':
                            values.update({
                                'code': product_barcode,
                                'serial_no': lot_name,
                                'quantity': row_no[1],
                                # 'uom':line[2],
                                # 'description':line[3],
                                'price': row_no[2],
                                # 'tax':line[5],
                            })
                        else:
                            values.update({
                                'code': product_barcode,
                                'quantity': row_no[1],
                                'serial_no': lot_name,
                            })
                        res = self.create_inv_line(values)
            else:
                try:
                    wb = openpyxl.load_workbook(
                        filename=io.BytesIO(base64.b64decode(self.invoice_file)), read_only=True
                    )
                    ws = wb.active
                    values = {}
                except Exception:
                    raise exceptions.ValidationError(_("Invalid file!"))
                for row_no in ws.iter_rows(min_row=2, max_row=None, min_col=None,
                                           max_col=None, values_only=True):
                    counter+=1
                    # for row_no in range(sheet.nrows):
                    val = {}
                    product_barcode = ''
                    lot_name = ''
                    if row_no[0] == None:
                        continue
                    else:
                        if len(row_no[0]) > 13:
                            if row_no[0][0] == '0' and row_no[0][1] == '1' and row_no[0][16] == '2' and row_no[0][
                                17] == '1':
                                for i in range(0, len(row_no[0])):
                                    if i > 1 and i < 16:
                                        product_barcode += row_no[0][i]
                                    elif i > 17 and i < len(row_no[0]):
                                        lot_name += row_no[0][i]
                                        continue
                        else:
                            product_barcode = row_no[0]
                        if row_no[1] == None or row_no[1] <= 0:
                            raise ValidationError(_('%s Quantity must be greater than zero.') % (row_no[0]))
                        # if row_no[2] == None or row_no[2] <= 0:
                        #     raise ValidationError(_('%s Price Must be Greater than Zero.') % (row_no[0]))
                        if self.product_details_option == 'from_product':
                            values.update({
                                'code': product_barcode,
                                'serial_no': lot_name,
                                'quantity': row_no[1]
                            })
                        elif self.product_details_option == 'from_xls':
                            values.update({
                                'code': product_barcode,
                                'serial_no': lot_name,
                                'quantity': row_no[1],
                                # 'uom':line[2],
                                # 'description':line[3],
                                'price': row_no[2],
                                # 'tax':line[5],
                            })
                        else:
                            values.update({
                                'code': product_barcode,
                                'quantity': row_no[1],
                                'serial_no': lot_name,
                            })
                        res = self.create_inv_line(values)
        view_id = self.env.ref('bi_import_all_orders_lines.message_wizard_popup')
        context = dict(self._context or {})
        dict_msg = str(counter) + " Records Imported Successfully."
        context['message'] = dict_msg
        return {
            'name': _('Success'),
            'type': 'ir.actions.act_window',
            'view_mode': 'form',
            'res_model': 'message.wizard',
            'views': [(view_id.id, 'form')],
            # pass the id
            'view_id': view_id.id,
            'target': 'new',
            'context': context,
        }

    def create_inv_line(self, values):
        analytic_account_search = False
        account_inv_brw = self.env['account.move'].browse(self._context.get('active_id'))
        product = values.get('code'
                             '')
        serial_no = self.env['stock.lot']
        product_obj_search = self.env['product.product']
        if self.product_details_option == 'from_product':
            if self.import_prod_option == 'barcode':
                if len(values['code']) > 13:
                    product_obj_search = self.env['product.product'].search([('barcode', '=', values['code'])])
                    serial_no = self.env['stock.lot'].search([('name', '=', values['serial_no'])])
                    if not serial_no:
                        raise ValidationError(_('The serial number for this is not found in the database.'))
                else:
                    product_barcodes = self.env['product.barcode'].search([('barcode', '=', values['code'])])
                    if len(product_barcodes) > 0:
                        product_obj_search = product_barcodes[0].product_id
            elif self.import_prod_option == 'code':
                raise ValidationError(_('Please set the import Option to Barcode.'))
                # product_obj_search=self.env['product.product'].search([('default_code', '=',values['product'])])
            else:
                raise ValidationError(_('Please set the import Option to Barcode.'))
                # product_obj_search=self.env['product.product'].search([('name', '=',values['product'])])
            if product_obj_search:
                product_id = product_obj_search
            else:
                raise ValidationError(_('%s Product was not found in the Database.') % values.get('code'))

            if self.import_analytic_account_tags == True:
                analytic_account_search = False
                if values.get('analytic_distribution'):
                    analytic_account_search = self.env['account.analytic.account'].search(
                        [('name', '=', values.get('analytic_account'))], limit=1)
                    if not analytic_account_search:
                        raise ValidationError(_('%s Analytic Account is not found".') % values.get('analytic_account'))

                tag_id_lst = []
                if values.get('analytic_tags'):
                    if ';' in values.get('analytic_tags'):
                        analytic_tags_names = values.get('analytic_tags').split(';')
                        for name in analytic_tags_names:
                            tag = self.env['account.analytic.plan'].search([('name', '=', name)])
                            if not tag:
                                raise ValidationError(_('"%s" Tag not in your system') % name)
                            tag_id_lst.append(tag.id)
                    elif ',' in values.get('analytic_tags'):
                        analytic_tags_names = values.get('analytic_tags').split(',')
                        for name in analytic_tags_names:
                            tag = self.env['account.analytic.plan'].search([('name', '=', name)])
                            if not tag:
                                raise ValidationError(_('"%s" Tag not in your system') % name)
                            tag_id_lst.append(tag.id)
                    else:
                        analytic_tags_names = values.get('analytic_tags').split(',')
                        tag = self.env['account.analytic.plan'].search([('name', '=', analytic_tags_names)])
                        if not tag:
                            raise ValidationError(_('"%s" Tag not in your system') % analytic_tags_names)
                        tag_id_lst.append(tag.id)

                    if not tag_id_lst:
                        raise ValidationError(_('%s Analytic Tag is not found".') % values.get('analytic_tags'))

            if account_inv_brw.move_type == "out_invoice" and account_inv_brw.state == 'draft':
                cust_account_id = product_id.property_account_income_id.id
                if cust_account_id:
                    account_id = cust_account_id
                else:
                    account_id = product_id.categ_id.property_account_income_categ_id.id
                    if analytic_account_search:
                        analytic_account_id = analytic_account_search.id
                    else:
                        analytic_account_id = False
                    if self.import_analytic_account_tags == True:
                        if product_id.nhcl_product_type == 'unbranded':
                            existing_move_line = account_inv_brw.invoice_line_ids.filtered(
                                lambda x: x.product_id == product_id)
                            existing_line = existing_move_line.prod_serial_no.filtered(lambda x:x.id == serial_no.id)
                        else:
                            existing_line = account_inv_brw.invoice_line_ids.filtered(
                                lambda x: x.product_id == product_id and x.prod_barcode == values.get('code'))
                        if existing_line:
                            raise ValidationError(_('%s The product already exists.') % values.get('code'))
                        existing_invoice_line = account_inv_brw.invoice_line_ids.filtered(
                            lambda x: x.product_id == product_id)
                        if existing_invoice_line:
                            existing_invoice_line.quantity += values.get('quantity')
                            if product_id.nhcl_product_type == 'unbranded':
                                existing_invoice_line.prod_serial_no = [(4, serial_no.id)]
                        else:
                            vals = {
                                'account_id': account_id,
                                'product_id': product_id.id,
                                'prod_serial_no': serial_no.ids,
                                'prod_barcode': values.get('code'),
                                'name': product_id.display_name,
                                'quantity': values.get('quantity'),
                                'product_uom_id': product_id.uom_id.id,
                                'price_unit': product_id.lst_price,
                                # 'analytic_distribution' : analytic_distribution,
                                # 'analytic_tag_ids' : [(6,0,tag_id_lst)],
                                # 'analytic_distribution': self.analytic_distribution,
                            }
                            account_inv_brw.write({'invoice_line_ids': ([(0, 0, vals)])})
                        return True
                    else:
                        if product_id.nhcl_product_type == 'unbranded':
                            existing_move_line = account_inv_brw.invoice_line_ids.filtered(
                                lambda x: x.product_id == product_id)
                            existing_line = existing_move_line.prod_serial_no.filtered(lambda x:x.id == serial_no.id)
                        else:
                            existing_line = account_inv_brw.invoice_line_ids.filtered(
                                lambda x: x.product_id == product_id and x.prod_barcode == values.get('code'))
                        if existing_line:
                            raise ValidationError(_('%s The product already exists.') % values.get('code'))
                        existing_invoice_line = account_inv_brw.invoice_line_ids.filtered(
                            lambda x: x.product_id == product_id)
                        if existing_invoice_line:
                            existing_invoice_line.quantity += values.get('quantity')
                            if product_id.nhcl_product_type == 'unbranded':
                                existing_invoice_line.prod_serial_no = [(4, serial_no.id)]
                        else:
                            vals = {
                                'account_id': account_id,
                                'product_id': product_id.id,
                                'prod_serial_no': serial_no.ids,
                                'prod_barcode': values.get('code'),
                                'name': product_id.display_name,
                                'quantity': values.get('quantity'),
                                'product_uom_id': product_id.uom_id.id,
                                'price_unit': product_id.lst_price,
                            }
                            account_inv_brw.write({'invoice_line_ids': ([(0, 0, vals)])})

            elif account_inv_brw.move_type == "in_invoice" and account_inv_brw.state == 'draft':
                vendor_account_id = product_id.property_account_expense_id.id
                if vendor_account_id:
                    account_id = vendor_account_id
                else:
                    account_id = product_id.categ_id.property_account_expense_categ_id.id
                if analytic_account_search:
                    analytic_account_id = analytic_account_search.id
                else:
                    analytic_account_id = False
                if self.import_analytic_account_tags == True:
                    if product_id.nhcl_product_type == 'unbranded':
                        existing_move_line = account_inv_brw.invoice_line_ids.filtered(
                            lambda x: x.product_id == product_id)
                        existing_line = existing_move_line.prod_serial_no.filtered(lambda x:x.id == serial_no.id)
                    else:
                        existing_line = account_inv_brw.invoice_line_ids.filtered(
                            lambda x: x.product_id == product_id and x.prod_barcode == values.get('code'))
                    if existing_line:
                        raise ValidationError(_('%s The product already exists.') % values.get('code'))
                    existing_invoice_line = account_inv_brw.invoice_line_ids.filtered(
                        lambda x: x.product_id == product_id)
                    if existing_invoice_line:
                        existing_invoice_line.quantity += values.get('quantity')
                        if product_id.nhcl_product_type == 'unbranded':
                            existing_invoice_line.prod_serial_no = [(4, serial_no.id)]
                    else:
                        vals = {
                            'account_id': account_id,
                            'product_id': product_id.id,
                            'prod_serial_no': serial_no.ids,
                            'prod_barcode': values.get('code'),
                            'name': product_id.display_name,
                            'quantity': values.get('quantity'),
                            'product_uom_id': product_id.uom_id.id,
                            'price_unit': product_id.lst_price,
                            # 'analytic_account_id' : analytic_account_id,
                            # 'analytic_tag_ids' : [(6,0,tag_id_lst)]
                        }
                        account_inv_brw.write({'invoice_line_ids': ([(0, 0, vals)])})
                    return True
                else:
                    if product_id.nhcl_product_type == 'unbranded':
                        existing_move_line = account_inv_brw.invoice_line_ids.filtered(
                            lambda x: x.product_id == product_id)
                        existing_line = existing_move_line.prod_serial_no.filtered(lambda x:x.id == serial_no.id)
                    else:
                        existing_line = account_inv_brw.invoice_line_ids.filtered(
                            lambda x: x.product_id == product_id and x.prod_barcode == values.get('code'))
                    if existing_line:
                        raise ValidationError(_('%s The product already exists.') % values.get('code'))
                    existing_invoice_line = account_inv_brw.invoice_line_ids.filtered(
                        lambda x: x.product_id == product_id)
                    if existing_invoice_line:
                        existing_invoice_line.quantity += values.get('quantity')
                        if product_id.nhcl_product_type == 'unbranded':
                            existing_invoice_line.prod_serial_no = [(4, serial_no.id)]
                    else:
                        vals = {
                            'account_id': account_id,
                            'product_id': product_id.id,
                            'prod_serial_no': serial_no.ids,
                            'prod_barcode': values.get('code'),
                            'name': product_id.display_name,
                            'quantity': values.get('quantity'),
                            'product_uom_id': product_id.uom_id.id,
                            'price_unit': product_id.lst_price,
                        }
                        account_inv_brw.write({'invoice_line_ids': ([(0, 0, vals)])})
                    return True

            elif account_inv_brw.state != 'draft':
                raise UserError(_('We cannot import data in validated or confirmed Invoice.'))

        else:
            if self.import_prod_option == 'barcode':
                if len(values['code']) > 13:
                    product_obj_search = self.env['product.product'].search([('barcode', '=', values['code'])])
                    serial_no = self.env['stock.lot'].search([('name', '=', values['serial_no'])])
                    if not serial_no:
                        raise ValidationError(_('The serial number for this is not found in the database.'))
                else:
                    product_barcodes = self.env['product.barcode'].search([('barcode', '=', values['code'])])
                    if len(product_barcodes) > 0:
                        product_obj_search = product_barcodes[0].product_id
            elif self.import_prod_option == 'code':
                raise ValidationError(_('Please set the import Option to Barcode.'))
                # product_obj_search=self.env['product.product'].search([('default_code', '=',values['product'])])
            else:
                raise ValidationError(_('Please set the import Option to Barcode.'))
                # product_obj_search=self.env['product.product'].search([('name', '=',values['product'])])

            # uom_obj_search=self.env['uom.uom'].search([('name','=',uom)])
            # if not uom_obj_search:
            # raise ValidationError(_('UOM "%s" is Not Available') % uom)

            if self.import_analytic_account_tags == True:
                analytic_account_search = False
                if values.get('analytic_account'):
                    analytic_account_search = self.env['account.analytic.account'].search(
                        [('name', '=', values.get('analytic_account'))], limit=1)
                    if not analytic_account_search:
                        raise ValidationError(_('%s Analytic Account is not found".') % values.get('analytic_account'))

                tag_id_lst = []
                if values.get('analytic_tags'):
                    if ';' in values.get('analytic_tags'):
                        analytic_tags_names = values.get('analytic_tags').split(';')
                        for name in analytic_tags_names:
                            tag = self.env['account.analytic.tag'].search([('name', '=', name)])
                            if not tag:
                                raise ValidationError(_('"%s" Tag not in your system') % name)
                            tag_id_lst.append(tag.id)
                    elif ',' in values.get('analytic_tags'):
                        analytic_tags_names = values.get('analytic_tags').split(',')
                        for name in analytic_tags_names:
                            tag = self.env['account.analytic.plan'].search([('name', '=', name)])
                            if not tag:
                                raise ValidationError(_('"%s" Tag not in your system') % name)
                            tag_id_lst.append(tag.id)
                    else:
                        analytic_tags_names = values.get('analytic_tags')
                        tag = self.env['account.analytic.plan'].search([('name', '=', analytic_tags_names)])
                        if not tag:
                            raise ValidationError(_('"%s" Tag not in your system') % analytic_tags_names)
                        tag_id_lst.append(tag.id)

                    if not tag_id_lst:
                        raise ValidationError(_('%s Analytic Tag is not found".') % values.get('analytic_tags'))

            if product_obj_search:
                product_id = product_obj_search
            else:
                if self.import_prod_option == 'name':
                    product_id = self.env['product.product'].create({'name': product, 'lst_price': values.get('price')})
                else:
                    raise ValidationError(
                        _('%s Product was not found in the Database.') % values.get(
                            'product'))

            if account_inv_brw.move_type == "out_invoice" and account_inv_brw.state == 'draft':
                tax_id_lst = []
                if values.get('tax'):
                    if ';' in values.get('tax'):
                        tax_names = values.get('tax').split(';')
                        for name in tax_names:
                            tax = self.env['account.tax'].search([('name', '=', name), ('type_tax_use', '=', 'sale')])
                            if not tax:
                                raise ValidationError(_('"%s" Tax is not in your system') % name)
                            tax_id_lst.append(tax.id)
                    elif ',' in values.get('tax'):
                        tax_names = values.get('tax').split(',')
                        for name in tax_names:
                            tax = self.env['account.tax'].search([('name', '=', name), ('type_tax_use', '=', 'sale')])
                            if not tax:
                                raise ValidationError(_('"%s" Tax is not in your system') % name)
                            tax_id_lst.append(tax.id)
                    else:
                        tax_names = values.get('tax').split(',')
                        tax = self.env['account.tax'].search([('name', '=', tax_names), ('type_tax_use', '=', 'sale')])
                        if not tax:
                            raise ValidationError(_('"%s" Tax is not in your system') % tax_names)
                        tax_id_lst.append(tax.id)

                cust_account_id = product_id.property_account_income_id.id
                if cust_account_id:
                    account_id = cust_account_id
                else:
                    account_id = product_id.categ_id.property_account_income_categ_id.id

                if analytic_account_search:
                    analytic_account_id = analytic_account_search.id
                else:
                    analytic_account_id = False
                if self.import_analytic_account_tags == True:
                    if product_id.nhcl_product_type == 'unbranded':
                        existing_move_line = account_inv_brw.invoice_line_ids.filtered(
                            lambda x: x.product_id == product_id)
                        existing_line = existing_move_line.prod_serial_no.filtered(lambda x:x.id == serial_no.id)
                    else:
                        existing_line = account_inv_brw.invoice_line_ids.filtered(
                            lambda x: x.product_id == product_id and x.prod_barcode == values.get('code'))
                    if existing_line:
                        raise ValidationError(_('%s The product already exists.') % values.get('code'))
                    existing_invoice_line = account_inv_brw.invoice_line_ids.filtered(
                        lambda x: x.product_id == product_id)
                    if existing_invoice_line:
                        existing_invoice_line.quantity += values.get('quantity')
                        if product_id.nhcl_product_type == 'unbranded':
                            existing_invoice_line.prod_serial_no = [(4, serial_no.id)]
                    else:
                        vals = {
                            'account_id': account_id,
                            'product_id': product_id.id,
                            'prod_serial_no': serial_no.ids,
                            'prod_barcode': values.get('code'),
                            'name': product_id.display_name,
                            'quantity': values.get('quantity'),
                            'product_uom_id': product_id.uom_id.id,
                            # 'price_unit': values.get('price'),
                            # 'analytic_account_id' : analytic_account_id,
                            # 'analytic_tag_ids' : [(6,0,tag_id_lst)],
                            'tax_ids': ([(6, 0, tax_id_lst)])
                        }
                        account_inv_brw.write({'invoice_line_ids': ([(0, 0, vals)])})
                    return True
                else:
                    if product_id.nhcl_product_type == 'unbranded':
                        existing_move_line = account_inv_brw.invoice_line_ids.filtered(
                            lambda x: x.product_id == product_id)
                        existing_line = existing_move_line.prod_serial_no.filtered(lambda x:x.id == serial_no.id)
                    else:
                        existing_line = account_inv_brw.invoice_line_ids.filtered(
                            lambda x: x.product_id == product_id and x.prod_barcode == values.get('code'))
                    if existing_line:
                        raise ValidationError(_('%s The product already exists.') % values.get('code'))
                    existing_invoice_line = account_inv_brw.invoice_line_ids.filtered(
                        lambda x: x.product_id == product_id)
                    if existing_invoice_line:
                        existing_invoice_line.quantity += values.get('quantity')
                        if product_id.nhcl_product_type == 'unbranded':
                            existing_invoice_line.prod_serial_no = [(4, serial_no.id)]
                    else:
                        vals = {
                            'account_id': account_id,
                            'product_id': product_id.id,
                            'prod_serial_no': serial_no.ids,
                            'prod_barcode': values.get('code'),
                            'name': product_id.display_name,
                            'quantity': values.get('quantity'),
                            'product_uom_id': product_id.uom_id.id,
                            # 'price_unit': values.get('price'),
                            'tax_ids': ([(6, 0, tax_id_lst)])
                        }
                        account_inv_brw.write({'invoice_line_ids': ([(0, 0, vals)])})
                    return True

            elif account_inv_brw.move_type == "in_invoice" and account_inv_brw.state == 'draft':
                tax_id_lst = []
                if values.get('tax'):
                    if ';' in values.get('tax'):
                        tax_names = values.get('tax').split(';')
                        for name in tax_names:
                            tax = self.env['account.tax'].search(
                                [('name', '=', name), ('type_tax_use', '=', 'purchase')])
                            if not tax:
                                raise ValidationError(_('"%s" Tax is not in your system') % name)
                            tax_id_lst.append(tax.id)
                    elif ',' in values.get('tax'):
                        tax_names = values.get('tax').split(',')
                        for name in tax_names:
                            tax = self.env['account.tax'].search(
                                [('name', '=', name), ('type_tax_use', '=', 'purchase')])
                            if not tax:
                                raise ValidationError(_('"%s" Tax is not in your system') % name)
                            tax_id_lst.append(tax.id)
                    else:
                        tax_names = values.get('tax').split(',')
                        tax = self.env['account.tax'].search(
                            [('name', '=', tax_names), ('type_tax_use', '=', 'purchase')])
                        if not tax:
                            raise ValidationError(_('"%s" Tax is not in your system') % tax_names)
                        tax_id_lst.append(tax.id)

                vendor_account_id = product_id.property_account_expense_id.id
                if vendor_account_id:
                    account_id = vendor_account_id
                else:
                    account_id = product_id.categ_id.property_account_expense_categ_id.id

                if analytic_account_search:
                    analytic_account_id = analytic_account_search.id
                else:
                    analytic_account_id = False
                if self.import_analytic_account_tags == True:
                    if product_id.nhcl_product_type == 'unbranded':
                        existing_move_line = account_inv_brw.invoice_line_ids.filtered(
                            lambda x: x.product_id == product_id)
                        existing_line = existing_move_line.prod_serial_no.filtered(lambda x:x.id == serial_no.id)
                    else:
                        existing_line = account_inv_brw.invoice_line_ids.filtered(
                            lambda x: x.product_id == product_id and x.prod_barcode == values.get('code'))
                    if existing_line:
                        raise ValidationError(_('%s The product already exists.') % values.get('code'))
                    existing_invoice_line = account_inv_brw.invoice_line_ids.filtered(
                        lambda x: x.product_id == product_id)
                    if existing_invoice_line:
                        existing_invoice_line.quantity += values.get('quantity')
                        if product_id.nhcl_product_type == 'unbranded':
                            existing_invoice_line.prod_serial_no = [(4, serial_no.id)]
                    else:
                        vals = {
                            'account_id': account_id,
                            'product_id': product_id.id,
                            'prod_serial_no': serial_no.ids,
                            'prod_barcode': values.get('code'),
                            'name': product_id.display_name,
                            'quantity': values.get('quantity'),
                            'product_uom_id': product_id.uom_id.id,
                            # 'price_unit': values.get('price'),
                            # 'analytic_distribution' : analytic_distribution,
                            # 'analytic_tag_ids' : [(6,0,tag_id_lst)],
                            'tax_ids': ([(6, 0, tax_id_lst)])
                        }
                        account_inv_brw.write({'invoice_line_ids': ([(0, 0, vals)])})
                    return True
                else:
                    if product_id.nhcl_product_type == 'unbranded':
                        existing_move_line = account_inv_brw.invoice_line_ids.filtered(
                            lambda x: x.product_id == product_id)
                        existing_line = existing_move_line.prod_serial_no.filtered(lambda x:x.id == serial_no.id)
                    else:
                        existing_line = account_inv_brw.invoice_line_ids.filtered(
                            lambda x: x.product_id == product_id and x.prod_barcode == values.get('code'))
                    if existing_line:
                        raise ValidationError(_('%s The product already exists.') % values.get('code'))
                    existing_invoice_line = account_inv_brw.invoice_line_ids.filtered(
                        lambda x: x.product_id == product_id)
                    if existing_invoice_line:
                        existing_invoice_line.quantity += values.get('quantity')
                        if product_id.nhcl_product_type == 'unbranded':
                            existing_invoice_line.prod_serial_no = [(4, serial_no.id)]
                    else:
                        vals = {
                            'account_id': account_id,
                            'product_id': product_id.id,
                            'prod_serial_no': serial_no.ids,
                            'prod_barcode': values.get('code'),
                            'name': product_id.display_name,
                            'quantity': values.get('quantity'),
                            'product_uom_id': product_id.uom_id.id,
                            # 'price_unit': values.get('price'),
                            'tax_ids': ([(6, 0, tax_id_lst)])
                        }

                        account_inv_brw.write({'invoice_line_ids': ([(0, 0, vals)])})
                    return True

            elif account_inv_brw.state != 'draft':
                raise UserError(_('We cannot import data in validated or confirmed Invoice.'))

