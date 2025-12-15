from odoo import models,fields,_,api
from odoo.exceptions import ValidationError, UserError
import logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)
import base64
import io

from dateutil.relativedelta import relativedelta
from datetime import datetime
import requests
from datetime import timedelta

class ProductMarginApproval(models.Model):
    _name = 'product.margin.approval'
    _inherit = ['mail.thread','mail.activity.mixin']
    _rec_name = 'name'

    name = fields.Char(string="Name", copy=False, index=True, default=lambda self: _('New'), tracking=True)
    lot_name = fields.Char(string="Enter Serial Number")
    lot_ref = fields.Char(string="Enter Barcode")
    product_margin_approval_ids = fields.One2many('product.margin.approval.line','product_margin_approval_id')
    state = fields.Selection([
        ('draft', 'Draft'),
        ('submitted', 'Submitted'),
        ('approved', 'Approved'), ('refused', 'Refused'),

    ], default='draft', string="Status")
    user_id = fields.Many2one('res.users', string="Requester", default=lambda self: self.env.user)
    manager_id = fields.Many2one('res.users', string="Manager", compute='compute_get_manager')
    activity_user_id = fields.Many2one('res.users', string='Activity User')
    is_manager = fields.Boolean(string="Approve", compute='compute_is_manager')
    file_data = fields.Binary("File", attachment=True)
    file_name = fields.Char("File Name")
    from_date = fields.Date("From Date", copy=False)
    to_date = fields.Date("To Date", copy=False)
    filter_type = fields.Selection([
        ('barcode', 'Barcode'), ('document', 'Document'),
        ('condition', 'Conditions'), ('update', 'Update Document'),], string="Filter Type", copy=False)
    pm_category_1_ids = fields.Many2many('product.attribute.value', 'pm_cat_1', string='Color', copy=False,
                                         domain=[('attribute_id.name', '=', 'Color')])
    pm_category_2_ids = fields.Many2many('product.attribute.value', 'pm_cat_2', string='Fit', copy=False,
                                         domain=[('attribute_id.name', '=', 'Fit')])
    pm_category_3_ids = fields.Many2many('product.attribute.value', 'pm_cat_3', string='Brand', copy=False,
                                         domain=[('attribute_id.name', '=', 'Brand')])
    pm_category_4_ids = fields.Many2many('product.attribute.value', 'pm_cat_4', string='Pattern', copy=False,
                                         domain=[('attribute_id.name', '=', 'Pattern')])
    pm_category_5_ids = fields.Many2many('product.attribute.value', 'pm_cat_5', string='Border Type', copy=False,
                                         domain=[('attribute_id.name', '=', 'Border Type')])
    pm_category_6_ids = fields.Many2many('product.attribute.value', 'pm_cat_6', string='Border Size', copy=False,
                                         domain=[('attribute_id.name', '=', 'Border Size')])
    pm_category_7_ids = fields.Many2many('product.attribute.value', 'pm_cat_7', string='Size', copy=False,
                                         domain=[('attribute_id.name', '=', 'Size')])
    pm_category_8_ids = fields.Many2many('product.attribute.value', 'pm_cat_8', string='Category 8', copy=False)

    pm_description_1_ids = fields.Many2many('product.aging.line', string='Product Ageing', copy=False)
    pm_description_2_ids = fields.Many2many('product.attribute.value', 'pm_des_2', string='Range', copy=False,
                                            domain=[('attribute_id.name', '=', 'Range')])
    pm_description_3_ids = fields.Many2many('product.attribute.value', 'pm_des_3', string='Collection', copy=False,
                                            domain=[('attribute_id.name', '=', 'Collection')])
    pm_description_4_ids = fields.Many2many('product.attribute.value', 'pm_des_4', string='Fabric', copy=False,
                                            domain=[('attribute_id.name', '=', 'Fabric')])
    pm_description_5_ids = fields.Many2many('product.attribute.value', 'pm_des_5', string='Exclusive', copy=False,
                                            domain=[('attribute_id.name', '=', 'Exclusive')])
    pm_description_6_ids = fields.Many2many('product.attribute.value', 'pm_des_6', string='Print', copy=False,
                                            domain=[('attribute_id.name', '=', 'Print')])
    pm_description_7_ids = fields.Many2many('product.attribute.value', 'pm_des_7', string='Days Ageing', copy=False,
                                            domain=[('attribute_id.name', '=', 'Days Ageing')])
    pm_description_8_ids = fields.Many2many('product.attribute.value', 'pm_des_8', string='Description 8', copy=False)
    pm_day_ageing_slab = fields.Selection([('1', '0-30'), ('2', '30-60'),
                                           ('3', '60-90'), ('4', '90-120'),
                                           ('5', '120-150'), ('6', '150-180'),
                                           ('7', '180-210'), ('8', '210-240'),
                                           ('9', '240-270'), ('10', '270-300'),
                                           ('11', '300-330'), ('12', '330-360')
                                           ], string="Days Ageing Slab")
    pm_receipt = fields.Many2one('stock.picking', string="G.R.C", copy=False,
                                 domain=[('picking_type_id.name', '=', 'Receipts'),
                                         ('company_id.nhcl_company_bool', '=', True), ('state', '=', 'done')])
    nhcl_integration_status = fields.Boolean(string="Integration Status")
    margin_id = fields.Many2one('product.margin.approval', string="Margin Document", copy=False)

    @api.constrains('from_date', 'to_date')
    def _check_date_difference(self):
        for record in self:
            if record.from_date and record.to_date:
                if record.to_date < record.from_date:
                    raise ValidationError("To Date cannot be earlier than From Date.")


    def action_import_excel(self):
        return {
            'type': 'ir.actions.act_window',
            'name': 'Import Excel',
            'res_model': 'product.margin.approval.import.wizard',
            'view_mode': 'form',
            'target': 'new',
            'context': {'default_product_margin_approval_id': self.id},
        }

    @api.model
    def create(self, vals_list):
        def get_unique_sequence(sequence_code):
            while True:
                seq_number = self.env['ir.sequence'].next_by_code(sequence_code) or 'New'
                if not self.env['product.margin.approval'].search([('name', '=', seq_number)]):
                    return seq_number
                else:
                    # If the sequence number is already used, log a warning and regenerate
                    logger.warning(f"Sequence number {seq_number} already exists. Regenerating...")

        if vals_list.get('name', 'New') == 'New':
            vals_list['name'] = get_unique_sequence('product.margin.approval')
        res = super(ProductMarginApproval, self).create(vals_list)
        return res

    @api.depends('user_id')
    def compute_is_manager(self):
        for record in self:
            if record.manager_id == self.env.user:
                record.is_manager = True
            else:
                record.is_manager = False


    @api.depends('user_id')
    def compute_get_manager(self):
        self.manager_id = self.user_id.employee_parent_id.user_id

    @api.constrains('product_margin_approval_ids')
    def _check_duplicate_lot_ids(self):
        for rec in self:
            seen = set()
            for line in rec.product_margin_approval_ids:
                if line.lot_id.id in seen:
                    raise ValidationError(f"The serial number '{line.lot_id.name}' is duplicated.")
                seen.add(line.lot_id.id)

    @api.onchange('lot_name', 'lot_ref')
    def _onchange_lot_fields(self):
        Lot = self.env['stock.lot']
        if self.lot_name:
            lot = Lot.search([
                ('name', '=', self.lot_name),
                ('company_id.nhcl_company_bool', '=', True)
            ], limit=1)
            if lot:
                self.product_margin_approval_ids = [(0, 0, {
                    'lot_id': lot.id,
                    'product_id': lot.product_id.id,
                    'actual_cost_price': lot.cost_price,
                    'actual_margin': lot.nhcl_margin_lot,
                    'actual_rsp_price': lot.rs_price,
                    'from_company': lot.company_id.id,
                })]
                self.lot_name = False
        elif self.lot_ref:
            matching_lots = Lot.search([
                ('ref', '=', self.lot_ref),
                ('company_id.nhcl_company_bool', '=', True)
            ])
            if not matching_lots:
                raise ValidationError(f"No serials found for barcode {self.lot_ref}.")
            new_lines = [(0, 0, {
                'lot_id': lot.id,
                'actual_cost_price': lot.cost_price,
                'actual_margin': lot.nhcl_margin_lot,
                'actual_rsp_price': lot.rs_price,
                'from_company': lot.company_id.id,
            }) for lot in matching_lots]
            self.product_margin_approval_ids = new_lines
            self.lot_ref = False

    def action_submit(self):
        self.ensure_one()
        if self.user_id != self.env.user:
            raise UserError('You can only submit your own requests.')
        for rec in self.product_margin_approval_ids:
            if rec.new_margin == 0:
                raise UserError('New Margin Greater Than 0')
        self.state = 'submitted'
        self.create_activity_for_manager()

    def create_activity_for_manager(self):
        activity_type = self.env.ref('mail.mail_activity_data_todo')
        self.env['mail.activity'].create({
            'res_model_id': self.env.ref('cmr_customizations.model_product_margin_approval').id,
            'res_id': self.id,
            'activity_type_id': activity_type.id,
            'user_id': self.manager_id.id,
            'summary': 'Review Approval Request: ' + self.product_margin_approval_ids[0].lot_id.name,
            'date_deadline': fields.Datetime.now(),
        })

    def action_approve(self):
        self.ensure_one()
        if self.state != 'submitted':
            raise UserError('Only submitted requests can be approved.')
        for rec in self.product_margin_approval_ids:
            rec.lot_id.shortage = rec.shortage
            rec.lot_id.nhcl_updated_margin_lot = rec.new_margin
            rec.lot_id.rs_price = rec.new_rsp_price
            branch_lot = self.env['stock.lot'].sudo().search([('name','=',rec.lot_id.name),('product_qty','=',1)])
            if rec.new_rsp_price != 0:
                branch_lot.rs_price = rec.new_rsp_price
            else:
                branch_lot.rs_price = rec.enter_new_rsp
        self.state = 'approved'
        self._mark_activity_done()
        self.ho_store_update_rsp()


    def _mark_activity_done(self):
        activities = self.env['mail.activity'].search([
            ('res_model', '=', 'product.margin.approval'),
            ('res_id', '=', self.id),
            ('user_id', '=', self.manager_id.id),
            ('state', '=', 'draft')
        ])
        for activity in activities:
            activity.action_done()

    def action_refused(self):
        if self.state == 'submitted':
            self.state = 'refused'

    def action_reset_to_draft(self):
        if self.state == 'refused':
            self.state = 'draft'

    @api.model
    def create_manager_activity(self):
        today = fields.Date.today()
        reminder_date = today + timedelta(days=3)
        records = self.search([('to_date', '=', reminder_date), ('manager_id', '!=', False)])
        for record in records:
            existing_activity = self.env['mail.activity'].search([
                ('res_model', '=', 'product.margin.approval'),
                ('res_id', '=', record.id),
                ('user_id', '=', record.manager_id.id),
                ('activity_type_id', '=', self.env.ref('mail.mail_activity_data_todo').id)
            ], limit=1)
            if not existing_activity:
                self.env['mail.activity'].create({
                    'res_model_id': self.env['ir.model']._get('product.margin.approval').id,
                    'res_id': record.id,
                    'activity_type_id': self.env.ref('mail.mail_activity_data_todo').id,
                    'summary': 'Margin Exping',
                    'note': 'Please review this Product Margin before the deadline.',
                    'user_id': record.manager_id.id,
                    'date_deadline': record.to_date,
                })

    def get_filtered_serials(self):
        matching_lots = []
        if self.filter_type == 'condition':
            self.product_margin_approval_ids.unlink()
            stock_lot_obj = self.env['stock.lot']
            # Fetch new serials based on the current rule
            matching_lots = stock_lot_obj.search_by_product_margin(self)
            for lot in matching_lots:
                self.product_margin_approval_ids = [(0, 0, {
                    'lot_id': lot.id,
                    'product_id': lot.lot_id.product_id.id,
                    'actual_cost_price': lot.cost_price,
                    'actual_margin': lot.nhcl_margin_lot,
                    'actual_rsp_price': lot.rs_price,
                    'from_company': lot.company_id.id,
                })]
        elif self.filter_type == 'document':
            self.product_margin_approval_ids.unlink()
            if self.pm_receipt:
                if self.pm_receipt and self.pm_receipt.is_landed_cost == 'yes':
                    landed_cost = self.env['stock.landed.cost'].search([('picking_ids','=',self.pm_receipt.id)])
                    if landed_cost.state == 'draft':
                        raise ValidationError("Landed not added to this G.R.C")
                for lots in self.pm_receipt.move_line_ids_without_package:
                    self.product_margin_approval_ids = [(0, 0, {
                        'lot_id': lots.lot_id.id,
                        'product_id': lots.lot_id.product_id.id,
                        'actual_cost_price': lots.lot_id.cost_price,
                        'actual_margin': lots.lot_id.nhcl_margin_lot,
                        'actual_rsp_price': lots.lot_id.rs_price,
                        'from_company': lots.lot_id.company_id.id,
                    })]

        elif self.filter_type == 'update':
            self.product_margin_approval_ids.unlink()
            if self.margin_id:
                for lots in self.margin_id.product_margin_approval_ids:
                    self.product_margin_approval_ids = [(0, 0, {
                        'lot_id': lots.lot_id.id,
                        'product_id': lots.lot_id.product_id.id,
                        'actual_cost_price': lots.lot_id.cost_price,
                        'actual_margin': lots.lot_id.nhcl_updated_margin_lot,
                        'enter_new_margin': lots.lot_id.nhcl_margin_lot,
                        'new_margin': lots.lot_id.nhcl_margin_lot,
                        'actual_rsp_price': lots.lot_id.rs_price,
                        'new_rsp_price': lots.lot_id.rs_price,
                        'from_company': lots.lot_id.company_id.id,
                    })]

    def ho_store_update_rsp(self):
        # pass
        for rec in self.product_margin_approval_ids:
            main_company = self.env['res.company'].search([('nhcl_company_bool', '=', True)])
            branch_lot = self.env['stock.lot'].sudo().search([('name', '=', rec.lot_id.name), ('product_qty', '>=', 1), ('company_id', '!=', main_company.id)])
            ho_store_id = self.env['nhcl.ho.store.master'].search(
                [('nhcl_store_type', '!=', 'ho'), ('nhcl_active', '=', True),
                 ('nhcl_store_name', '=', branch_lot.company_id.name)
                 ]
            )
            try:
                store_ip = ho_store_id.nhcl_terminal_ip
                store_port = ho_store_id.nhcl_port_no
                store_api_key = ho_store_id.nhcl_api_key
                headers_source = {'api-key': f"{store_api_key}", 'Content-Type': 'application/json'}
                serial_number_data_url = f"http://{store_ip}:{store_port}/api/stock.lot/search"
                lot_domain = [('name', '=', branch_lot.name)]
                lot_data_url = f"{serial_number_data_url}?domain={lot_domain}"
                transporter_data = requests.get(lot_data_url, headers=headers_source).json()
                lot_ids = transporter_data.get("data")
                if lot_ids:
                    lot_id = lot_ids[0]["id"]
                    if rec.new_rsp_price != 0:
                        lot_data = {
                            'rs_price': rec.new_rsp_price,

                        }
                    else:
                        lot_data = {
                            'rs_price': rec.enter_new_rsp,

                        }
                    store_url_data = f"http://{store_ip}:{store_port}/api/stock.lot/{lot_id}"
                    response = requests.put(store_url_data, headers=headers_source, json=lot_data)
                    response.raise_for_status()
                    response_json = response.json()
                    message = response_json.get("message", "No message provided")
                    response_code = response_json.get("responseCode", "No response code provided")
                    if not response_json.get("success", True):
                        ho_store_id.create_cmr_transaction_server_replication_log('success', message)
                        ho_store_id.create_cmr_transaction_replication_log(response_json['object_name'], self.id, 200,
                                                                           'add', 'failure', message)
                    else:
                        self.nhcl_integration_status = True
                        ho_store_id.create_cmr_transaction_server_replication_log('success', message)
                        ho_store_id.create_cmr_transaction_replication_log(response_json['object_name'], self.id, 200,
                                                                           'add', 'success',
                                                                           f"Successfully Updated RSP Price: {message}")

                else:
                    ho_store_id.create_cmr_transaction_replication_log('stock.lot', self.id, 200,
                                                                       'add', 'failure',
                                                                       f"{self.name, rec.lot_id.name}Lot Not found")

            except requests.exceptions.RequestException as e:
                ho_store_id.create_cmr_transaction_server_replication_log('failure', e)


class ProductMarginApprovalLine(models.Model):
    _name = 'product.margin.approval.line'

    product_margin_approval_id = fields.Many2one('product.margin.approval', string="Margin approval")
    lot_id = fields.Many2one('stock.lot', string="Serial Numbers")
    from_company = fields.Many2one('res.company', string="From Company", copy=False)
    actual_cost_price = fields.Float(string='Actual CP', copy=False,)
    actual_margin = fields.Integer(string='Actual Margin', copy=False,)
    actual_rsp_price = fields.Float(string='Actual RSP', copy=False,)
    new_cost_price = fields.Float(string='New CP', copy=False,)
    new_margin = fields.Float(string='New Margin', copy=False,)
    new_rsp_price = fields.Float(string='New RSP', copy=False,)

    enter_new_margin = fields.Float(string='Enter Margin', copy=False, )
    enter_new_rsp = fields.Float(string='Enter Rsp', copy=False, )
    rsp_before_roundup = fields.Float(string='rsp round', copy=False)
    margin_before_roundup = fields.Float(string='margin roundup', copy=False)

    shortage = fields.Float(string='Excess/Shortage', copy=False,)
    product_id = fields.Many2one('product.product', string="Product", copy=False)

    @api.constrains('enter_new_margin', 'enter_new_rsp')
    def _check_exclusive_fields(self):
        for record in self:
            if record.enter_new_margin and record.enter_new_rsp:
                raise ValidationError("You cannot enter both Margin and RSP. Delete the line to enter a new value.")

    @api.onchange('enter_new_margin')
    @api.constrains('enter_new_margin')
    def _compute_new_rsp(self):
        for rec in self:
            if rec.enter_new_margin != 0:
                rec.rsp_before_roundup = rec.actual_cost_price + ((rec.enter_new_margin / 100) * rec.actual_cost_price)
                rec.shortage = rec.enter_new_margin - rec.actual_margin
                rec.new_margin = rec.enter_new_margin
                # Extract last two digits of the price
                last_two_digits = int(rec.rsp_before_roundup) % 100
                # Adjust last two digits based on condition
                if last_two_digits <= 49:
                    rec.new_rsp_price = (int(rec.rsp_before_roundup) // 100) * 100 + 49
                else:
                    rec.new_rsp_price = (int(rec.rsp_before_roundup) // 100) * 100 + 99
            else:
                rec.new_rsp_price = 0
                rec.shortage = rec.new_margin - rec.actual_margin


    @api.onchange('enter_new_rsp')
    def _inverse_new_margin(self):
        for rec in self:
            if rec.enter_new_rsp != 0:
                rec.new_margin = ((rec.enter_new_rsp - rec.actual_cost_price) / rec.actual_cost_price) * 100
                rec.shortage = rec.new_margin - rec.actual_margin
                rec.new_rsp_price = rec.enter_new_rsp
            else:
                rec.new_margin = 0
                rec.shortage = rec.new_margin - rec.actual_margin


class ProductAging(models.Model):
    _name = 'product.aging'

    product_aging_ids = fields.One2many(
        'product.aging.line', 'product_aging_id',string="Product Aging",
        ondelete='cascade')
    name = fields.Char(string="Prefix Name")
    from_date = fields.Date(string="From Date")
    to_date = fields.Date(string="To Date")

    @api.constrains('from_date', 'to_date')
    def _check_duplicate_aging_lines(self):
        for record in self:
            overlapping_records = self.env['product.aging'].search([
                ('id', '!=', record.id),
                ('from_date', '<=', record.to_date),
                ('to_date', '>=', record.from_date)
            ])
            if overlapping_records:
                raise ValidationError(
                    "The date range {} - {} overlaps with another existing record.".format(
                        record.from_date, record.to_date
                    )
                )

    def create_monthly_lines(self):
        from_date = self.from_date
        to_date = self.to_date

        current_date = datetime.strptime(str(from_date), '%Y-%m-%d')
        end_date = datetime.strptime(str(to_date), '%Y-%m-%d')

        for i in range(12):
            month_start = current_date
            month_end = current_date + relativedelta(day=31)

            # Create a line for each month
            self.env['product.aging.line'].create({
                'product_aging_id': self.id,
                'name': f"{self.name}{i + 1}",
                'from_date': month_start.date(),
                'to_date': month_end.date(),
            })

            # Move to the next month
            current_date = current_date + relativedelta(months=1)

    @api.model
    def create(self, vals):
        if 'name' in vals:
            name_lower = vals['name'].strip().lower()
            existing_names = self.env['product.aging'].search([]).mapped('name')
            existing_names_lower = [n.strip().lower() for n in existing_names]
            if name_lower in existing_names_lower:
                raise ValidationError(f"Aging Prefix '{vals['name']}' already exists (case-insensitive).")
        if 'from_date' in vals:
            from_date = datetime.strptime(vals['from_date'], "%Y-%m-%d").date()
            if from_date.day != 1:
                raise ValidationError("From Date must be the first day of the month.")
        record = super(ProductAging, self).create(vals)
        if record.from_date > record.to_date:
            raise ValidationError("From Date cannot be later than To Date.")
        expected_to_date = record.from_date + relativedelta(years=1, days=-1)
        if record.to_date != expected_to_date:
            raise ValidationError("To Date must be exactly 1 year after From Date.")
        record.create_monthly_lines()
        return record

    def write(self, vals):
        fields_to_check = {'from_date', 'to_date', 'name'}
        if any(field in vals for field in fields_to_check):
            for record in self:
                # Step 1: Delete existing aging lines
                existing_lines = self.env['product.aging.line'].search([('product_aging_id', '=', record.id)])
                if existing_lines:
                    existing_lines.unlink()
                from_date = vals.get('from_date', record.from_date)
                to_date = vals.get('to_date', record.to_date)
                name = vals.get('name', record.name)
                if isinstance(from_date, str):
                    from_date = datetime.strptime(from_date, "%Y-%m-%d").date()
                    if from_date.day != 1:
                        raise ValidationError("From Date must be the first day of the month.")
                if isinstance(to_date, str):
                    to_date = datetime.strptime(to_date, "%Y-%m-%d").date()
                if from_date > to_date:
                    raise ValidationError("From Date cannot be later than To Date.")
                expected_to_date = from_date + relativedelta(years=1, days=-1)
                if to_date != expected_to_date:
                    raise ValidationError("To Date must be exactly 1 year after From Date.")
                res = super(ProductAging, self).write(vals)
                record.create_monthly_lines()
                return res
        return super(ProductAging, self).write(vals)

    def _update_stock_lot_aging(self):
        self.env['stock.lot'].search([])._get_product_aging()


class ProductAgingLine(models.Model):
    _name = 'product.aging.line'
    _inherit = ['mail.thread', 'mail.activity.mixin']

    product_aging_id = fields.Many2one('product.aging',string="Aging", ondelete='cascade')
    name = fields.Char(string="Name")
    from_date = fields.Date(string="From Date")
    to_date = fields.Date(string="To Date")
    nhcl_id = fields.Integer(string="Nhcl Id", copy=False, index=True, tracking=True)

    @api.model
    def create(self, vals):
        self.env.cr.execute("SELECT MAX(nhcl_id) FROM product_aging_line")
        max_nhcl_id = self.env.cr.fetchone()[0] or 0
        vals['nhcl_id'] = max_nhcl_id + 1
        return super(ProductAgingLine, self).create(vals)

class ProductMarginApprovalImportWizard(models.TransientModel):
    _name = 'product.margin.approval.import.wizard'

    product_margin_approval_id = fields.Many2one('product.margin.approval', string="Margin Approval")
    file_data = fields.Binary("Excel File", required=True)
    file_name = fields.Char("File Name")

    def is_valid_file_extension(self, file_name):
        """
        Validate if the file extension is one of the supported formats.
        """
        valid_extensions = ['.xls', '.xlsx', '.ods']
        return any(file_name.lower().endswith(ext) for ext in valid_extensions)

    def action_import(self):
        # Ensure that file_name is not None, False, or a boolean value
        if not self.file_name or isinstance(self.file_name, bool):
            raise UserError("File name is missing or invalid.")

        # Optional: Set a default file name if it's still False or missing
        if not self.file_name:
            self.file_name = self._context.get('default_file_name', 'uploaded_file')  # Default name if none provided

        try:
            file_content = base64.b64decode(self.file_data)

            if self.file_name.lower().endswith('.xls'):  # Handle .xls with xlrd
                import xlrd
                workbook = xlrd.open_workbook(file_contents=file_content)
                sheet = workbook.sheet_by_index(0)

                for row in range(1, sheet.nrows):  # Start from row 1 to skip header
                    lot_name = sheet.cell_value(row, 0)
                    if lot_name:
                        self.create_margin_approval_line(lot_name)

            elif self.file_name.lower().endswith('.xlsx'):  # Handle .xlsx with openpyxl
                import openpyxl
                workbook = openpyxl.load_workbook(io.BytesIO(file_content))
                sheet = workbook.active  # Get the active sheet

                for row in range(1, sheet.max_row):  # Start from row 1 to skip header
                    lot_name = sheet.cell(row=row + 1, column=1).value  # Openpyxl uses 1-based indexing
                    if lot_name:
                        self.create_margin_approval_line(lot_name)

            elif self.file_name.lower().endswith('.ods'):  # Handle .ods with odfpy
                from odf.opendocument import load
                from odf.opendocument import Table, TableRow, TableCell

                document = load(io.BytesIO(file_content))
                table = document.getElementsByType(Table)[0]  # Assuming the first table is the one to process
                for row in table.getElementsByType(TableRow)[1:]:  # Skip header row
                    cells = row.getElementsByType(TableCell)
                    lot_name = cells[0].firstChild.data if cells else ''
                    if lot_name:
                        self.create_margin_approval_line(lot_name)

            else:
                raise UserError("Unsupported file format. Only XLS, XLSX, and ODS are supported.")

            return {'type': 'ir.actions.act_window_close'}

        except Exception as e:
            raise UserError(f"Error importing file: {str(e)}")


    def create_margin_approval_line(self, lot_name):
        """
        Helper function to create a margin approval line from a given lot name.
        """
        lot = self.env['stock.lot'].search([('name', '=', lot_name)], limit=1)
        if lot:
            self.env['product.margin.approval.line'].create({
                'product_margin_approval_id': self.product_margin_approval_id.id,
                'lot_id': lot.id,
                'actual_cost_price': lot.cost_price,
                'actual_margin': lot.nhcl_margin_lot,
                'actual_rsp_price': lot.rs_price,
            })
        else:
            raise UserError(f"Serial number {lot_name} not found.")