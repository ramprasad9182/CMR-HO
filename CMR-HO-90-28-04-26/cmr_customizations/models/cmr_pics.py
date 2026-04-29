from odoo import models,fields, _, api
from odoo.exceptions import ValidationError



class GRCMaster(models.Model):
    _name = 'grc.master'
    _description = "GRC Master"

    name = fields.Char(string="Name")

    @api.constrains('name')
    def _check_unique_name(self):
        for rec in self:
            if self.search_count([('name', '=', rec.name), ('id', '!=', rec.id)]):
                raise ValidationError(f"This GRC is {rec.name} Already Used.")


class ExcelQtySplit(models.Model):
    _name = 'excel.qty.split'
    _description = 'Excel Qty Split'

    name = fields.Char(default="Qty Split")
    file = fields.Binary(string="Upload Excel", required=True)
    filename = fields.Char()
    folder_path = fields.Char(string="Folder Path")
    max_qty = fields.Integer(string="Max Qty per File", required=True)
    line_ids = fields.One2many('excel.qty.split.line','parent_id',string="Split Files")


    def action_process_from_path(self):
        import os
        import base64
        import re

        if not self.folder_path:
            raise ValidationError("Please provide folder path")

        if not os.path.exists(self.folder_path):
            raise ValidationError("Invalid folder path")

        # ⚡ Fast delete (no ORM overhead)
        self.env.cr.execute(
            "DELETE FROM excel_qty_split_line WHERE parent_id = %s",
            [self.id]
        )

        def extract_number(filename):
            match = re.search(r'(\d+)', filename)
            return int(match.group(1)) if match else 0

        files = sorted(
            [f for f in os.listdir(self.folder_path) if f.endswith('.xlsx')],
            key=extract_number
        )

        if not files:
            raise ValidationError("No Excel files found in folder")

        batch_vals = []
        sequence = 1
        BATCH_SIZE = 50  # 🔥 tune based on system (50–200 ideal)

        for file_name in files:
            file_path = os.path.join(self.folder_path, file_name)

            # ⚡ Direct binary read (fastest possible)
            with open(file_path, 'rb') as f:
                file_data = base64.b64encode(f.read())

            clean_name = re.sub(r'[^A-Za-z0-9_.-]', '_', file_name)

            batch_vals.append({
                'parent_id': self.id,
                'sequence': sequence,
                'output_file': file_data,
                'output_filename': clean_name,
            })

            sequence += 1

            # ⚡ Batch insert
            if len(batch_vals) >= BATCH_SIZE:
                self.env['excel.qty.split.line'].create(batch_vals)
                self.env.cr.commit()  # 🔥 prevent timeout
                batch_vals = []

        # remaining records
        if batch_vals:
            self.env['excel.qty.split.line'].create(batch_vals)
            self.env.cr.commit()

        return True

    def action_create_po_from_lines(self):
        import base64
        from io import BytesIO
        from openpyxl import load_workbook

        if not self.line_ids:
            raise ValidationError("No split lines found.")

        # ✅ preload products (VERY IMPORTANT)
        products = self.env['product.product'].search([])
        barcode_map = {p.barcode: p for p in products if p.barcode}
        default_code_map = {p.default_code: p for p in products if p.default_code}
        name_map = {p.name.strip().lower(): p for p in products if p.name}

        partner = self.env.ref('base.partner_admin')
        picking = self.env['stock.picking.type'].search([
            ('stock_picking_type', '=', 'receipt'),
            ('company_id', '=', self.env.company.id)
        ], limit=1)

        lines = self.env['excel.qty.split.line'].search([
            ('parent_id', '=', self.id),
            ('is_used', '=', False),
            ('output_file', '!=', False),
        ], limit=5)

        created_pos = []

        for line in lines:
            try:
                file_content = base64.b64decode(line.output_file)
                wb = load_workbook(BytesIO(file_content), read_only=True, data_only=True)
                ws = wb.active
            except Exception:
                raise ValidationError(f"Failed to read file for line {line.sequence}")

            # ✅ Read header
            headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
            header_index = {h: i for i, h in enumerate(headers)}

            required_columns = [
                'Odoo_Product', 'Item_Code', 'Barcode',
                'RSP', 'Cost_Rate', 'Closing_Qty'
            ]

            missing = [c for c in required_columns if c not in header_index]
            if missing:
                raise ValidationError(f"Missing columns in line {line.sequence}: {', '.join(missing)}")

            # ✅ Create PO
            po = self.env['purchase.order'].create({
                'partner_id': partner.id,
                'date_order': fields.Datetime.now(),
                'nhcl_po_type': 'data_import',
                'company_id': self.env.company.id,
                'picking_type_id': picking.id
            })

            line_buffer = []

            # ✅ FAST streaming rows
            for row in ws.iter_rows(min_row=2, values_only=True):

                qty = float(str(row[header_index['Closing_Qty']] or 0).replace(',', ''))
                if qty <= 0:
                    continue

                barcode = str(row[header_index['Barcode']] or '').strip()
                item_code = str(row[header_index['Item_Code']] or '').strip()
                odoo_product_name = str(row[header_index['Odoo_Product']] or '').strip().lower()

                product = False

                # ✅ ZERO DB CALLS
                if barcode and barcode != '-':
                    product = barcode_map.get(barcode)

                if not product and item_code:
                    product = default_code_map.get(item_code)

                if not product and odoo_product_name:
                    product = name_map.get(odoo_product_name)

                if not product:
                    continue  # skip instead of crash

                cp = float(row[header_index['Cost_Rate']] or 0)
                rsp = float(row[header_index['RSP']] or 0)

                if cp <= 0 or rsp <= 0:
                    continue

                move_brand_barcode = barcode if barcode and barcode != '-' else item_code

                line_buffer.append({
                    'order_id': po.id,
                    'product_id': product.id,
                    'name': product.name,
                    'product_qty': qty,
                    'price_unit': cp,
                    'date_planned': fields.Datetime.now(),
                    'move_brand_barcode': move_brand_barcode,
                    'move_cp': cp,
                    'move_mrp': 10000,
                    'move_rsp': rsp,
                    'type_product': 'brand',
                    'company_id': self.env.company.id,
                })

            # ✅ bulk create
            if line_buffer:
                self.env['purchase.order.line'].create(line_buffer)

            created_pos.append(po)

            line.write({
                'is_used': True,
                'po_id': po.id,
            })

        # ✅ confirm AFTER all creation
        for po in created_pos:
            po.button_confirm()

        return {
            'type': 'ir.actions.client',
            'tag': 'display_notification',
            'params': {
                'title': 'Success',
                'message': f'{len(created_pos)} Purchase Orders created & confirmed',
                'type': 'success',
                'sticky': False,
            }
        }

    def action_validate_receipts(self):
        lines = self.env['excel.qty.split.line'].search([
            ('parent_id', '=', self.id),('po_id', '!=', False)], limit=3)
        # Collect all pickings in one go
        pickings = lines.mapped('po_id.picking_ids').filtered(lambda p: p.state != 'cancel')
        if not pickings:
            return
        pickings.write({'is_landed_cost': 'no'})
        # Filter only non-done pickings
        pending_pickings = pickings.filtered(lambda p: p.state != 'done')
        # Validate in batch loop (minimal overhead)
        for picking in pending_pickings:
            picking.button_validate()



class ExcelQtySplitLine(models.Model):
    _name = 'excel.qty.split.line'
    _description = 'Excel Qty Split Line'

    parent_id = fields.Many2one('excel.qty.split', ondelete='cascade')

    sequence = fields.Integer()
    is_used = fields.Boolean()

    output_file = fields.Binary(string="Excel File")
    output_filename = fields.Char()
    po_id = fields.Many2one('purchase.order', string="Purchase Order")